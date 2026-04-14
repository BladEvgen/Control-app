import datetime
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from monitoring_app.cache_conf import Cache
from monitoring_app.models import (
    APIKey,
    ChildDepartment,
    ClassLocation,
    LessonAttendance,
    Position,
    Staff,
    StaffAttendance,
)
from monitoring_app.services import building_attendance_report
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

ABILAI_ADDRESS = "Проспект Абылай хана, 51/53"
TOREKULOVA_ADDRESS = "Улица Торекулова, 71"
FACEID_FILL_RGB = "DC2626"
GPS_SPOOF_FILL_RGB = "1D4ED8"
FACEID_GPS_FILL_RGB = "7E22CE"


def _aware_dt(day: datetime.date, hour: int, minute: int = 0) -> datetime.datetime:
    naive = datetime.datetime(day.year, day.month, day.day, hour, minute)
    return timezone.make_aware(naive, timezone.get_current_timezone())


def _workbook_snapshot(raw_bytes: bytes) -> dict[str, list[tuple[object, ...]]]:
    workbook = load_workbook(BytesIO(raw_bytes), data_only=True)
    snapshot: dict[str, list[tuple[object, ...]]] = {}
    for sheet in workbook.worksheets:
        rows: list[tuple[object, ...]] = []
        for row in sheet.iter_rows(values_only=True):
            normalized = tuple("" if value is None else value for value in row)
            rows.append(normalized)
        snapshot[sheet.title] = rows
    return snapshot


def _find_attendance_cell(
    raw_bytes: bytes,
    *,
    staff_fio: str,
    target_day: datetime.date,
):
    workbook = load_workbook(BytesIO(raw_bytes))
    worksheet = workbook["Отчет посещаемости"]
    target_header = target_day.strftime("%d.%m.%Y")

    header_row_idx = None
    target_col_idx = None
    for row in worksheet.iter_rows():
        values = [cell.value for cell in row]
        if "ФИО" not in values or target_header not in values:
            continue
        header_row_idx = row[0].row
        target_col_idx = values.index(target_header) + 1
        break

    if header_row_idx is None or target_col_idx is None:
        raise AssertionError(f"Date column {target_header} not found in workbook")

    for row_idx in range(header_row_idx + 1, worksheet.max_row + 1):
        if worksheet.cell(row=row_idx, column=1).value != staff_fio:
            continue
        return worksheet.cell(row=row_idx, column=target_col_idx)

    raise AssertionError(f"Staff row {staff_fio} not found in workbook")


def _get_attendance_cell_value(
    raw_bytes: bytes,
    *,
    staff_fio: str,
    target_day: datetime.date,
) -> object:
    return _find_attendance_cell(
        raw_bytes,
        staff_fio=staff_fio,
        target_day=target_day,
    ).value


def _get_attendance_cell_fill_rgb(
    raw_bytes: bytes,
    *,
    staff_fio: str,
    target_day: datetime.date,
) -> str:
    cell = _find_attendance_cell(
        raw_bytes,
        staff_fio=staff_fio,
        target_day=target_day,
    )
    start_color = getattr(cell.fill, "start_color", None)
    rgb = ""
    if start_color is not None:
        rgb = str(getattr(start_color, "rgb", "") or getattr(start_color, "index", ""))
    return rgb.upper()


class BuildingAttendanceReportServiceTests(TestCase):
    def setUp(self):
        self.student_position = Position.objects.create(name="Студент")
        self.dept_a = ChildDepartment.objects.create(id="D-A", name="Кафедра А")
        self.dept_b = ChildDepartment.objects.create(id="D-B", name="Кафедра Б")
        self.class_location = ClassLocation.objects.create(
            name="Абылай",
            address=ABILAI_ADDRESS,
            latitude=43.2389,
            longitude=76.8897,
        )
        ClassLocation.objects.create(
            name="Торекулова",
            address=TOREKULOVA_ADDRESS,
            latitude=43.255,
            longitude=76.93,
        )

    def _create_student(self, pin: str, department: ChildDepartment) -> Staff:
        staff = Staff.objects.create(
            pin=pin,
            name=f"Name-{pin}",
            surname=f"Surname-{pin}",
            department=department,
        )
        staff.positions.add(self.student_position)
        return staff

    def _create_sa(
        self, staff: Staff, event_day: datetime.date, area_name: str
    ) -> None:
        StaffAttendance.objects.create(
            staff=staff,
            date_at=event_day + datetime.timedelta(days=1),
            first_in=_aware_dt(event_day, 9, 0),
            last_out=_aware_dt(event_day, 15, 0),
            area_name_in=area_name,
            area_name_out=area_name,
        )

    def _create_la(
        self,
        staff: Staff,
        event_day: datetime.date,
        *,
        latitude: float,
        longitude: float,
        auto_status: str = LessonAttendance.PHOTO_SPOOF_STATUS_PENDING,
        manual_verdict: str = LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
    ) -> None:
        LessonAttendance.objects.create(
            staff=staff,
            subject_name="Math",
            tutor_id=1,
            tutor="Tutor",
            first_in=_aware_dt(event_day, 10, 0),
            last_out=_aware_dt(event_day, 11, 30),
            latitude=latitude,
            longitude=longitude,
            date_at=event_day,
            photo_spoof_status=auto_status,
            photo_manual_verdict=manual_verdict,
        )

    def test_default_selects_last_7_dates_with_data(self):
        student = self._create_student("S100S", self.dept_a)
        for day in range(1, 11):
            event_day = datetime.date(2026, 3, day)
            self._create_sa(student, event_day, "цос")

        result = building_attendance_report.build_building_attendance_report_excel(
            date_from=None,
            date_to=None,
            days_with_data=7,
        )

        expected = [datetime.date(2026, 3, day) for day in range(4, 11)]
        self.assertEqual(result.selected_dates, expected)

    def test_range_uses_only_dates_with_data(self):
        student = self._create_student("S101S", self.dept_a)
        self._create_sa(student, datetime.date(2026, 3, 1), "цос")
        self._create_sa(student, datetime.date(2026, 3, 3), "цос")

        result = building_attendance_report.build_building_attendance_report_excel(
            date_from=datetime.date(2026, 3, 1),
            date_to=datetime.date(2026, 3, 5),
            days_with_data=7,
        )

        self.assertEqual(
            result.selected_dates,
            [datetime.date(2026, 3, 1), datetime.date(2026, 3, 3)],
        )

    def test_parse_date_from_to_today(self):
        params = building_attendance_report.parse_report_request_params(
            date_from_raw="2026-03-01",
            date_to_raw=None,
            days_with_data_raw=None,
            today=datetime.date(2026, 3, 5),
        )
        self.assertEqual(params.date_from, datetime.date(2026, 3, 1))
        self.assertEqual(params.date_to, datetime.date(2026, 3, 5))
        self.assertEqual(params.days_with_data, 7)

    def test_effective_suspicious_filter_with_manual_clean_override(self):
        excluded_student = self._create_student("S102S", self.dept_a)
        included_student = self._create_student("S103S", self.dept_a)
        target_day = datetime.date(2026, 3, 10)

        self._create_la(
            excluded_student,
            target_day,
            latitude=43.2389,
            longitude=76.8897,
            auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS,
            manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
        )
        self._create_la(
            included_student,
            target_day,
            latitude=43.2389,
            longitude=76.8897,
            auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS,
            manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_CLEAN,
        )

        result = building_attendance_report.build_building_attendance_report_excel(
            date_from=target_day,
            date_to=target_day,
            days_with_data=7,
        )

        self.assertEqual(len(result.daily_rows), 1)
        row = result.daily_rows[0]
        self.assertEqual(row["students_count"], 1)
        self.assertEqual(row["day_total_students"], 1)
        self.assertEqual(row["pct"], 100.0)

    def test_includes_pending_review_error_and_excludes_only_rejected(self):
        target_day = datetime.date(2026, 3, 13)

        included_statuses = [
            LessonAttendance.PHOTO_SPOOF_STATUS_PENDING,
            LessonAttendance.PHOTO_SPOOF_STATUS_REVIEW,
            LessonAttendance.PHOTO_SPOOF_STATUS_ERROR,
            LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN,
        ]
        for idx, status_value in enumerate(included_statuses, start=1):
            student = self._create_student(f"S13{idx:02d}S", self.dept_a)
            self._create_la(
                student,
                target_day,
                latitude=43.2389,
                longitude=76.8897,
                auto_status=status_value,
                manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
            )

        auto_suspicious_student = self._create_student("S1390S", self.dept_a)
        self._create_la(
            auto_suspicious_student,
            target_day,
            latitude=43.2389,
            longitude=76.8897,
            auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS,
            manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
        )

        manual_clean_override_student = self._create_student("S1391S", self.dept_a)
        self._create_la(
            manual_clean_override_student,
            target_day,
            latitude=43.2389,
            longitude=76.8897,
            auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_REVIEW,
            manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_CLEAN,
        )
        manual_suspicious_student = self._create_student("S1392S", self.dept_a)
        self._create_la(
            manual_suspicious_student,
            target_day,
            latitude=43.2389,
            longitude=76.8897,
            auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN,
            manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS,
        )

        result = building_attendance_report.build_building_attendance_report_excel(
            date_from=target_day,
            date_to=target_day,
            days_with_data=7,
        )

        self.assertEqual(len(result.daily_rows), 1)
        row = result.daily_rows[0]
        self.assertEqual(row["students_count"], 5)
        self.assertEqual(row["day_total_students"], 5)
        self.assertEqual(row["pct"], 100.0)

    def test_excludes_entire_staff_day_when_any_lesson_is_suspicious(self):
        student = self._create_student("S1400S", self.dept_a)
        target_day = datetime.date(2026, 3, 14)

        self._create_la(
            student,
            target_day,
            latitude=43.2389,
            longitude=76.8897,
            auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN,
        )
        self._create_la(
            student,
            target_day,
            latitude=43.2389,
            longitude=76.8897,
            auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS,
        )

        result = building_attendance_report.build_building_attendance_report_excel(
            date_from=target_day,
            date_to=target_day,
            days_with_data=7,
        )

        self.assertEqual(result.selected_dates, [])
        self.assertEqual(result.daily_rows, [])
        self.assertEqual(result.summary_rows, [])

    def test_keeps_day_when_staff_attendance_exists_even_if_lesson_is_suspicious(self):
        student = self._create_student("S1401S", self.dept_a)
        target_day = datetime.date(2026, 3, 15)

        self._create_sa(student, target_day, "цос")
        self._create_la(
            student,
            target_day,
            latitude=43.2389,
            longitude=76.8897,
            manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS,
        )

        result = building_attendance_report.build_building_attendance_report_excel(
            date_from=target_day,
            date_to=target_day,
            days_with_data=7,
        )

        self.assertEqual(result.selected_dates, [target_day])
        self.assertEqual(len(result.daily_rows), 1)
        self.assertEqual(result.daily_rows[0]["students_count"], 1)

    def test_deduplicates_staff_between_sa_and_la_with_sa_priority(self):
        student = self._create_student("S104S", self.dept_a)
        target_day = datetime.date(2026, 3, 10)

        self._create_sa(student, target_day, "цос")
        self._create_la(
            student,
            target_day,
            latitude=43.255,
            longitude=76.93,
        )

        result = building_attendance_report.build_building_attendance_report_excel(
            date_from=target_day,
            date_to=target_day,
            days_with_data=7,
        )

        self.assertEqual(len(result.daily_rows), 1)
        self.assertEqual(result.daily_rows[0]["building_address"], ABILAI_ADDRESS)

    def test_la_location_fallback_uses_haversine_when_kdtree_radius_misses(self):
        student = self._create_student("S104B", self.dept_a)
        target_day = datetime.date(2026, 3, 14)

        # Point is within ~161m by Haversine, but lon-delta can miss KDTree
        # query_radius in degree space.
        self._create_la(
            student,
            target_day,
            latitude=43.2389,
            longitude=76.8917,
        )

        result = building_attendance_report.build_building_attendance_report_excel(
            date_from=target_day,
            date_to=target_day,
            days_with_data=7,
        )

        self.assertEqual(len(result.daily_rows), 1)
        self.assertEqual(result.daily_rows[0]["building_address"], ABILAI_ADDRESS)
        self.assertNotEqual(
            result.daily_rows[0]["building_name"],
            building_attendance_report.UNKNOWN_LOCATION_NAME,
        )

    def test_pct_calculated_from_students_present_per_day(self):
        student_a = self._create_student("S105S", self.dept_a)
        student_b = self._create_student("S106S", self.dept_a)
        target_day = datetime.date(2026, 3, 11)

        self._create_sa(student_a, target_day, "цос")
        self._create_sa(student_b, target_day, "торекулова турникет")

        result = building_attendance_report.build_building_attendance_report_excel(
            date_from=target_day,
            date_to=target_day,
            days_with_data=7,
        )

        self.assertEqual(len(result.daily_rows), 2)
        for row in result.daily_rows:
            self.assertEqual(row["day_total_students"], 2)
            self.assertEqual(row["pct"], 50.0)

    def test_excel_headers_are_russian_and_without_department_id(self):
        student = self._create_student("S107S", self.dept_a)
        target_day = datetime.date(2026, 3, 12)
        self._create_sa(student, target_day, "цос")

        result = building_attendance_report.build_building_attendance_report_excel(
            date_from=target_day,
            date_to=target_day,
            days_with_data=7,
        )
        workbook = load_workbook(BytesIO(result.excel_bytes), data_only=True)

        self.assertIn("Дневной_срез", workbook.sheetnames)
        self.assertIn("Свод_локации", workbook.sheetnames)
        self.assertIn("Графики", workbook.sheetnames)

        daily_headers = [cell.value for cell in workbook["Дневной_срез"][1]]
        summary_headers = [cell.value for cell in workbook["Свод_локации"][1]]
        self.assertNotIn("department_id", daily_headers)
        self.assertNotIn("department_id", summary_headers)
        self.assertIn("Кафедра", daily_headers)
        self.assertIn("Локация", summary_headers)


class BuildingAttendanceReportApiAndCommandTests(APITestCase):
    def setUp(self):
        super().setUp()
        Cache.clear()
        self.student_position = Position.objects.create(name="Студент")
        self.department = ChildDepartment.objects.create(id="D-API", name="Кафедра API")
        self.staff = Staff.objects.create(
            pin="S200S",
            name="Api",
            surname="Student",
            department=self.department,
        )
        self.staff.positions.add(self.student_position)
        self.class_location = ClassLocation.objects.create(
            name="Абылай",
            address=ABILAI_ADDRESS,
            latitude=43.2389,
            longitude=76.8897,
        )

        event_day = datetime.date(2026, 3, 10)
        StaffAttendance.objects.create(
            staff=self.staff,
            date_at=event_day + datetime.timedelta(days=1),
            first_in=_aware_dt(event_day, 9, 0),
            last_out=_aware_dt(event_day, 14, 0),
            area_name_in="цос",
            area_name_out="цос",
        )

        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="report-user",
            password="strong-pass-123",
        )
        self.api_key = APIKey.objects.create(
            key_name="report-api-key",
            created_by=self.user,
        )
        self.client.credentials(HTTP_X_API_KEY=self.api_key.key)

    def test_endpoint_returns_excel(self):
        response = self.client.get(
            reverse("building-attendance-report"),
            {
                "date_from": "2026-03-10",
                "date_to": "2026-03-10",
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.headers.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(".xlsx", response.headers.get("Content-Disposition", ""))

    def test_command_and_endpoint_have_same_result(self):
        endpoint_response = self.client.get(
            reverse("building-attendance-report"),
            {
                "date_from": "2026-03-10",
                "date_to": "2026-03-10",
            },
        )
        self.assertEqual(endpoint_response.status_code, status.HTTP_200_OK)

        with TemporaryDirectory() as tmp_dir:
            output_path = f"{tmp_dir}/building_report.xlsx"
            call_command(
                "export_building_attendance_report",
                "--date-from",
                "2026-03-10",
                "--date-to",
                "2026-03-10",
                "--output",
                output_path,
            )
            command_bytes = Path(output_path).read_bytes()

        endpoint_snapshot = _workbook_snapshot(endpoint_response.content)
        command_snapshot = _workbook_snapshot(command_bytes)
        self.assertEqual(endpoint_snapshot, command_snapshot)

    def test_command_output_forces_xlsx_extension(self):
        with TemporaryDirectory() as tmp_dir:
            output_path_without_ext = f"{tmp_dir}/my_report"
            call_command(
                "export_building_attendance_report",
                "--date-from",
                "2026-03-10",
                "--date-to",
                "2026-03-10",
                "--output",
                output_path_without_ext,
            )
            forced_path = Path(f"{output_path_without_ext}.xlsx")
            self.assertTrue(forced_path.exists())

            output_path_wrong_ext = f"{tmp_dir}/my_report.csv"
            call_command(
                "export_building_attendance_report",
                "--date-from",
                "2026-03-10",
                "--date-to",
                "2026-03-10",
                "--output",
                output_path_wrong_ext,
            )
            forced_from_wrong = Path(f"{tmp_dir}/my_report.xlsx")
            self.assertTrue(forced_from_wrong.exists())


class DepartmentAttendanceExcelTests(APITestCase):
    def setUp(self):
        super().setUp()
        Cache.clear()
        self.student_position = Position.objects.create(name="Студент")
        self.department = ChildDepartment.objects.create(
            id="D-EXCEL",
            name="Кафедра Excel",
        )
        self.staff = Staff.objects.create(
            pin="S300S",
            name="Api",
            surname="Student",
            department=self.department,
        )
        self.staff.positions.add(self.student_position)
        self.class_location = ClassLocation.objects.create(
            name="Абылай",
            address=ABILAI_ADDRESS,
            latitude=43.2389,
            longitude=76.8897,
        )
        self.target_day = datetime.date(2026, 3, 10)

        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="department-excel-user",
            password="strong-pass-123",
        )
        self.api_key = APIKey.objects.create(
            key_name="department-excel-api-key",
            created_by=self.user,
        )
        self.client.credentials(HTTP_X_API_KEY=self.api_key.key)

    def _create_lesson(
        self,
        *,
        hour: int,
        event_day: datetime.date | None = None,
        latitude: float | None = None,
        longitude: float | None = None,
        staff: Staff | None = None,
        auto_status: str = LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN,
        manual_verdict: str = LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
    ) -> None:
        lesson_day = event_day or self.target_day
        lesson_staff = staff or self.staff
        LessonAttendance.objects.create(
            staff=lesson_staff,
            subject_name="Math",
            tutor_id=1,
            tutor="Tutor",
            first_in=_aware_dt(lesson_day, hour, 0),
            last_out=_aware_dt(lesson_day, hour + 1, 0),
            latitude=self.class_location.latitude if latitude is None else latitude,
            longitude=self.class_location.longitude if longitude is None else longitude,
            date_at=lesson_day,
            photo_spoof_status=auto_status,
            photo_manual_verdict=manual_verdict,
        )

    def _create_staff_member(self, *, pin: str, name: str, surname: str) -> Staff:
        staff = Staff.objects.create(
            pin=pin,
            name=name,
            surname=surname,
            department=self.department,
        )
        staff.positions.add(self.student_position)
        return staff

    def _create_sa(self) -> None:
        StaffAttendance.objects.create(
            staff=self.staff,
            date_at=self.target_day + datetime.timedelta(days=1),
            first_in=_aware_dt(self.target_day, 8, 30),
            last_out=_aware_dt(self.target_day, 14, 30),
            area_name_in="цос",
            area_name_out="цос",
        )

    def _download_excel(
        self,
        *,
        start_day: datetime.date | None = None,
        end_day: datetime.date | None = None,
    ):
        range_start = start_day or self.target_day
        range_end = end_day or self.target_day
        response = self.client.get(
            reverse("sent_excel", kwargs={"department_id": self.department.id}),
            {
                "startDate": range_start.isoformat(),
                "endDate": range_end.isoformat(),
            },
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        return response.content

    def test_department_excel_marks_day_absent_when_any_lesson_is_suspicious(self):
        self._create_lesson(hour=9)
        self._create_lesson(
            hour=11,
            auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS,
        )

        excel_bytes = self._download_excel()
        attendance_value = _get_attendance_cell_value(
            excel_bytes,
            staff_fio="Student Api",
            target_day=self.target_day,
        )

        self.assertIn("Отсутствие", str(attendance_value))
        self.assertIn("Подозрение на обман FaceID", str(attendance_value))
        self.assertTrue(
            _get_attendance_cell_fill_rgb(
                excel_bytes,
                staff_fio="Student Api",
                target_day=self.target_day,
            ).endswith(FACEID_FILL_RGB)
        )

    def test_department_excel_keeps_staff_attendance_when_lesson_day_is_suspicious(
        self,
    ):
        self._create_sa()
        self._create_lesson(
            hour=11,
            manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS,
        )

        excel_bytes = self._download_excel()
        attendance_value = _get_attendance_cell_value(
            excel_bytes,
            staff_fio="Student Api",
            target_day=self.target_day,
        )

        self.assertIn("08:30:00 - 14:30:00", str(attendance_value))
        self.assertIn("Подозрение на обман FaceID", str(attendance_value))
        self.assertTrue(
            _get_attendance_cell_fill_rgb(
                excel_bytes,
                staff_fio="Student Api",
                target_day=self.target_day,
            ).endswith(FACEID_FILL_RGB)
        )

    def test_department_excel_manual_clean_override_removes_faceid_flag(self):
        self._create_lesson(
            hour=9,
            auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS,
            manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_CLEAN,
        )

        excel_bytes = self._download_excel()
        attendance_value = _get_attendance_cell_value(
            excel_bytes,
            staff_fio="Student Api",
            target_day=self.target_day,
        )

        self.assertIn("10:00:00 - 11:00:00", str(attendance_value))
        self.assertNotIn("Подозрение на обман FaceID", str(attendance_value))
        self.assertFalse(
            _get_attendance_cell_fill_rgb(
                excel_bytes,
                staff_fio="Student Api",
                target_day=self.target_day,
            ).endswith(FACEID_FILL_RGB)
        )

    def test_department_excel_marks_gps_spoof_when_same_micro_point_repeats(self):
        start_day = datetime.date(2026, 3, 10)
        end_day = datetime.date(2026, 3, 12)
        repeated_points = [
            (43.2389050, 76.8897050),
            (43.2389054, 76.8897055),
            (43.2389058, 76.8897060),
        ]
        for index, point in enumerate(repeated_points):
            self._create_lesson(
                hour=9,
                event_day=start_day + datetime.timedelta(days=index),
                latitude=point[0],
                longitude=point[1],
            )

        excel_bytes = self._download_excel(start_day=start_day, end_day=end_day)

        for offset in range(3):
            current_day = start_day + datetime.timedelta(days=offset)
            attendance_value = _get_attendance_cell_value(
                excel_bytes,
                staff_fio="Student Api",
                target_day=current_day,
            )
            self.assertIn("Подозрение на подмену локации", str(attendance_value))
            self.assertTrue(
                _get_attendance_cell_fill_rgb(
                    excel_bytes,
                    staff_fio="Student Api",
                    target_day=current_day,
                ).endswith(GPS_SPOOF_FILL_RGB)
            )

    def test_department_excel_does_not_mark_gps_spoof_below_share_threshold(self):
        start_day = datetime.date(2026, 3, 10)
        end_day = datetime.date(2026, 3, 16)
        repeated_points = [
            (43.2389050, 76.8897050),
            (43.2389054, 76.8897055),
            (43.2389058, 76.8897060),
        ]
        for index, point in enumerate(repeated_points):
            self._create_lesson(
                hour=9,
                event_day=start_day + datetime.timedelta(days=index),
                latitude=point[0],
                longitude=point[1],
            )

        unique_points = [
            (43.2389400, 76.8897400),
            (43.2389850, 76.8897850),
            (43.2390300, 76.8898300),
            (43.2390750, 76.8898750),
        ]
        for index, point in enumerate(unique_points, start=3):
            self._create_lesson(
                hour=9,
                event_day=start_day + datetime.timedelta(days=index),
                latitude=point[0],
                longitude=point[1],
            )

        excel_bytes = self._download_excel(start_day=start_day, end_day=end_day)

        self.assertFalse(
            _get_attendance_cell_fill_rgb(
                excel_bytes,
                staff_fio="Student Api",
                target_day=start_day,
            ).endswith(GPS_SPOOF_FILL_RGB)
        )

    def test_department_excel_does_not_mark_gps_spoof_for_only_two_days(self):
        start_day = datetime.date(2026, 3, 10)
        end_day = datetime.date(2026, 3, 11)
        for offset, point in enumerate(
            [
                (43.2389050, 76.8897050),
                (43.2389054, 76.8897055),
            ]
        ):
            self._create_lesson(
                hour=9,
                event_day=start_day + datetime.timedelta(days=offset),
                latitude=point[0],
                longitude=point[1],
            )

        excel_bytes = self._download_excel(start_day=start_day, end_day=end_day)

        self.assertFalse(
            _get_attendance_cell_fill_rgb(
                excel_bytes,
                staff_fio="Student Api",
                target_day=start_day,
            ).endswith(GPS_SPOOF_FILL_RGB)
        )

    def test_department_excel_marks_shared_micro_point_per_staff_independently(self):
        start_day = datetime.date(2026, 3, 10)
        end_day = datetime.date(2026, 3, 12)
        repeated_points = [
            (43.2389050, 76.8897050),
            (43.2389054, 76.8897055),
            (43.2389058, 76.8897060),
        ]
        peer_staff = [
            self._create_staff_member(
                pin="S301S",
                name="PeerOne",
                surname="Student",
            )
        ]

        for staff in [self.staff, *peer_staff]:
            for index, point in enumerate(repeated_points):
                self._create_lesson(
                    hour=9,
                    event_day=start_day + datetime.timedelta(days=index),
                    latitude=point[0],
                    longitude=point[1],
                    staff=staff,
                )

        excel_bytes = self._download_excel(start_day=start_day, end_day=end_day)

        for staff_fio in ["Student Api", "Student PeerOne"]:
            for offset in range(3):
                current_day = start_day + datetime.timedelta(days=offset)
                attendance_value = _get_attendance_cell_value(
                    excel_bytes,
                    staff_fio=staff_fio,
                    target_day=current_day,
                )
                self.assertIn(
                    "Подозрение на подмену локации",
                    str(attendance_value),
                )
                self.assertTrue(
                    _get_attendance_cell_fill_rgb(
                        excel_bytes,
                        staff_fio=staff_fio,
                        target_day=current_day,
                    ).endswith(GPS_SPOOF_FILL_RGB)
                )

    def test_department_excel_combines_faceid_and_gps_spoof(self):
        start_day = datetime.date(2026, 3, 10)
        end_day = datetime.date(2026, 3, 12)
        repeated_points = [
            (43.2389050, 76.8897050),
            (43.2389054, 76.8897055),
            (43.2389058, 76.8897060),
        ]
        for index, point in enumerate(repeated_points):
            lesson_kwargs = {}
            if index == 1:
                lesson_kwargs["auto_status"] = (
                    LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS
                )
            self._create_lesson(
                hour=9,
                event_day=start_day + datetime.timedelta(days=index),
                latitude=point[0],
                longitude=point[1],
                **lesson_kwargs,
            )

        excel_bytes = self._download_excel(start_day=start_day, end_day=end_day)

        attendance_value = _get_attendance_cell_value(
            excel_bytes,
            staff_fio="Student Api",
            target_day=start_day + datetime.timedelta(days=1),
        )
        self.assertIn(
            "Подозрение на обман FaceID и подмену локации",
            str(attendance_value),
        )
        self.assertTrue(
            _get_attendance_cell_fill_rgb(
                excel_bytes,
                staff_fio="Student Api",
                target_day=start_day + datetime.timedelta(days=1),
            ).endswith(FACEID_GPS_FILL_RGB)
        )

    def test_department_excel_legend_uses_neutral_labels(self):
        excel_bytes = self._download_excel()
        snapshot = _workbook_snapshot(excel_bytes)["Отчет посещаемости"]
        legend_labels = {
            row[1]
            for row in snapshot[4:13]
            if len(row) > 1 and isinstance(row[1], str) and row[1]
        }

        self.assertIn("Активность в выходной день", legend_labels)
        self.assertIn("Удаленный формат", legend_labels)
        self.assertIn("Согласованная причина отсутствия", legend_labels)
        self.assertIn("Несогласованное отсутствие / отсутствие", legend_labels)
        self.assertIn("Только событие лифта", legend_labels)
        self.assertIn("Попытка обмана FaceID", legend_labels)
        self.assertIn("Подозрение на подмену локации", legend_labels)
        self.assertIn("FaceID + подмена локации", legend_labels)

    def test_department_excel_cache_refreshes_after_verdict_change(self):
        lesson = LessonAttendance.objects.create(
            staff=self.staff,
            subject_name="Math",
            tutor_id=1,
            tutor="Tutor",
            first_in=_aware_dt(self.target_day, 9, 0),
            last_out=_aware_dt(self.target_day, 10, 0),
            latitude=self.class_location.latitude,
            longitude=self.class_location.longitude,
            date_at=self.target_day,
            photo_spoof_status=LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN,
            photo_manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
        )

        initial_excel = self._download_excel()
        self.assertFalse(
            _get_attendance_cell_fill_rgb(
                initial_excel,
                staff_fio="Student Api",
                target_day=self.target_day,
            ).endswith(FACEID_FILL_RGB)
        )

        lesson.photo_manual_verdict = LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS
        lesson.save(update_fields=["photo_manual_verdict"])

        refreshed_excel = self._download_excel()
        self.assertTrue(
            _get_attendance_cell_fill_rgb(
                refreshed_excel,
                staff_fio="Student Api",
                target_day=self.target_day,
            ).endswith(FACEID_FILL_RGB)
        )

    def test_department_excel_cache_refreshes_after_classlocation_change(self):
        start_day = datetime.date(2026, 3, 10)
        end_day = datetime.date(2026, 3, 12)
        for offset, point in enumerate(
            [
                (43.2389050, 76.8897050),
                (43.2389054, 76.8897055),
                (43.2389058, 76.8897060),
            ]
        ):
            self._create_lesson(
                hour=9,
                event_day=start_day + datetime.timedelta(days=offset),
                latitude=point[0],
                longitude=point[1],
            )

        initial_excel = self._download_excel(start_day=start_day, end_day=end_day)
        self.assertTrue(
            _get_attendance_cell_fill_rgb(
                initial_excel,
                staff_fio="Student Api",
                target_day=start_day,
            ).endswith(GPS_SPOOF_FILL_RGB)
        )

        self.class_location.latitude = 43.1000
        self.class_location.longitude = 76.7000
        self.class_location.save(update_fields=["latitude", "longitude"])

        refreshed_excel = self._download_excel(start_day=start_day, end_day=end_day)
        self.assertFalse(
            _get_attendance_cell_fill_rgb(
                refreshed_excel,
                staff_fio="Student Api",
                target_day=start_day,
            ).endswith(GPS_SPOOF_FILL_RGB)
        )


class SuspiciousLessonAttendanceExportCommandTests(TestCase):
    def setUp(self):
        self.department = ChildDepartment.objects.create(
            id="D-SUSP",
            name="Кафедра Suspicious",
        )
        self.staff = Staff.objects.create(
            pin="S400S",
            name="Api",
            surname="Student",
            department=self.department,
        )
        ClassLocation.objects.create(
            name="Абылай",
            address=ABILAI_ADDRESS,
            latitude=43.2389,
            longitude=76.8897,
        )
        self.target_day = datetime.date(2026, 3, 10)

    def _create_lesson(
        self,
        *,
        hour: int,
        auto_status: str = LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN,
    ) -> None:
        LessonAttendance.objects.create(
            staff=self.staff,
            subject_name="Math",
            tutor_id=1,
            tutor="Tutor",
            first_in=_aware_dt(self.target_day, hour, 0),
            last_out=_aware_dt(self.target_day, hour + 1, 0),
            latitude=43.2389,
            longitude=76.8897,
            date_at=self.target_day,
            photo_spoof_status=auto_status,
            photo_manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
        )

    def test_suspicious_export_command_keeps_only_truly_suspicious_rows(self):
        self._create_lesson(
            hour=9, auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN
        )
        self._create_lesson(
            hour=11,
            auto_status=LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS,
        )

        with TemporaryDirectory() as tmp_dir:
            output_path = f"{tmp_dir}/suspicious_report.xlsx"
            call_command(
                "export_suspicious_lesson_attendance",
                "--date-from",
                self.target_day.isoformat(),
                "--date-to",
                self.target_day.isoformat(),
                "--output",
                output_path,
            )
            workbook = load_workbook(output_path, data_only=True)

        rows = list(workbook["Suspicious"].iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][1], "Student Api")
        self.assertEqual(rows[1][3], self.target_day.strftime("%d.%m.%Y"))
