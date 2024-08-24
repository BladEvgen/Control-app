import requests
import datetime
import pandas as pd
from dateutil import parser
from openpyxl import Workbook
from django.conf import settings
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


API_BASE_URL = settings.MAIN_IP
API_KEY = settings.MINE_API
DATA_URL = f"{API_BASE_URL}/api/staff/s00260/?start_date=2024-06-29&end_date=2024-07-29"

CONTRACT_TYPE_CHOICES = {
    "full_time": "Полная занятость",
    "part_time": "Частичная занятость",
    "gph": "ГПХ",
}


def fetch_data(url: str, api_key: str) -> dict:
    headers = {"x-api-key": api_key}
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json()


def format_datetime(dt_str: str, output_format: str) -> str:
    return parser.parse(dt_str).strftime(output_format) if dt_str else None


def parse_attendance_data(data: dict) -> tuple:
    name = f"{data.get('name', '')} {data.get('surname', '')}"
    department = data.get("department", "").replace("_", " ").capitalize()
    contract_type = data.get("contract_type", None)
    contract_type_display = CONTRACT_TYPE_CHOICES.get(contract_type, None)
    percent_for_period = data.get("percent_for_period", 0)
    attendance = data.get("attendance", {})

    rows = []
    for date, details in sorted(
        attendance.items(),
        key=lambda x: datetime.datetime.strptime(x[0], "%d-%m-%Y"),
        reverse=True,
    ):
        date_str = datetime.datetime.strptime(date, "%d-%m-%Y").strftime("%d.%m.%Y")
        is_weekend = details.get("is_weekend", False)
        first_in = format_datetime(details.get("first_in"), "%H:%M:%S %Z")
        last_out = format_datetime(details.get("last_out"), "%H:%M:%S %Z")
        percent_day = details.get("percent_day", 0)

        attendance_info = (
            "Выходной"
            if is_weekend and not (first_in or last_out)
            else f"{first_in} - {last_out}" if first_in and last_out else "Отсутствие"
        )
        rows.append([date_str, attendance_info, percent_day])

    rows.append(["Процент за период", "", percent_for_period])
    df = pd.DataFrame(rows, columns=["Дата", "Посещаемость", "Процент дня"])
    return name, department, contract_type_display, percent_for_period, df


def auto_fit_column_widths(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width


def save_to_excel(
    name: str,
    department: str,
    contract_type: str | None,
    percent_for_period: float,
    df: pd.DataFrame,
) -> Workbook:
    wb = Workbook()
    ws = wb.active

    data_font = Font(name="Roboto", size=12)
    title_font = Font(name="Roboto", size=12, bold=True)
    data_alignment = Alignment(horizontal="center", vertical="center")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(["ФИО", name])
    ws.append(["Отдел", department])
    if contract_type:
        ws.append(["Тип занятости", contract_type])
    ws.append(["Процент за период", percent_for_period])
    ws.append([])
    ws.append(list(df.columns))

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 6):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = data_alignment
            cell.font = data_font

    for row in ws.iter_rows(min_row=1, max_row=4 if contract_type else 3, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = header_alignment
            cell.font = title_font if cell.column == 1 else data_font

    for cell in ws.iter_rows(
        min_row=4 if contract_type else 3,
        max_row=4 if contract_type else 3,
        min_col=1,
        max_col=2,
    ):
        for c in cell:
            c.alignment = header_alignment
            c.font = title_font if c.column == 1 else data_font

    for row in ws.iter_rows(min_row=6 if contract_type else 5, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = data_alignment
            cell.font = data_font

    auto_fit_column_widths(ws)

    return wb


def main():
    try:
        data = fetch_data(DATA_URL, API_KEY)
        if data:
            name, department, contract_type, percent_for_period, df = parse_attendance_data(data)
            workbook = save_to_excel(name, department, contract_type, percent_for_period, df)
            filename = f"Отчет_{department}_{name}.xlsx".replace(" ", "_")
            workbook.save(filename)
    except requests.RequestException as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    main()
