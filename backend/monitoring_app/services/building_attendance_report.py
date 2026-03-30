from __future__ import annotations

import datetime
from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Optional

from django.db.models import Prefetch
from django.utils import timezone
from monitoring_app import models, utils
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font

DEFAULT_DAYS_WITH_DATA = 7
UNKNOWN_LOCATION_NAME = "Неизвестная локация"
CIS_DATE_FORMAT = "%d.%m.%Y"
WEEKDAY_RU = (
    "Понедельник",
    "Вторник",
    "Среда",
    "Четверг",
    "Пятница",
    "Суббота",
    "Воскресенье",
)


@dataclass(frozen=True)
class ReportRequestParams:
    date_from: Optional[datetime.date]
    date_to: Optional[datetime.date]
    days_with_data: int


@dataclass(frozen=True)
class BuildingAttendanceReportResult:
    excel_bytes: bytes
    selected_dates: list[datetime.date]
    daily_rows: list[dict[str, Any]]
    summary_rows: list[dict[str, Any]]


def parse_report_request_params(
    *,
    date_from_raw: Optional[str],
    date_to_raw: Optional[str],
    days_with_data_raw: Optional[Any],
    today: Optional[datetime.date] = None,
    default_days_with_data: int = DEFAULT_DAYS_WITH_DATA,
) -> ReportRequestParams:
    """Parse and validate input params for building attendance report."""
    if default_days_with_data <= 0:
        raise ValueError("default_days_with_data must be positive")

    now_date = today or timezone.localdate()

    if days_with_data_raw in (None, ""):
        days_with_data = default_days_with_data
    else:
        try:
            days_with_data = int(days_with_data_raw)
        except (TypeError, ValueError) as exc:
            raise ValueError("days_with_data must be a positive integer") from exc

    if days_with_data <= 0:
        raise ValueError("days_with_data must be greater than 0")

    if date_to_raw and not date_from_raw:
        raise ValueError("date_to can be used only together with date_from")

    if date_from_raw:
        date_from = _parse_iso_date(date_from_raw, "date_from")
        if date_to_raw:
            date_to = _parse_iso_date(date_to_raw, "date_to")
        else:
            date_to = now_date

        if date_from > date_to:
            raise ValueError("date_from must be less than or equal to date_to")
        return ReportRequestParams(
            date_from=date_from,
            date_to=date_to,
            days_with_data=days_with_data,
        )

    return ReportRequestParams(
        date_from=None,
        date_to=None,
        days_with_data=days_with_data,
    )


def build_report_filename(
    selected_dates: list[datetime.date], prefix: str = "building_attendance"
) -> str:
    if not selected_dates:
        return f"{prefix}_no_data.xlsx"
    date_from = selected_dates[0].strftime("%Y-%m-%d")
    date_to = selected_dates[-1].strftime("%Y-%m-%d")
    if date_from == date_to:
        return f"{prefix}_{date_from}.xlsx"
    return f"{prefix}_{date_from}_{date_to}.xlsx"


def build_building_attendance_report_excel(
    *,
    date_from: Optional[datetime.date],
    date_to: Optional[datetime.date],
    days_with_data: int = DEFAULT_DAYS_WITH_DATA,
) -> BuildingAttendanceReportResult:
    """Build report rows and Excel bytes for building attendance by departments."""
    staff_candidates = list(
        models.Staff.objects.select_related("department")
        .prefetch_related(
            Prefetch("positions", queryset=models.Position.objects.only("name"))
        )
        .only("id", "pin", "department_id", "department__name")
    )
    students = [staff for staff in staff_candidates if _is_student_staff(staff)]

    if not students:
        excel_bytes = _build_excel_file([], [])
        return BuildingAttendanceReportResult(
            excel_bytes=excel_bytes,
            selected_dates=[],
            daily_rows=[],
            summary_rows=[],
        )

    staff_ids = [staff.id for staff in students]
    selected_dates = _select_event_dates(
        staff_ids=staff_ids,
        date_from=date_from,
        date_to=date_to,
        days_with_data=days_with_data,
    )

    if not selected_dates:
        excel_bytes = _build_excel_file([], [])
        return BuildingAttendanceReportResult(
            excel_bytes=excel_bytes,
            selected_dates=[],
            daily_rows=[],
            summary_rows=[],
        )

    daily_rows, summary_rows = _collect_daily_and_summary_rows(
        students=students,
        selected_dates=selected_dates,
    )
    excel_bytes = _build_excel_file(daily_rows, summary_rows)
    return BuildingAttendanceReportResult(
        excel_bytes=excel_bytes,
        selected_dates=selected_dates,
        daily_rows=daily_rows,
        summary_rows=summary_rows,
    )


def _parse_iso_date(value: str, field_name: str) -> datetime.date:
    try:
        return datetime.date.fromisoformat(str(value))
    except ValueError as exc:
        raise ValueError(f"{field_name} must be in YYYY-MM-DD format") from exc


def _select_event_dates(
    *,
    staff_ids: list[int],
    date_from: Optional[datetime.date],
    date_to: Optional[datetime.date],
    days_with_data: int,
) -> list[datetime.date]:
    all_dates = _get_dates_with_data(
        staff_ids=staff_ids,
        date_from=date_from,
        date_to=date_to,
    )
    if not all_dates:
        return []

    if date_from is not None and date_to is not None:
        return sorted(all_dates)

    desc_dates = sorted(all_dates, reverse=True)
    selected = desc_dates[:days_with_data]
    return sorted(selected)


def _get_dates_with_data(
    *,
    staff_ids: list[int],
    date_from: Optional[datetime.date],
    date_to: Optional[datetime.date],
) -> set[datetime.date]:
    if not staff_ids:
        return set()

    sa_dates = _get_sa_event_dates(
        staff_ids=staff_ids,
        date_from=date_from,
        date_to=date_to,
    )
    la_dates = _get_la_event_dates(
        staff_ids=staff_ids,
        date_from=date_from,
        date_to=date_to,
    )
    return sa_dates | la_dates


def _get_sa_event_dates(
    *,
    staff_ids: list[int],
    date_from: Optional[datetime.date],
    date_to: Optional[datetime.date],
) -> set[datetime.date]:
    one_day = datetime.timedelta(days=1)
    qs = models.StaffAttendance.objects.filter(
        staff_id__in=staff_ids,
        first_in__isnull=False,
    )
    if date_from is not None and date_to is not None:
        qs = qs.filter(
            date_at__range=(date_from + one_day, date_to + one_day),
        )

    result: set[datetime.date] = set()
    for value in qs.values_list("date_at", flat=True).distinct():
        date_value = _normalize_to_date(value)
        if date_value is None:
            continue
        result.add(date_value - one_day)
    return result


def _get_la_event_dates(
    *,
    staff_ids: list[int],
    date_from: Optional[datetime.date],
    date_to: Optional[datetime.date],
) -> set[datetime.date]:
    qs = models.LessonAttendance.exclude_report_invalid_days(
        models.LessonAttendance.objects.filter(staff_id__in=staff_ids)
    )
    if date_from is not None and date_to is not None:
        qs = qs.filter(date_at__range=(date_from, date_to))

    result: set[datetime.date] = set()
    for value in qs.values_list("date_at", flat=True).distinct():
        date_value = _normalize_to_date(value)
        if date_value is None:
            continue
        result.add(date_value)
    return result


def _normalize_to_date(value: Any) -> Optional[datetime.date]:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


def _is_student_staff(staff: models.Staff) -> bool:
    for position in staff.positions.all():
        normalized = str(getattr(position, "name", "") or "").casefold()
        if "студент" in normalized or "student" in normalized:
            return True
    return False


def _format_cis_date(value: datetime.date) -> str:
    return value.strftime(CIS_DATE_FORMAT)


def _weekday_ru(value: datetime.date) -> str:
    return WEEKDAY_RU[value.weekday()]


def _building_sort_key(building_name: str, building_address: str) -> tuple[str, str]:
    return (
        str(building_name or "").strip().casefold(),
        str(building_address or "").strip().casefold(),
    )


def _collect_daily_and_summary_rows(
    *,
    students: list[models.Staff],
    selected_dates: list[datetime.date],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    selected_set = set(selected_dates)
    date_from = selected_dates[0]
    date_to = selected_dates[-1]
    one_day = datetime.timedelta(days=1)

    staff_ids = [staff.id for staff in students]
    staff_map = {
        staff.id: {
            "pin": staff.pin,
            "department_name": staff.department.name if staff.department else "N/A",
        }
        for staff in students
    }

    sa_by_key: dict[tuple[datetime.date, int], list[dict[str, Any]]] = defaultdict(list)
    la_by_key: dict[tuple[datetime.date, int], list[dict[str, Any]]] = defaultdict(list)

    sa_qs = models.StaffAttendance.objects.filter(
        staff_id__in=staff_ids,
        date_at__range=(date_from + one_day, date_to + one_day),
        first_in__isnull=False,
    ).values("staff_id", "date_at", "first_in", "area_name_in", "area_name_out")

    for record in sa_qs:
        event_date = _normalize_to_date(record.get("date_at"))
        if event_date is None:
            continue
        event_date = event_date - one_day
        if event_date not in selected_set:
            continue
        key = (event_date, int(record["staff_id"]))
        sa_by_key[key].append(record)

    la_qs = models.LessonAttendance.exclude_report_invalid_days(
        models.LessonAttendance.objects.filter(
            staff_id__in=staff_ids,
            date_at__range=(date_from, date_to),
        )
    ).values("staff_id", "date_at", "first_in", "latitude", "longitude")

    for record in la_qs:
        event_date = _normalize_to_date(record.get("date_at"))
        if event_date is None or event_date not in selected_set:
            continue
        key = (event_date, int(record["staff_id"]))
        la_by_key[key].append(record)

    address_to_name, location_points = _build_location_maps()

    holidays = {
        holiday.date: holiday.is_working_day
        for holiday in models.PublicHoliday.objects.filter(
            date__range=(date_from, date_to)
        )
    }

    daily_bucket: dict[tuple[datetime.date, str], dict[str, Any]] = defaultdict(
        lambda: {"staff_ids": set(), "buildings": defaultdict(set)}
    )
    global_day_staff_ids: dict[datetime.date, set[int]] = defaultdict(set)
    global_location_day_staff: dict[tuple[datetime.date, str, str], set[int]] = (
        defaultdict(set)
    )

    all_keys = set(sa_by_key.keys()) | set(la_by_key.keys())
    for event_date, staff_id in all_keys:
        staff_info = staff_map.get(staff_id)
        if staff_info is None:
            continue

        sa_records = sa_by_key.get((event_date, staff_id), [])
        la_records = la_by_key.get((event_date, staff_id), [])
        if not sa_records and not la_records:
            continue

        location_name, location_address = _choose_location(
            sa_records=sa_records,
            la_records=la_records,
            address_to_name=address_to_name,
            location_points=location_points,
        )

        dept_name = staff_info["department_name"]
        bucket_key = (event_date, dept_name)
        bucket = daily_bucket[bucket_key]
        bucket["staff_ids"].add(staff_id)
        bucket["buildings"][(location_name, location_address)].add(staff_id)
        global_day_staff_ids[event_date].add(staff_id)
        global_location_day_staff[(event_date, location_name, location_address)].add(
            staff_id
        )

    daily_rows: list[dict[str, Any]] = []
    summary_acc: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {"total_visits": 0, "days": 0, "pct_sum": 0.0}
    )
    summary_days: dict[tuple[str, str], set[datetime.date]] = defaultdict(set)

    sorted_daily_keys = sorted(
        daily_bucket.keys(),
        key=lambda item: (item[0], str(item[1]).strip().casefold()),
    )

    for event_date, dept_name in sorted_daily_keys:
        bucket = daily_bucket[(event_date, dept_name)]
        day_total_students = len(bucket["staff_ids"])

        buildings_items = sorted(
            bucket["buildings"].items(),
            key=lambda item: _building_sort_key(item[0][0], item[0][1]),
        )
        for (building_name, building_address), staff_set in buildings_items:
            students_count = len(staff_set)
            pct = round((students_count / day_total_students) * 100, 2)

            row = {
                "date": _format_cis_date(event_date),
                "day_of_week": _weekday_ru(event_date),
                "is_weekend": event_date.weekday() >= 5,
                "is_public_holiday": event_date in holidays,
                "is_working_day_override": (
                    holidays.get(event_date) if event_date in holidays else None
                ),
                "department_name": dept_name,
                "building_name": building_name,
                "building_address": building_address,
                "students_count": students_count,
                "day_total_students": day_total_students,
                "pct": pct,
            }
            daily_rows.append(row)

    for (event_date, building_name, building_address), staff_set in sorted(
        global_location_day_staff.items(),
        key=lambda item: (
            item[0][0],
            _building_sort_key(item[0][1], item[0][2]),
        ),
    ):
        day_total_students = len(global_day_staff_ids.get(event_date, set()))
        students_count = len(staff_set)
        pct = (
            round((students_count / day_total_students) * 100, 2)
            if day_total_students
            else 0.0
        )
        summary_key = (building_name, building_address)
        summary_acc[summary_key]["total_visits"] += students_count
        summary_acc[summary_key]["days"] += 1
        summary_acc[summary_key]["pct_sum"] += pct
        summary_days[summary_key].add(event_date)

    summary_rows: list[dict[str, Any]] = []
    for summary_key in sorted(
        summary_acc.keys(),
        key=lambda item: _building_sort_key(item[0], item[1]),
    ):
        building_name, building_address = summary_key
        acc = summary_acc[summary_key]
        days = acc["days"]
        total_visits = acc["total_visits"]

        summary_rows.append(
            {
                "building_name": building_name,
                "building_address": building_address,
                "total_visits": total_visits,
                "avg_daily_count": round(total_visits / days, 2) if days else 0.0,
                "avg_pct": round(acc["pct_sum"] / days, 2) if days else 0.0,
                "days_with_location_data": len(summary_days[summary_key]),
            }
        )

    return daily_rows, summary_rows


def _build_location_maps() -> tuple[dict[str, str], list[dict[str, Any]]]:
    class_locations = list(
        models.ClassLocation.objects.only(
            "name", "address", "latitude", "longitude"
        ).values("name", "address", "latitude", "longitude")
    )

    address_to_name: dict[str, str] = {}
    location_points: list[dict[str, Any]] = []

    for loc in class_locations:
        name = str(loc.get("name") or "").strip()
        address = str(loc.get("address") or "").strip()
        lat = loc.get("latitude")
        lon = loc.get("longitude")
        if not name or not address or lat is None or lon is None:
            continue

        address_to_name.setdefault(address, name)

        point = {
            "name": name,
            "address": address,
            "latitude": float(lat),
            "longitude": float(lon),
        }
        location_points.append(point)

    return address_to_name, location_points


def _choose_location(
    *,
    sa_records: list[dict[str, Any]],
    la_records: list[dict[str, Any]],
    address_to_name: dict[str, str],
    location_points: list[dict[str, Any]],
) -> tuple[str, str]:
    sa_address = _resolve_sa_address(sa_records)
    if sa_address:
        return (address_to_name.get(sa_address) or sa_address, sa_address)

    earliest_la = _pick_earliest_la(la_records)
    la_address = _resolve_la_address(
        record=earliest_la,
        location_points=location_points,
    )
    if la_address:
        return (address_to_name.get(la_address) or la_address, la_address)

    return (UNKNOWN_LOCATION_NAME, UNKNOWN_LOCATION_NAME)


def _resolve_sa_address(sa_records: list[dict[str, Any]]) -> Optional[str]:
    for record in sa_records:
        for field_name in ("area_name_in", "area_name_out"):
            raw = record.get(field_name)
            if not raw:
                continue
            resolved = utils.resolve_area_address(raw)
            if resolved:
                return resolved
    return None


def _pick_earliest_la(la_records: list[dict[str, Any]]) -> Optional[dict[str, Any]]:
    if not la_records:
        return None
    with_time = [record for record in la_records if record.get("first_in") is not None]
    if with_time:
        return min(with_time, key=lambda record: record["first_in"])
    return la_records[0]


def _resolve_la_address(
    *,
    record: Optional[dict[str, Any]],
    location_points: list[dict[str, Any]],
) -> Optional[str]:
    return _nearest_address_by_haversine(record=record, location_points=location_points)


def _nearest_address_by_haversine(
    *,
    record: Optional[dict[str, Any]],
    location_points: list[dict[str, Any]],
) -> Optional[str]:
    if record is None or not location_points:
        return None

    lat = record.get("latitude")
    lon = record.get("longitude")
    if lat is None or lon is None:
        return None

    nearest_address: Optional[str] = None
    nearest_distance = float("inf")
    for point in location_points:
        distance_m = utils.calculate_distance_haversine(
            lat,
            lon,
            point["latitude"],
            point["longitude"],
        )
        if distance_m < nearest_distance:
            nearest_distance = distance_m
            nearest_address = point["address"]

    return nearest_address


def _build_excel_file(
    daily_rows: list[dict[str, Any]],
    summary_rows: list[dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    ws_daily = workbook.active
    ws_daily.title = "Дневной_срез"

    daily_headers = [
        "Дата",
        "День недели",
        "Выходной",
        "Праздник",
        "Рабочий праздник (override)",
        "Кафедра",
        "Локация",
        "Адрес",
        "Студентов в локации",
        "Всего студентов в кафедре за день",
        "Доля локации, %",
    ]
    ws_daily.append(daily_headers)

    for row in daily_rows:
        ws_daily.append(
            [
                row["date"],
                row["day_of_week"],
                row["is_weekend"],
                row["is_public_holiday"],
                (
                    ""
                    if row["is_working_day_override"] is None
                    else row["is_working_day_override"]
                ),
                row["department_name"],
                row["building_name"],
                row["building_address"],
                row["students_count"],
                row["day_total_students"],
                row["pct"],
            ]
        )

    ws_summary = workbook.create_sheet("Свод_локации")
    summary_headers = [
        "Локация",
        "Адрес",
        "Всего посещений",
        "Среднее кол-во студентов в день",
        "Средняя доля, %",
        "Дней с данными по локации",
    ]
    ws_summary.append(summary_headers)

    for row in summary_rows:
        ws_summary.append(
            [
                row["building_name"],
                row["building_address"],
                row["total_visits"],
                row["avg_daily_count"],
                row["avg_pct"],
                row["days_with_location_data"],
            ]
        )

    ws_graph = workbook.create_sheet("Графики")
    ws_graph.append(["Как читать график"])
    ws_graph.append(
        [
            "Столбцы ниже показывают среднюю долю посещаемости по локациям за выбранный период."
        ]
    )
    ws_graph.append([])
    graph_headers = [
        "Локация",
        "Адрес",
        "Средняя доля, %",
        "Среднее кол-во студентов в день",
    ]
    ws_graph.append(graph_headers)

    chart_source_rows = sorted(
        summary_rows,
        key=lambda row: (
            -float(row.get("avg_pct") or 0),
            _building_sort_key(
                str(row.get("building_name") or ""),
                str(row.get("building_address") or ""),
            ),
        ),
    )
    for row in chart_source_rows:
        ws_graph.append(
            [
                row["building_name"],
                row["building_address"],
                row["avg_pct"],
                row["avg_daily_count"],
            ]
        )

    data_rows_count = len(chart_source_rows)
    if data_rows_count > 0:
        chart = BarChart()
        chart.title = "Средняя доля посещаемости по локациям"
        chart.y_axis.title = "Процент"
        chart.x_axis.title = "Локация"
        min_row = 5
        max_row = min_row + data_rows_count
        data_ref = Reference(
            ws_graph,
            min_col=3,
            max_col=3,
            min_row=min_row - 1,
            max_row=max_row,
        )
        category_ref = Reference(
            ws_graph,
            min_col=1,
            max_col=1,
            min_row=min_row,
            max_row=max_row,
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(category_ref)
        chart.height = 11
        chart.width = 24
        ws_graph.add_chart(chart, "F5")

    _style_worksheet(ws_daily)
    _style_worksheet(ws_summary)
    _style_worksheet(ws_graph)
    ws_daily.freeze_panes = "A2"
    ws_summary.freeze_panes = "A2"
    ws_graph.freeze_panes = "A5"

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _style_worksheet(worksheet) -> None:
    header_font = Font(bold=True)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = cell_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = cell_alignment

    for column_cells in worksheet.columns:
        max_width = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_width = max(max_width, len(str(value)))
        worksheet.column_dimensions[column_letter].width = min(max_width + 2, 50)
