import base64
import datetime
import hashlib
import json
import logging
import math
import mimetypes
import os
import re
import time
import zipfile
from collections import Counter, defaultdict
from contextlib import AbstractContextManager, contextmanager
from io import BytesIO
from pathlib import Path
from typing import Any, Generator, List, Optional, Tuple, cast

import monitoring_app.tasks as tasks
from asgiref.sync import async_to_sync
from celery.result import AsyncResult
from channels.layers import get_channel_layer
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, get_user_model, login, logout
from django.core.cache import cache
from django.core.files.base import ContentFile
from django.db import IntegrityError, connection, transaction
from django.db.models import Count, Prefetch, Q
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template import TemplateDoesNotExist
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.generic import View
from drf_yasg import openapi
from drf_yasg.inspectors import SwaggerAutoSchema
from drf_yasg.utils import merge_params, no_body, swagger_auto_schema
from monitoring_app import (
    async_logic,
    attendance_fetcher,
    ml,
    models,
    permissions,
    serializers,
    utils,
)
from monitoring_app.cache_conf import Cache, get_cache
from monitoring_app.lesson_locations_conf import (
    ACCEPTANCE_R_CLUSTER,
    ACCEPTANCE_R_SAME_POINT,
    ACCEPTANCE_R_STANDALONE,
    CLASS_LOCATION_ACCEPTANCE_RADII_CACHE_KEY,
    CLASS_LOCATION_ACCEPTANCE_RADII_CACHE_TTL,
    CLASS_LOCATION_LIST_CACHE_KEY,
    CLASS_LOCATION_LIST_CACHE_TTL,
    CLUSTER_THRESHOLD_M,
    DEFAULT_ACCEPTANCE_RADIUS_M,
    SAME_POINT_THRESHOLD_M,
)
from monitoring_app.services import building_attendance_report
from monitoring_app.signals import invalidate_class_location_cache_impl
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.decorators import (
    api_view,
    authentication_classes,
    permission_classes,
)
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import AllowAny, IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication


def _db_atomic() -> AbstractContextManager[None]:
    """Типизированная обёртка над transaction.atomic() для статического анализа."""
    return cast(AbstractContextManager[None], transaction.atomic())


def _drf_validation_error_text(exc: ValidationError) -> str:
    """Текст для JSON без str(exc) → «[ErrorDetail(string='…')]»."""
    detail = exc.detail
    if isinstance(detail, (list, tuple)):
        return " ".join(str(x) for x in detail).strip()
    if isinstance(detail, dict):
        parts: list[str] = []
        for key, val in detail.items():
            if isinstance(val, (list, tuple)):
                parts.append(f"{key}: {', '.join(str(x) for x in val)}")
            else:
                parts.append(f"{key}: {val}")
        return "; ".join(parts).strip()
    return str(detail).strip()


def _get_manual_parameters_for_inspector(view, method, overrides):
    """manual_parameters из overrides или из _swagger_auto_schema исходной функции (для @api_view)."""
    manual = overrides.get("manual_parameters") or []
    if not manual and method:
        action_method = getattr(view, method.lower(), None)
        if action_method:
            func = getattr(action_method, "__func__", action_method)
            schema = getattr(func, "_swagger_auto_schema", None)
            if isinstance(schema, dict):
                method_data = schema.get(method.lower()) or schema.get(method.upper())
                if isinstance(method_data, dict):
                    manual = method_data.get("manual_parameters") or []
    return manual


class FormOnlySwaggerAutoSchema(SwaggerAutoSchema):
    """Инспектор для form-only ручных параметров.

    Если в manual_parameters есть form-поля, body-параметр удаляется,
    чтобы не падала генерация схемы.
    """

    def add_manual_parameters(self, parameters):
        manual = _get_manual_parameters_for_inspector(
            self.view, self.method, self.overrides
        )
        has_form = any(getattr(p, "in_", None) == openapi.IN_FORM for p in manual)
        if has_form:
            parameters = [
                p for p in parameters if getattr(p, "in_", None) != openapi.IN_BODY
            ]
            return merge_params(parameters, manual)
        return super().add_manual_parameters(parameters)


logger = logging.getLogger(__name__)
lesson_attendance_logger = logging.getLogger("monitoring_app.lesson_attendance")
photo_verdict_logger = logging.getLogger("monitoring_app.photo_verdict")

ExcelRow = Tuple[Any, ...]
User = get_user_model()


@contextmanager
def atomic_block() -> Generator[None, None, None]:
    with transaction.atomic():  # type: ignore[misc]
        yield


LUNCH_BREAK_START = datetime.time(hour=12, minute=55)
LUNCH_BREAK_END = datetime.time(hour=14, minute=5)

CLASS_LOCATION_CACHE_TTL = datetime.timedelta(minutes=60)
DEPARTMENT_CONFIRMATION_CACHE_TTL = 4 * 60 * 60
DEPARTMENT_CONFIRMATION_EPOCH_CACHE_KEY = "department_confirmation_epoch_hour"
DEPARTMENT_CONFIRMATION_EPOCH_TTL = DEPARTMENT_CONFIRMATION_CACHE_TTL + 60 * 60
STAFF_PINS_HEADER_NAME = "X-Staff-Pins"
LESSON_REPORT_CACHE_VERSION = models.LessonAttendance.REPORT_FILTER_CACHE_VERSION
SUSPICIOUS_LOCATION_PATTERNS_EPOCH_CACHE_KEY = "suspicious_location_patterns_epoch"
SUSPICIOUS_LOCATION_PATTERNS_CACHE_VERSION = "v9"
SUSPICIOUS_LOCATION_PATTERNS_CACHE_TTL = 60 * 60
# Кластер GPS-записей за день у одного человека в один дневной якорь.
SUSPICIOUS_LOCATION_PERSON_DAY_RADIUS_M = 10
# Склейка дневных якорей для «person_repeat» .
SUSPICIOUS_LOCATION_PERSON_REPEAT_RADIUS_M = 2
# Кластеризация near-сигналов shared_point внутри одного дня.
SUSPICIOUS_LOCATION_GROUP_CLUSTER_RADIUS_M = 10
SUSPICIOUS_LOCATION_PERSON_REPEAT_MIN_ACTIVE_DAYS = 7
SUSPICIOUS_LOCATION_PERSON_REPEAT_MIN_PCT = 0.70
SUSPICIOUS_LOCATION_REASON_LEGEND = {
    "shared_point": (
        "В эту дату несколько людей переиспользовали одну и ту же "
        "микроточку или микрозону."
    ),
    "person_repeat": (
        "Один и тот же человек слишком стабильно повторяет одну и ту же "
        "микрозону по дням."
    ),
    "multi_day_pattern": "Такой же паттерн повторялся в несколько разных дней.",
}

_STAFF_PIN_WRAPPED_RE = re.compile(r"^S\d+S$")
_STAFF_PIN_NUMERIC_RE = re.compile(r"^\d+$")


def _normalize_staff_pin_token(token: Any) -> Optional[str]:
    raw = str(token or "").strip().upper()
    if not raw:
        return None
    if _STAFF_PIN_NUMERIC_RE.fullmatch(raw):
        return f"S{raw}S"
    if _STAFF_PIN_WRAPPED_RE.fullmatch(raw):
        return raw
    return None


def _parse_staff_pins_header(raw_header_value: Optional[str]) -> List[str]:
    if raw_header_value is None:
        return []

    parsed: List[str] = []
    seen: set[str] = set()
    for token in raw_header_value.split(","):
        normalized = _normalize_staff_pin_token(token)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        parsed.append(normalized)
    return parsed


def _parse_staff_pins_csv(raw_value: Optional[str]) -> List[str]:
    if raw_value is None:
        return []
    return _parse_staff_pins_header(raw_value)


def _parse_query_bool(raw_value: Optional[str]) -> bool:
    if raw_value is None:
        return False
    return str(raw_value).strip().lower() in {"1", "true", "yes", "on"}


def _department_confirmation_hour_bucket(
    now_dt: Optional[datetime.datetime] = None,
) -> str:
    current = now_dt or timezone.localtime()
    fallback = current.strftime("%Y%m%d%H")
    epoch = Cache.get(DEPARTMENT_CONFIRMATION_EPOCH_CACHE_KEY)
    if epoch is not None:
        return str(epoch)
    Cache.set(
        DEPARTMENT_CONFIRMATION_EPOCH_CACHE_KEY,
        fallback,
        DEPARTMENT_CONFIRMATION_EPOCH_TTL,
    )
    return fallback


def _get_suspicious_location_patterns_epoch() -> str:
    epoch = Cache.get(SUSPICIOUS_LOCATION_PATTERNS_EPOCH_CACHE_KEY)
    return str(epoch) if epoch is not None else "0"


def _build_department_confirmation_cache_key(
    *,
    child_department_id: Optional[str],
    use_range: bool,
    date_str: Optional[str],
    date_from_str: Optional[str],
    date_to_str: Optional[str],
    use_staff_pins_mode: bool,
    staff_pins: List[str],
    hour_bucket: str,
) -> str:
    suffix = f"{LESSON_REPORT_CACHE_VERSION}_hour_{hour_bucket}"
    if not use_staff_pins_mode:
        if use_range:
            return (
                f"department_confirmation_{child_department_id}_{date_from_str}_"
                f"{date_to_str}_{suffix}"
            )
        return f"department_confirmation_{child_department_id}_{date_str}_{suffix}"

    sorted_unique_pins = sorted({str(pin).strip().upper() for pin in staff_pins if pin})
    digest_source = ",".join(sorted_unique_pins)
    pins_hash = hashlib.sha1(digest_source.encode("utf-8")).hexdigest()
    if use_range:
        return (
            f"department_confirmation_pins_{pins_hash}_{date_from_str}_"
            f"{date_to_str}_{suffix}"
        )
    return f"department_confirmation_pins_{pins_hash}_{date_str}_{suffix}"


def get_confirmable_threshold(total_group: int) -> int:
    """Возвращает минимальный порог присутствующих в основной локации.

    Правило для малых групп фиксированное и более строгое:
    - ``n == 1`` -> ``1``
    - ``n == 2`` -> ``2``
    - ``n == 3`` -> ``2``
    - ``n == 4`` -> ``3``

    Для групп ``n >= 5`` используется динамический порог:
    ``max(2, ceil(0.20 * n + 0.70 * sqrt(n)))``.

    Args:
        total_group (int): Размер группы (общее число студентов).

    Returns:
        int: Минимальное число отметившихся в главной локации, требуемое
            для подтверждения посещаемости.
    """
    n = max(1, total_group)
    if n == 1:
        return 1
    if n == 2:
        return 2
    if n == 3:
        return 2
    if n == 4:
        return 3
    return max(2, math.ceil(0.20 * n + 0.70 * math.sqrt(n)))


def get_min_leader_share(total_group: int) -> float:
    """Возвращает минимальную долю главной локации среди отметившихся.

    Args:
        total_group (int): Размер группы (общее число студентов).

    Returns:
        float: Минимальная доля в диапазоне ``0..1``:
            ``0.60`` при ``n <= 5``, ``0.55`` при ``n <= 12``, иначе ``0.50``.
    """
    if total_group <= 5:
        return 0.60
    if total_group <= 12:
        return 0.55
    return 0.50


def is_main_location_confirmable(
    leader_count: int,
    total_group: int,
    total_with_attendance: int,
) -> bool:
    """Определяет, можно ли считать главную локацию подтверждающей.

    Подтверждение возможно только при одновременном выполнении условий:
    1. Есть хотя бы одна отметка посещаемости за день.
    2. Число студентов в главной локации не меньше порога
       ``get_confirmable_threshold(total_group)``.
    3. Доля главной локации среди отметившихся не меньше
       ``get_min_leader_share(total_group)``.

    Args:
        leader_count (int): Количество отметившихся в главной локации.
        total_group (int): Размер группы (общее число студентов).
        total_with_attendance (int): Общее число отметившихся за день.

    Returns:
        bool: ``True``, если локация проходит порог и долю; иначе ``False``.
    """
    if total_with_attendance <= 0:
        return False
    threshold = get_confirmable_threshold(total_group)
    if leader_count < threshold:
        return False
    leader_share = leader_count / total_with_attendance
    if leader_share < get_min_leader_share(total_group):
        return False
    return True


CLASS_LOCATION_CACHE = {
    "expires_at": None,
    "kd_tree": None,
    "class_names": [],
    "searcher_payload": [],
    "searcher": None,
}


def get_class_location_cache():
    """
    Кэш локаций: KDTree, LocationSearcher, location_acceptance_radius_m.
    R_loc (60–80 м по умолчанию или acceptance_radius_m из БД) — в Redis и in-memory;
    Celery Beat / warmup_class_location_buffers обновляют.
    """
    now = timezone.now()
    cache_expired = (
        CLASS_LOCATION_CACHE["expires_at"] is None
        or CLASS_LOCATION_CACHE["expires_at"] <= now
    )

    if cache_expired:
        try:
            locations = list(
                models.ClassLocation.objects.only(
                    "id", "name", "latitude", "longitude", "acceptance_radius_m"
                )
            )
            payload = [
                {
                    "name": loc.name,
                    "latitude": loc.latitude,
                    "longitude": loc.longitude,
                }
                for loc in locations
                if loc.latitude is not None and loc.longitude is not None
            ]

            kd_tree = None
            class_names = []
            if payload:
                try:
                    from sklearn.neighbors import KDTree

                    coords = [(item["latitude"], item["longitude"]) for item in payload]
                    kd_tree = KDTree(coords, metric="euclidean")
                    class_names = [item["name"] for item in payload]
                except Exception as exc:
                    logger.warning(f"KDTree initialization failed: {exc}")
                    kd_tree = None
                    class_names = []

            searcher = None
            if payload:
                try:
                    searcher = utils.LocationSearcher(payload)
                except Exception as exc:
                    logger.warning(f"LocationSearcher initialization failed: {exc}")

            locs_with_coords = [
                loc
                for loc in locations
                if loc.latitude is not None and loc.longitude is not None
            ]
            location_acceptance_radius_m = Cache.get(
                CLASS_LOCATION_ACCEPTANCE_RADII_CACHE_KEY
            )
            if location_acceptance_radius_m is None:
                location_acceptance_radius_m = (
                    utils.compute_class_location_acceptance_radii(
                        locs_with_coords,
                        r_same_point=ACCEPTANCE_R_SAME_POINT,
                        r_cluster=ACCEPTANCE_R_CLUSTER,
                        r_standalone=ACCEPTANCE_R_STANDALONE,
                        same_point_threshold=SAME_POINT_THRESHOLD_M,
                        cluster_threshold=CLUSTER_THRESHOLD_M,
                    )
                )
                Cache.set(
                    CLASS_LOCATION_ACCEPTANCE_RADII_CACHE_KEY,
                    location_acceptance_radius_m,
                    CLASS_LOCATION_ACCEPTANCE_RADII_CACHE_TTL,
                )

            CLASS_LOCATION_CACHE.update(
                {
                    "expires_at": now + CLASS_LOCATION_CACHE_TTL,
                    "kd_tree": kd_tree,
                    "class_names": class_names,
                    "searcher_payload": payload,
                    "searcher": searcher,
                    "location_acceptance_radius_m": location_acceptance_radius_m,
                }
            )
        except Exception as exc:
            logger.exception("get_class_location_cache failed: %s", exc)
            CLASS_LOCATION_CACHE["expires_at"] = None
            raise

    return CLASS_LOCATION_CACHE


def _to_date(dt):
    """Normalize date_at from DB (date or datetime) to date.

    Args:
        dt: Value from date_at field (date or datetime).

    Returns:
        date or None: The calendar date, or None if dt is None.
    """
    if dt is None:
        return None
    return dt.date() if hasattr(dt, "date") and callable(getattr(dt, "date")) else dt


def fetch_attendance_by_event_dates(staff_ids, date_from, date_to):
    """Загружает StaffAttendance и LessonAttendance по датам событий одним проходом.

    StaffAttendance.date_at — день выгрузки (обычно календарный день после
    рабочего дня смены). В админке по event_date «2 апреля» ищите строку с
    date_at «3 апреля». LessonAttendance.date_at — календарный день занятия.
    Результаты используются
    в api/staff/{pin}/ и api/attendance/department-confirmation/.

    Args:
        staff_ids: Список id сотрудников (первичные ключи Staff).
        date_from: Начало диапазона дат событий (включительно), date.
        date_to: Конец диапазона дат событий (включительно), date.

    Returns:
        Кортеж (sa_by_event_date, la_by_event_date): каждый элемент — словарь
        {event_date: list[dict]} с записями из .values().
    """
    date_from_plus1 = date_from + datetime.timedelta(days=1)
    date_to_plus1 = date_to + datetime.timedelta(days=1)
    one_day = datetime.timedelta(days=1)

    sa_by_event_date = defaultdict(list)
    for r in models.StaffAttendance.objects.filter(
        staff_id__in=staff_ids,
        date_at__gte=date_from_plus1,
        date_at__lte=date_to_plus1,
    ).values(
        "staff_id",
        "date_at",
        "first_in",
        "last_out",
        "area_name_in",
        "area_name_out",
        "effective_work_seconds",
        "area_sequence",
        "effective_work_intervals",
    ):
        d = _to_date(r["date_at"])
        if d is not None:
            sa_by_event_date[d - one_day].append(r)

    la_by_event_date = defaultdict(list)
    lesson_attendance_qs = models.LessonAttendance.exclude_report_invalid_days(
        models.LessonAttendance.objects.filter(
            staff_id__in=staff_ids,
            date_at__gte=date_from,
            date_at__lte=date_to,
        )
    )
    for r in lesson_attendance_qs.values(
        "staff_id",
        "date_at",
        "first_in",
        "last_out",
        "latitude",
        "longitude",
        "duration_seconds",
    ):
        d = _to_date(r["date_at"])
        if d is not None:
            la_by_event_date[d].append(r)

    return dict(sa_by_event_date), dict(la_by_event_date)


def _sort_datetime_value(value: Any) -> datetime.datetime:
    if isinstance(value, datetime.datetime):
        if timezone.is_naive(value):
            return value.replace(tzinfo=datetime.timezone.utc)
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(
            value,
            datetime.time.min,
            tzinfo=datetime.timezone.utc,
        )
    return datetime.datetime.min.replace(tzinfo=datetime.timezone.utc)


def _cluster_geo_items(
    items: list[dict[str, Any]],
    radius_m: int,
    *,
    lat_key: str = "lat",
    lon_key: str = "lon",
    sort_time_key: str = "sort_time",
    sort_id_key: str = "sort_id",
) -> list[dict[str, Any]]:
    if not items:
        return []

    parents = list(range(len(items)))

    def find(index: int) -> int:
        while parents[index] != index:
            parents[index] = parents[parents[index]]
            index = parents[index]
        return index

    def union(left: int, right: int) -> None:
        left_root = find(left)
        right_root = find(right)
        if left_root != right_root:
            parents[right_root] = left_root

    for left, left_item in enumerate(items):
        for right, right_item in enumerate(items[left + 1 :], start=left + 1):
            distance_m = utils.calculate_distance_haversine(
                float(left_item[lat_key]),
                float(left_item[lon_key]),
                float(right_item[lat_key]),
                float(right_item[lon_key]),
            )
            if distance_m <= radius_m:
                union(left, right)

    groups: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for index, item in enumerate(items):
        groups[find(index)].append(item)

    clusters: list[dict[str, Any]] = []
    for group_items in groups.values():
        group_items.sort(
            key=lambda item: (
                _sort_datetime_value(item.get(sort_time_key)),
                item.get(sort_id_key, 0),
            )
        )
        center_lat = sum(float(item[lat_key]) for item in group_items) / len(
            group_items
        )
        center_lon = sum(float(item[lon_key]) for item in group_items) / len(
            group_items
        )
        clusters.append(
            {
                "items": group_items,
                "count": len(group_items),
                "center_lat": center_lat,
                "center_lon": center_lon,
            }
        )

    clusters.sort(
        key=lambda cluster: (
            -cluster["count"],
            _sort_datetime_value(cluster["items"][0].get(sort_time_key)),
            cluster["items"][0].get(sort_id_key, 0),
            round(cluster["center_lat"], 7),
            round(cluster["center_lon"], 7),
        )
    )
    return clusters


def _get_nearest_class_location_context(
    latitude: float,
    longitude: float,
    class_locations: list[models.ClassLocation],
    location_radii: dict[Any, Any],
) -> dict[str, Any]:
    nearest_location = None
    nearest_distance = float("inf")
    for location in class_locations:
        loc_lat = cast(float, location.latitude)
        loc_lon = cast(float, location.longitude)
        distance_m = utils.calculate_distance_haversine(
            latitude,
            longitude,
            loc_lat,
            loc_lon,
        )
        if distance_m < nearest_distance:
            nearest_distance = distance_m
            nearest_location = location

    if nearest_location is None:
        return {
            "location_name": None,
            "location_address": None,
            "distance_m": None,
            "inside_known_location": False,
        }

    location_radius = utils.get_location_radius(nearest_location, location_radii)
    return {
        "location_name": nearest_location.name,
        "location_address": nearest_location.address,
        "distance_m": round(nearest_distance, 2),
        "inside_known_location": nearest_distance <= location_radius,
    }


def _build_day_anchor(records: list[dict[str, Any]]) -> Optional[dict[str, Any]]:
    if not records:
        return None

    cluster_items = [
        {
            "lat": float(record["latitude"]),
            "lon": float(record["longitude"]),
            "sort_time": record.get("first_in"),
            "sort_id": int(record["id"]),
            "record": record,
            "signature": f"{float(record['latitude']):.7f}|{float(record['longitude']):.7f}",
        }
        for record in records
        if record.get("latitude") is not None and record.get("longitude") is not None
    ]
    if not cluster_items:
        return None

    dominant_cluster = _cluster_geo_items(
        cluster_items,
        SUSPICIOUS_LOCATION_PERSON_DAY_RADIUS_M,
    )[0]

    signature_stats: dict[str, dict[str, Any]] = {}
    for item in dominant_cluster["items"]:
        signature = str(item["signature"])
        stats = signature_stats.setdefault(
            signature,
            {
                "count": 0,
                "sort_time": item.get("sort_time"),
                "sort_id": item.get("sort_id", 0),
            },
        )
        stats["count"] += 1
        if _sort_datetime_value(item.get("sort_time")) < _sort_datetime_value(
            stats.get("sort_time")
        ):
            stats["sort_time"] = item.get("sort_time")
            stats["sort_id"] = item.get("sort_id", 0)
        elif _sort_datetime_value(item.get("sort_time")) == _sort_datetime_value(
            stats.get("sort_time")
        ) and item.get("sort_id", 0) < stats.get("sort_id", 0):
            stats["sort_id"] = item.get("sort_id", 0)

    dominant_signature = sorted(
        signature_stats.items(),
        key=lambda item: (
            -item[1]["count"],
            _sort_datetime_value(item[1].get("sort_time")),
            item[1].get("sort_id", 0),
            item[0],
        ),
    )[0][0]
    exact_latitude, exact_longitude = map(float, dominant_signature.split("|"))

    attendance_ids = sorted(
        {int(item["record"]["id"]) for item in dominant_cluster["items"]}
    )
    first_in = dominant_cluster["items"][0]["record"].get("first_in")
    base_record = dominant_cluster["items"][0]["record"]
    return {
        "staff_id": int(base_record["staff_id"]),
        "date": base_record["date_at"],
        "exact_signature": dominant_signature,
        "exact_lat": exact_latitude,
        "exact_lon": exact_longitude,
        "center_lat": dominant_cluster["center_lat"],
        "center_lon": dominant_cluster["center_lon"],
        "attendance_ids": attendance_ids,
        "first_in": first_in,
    }


def _build_person_repeat_profile(
    anchors: list[dict[str, Any]],
    class_locations: list[models.ClassLocation],
    location_radii: dict[Any, Any],
) -> Optional[dict[str, Any]]:
    if not anchors:
        return None

    repeat_items = [
        {
            "lat": float(anchor["center_lat"]),
            "lon": float(anchor["center_lon"]),
            "sort_time": anchor["date"],
            "sort_id": (
                int(anchor["attendance_ids"][0]) if anchor["attendance_ids"] else 0
            ),
            "anchor": anchor,
        }
        for anchor in anchors
    ]
    repeat_cluster = _cluster_geo_items(
        repeat_items,
        SUSPICIOUS_LOCATION_PERSON_REPEAT_RADIUS_M,
    )[0]
    repeat_dates = sorted(
        {item["anchor"]["date"].isoformat() for item in repeat_cluster["items"]}
    )
    attendance_ids = sorted(
        {
            attendance_id
            for item in repeat_cluster["items"]
            for attendance_id in item["anchor"]["attendance_ids"]
        }
    )
    location_context = _get_nearest_class_location_context(
        repeat_cluster["center_lat"],
        repeat_cluster["center_lon"],
        class_locations,
        location_radii,
    )
    active_days = len(anchors)
    repeat_days = len(repeat_dates)
    repeat_pct = round(100.0 * repeat_days / active_days, 1) if active_days else 0.0
    return {
        "active_days": active_days,
        "repeat_days": repeat_days,
        "repeat_pct": repeat_pct,
        "center_lat": repeat_cluster["center_lat"],
        "center_lon": repeat_cluster["center_lon"],
        "dates": repeat_dates,
        "attendance_ids": attendance_ids,
        "location_name": location_context["location_name"],
        "location_address": location_context["location_address"],
        "distance_m": location_context["distance_m"],
        "inside_known_location": location_context["inside_known_location"],
        "is_actionable": (
            active_days >= SUSPICIOUS_LOCATION_PERSON_REPEAT_MIN_ACTIVE_DAYS
            and (repeat_days / active_days) >= SUSPICIOUS_LOCATION_PERSON_REPEAT_MIN_PCT
        ),
    }


def _build_suspicious_location_patterns_cache_key(
    *,
    date_from_str: str,
    date_to_str: str,
    child_department_id: Optional[str],
    staff_pins: list[str],
    include_medium: bool,
) -> str:
    epoch = _get_suspicious_location_patterns_epoch()
    suffix = (
        f"{SUSPICIOUS_LOCATION_PATTERNS_CACHE_VERSION}_"
        f"{LESSON_REPORT_CACHE_VERSION}_{epoch}_{date_from_str}_{date_to_str}"
    )
    if staff_pins:
        normalized_pins = sorted({pin for pin in staff_pins if pin})
        pins_hash = hashlib.sha1(",".join(normalized_pins).encode("utf-8")).hexdigest()
        return (
            "suspicious_location_patterns_"
            f"pins_{pins_hash}_{int(include_medium)}_{suffix}"
        )
    return (
        "suspicious_location_patterns_"
        f"dept_{child_department_id}_{int(include_medium)}_{suffix}"
    )


def _sort_reason_codes(reason_codes: set[str]) -> list[str]:
    priority = {
        "shared_point": 0,
        "person_repeat": 1,
        "multi_day_pattern": 2,
    }
    return sorted(reason_codes, key=lambda code: (priority.get(code, 99), code))


def _suspicious_candidate_priority(
    candidate: dict[str, Any],
    repeat_pct: float,
    repeat_days: int,
) -> tuple[Any, ...]:
    pattern_priority = {
        "shared_point_exact": 0,
        "shared_point_near": 1,
        "person_repeat": 2,
    }
    return (
        candidate["severity_rank"],
        candidate["group_days"],
        candidate.get("staff_count", 0),
        repeat_pct,
        repeat_days,
        -pattern_priority.get(candidate["pattern_type"], 99),
        len(candidate["dates"]),
        len(candidate["reason"]),
        round(float(candidate["lat"]), 7),
        round(float(candidate["lon"]), 7),
    )


def _resolve_la_location(lat, lon, kd_tree, class_names):
    """Определяет название локации по координатам через KD-дерево.

    Args:
        lat: Широта (float или None).
        lon: Долгота (float или None).
        kd_tree: KDTree для поиска по координатам или None.
        class_names: Список названий локаций по индексам дерева.

    Returns:
        Название локации (str) или None при отсутствии данных или ошибке.
    """
    if not kd_tree or not class_names or lat is None or lon is None:
        return None
    try:
        _distances, indices = kd_tree.query([[lat, lon]], k=1)
        if hasattr(indices, "ndim") and indices.ndim > 1:
            indices = indices.flatten()
        return class_names[int(indices[0])] if len(indices) > 0 else None
    except Exception as e:
        logger.warning("Error resolving LA location: %s", e)
        return None


def _merge_attendance_for_date(sa_records, la_records, kd_tree, class_names):
    """Объединяет StaffAttendance и LessonAttendance за одну дату событий.

    Границы first_in/last_out берутся по минимуму/максимуму из SA и LA; при
    совпадении приоритет у SA. Зоны для LA определяются по координатам через
    kd_tree. effective_work_seconds считается объединением интервалов SA и LA
    с вычитанием пересечений (merge_work_intervals_to_total_seconds). area_sequence
    возвращается только когда обе границы из SA.

    Args:
        sa_records: Список словарей SA (staff_id, first_in, last_out,
            area_name_in, area_name_out, effective_work_seconds, effective_work_intervals).
        la_records: Список словарей LA (staff_id, first_in, last_out,
            latitude, longitude, duration_seconds).
        kd_tree: KDTree для поиска локации по координатам или None.
        class_names: Список названий локаций по индексам дерева.

    Returns:
        Словарь: first_in, last_out, area_name_in, area_name_out,
        first_in_source, last_out_source, effective_work_seconds, area_sequence.
    """
    combined: dict[str, Any] = {
        "first_in": None,
        "last_out": None,
        "area_name_in": None,
        "area_name_out": None,
        "first_in_source": None,
        "last_out_source": None,
        "effective_work_seconds": None,
        "area_sequence": None,
    }
    if sa_records:
        first_sa = sa_records[0]
        combined["effective_work_seconds"] = first_sa.get("effective_work_seconds")
        combined["area_sequence"] = first_sa.get("area_sequence")
    for r in sa_records:
        if r.get("first_in") and (
            combined["first_in"] is None or r["first_in"] < combined["first_in"]
        ):
            combined["first_in"] = r["first_in"]
            combined["first_in_source"] = "staff_attendance"
            if r.get("area_name_in"):
                combined["area_name_in"] = (
                    utils.resolve_area_address(r["area_name_in"]) or r["area_name_in"]
                )
        if r.get("last_out") and (
            combined["last_out"] is None or r["last_out"] > combined["last_out"]
        ):
            combined["last_out"] = r["last_out"]
            combined["last_out_source"] = "staff_attendance"
            if r.get("area_name_out"):
                combined["area_name_out"] = (
                    utils.resolve_area_address(r["area_name_out"]) or r["area_name_out"]
                )

    earliest_la = None
    latest_la = None
    for r in la_records:
        if r.get("first_in"):
            if earliest_la is None:
                earliest_la = r
            elif r["first_in"] < earliest_la["first_in"]:
                earliest_la = r
        if r.get("last_out"):
            if latest_la is None:
                latest_la = r
            elif r["last_out"] > latest_la["last_out"]:
                latest_la = r

    if earliest_la is not None and (
        combined["first_in"] is None or earliest_la["first_in"] < combined["first_in"]
    ):
        combined["first_in"] = earliest_la["first_in"]
        combined["first_in_source"] = "lesson_attendance"
        name = _resolve_la_location(
            earliest_la.get("latitude"),
            earliest_la.get("longitude"),
            kd_tree,
            class_names,
        )
        if name:
            combined["area_name_in"] = name
    if latest_la is not None and (
        combined["last_out"] is None or latest_la["last_out"] > combined["last_out"]
    ):
        combined["last_out"] = latest_la["last_out"]
        combined["last_out_source"] = "lesson_attendance"
        name = _resolve_la_location(
            latest_la.get("latitude"), latest_la.get("longitude"), kd_tree, class_names
        )
        if name:
            combined["area_name_out"] = name

    intervals: List[Tuple[datetime.datetime, datetime.datetime]] = []
    if sa_records:
        for raw in sa_records[0].get("effective_work_intervals") or []:
            try:
                s = raw.get("start") and datetime.datetime.fromisoformat(
                    raw["start"].replace("Z", "+00:00")
                )
                e = raw.get("end") and datetime.datetime.fromisoformat(
                    raw["end"].replace("Z", "+00:00")
                )
                if s is not None and e is not None and e > s:
                    intervals.append((s, e))
            except (ValueError, TypeError, AttributeError):
                continue
    for la in la_records:
        fi, lo = la.get("first_in"), la.get("last_out")
        if fi is not None and lo is not None and lo > fi:
            intervals.append((fi, lo))
    total_effective = utils.merge_work_intervals_to_total_seconds(intervals)
    combined["effective_work_seconds"] = (
        total_effective if total_effective > 0 else None
    )
    if (
        combined["first_in_source"] != "staff_attendance"
        or combined["last_out_source"] != "staff_attendance"
    ):
        combined["area_sequence"] = None

    return combined


def calculate_effective_minutes_with_lunch(first_in, last_out):
    """Считает минуты между первым входом и последним выходом (fallback без событий СКУД).

    Обед не вычитается: обеденный перерыв учитывается только при наличии события
    выхода через турникет в окне 12:55–14:05 (логика в фетчере — effective_work_seconds).
    Без событий считаем, что сотрудник не выходил (обедал на месте / работал).
    """
    if not first_in or not last_out:
        return 0.0

    current_tz = timezone.get_current_timezone()
    start = timezone.localtime(first_in, current_tz)
    end = timezone.localtime(last_out, current_tz)

    if end <= start:
        return 0.0

    return (end - start).total_seconds() / 60


class StaffAttendancePagination(PageNumberPagination):
    page_size = 5000
    page_size_query_param = "page_size"
    max_page_size = 20000


token_param_config = openapi.Parameter(
    "Authorization",
    in_=openapi.IN_HEADER,
    description="Token [access_token]",
    type=openapi.TYPE_STRING,
)


@permission_classes([AllowAny])
@never_cache
def home(request):
    response = render(
        request,
        "index.html",
    )
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


def _get_frontend_app_version_path() -> Path:
    frontend_dir = Path(getattr(settings, "FRONTEND_DIR"))
    return frontend_dir / "dist" / "app-version.json"


@permission_classes([AllowAny])
@never_cache
def app_version(request):
    app_version_path = _get_frontend_app_version_path()
    try:
        payload = json.loads(app_version_path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        response = JsonResponse(
            {"error": "Build version metadata is unavailable."},
            status=503,
        )
    except (OSError, json.JSONDecodeError):
        logger.exception("Failed to read frontend app-version.json")
        response = JsonResponse(
            {"error": "Build version metadata is unavailable."},
            status=503,
        )
    else:
        response = JsonResponse(payload)

    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


@permission_classes([AllowAny])
@never_cache
def react_app(request):
    try:
        response = render(request, "index.html")
    except TemplateDoesNotExist:
        logger.exception("React app template index.html not found")
        return HttpResponse(
            b"Error loading React app",
            status=500,
            content_type="text/plain; charset=utf-8",
        )
    except Exception:
        logger.exception("React app render failed")
        return HttpResponse(
            b"Error loading React app",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


class StaffAttendanceStatsView(APIView):
    """
    Представление для получения статистики о посещаемости персонала.

    Это представление фильтрует данные, чтобы включать только сотрудников, относящихся к отделу с ID AUP.

    Параметры запроса:
        date (str): Дата, для которой запрашивается статистика посещаемости, в формате 'YYYY-MM-DD'. По умолчанию используется текущая дата.
        pin (str): ПИН-код сотрудника для получения конкретной статистики посещаемости.

    Возвращает:
        JSON-ответ, содержащий:
            department_name (str): Название отдела.
            total_staff_count (int): Общее количество сотрудников.
            present_staff_count (int): Количество присутствующих сотрудников.
            absent_staff_count (int): Количество отсутствующих сотрудников.
            present_between_9_to_18 (int): Количество сотрудников, присутствующих с 08:00 до 19:00.
            present_data (list): Список словарей с информацией о присутствующих сотрудниках, включая ПИН, имя, количество минут присутствия и индивидуальный процент.
            absent_data (list): Список словарей с информацией об отсутствующих сотрудниках, включая ПИН и имя.
            data_for_date (str): Дата, за которую предоставлены данные, в формате 'YYYY-MM-DD'.

    Примеры:
        GET /api/attendance/stats/?date=2024-07-20
        GET /api/attendance/stats/?pin=123456

    Примечание:
        Ответ кэшируется на 1 час, а информация о государственных праздниках кэшируется на 1 минуту.
    """

    permission_classes = [permissions.IsAuthenticatedOrAPIKey]

    @swagger_auto_schema(
        operation_summary="Получить список людей об их присутствии",
        operation_description="View для получения статистики о посещаемости персонала.",
        tags=["Attendance & Statistics"],
        responses={
            200: openapi.Response(
                description="Successful response",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "department_name": openapi.Schema(type=openapi.TYPE_STRING),
                        "total_staff_count": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "present_staff_count": openapi.Schema(
                            type=openapi.TYPE_INTEGER
                        ),
                        "absent_staff_count": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "present_between_9_to_18": openapi.Schema(
                            type=openapi.TYPE_INTEGER
                        ),
                        "present_data": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "staff_pin": openapi.Schema(
                                        type=openapi.TYPE_STRING
                                    ),
                                    "name": openapi.Schema(type=openapi.TYPE_STRING),
                                    "minutes_present": openapi.Schema(
                                        type=openapi.TYPE_NUMBER
                                    ),
                                    "individual_percentage": openapi.Schema(
                                        type=openapi.TYPE_NUMBER
                                    ),
                                },
                            ),
                        ),
                        "absent_data": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "staff_pin": openapi.TYPE_STRING,
                                    "name": openapi.Schema(type=openapi.TYPE_STRING),
                                },
                            ),
                        ),
                        "data_for_date": openapi.Schema(type=openapi.TYPE_STRING),
                    },
                ),
            ),
            404: "Not Found",
            500: "Internal Server Error",
        },
        manual_parameters=[
            openapi.Parameter(
                name="X-API-KEY",
                in_=openapi.IN_HEADER,
                type=openapi.TYPE_STRING,
                required=False,
                description="API ключ для аутентификации (альтернатива JWT токену).",
            ),
            openapi.Parameter(
                "date",
                openapi.IN_QUERY,
                description="Date in 'YYYY-MM-DD' format.",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "pin",
                openapi.IN_QUERY,
                description="Staff PIN.",
                type=openapi.TYPE_STRING,
            ),
        ],
    )
    def get(self, request):
        logger.info("Received request for staff attendance stats.")

        date_param = request.query_params.get(
            "date", timezone.now().date().strftime("%Y-%m-%d")
        )
        date_param = datetime.datetime.strptime(date_param, "%Y-%m-%d").date()
        pin_param = request.query_params.get("pin", None)

        logger.debug(f"Parsed date_param: {date_param}, pin_param: {pin_param}")

        try:
            target_date = self.get_last_working_day(date_param)
            next_date = target_date + datetime.timedelta(days=1)
            cache_key = (
                f"staff_attendance_stats_{LESSON_REPORT_CACHE_VERSION}_"
                f"{target_date}_{pin_param}"
            )

            logger.debug(f"Generated cache_key: {cache_key}")

            cached_data = get_cache(
                cache_key,
                query=lambda: self.query_data(target_date, next_date, pin_param),
                timeout=6 * 3600,
            )

            logger.info("Successfully retrieved staff attendance data.")
            response = Response(cached_data)
            response["Cache-Control"] = "public, max-age=21600"
            return response

        except Exception as e:
            logger.error(f"Error while processing request: {str(e)}")
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def get_last_working_day(self, date: datetime.date) -> datetime.date:
        """
        Определение последнего рабочего дня, с учетом выходных и государственных праздников.

        Args:
            date (datetime.date): Дата, для которой нужно найти последний рабочий день.

        Returns:
            datetime.date: Последний рабочий день.
        """
        logger.debug(f"Calculating last working day for date: {date}")

        holidays = (
            get_cache(
                "public_holidays",
                query=lambda: list(models.PublicHoliday.objects.all()),
                timeout=10 * 6,
            )
            or []
        )
        holiday_dates = {
            holiday.date: holiday.is_working_day for holiday in holidays if holiday
        }

        while date.weekday() >= 5 or (
            date in holiday_dates and not holiday_dates[date]
        ):
            logger.debug(f"{date} is not a working day, moving to previous day.")
            date -= datetime.timedelta(days=1)

        logger.debug(f"Last working day determined: {date}")
        return date

    def query_data(
        self,
        target_date: datetime.date,
        _next_date: datetime.date,
        pin_param: str | None,
    ) -> dict:
        """
        Запрашивает данные по целевой дате и отделу (или сотруднику).

        Args:
            target_date (datetime.date): Целевая дата.
            next_date (datetime.date): Следующая дата после целевой.
            pin_param (str, optional): ID родительского или дочернего отдела.

        Returns:
            dict: Данные о сотрудниках, их посещаемости и статистике.
        """
        logger.info(
            f"Querying data for target_date: {target_date}, pin_param: {pin_param}"
        )

        department_name = "Unknown Department"
        staff_queryset = None

        parent_department = (
            models.ParentDepartment.objects.filter(id=pin_param)
            .only("id", "name")
            .first()
        )
        child_department = (
            models.ChildDepartment.objects.filter(id=pin_param)
            .only("id", "name")
            .first()
        )

        match (parent_department, child_department):
            case (parent, None) if parent:
                staff_queryset = models.Staff.objects.filter(
                    department__parent_id=parent.id
                ).select_related("department")
                department_name = parent.name
                logger.info(
                    "StaffAttendanceStatsView: pin_param=%s → ParentDepartment id=%s, name=%s",
                    pin_param,
                    parent.id,
                    department_name,
                )
            case (None, child) if child:
                staff_queryset = models.Staff.objects.filter(
                    department=child
                ).select_related("department")
                department_name = child.name
                logger.info(
                    "StaffAttendanceStatsView: pin_param=%s → ChildDepartment id=%s, name=%s",
                    pin_param,
                    child.id,
                    department_name,
                )
            case _:
                staff_queryset = models.Staff.objects.filter(
                    Q(department__parent__name__icontains="AUP")
                    | Q(department__parent__name__icontains="АУП")
                ).select_related("department__parent")
                logger.info(
                    "StaffAttendanceStatsView: pin_param=%s → fallback AUP/АУП branch",
                    pin_param,
                )

        target_date_for_filter = target_date + datetime.timedelta(days=1)
        staff_queryset = (
            staff_queryset.select_related("department__parent")
            .prefetch_related(
                Prefetch(
                    "positions",
                    queryset=models.Position.objects.only("name"),
                )
            )
            .only(
                "id",
                "pin",
                "name",
                "surname",
                "department_id",
                "department__name",
                "department__parent__name",
            )
        )
        staff_members = list(staff_queryset)
        if department_name == "Unknown Department" and staff_members:
            parent = getattr(staff_members[0].department, "parent", None)
            if parent is not None:
                department_name = getattr(parent, "name", None) or department_name

        if not staff_members:
            return {
                "department_name": department_name,
                "total_staff_count": 0,
                "present_staff_count": 0,
                "absent_staff_count": 0,
                "present_between_9_to_18": 0,
                "present_data": [],
                "absent_data": [],
                "data_for_date": target_date.strftime("%Y-%m-%d"),
            }

        staff_ids = [s.id for s in staff_members]
        staff_id_to_pin = {s.id: s.pin for s in staff_members}

        staff_attendance_queryset = (
            models.StaffAttendance.objects.filter(
                date_at=target_date_for_filter,
                staff_id__in=staff_ids,
                first_in__isnull=False,
            )
            .select_related("staff")
            .only(
                "first_in",
                "last_out",
                "effective_work_seconds",
                "staff_id",
                "staff__pin",
                "staff__name",
                "staff__surname",
            )
        )
        present_staff_records = list(staff_attendance_queryset)
        attendance_by_pin = {
            record.staff.pin: record for record in present_staff_records
        }

        lesson_staff_ids = set(
            models.LessonAttendance.exclude_report_invalid_days(
                models.LessonAttendance.objects.filter(
                    date_at=target_date,
                    staff_id__in=staff_ids,
                )
            )
            .values_list("staff_id", flat=True)
            .distinct()
        )
        present_pins_from_lessons = {
            staff_id_to_pin[sid] for sid in lesson_staff_ids if sid in staff_id_to_pin
        }

        logger.info(
            "StaffAttendanceStatsView: target_date=%s, target_date_for_filter(SA)=%s, "
            "staff_count=%s, staff_ids_sample=%s, "
            "StaffAttendance(present)=%s, LessonAttendance(staff_ids)=%s, present_pins_from_lessons=%s",
            target_date,
            target_date_for_filter,
            len(staff_members),
            staff_ids[:5] if len(staff_ids) > 5 else staff_ids,
            len(present_staff_records),
            len(lesson_staff_ids),
            len(present_pins_from_lessons),
        )

        total_staff_count = len(staff_members)

        present_between_9_to_18 = 0
        for record in present_staff_records:
            first_in_time = record.first_in.time()
            if datetime.time(8, 0) <= first_in_time <= datetime.time(19, 0):
                present_between_9_to_18 += 1

        employee_position_name = "Сотрудник"
        employee_pins = {
            s.pin
            for s in staff_members
            if any(p.name == employee_position_name for p in s.positions.all())
        }
        logger.info(
            "StaffAttendanceStatsView: employee_pins(count)=%s, non_employee(count)=%s",
            len(employee_pins),
            total_staff_count - len(employee_pins),
        )
        present_data, absent_data = self.get_attendance_data(
            staff_members,
            attendance_by_pin,
            present_pins_from_lessons=present_pins_from_lessons,
            employee_pins=employee_pins,
        )
        absent_staff_count = total_staff_count - len(present_data)

        present_from_sa = sum(
            1 for p in present_data if attendance_by_pin.get(p["staff_pin"])
        )
        present_from_la_only = len(present_data) - present_from_sa
        logger.info(
            "StaffAttendanceStatsView: present_data=%s (from SA=%s, from LA only=%s), absent_data=%s, department=%s",
            len(present_data),
            present_from_sa,
            present_from_la_only,
            len(absent_data),
            department_name,
        )

        return {
            "department_name": department_name,
            "total_staff_count": total_staff_count,
            "present_staff_count": len(present_data),
            "absent_staff_count": absent_staff_count,
            "present_between_9_to_18": present_between_9_to_18,
            "present_data": present_data,
            "absent_data": absent_data,
            "data_for_date": target_date.strftime("%Y-%m-%d"),
        }

    def get_attendance_data(
        self,
        staff_members,
        attendance_by_pin,
        present_pins_from_lessons=None,
        employee_pins=None,
    ):
        """Формирует present_data и absent_data.

        Сотрудники (должность «Сотрудник»): только StaffAttendance, процент по времени.
        Студенты и др.: присутствие по StaffAttendance или LessonAttendance (удалённые локации),
        процент — факт присутствия (100%).
        """
        if present_pins_from_lessons is None:
            present_pins_from_lessons = set()
        if employee_pins is None:
            employee_pins = {
                s.pin
                for s in staff_members
                if any(p.name == "Сотрудник" for p in s.positions.all())
            }
        logger.debug("Generating attendance data.")
        present_data = []
        absent_data = []
        total_minutes = 8 * 60
        for staff in staff_members:
            is_employee = staff.pin in employee_pins
            attendance = attendance_by_pin.get(staff.pin)

            if attendance:
                if getattr(attendance, "effective_work_seconds", None) is not None:
                    minutes_present = attendance.effective_work_seconds / 60.0
                elif attendance.first_in and attendance.last_out:
                    minutes_present = (
                        attendance.last_out - attendance.first_in
                    ).total_seconds() / 60
                else:
                    minutes_present = 0
                if is_employee:
                    individual_percentage = (minutes_present / total_minutes) * 100
                else:
                    individual_percentage = 100.0
                present_data.append(
                    {
                        "staff_pin": staff.pin,
                        "name": f"{staff.surname} {staff.name}",
                        "minutes_present": round(minutes_present, 2),
                        "individual_percentage": round(individual_percentage, 2),
                    }
                )
                continue

            if not is_employee and staff.pin in present_pins_from_lessons:
                present_data.append(
                    {
                        "staff_pin": staff.pin,
                        "name": f"{staff.surname} {staff.name}",
                        "minutes_present": 0,
                        "individual_percentage": 100.0,
                    }
                )
                continue

            absent_data.append(
                {
                    "staff_pin": staff.pin,
                    "name": f"{staff.surname} {staff.name}",
                }
            )

        logger.info("Attendance data generation complete.")
        return present_data, absent_data


@swagger_auto_schema(
    method="get",
    operation_summary="Получить данные локаций для отображения на карте",
    operation_description=(
        "Эндпоинт для получения данных локаций с информацией о посещениях для заданной даты."
        " Опционально можно получить данные о сотрудниках, если задан параметр `employees=true`."
    ),
    tags=["Locations"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
        openapi.Parameter(
            name="date_at",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            format=openapi.FORMAT_DATE,
            description="Дата для фильтрации посещений (формат YYYY-MM-DD). Если не указана, используется текущая дата.",
            required=False,
        ),
        openapi.Parameter(
            name="employees",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_BOOLEAN,
            description="Флаг для включения данных о сотрудниках (true/false). По умолчанию false.",
            required=False,
        ),
    ],
    responses={
        200: openapi.Response(
            description="Успешный ответ с данными локаций для отображения на карте.",
            schema=openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "name": openapi.Schema(
                            type=openapi.TYPE_STRING, description="Название локации"
                        ),
                        "address": openapi.Schema(
                            type=openapi.TYPE_STRING, description="Адрес локации"
                        ),
                        "lat": openapi.Schema(
                            type=openapi.TYPE_NUMBER,
                            format=openapi.FORMAT_FLOAT,
                            description="Широта",
                        ),
                        "lng": openapi.Schema(
                            type=openapi.TYPE_NUMBER,
                            format=openapi.FORMAT_FLOAT,
                            description="Долгота",
                        ),
                        "employees": openapi.Schema(
                            type=openapi.TYPE_INTEGER,
                            description="Количество посещений или сотрудников (если указано employees=true)",
                            default=0,
                        ),
                    },
                ),
            ),
        ),
        500: openapi.Response(description="Внутренняя ошибка сервера"),
    },
)
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def map_location(request):
    """
    Возвращает данные о локациях и количестве сотрудников и студентов на основе турникетов и занятий за указанную дату.

    Args:
        request (HttpRequest): HTTP-запрос с параметром `date_at` для фильтрации данных и параметром `employees` для включения данных о сотрудниках.

    Returns:
        JsonResponse: JSON-ответ с данными о локациях, если запрос успешен, либо сообщение об ошибке.
    """
    try:
        date_at_str = request.GET.get("date_at", None)
        employees_required = request.GET.get("employees", "false").lower() == "true"

        if date_at_str:
            try:
                date_at = datetime.datetime.strptime(date_at_str, "%Y-%m-%d").date()
            except ValueError:
                logger.warning(f"Incorrect date format: {date_at_str}")
                return Response(
                    {"error": "Incorrect Date format, please use YYYY-MM-DD."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            date_at = datetime.datetime.now().date()

        logger.info(f"Using date: {date_at}")

        cache_key = (
            f"map_location_{LESSON_REPORT_CACHE_VERSION}_{date_at}_"
            f"{employees_required}"
        )

        def generate_map_data():
            locations = models.ClassLocation.objects.only(
                "address", "name", "latitude", "longitude"
            )
            return utils.generate_map_data(
                locations,
                date_at,
                search_staff_attendance=employees_required,
                filter_empty=not employees_required,
            )

        result = get_cache(
            cache_key,
            query=generate_map_data,
            timeout=5 * 60,
        )

        logger.info(f"Generated map data with employees: {result}")
        return Response(result, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Critical error in map_location: {str(e)}", exc_info=True)
        return Response(
            {"error": "A critical error occurred. Please try again later."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


def _build_one_day_confirmation(
    _dept: Optional[models.ChildDepartment],
    target_date: datetime.date,
    staff_list: list,
    location_searcher=None,
    name_to_address=None,
    address_to_name=None,
) -> dict[str, Any]:
    """Формирует ответ подтверждения посещаемости по отделу за один день (одиночный запрос по дате).

    Собирает отметки из StaffAttendance и LessonAttendance, строит список локаций
    (сортировка: по количеству отметившихся DESC, затем address ASC, name ASC).
    Основная локация дня — первая в списке. Подтверждение (confirmed=True) только
    у студентов из основной локации и только если она проходит динамический порог
    и минимальную долю среди отметившихся. pct считается от total_with_attendance.
    При отсутствии данных за день всем ставится waiting=True.

    Args:
        _dept: Экземпляр ChildDepartment (не используется, оставлен для совместимости API).
        target_date: Дата события (рабочий день).
        staff_list: Список сотрудников (объекты с id, pin).
        location_searcher: Опционально LocationSearcher для привязки LA к локациям.
        name_to_address: Опционально словарь название -> адрес.
        address_to_name: Опционально словарь адрес -> название.

    Returns:
        Словарь с ключами: date, data_available, total, locations (name, address, count, pct, pins_short),
        by_pin_short (confirmed, waiting, location, location_address, first_in по pin_short).
    """
    date_str = target_date.strftime("%Y-%m-%d")
    staff_ids = [s.id for s in staff_list]
    data_insert_date = target_date + datetime.timedelta(days=1)

    sa_qs = models.StaffAttendance.objects.filter(
        staff_id__in=staff_ids,
        date_at=data_insert_date,
        first_in__isnull=False,
    ).values("staff_id", "first_in", "area_name_in")
    sa_records = list(sa_qs)
    staff_with_sa = {r["staff_id"] for r in sa_records}

    la_qs = models.LessonAttendance.exclude_report_invalid_days(
        models.LessonAttendance.objects.filter(
            staff_id__in=staff_ids,
            date_at=target_date,
        )
    )
    la_qs = la_qs.exclude(staff_id__in=staff_with_sa).values(
        "staff_id", "first_in", "latitude", "longitude"
    )
    la_records = list(la_qs)

    if location_searcher is None or name_to_address is None or address_to_name is None:
        location_cache = get_class_location_cache()
        location_searcher = location_cache.get("searcher")
        if location_searcher is None and location_cache.get("searcher_payload"):
            try:
                location_searcher = utils.LocationSearcher(
                    location_cache["searcher_payload"]
                )
            except Exception:
                location_searcher = None
        class_locations = list(
            models.ClassLocation.objects.only("name", "address").values(
                "name", "address"
            )
        )
        name_to_address = {loc["name"]: loc["address"] for loc in class_locations}
        address_to_name = {loc["address"]: loc["name"] for loc in class_locations}

    staff_to_location: dict[int, str] = {}
    staff_first_in: dict[int, datetime.datetime] = {}
    for sa in sa_records:
        addr = utils.resolve_area_address(sa.get("area_name_in"))
        if addr:
            staff_to_location[sa["staff_id"]] = addr
        fi = sa.get("first_in")
        if fi:
            staff_first_in[sa["staff_id"]] = fi

    la_by_staff: dict[int, list[dict]] = defaultdict(list)
    for la in la_records:
        la_by_staff[la["staff_id"]].append(la)
    for sid, records in la_by_staff.items():
        earliest = min(
            (r for r in records if r.get("first_in")),
            key=lambda r: r["first_in"],
            default=None,
        )
        if (
            earliest
            and location_searcher
            and earliest.get("latitude") is not None
            and earliest.get("longitude") is not None
        ):
            nearest_name = location_searcher.find_nearest(
                earliest["latitude"], earliest["longitude"], radius=200
            )
            if nearest_name != "Unknown Area" and nearest_name in name_to_address:
                staff_to_location[sid] = name_to_address[nearest_name]
                staff_first_in[sid] = earliest["first_in"]

    location_counts: dict[str, list[str]] = defaultdict(list)
    for staff in staff_list:
        addr = staff_to_location.get(staff.id)
        if addr:
            location_counts[addr].append(staff.pin)

    data_available = bool(staff_to_location) or bool(la_records)
    if not data_available:
        data_available = (
            models.StaffAttendance.objects.filter(date_at=data_insert_date)
            .only("id")
            .exists()
        )

    total_with_attendance = sum(len(pins) for _, pins in location_counts.items())
    total_group = len(staff_list)
    locations_sorted = sorted(
        location_counts.items(),
        key=lambda x: (-len(x[1]), x[0], (address_to_name.get(x[0]) or x[0])),
    )
    main_address = locations_sorted[0][0] if locations_sorted else None
    first_count = len(locations_sorted[0][1]) if locations_sorted else 0
    confirmable_main_address = (
        main_address
        if main_address
        and is_main_location_confirmable(
            first_count, total_group, total_with_attendance
        )
        else None
    )

    locations_payload = []
    for addr, pins in locations_sorted:
        pct = (
            round(100.0 * len(pins) / total_with_attendance, 2)
            if total_with_attendance
            else 0
        )
        pins_short = [utils.pin_to_external_format(p) for p in pins]
        locations_payload.append(
            {
                "name": address_to_name.get(addr) or addr,
                "address": addr,
                "count": len(pins),
                "pct": pct,
                "pins_short": pins_short,
            }
        )

    by_pin_short: dict[str, dict[str, Any]] = {}
    for s in staff_list:
        addr = staff_to_location.get(s.id)
        if not data_available:
            confirmed = False
            waiting = True
        else:
            waiting = False
            if addr is None:
                confirmed = False
            elif confirmable_main_address is None:
                confirmed = False
            elif addr == confirmable_main_address:
                confirmed = True
            else:
                confirmed = False

        fi = staff_first_in.get(s.id)
        first_in_iso = (
            fi.astimezone(timezone.get_current_timezone()).isoformat() if fi else None
        )
        pin_short = utils.pin_to_external_format(s.pin)
        location_name = (address_to_name.get(addr) or addr) if addr else None
        by_pin_short[pin_short] = {
            "confirmed": confirmed,
            "waiting": waiting,
            "location": location_name,
            "location_address": addr if addr else None,
            "first_in": first_in_iso,
        }

    return {
        "date": date_str,
        "data_available": data_available,
        "total": total_group,
        "locations": locations_payload,
        "by_pin_short": by_pin_short,
    }


def _build_one_day_from_records(
    target_date: datetime.date,
    staff_list: list,
    sa_records: list,
    la_records: list,
    data_available: bool | None,
    location_searcher,
    name_to_address: dict,
    address_to_name: dict,
    dates_with_any_sa: set | None = None,
) -> dict[str, Any]:
    """Формирует ответ подтверждения за один день по уже загруженным SA/LA (без обращений к БД).

    Логика совпадает с _build_one_day_confirmation: основная локация — первая после
    сортировки (count DESC, address ASC, name ASC); подтверждение только для неё
    и только при прохождении динамического порога и минимальной доли. pct от числа
    отметившихся. Если передан dates_with_any_sa, data_available вычисляется как
    (есть отметки за день) или (data_insert_date в dates_with_any_sa).

    Args:
        target_date: Дата события (рабочий день).
        staff_list: Список сотрудников (id, pin).
        sa_records: Записи StaffAttendance за этот день (date_at = target_date + 1).
        la_records: Записи LessonAttendance за этот день.
        data_available: Используется только при dates_with_any_sa is None.
        location_searcher: LocationSearcher для привязки координат LA к локациям.
        name_to_address: Словарь название -> адрес.
        address_to_name: Словарь адрес -> название.
        dates_with_any_sa: Множество дат выгрузки, по которым есть хотя бы одна SA (опционально).

    Returns:
        Словарь: date, data_available, total, locations, by_pin_short.
    """
    date_str = target_date.strftime("%Y-%m-%d")
    staff_with_sa = frozenset(r["staff_id"] for r in sa_records if r.get("first_in"))

    staff_to_location: dict[int, str] = {}
    staff_first_in: dict[int, datetime.datetime] = {}
    for sa in sa_records:
        addr = utils.resolve_area_address(sa.get("area_name_in"))
        if addr:
            staff_to_location[sa["staff_id"]] = addr
        fi = sa.get("first_in")
        if fi:
            staff_first_in[sa["staff_id"]] = fi

    la_by_staff: dict[int, list[dict]] = defaultdict(list)
    for la in la_records:
        if la["staff_id"] not in staff_with_sa:
            la_by_staff[la["staff_id"]].append(la)
    for sid, records in la_by_staff.items():
        earliest = min(
            (r for r in records if r.get("first_in")),
            key=lambda r: r["first_in"],
            default=None,
        )
        if (
            earliest
            and location_searcher
            and earliest.get("latitude") is not None
            and earliest.get("longitude") is not None
        ):
            nearest_name = location_searcher.find_nearest(
                earliest["latitude"], earliest["longitude"], radius=200
            )
            if nearest_name != "Unknown Area" and nearest_name in name_to_address:
                staff_to_location[sid] = name_to_address[nearest_name]
                staff_first_in[sid] = earliest["first_in"]

    location_counts: dict[str, list[str]] = defaultdict(list)
    for staff in staff_list:
        addr = staff_to_location.get(staff.id)
        if addr:
            location_counts[addr].append(staff.pin)

    if dates_with_any_sa is not None:
        data_insert_date = target_date + datetime.timedelta(days=1)
        data_available = (
            bool(staff_to_location)
            or bool(la_records)
            or (data_insert_date in dates_with_any_sa)
        )
    elif data_available is None:
        data_available = False

    total_with_attendance = sum(len(pins) for _, pins in location_counts.items())
    total_group = len(staff_list)
    locations_sorted = sorted(
        location_counts.items(),
        key=lambda x: (-len(x[1]), x[0], (address_to_name.get(x[0]) or x[0])),
    )
    main_address = locations_sorted[0][0] if locations_sorted else None
    first_count = len(locations_sorted[0][1]) if locations_sorted else 0
    confirmable_main_address = (
        main_address
        if main_address
        and is_main_location_confirmable(
            first_count, total_group, total_with_attendance
        )
        else None
    )

    locations_payload = []
    for addr, pins in locations_sorted:
        pct = (
            round(100.0 * len(pins) / total_with_attendance, 2)
            if total_with_attendance
            else 0
        )
        pins_short = [utils.pin_to_external_format(p) for p in pins]
        locations_payload.append(
            {
                "name": address_to_name.get(addr) or addr,
                "address": addr,
                "count": len(pins),
                "pct": pct,
                "pins_short": pins_short,
            }
        )

    by_pin_short: dict[str, dict[str, Any]] = {}
    for s in staff_list:
        addr = staff_to_location.get(s.id)
        if not data_available:
            confirmed = False
            waiting = True
        else:
            waiting = False
            if addr is None:
                confirmed = False
            elif confirmable_main_address is None:
                confirmed = False
            elif addr == confirmable_main_address:
                confirmed = True
            else:
                confirmed = False

        fi = staff_first_in.get(s.id)
        first_in_iso = (
            fi.astimezone(timezone.get_current_timezone()).isoformat() if fi else None
        )
        pin_short = utils.pin_to_external_format(s.pin)
        location_name = (address_to_name.get(addr) or addr) if addr else None
        by_pin_short[pin_short] = {
            "confirmed": confirmed,
            "waiting": waiting,
            "location": location_name,
            "location_address": addr if addr else None,
            "first_in": first_in_iso,
        }

    return {
        "date": date_str,
        "data_available": data_available,
        "total": total_group,
        "locations": locations_payload,
        "by_pin_short": by_pin_short,
    }


@swagger_auto_schema(
    method="GET",
    operation_summary="Подтверждение оценок по посещаемости (отдел)",
    operation_description=(
        "Эндпоинт для интеграции с системой оценок: по отделу (группе) и дате возвращает, "
        "кто из сотрудников был на «основной» локации (где большинство отметившихся), а кто — нет.\n\n"
        "**Логика:**\n"
        "- Собираются отметки посещаемости за день (StaffAttendance и LessonAttendance).\n"
        "- Основная локация за день — та, где отметилось больше всего человек.\n"
        "- Проценты (pct) по локациям считаются от **числа отметившихся** (распределение пришедших).\n"
        "- Основная локация дня — первая в отсортированном списке (count DESC, address ASC, name ASC). "
        "Подтверждение только для неё и только если она проходит порог по размеру группы "
        "(n=1: 1, n=2: 2, n=3: 2, n=4: 3, иначе max(2, ceil(0.20*n + 0.70*sqrt(n)))) "
        "и минимальную долю среди отметившихся (60% при n<=5, 55% при n<=12, 50% иначе). Остальные локации — неподтверждение.\n"
        "- У каждой локации в ответе обязательно есть **name** и **address**.\n"
        "- В **by_pin_short**: **confirmed** = true только у тех, кто был в основной локации; "
        "из другой локации или без отметки — confirmed = false. **waiting** = данные ещё не выгружены.\n\n"
        "**Поиск по сотруднику:** ключи by_pin_short — PIN без обёртки S/T (S9614S → 9614). "
        "Оценка разрешена, если у pin_short **confirmed** = true."
    ),
    tags=["Attendance & Statistics"],
    manual_parameters=[
        openapi.Parameter(
            "child_department_id",
            openapi.IN_QUERY,
            description="ID подразделения (ChildDepartment), например группа «ЖМ 724 0/б»",
            type=openapi.TYPE_STRING,
            required=True,
        ),
        openapi.Parameter(
            STAFF_PINS_HEADER_NAME,
            openapi.IN_HEADER,
            description=(
                "Опциональный pins-mode: CSV список PIN в формате S{id}S "
                "(например: S9614S,S30108S). При передаче этого header расчёт "
                "идёт строго по переданным PIN и не зависит от существования "
                "child_department_id."
            ),
            type=openapi.TYPE_STRING,
            required=False,
        ),
        openapi.Parameter(
            "date",
            openapi.IN_QUERY,
            description="Одна дата YYYY-MM-DD. Используется либо date, либо пара date_from и date_to.",
            type=openapi.TYPE_STRING,
            required=False,
        ),
        openapi.Parameter(
            "date_from",
            openapi.IN_QUERY,
            description="Начало диапазона дат YYYY-MM-DD (включительно). В паре с date_to — один запрос на весь период (в т.ч. полгода).",
            type=openapi.TYPE_STRING,
            required=False,
        ),
        openapi.Parameter(
            "date_to",
            openapi.IN_QUERY,
            description="Конец диапазона дат YYYY-MM-DD (включительно). В паре с date_from.",
            type=openapi.TYPE_STRING,
            required=False,
        ),
    ],
    responses={
        200: openapi.Response(
            description="Успешный ответ: одна дата — те же поля (date, locations, by_pin_short и т.д.); диапазон — child_department_id, child_department_name, results: [{ date, data_available, total, locations, by_pin_short }, ...].",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "date": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Дата (при запросе одной даты).",
                    ),
                    "child_department_id": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="ID отдела (ChildDepartment).",
                    ),
                    "child_department_name": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Название отдела (группы).",
                    ),
                    "data_available": openapi.Schema(
                        type=openapi.TYPE_BOOLEAN,
                        description="True, если данные посещаемости за эту дату уже выгружены (при одной дате).",
                    ),
                    "total": openapi.Schema(
                        type=openapi.TYPE_INTEGER,
                        description="Число сотрудников в отделе (при одной дате).",
                    ),
                    "locations": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        description="Локации по убыванию pct (при одной дате).",
                        items=openapi.Schema(type=openapi.TYPE_OBJECT),
                    ),
                    "by_pin_short": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        description="По pin_short: confirmed, waiting, location, first_in (при одной дате).",
                        additional_properties=openapi.Schema(type=openapi.TYPE_OBJECT),
                    ),
                    "results": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        description="При запросе по диапазону: массив объектов по дням (date, data_available, total, locations, by_pin_short).",
                        items=openapi.Schema(type=openapi.TYPE_OBJECT),
                    ),
                },
            ),
        ),
        400: openapi.Response(
            description="Ошибка: не указаны date или date_from/date_to; неверный формат даты; date_from > date_to.",
        ),
        404: openapi.Response(
            description="Отдел с указанным child_department_id не найден.",
        ),
    },
)
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def department_attendance_confirmation(request):
    """Определяет подтверждение оценок по посещаемости для отдела на дату или диапазон дат.

    По переданному отделу (child_department_id) и дате (или диапазону date_from/date_to) возвращает:
    - список локаций с долей сотрудников (pct) и списком pins_short;
    - объект by_pin_short для поиска по pin_short (формат внешней системы: без S/T)
      с полями confirmed, waiting, location, first_in.

    Один день: query date=YYYY-MM-DD. Ответ как раньше (одна дата).
    Диапазон: query date_from=YYYY-MM-DD и date_to=YYYY-MM-DD (включительно). Ответ — results: [{ date, ... }, ...].

    Args:
        request: GET. Обязателен child_department_id. Либо date, либо оба date_from и date_to.
    """
    child_department_id = request.GET.get("child_department_id")
    date_str = request.GET.get("date")
    date_from_str = request.GET.get("date_from")
    date_to_str = request.GET.get("date_to")
    raw_staff_pins_header = request.headers.get(STAFF_PINS_HEADER_NAME)
    use_staff_pins_mode = raw_staff_pins_header is not None
    staff_pins = _parse_staff_pins_header(raw_staff_pins_header)

    if not child_department_id and not use_staff_pins_mode:
        return Response(
            {"error": "child_department_id is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if not child_department_id:
        child_department_id = "pins_mode"

    use_range = bool(date_from_str and date_to_str)
    if not use_range and not date_str:
        return Response(
            {"error": "Either date or both date_from and date_to are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if use_range and date_str:
        return Response(
            {"error": "Use either date or date_from/date_to, not both"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if use_range:
        try:
            date_from = datetime.datetime.strptime(date_from_str, "%Y-%m-%d").date()
            date_to = datetime.datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"error": "Invalid date format for date_from/date_to, use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if date_from > date_to:
            return Response(
                {"error": "date_from must be <= date_to"},
                status=status.HTTP_400_BAD_REQUEST,
            )
    else:
        try:
            target_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"error": "Invalid date format, use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    dept = None
    if not use_staff_pins_mode:
        dept = models.ChildDepartment.objects.filter(id=child_department_id).first()
        if not dept:
            return Response(
                {"error": f"ChildDepartment {child_department_id} not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
    else:
        dept = models.ChildDepartment.objects.filter(id=child_department_id).first()

    dept_name = dept.name if dept else ""

    cache_key = _build_department_confirmation_cache_key(
        child_department_id=child_department_id,
        use_range=use_range,
        date_str=date_str,
        date_from_str=date_from_str,
        date_to_str=date_to_str,
        use_staff_pins_mode=use_staff_pins_mode,
        staff_pins=staff_pins,
        hour_bucket=_department_confirmation_hour_bucket(),
    )
    cached = Cache.get(cache_key)
    if cached is not None:
        return Response(cached, status=status.HTTP_200_OK)

    if use_staff_pins_mode:
        staff_list = list(
            models.Staff.objects.filter(pin__in=staff_pins).only(
                "id", "pin", "name", "surname"
            )
        )
    else:
        staff_list = list(
            models.Staff.objects.filter(department_id=child_department_id).only(
                "id", "pin", "name", "surname"
            )
        )
    if not staff_list and not use_range:
        payload = {
            "date": date_str,
            "child_department_id": child_department_id,
            "child_department_name": dept_name,
            "data_available": False,
            "total": 0,
            "locations": [],
            "by_pin_short": {},
        }
        Cache.set(cache_key, payload, DEPARTMENT_CONFIRMATION_CACHE_TTL)
        return Response(payload, status=status.HTTP_200_OK)
    if not staff_list and use_range:
        results = []
        d = date_from
        while d <= date_to:
            results.append(
                {
                    "date": d.strftime("%Y-%m-%d"),
                    "data_available": False,
                    "total": 0,
                    "locations": [],
                    "by_pin_short": {},
                }
            )
            d += datetime.timedelta(days=1)
        payload = {
            "child_department_id": child_department_id,
            "child_department_name": dept_name,
            "results": results,
        }
        Cache.set(cache_key, payload, DEPARTMENT_CONFIRMATION_CACHE_TTL)
        return Response(payload, status=status.HTTP_200_OK)

    if use_range:
        staff_ids = [s.id for s in staff_list]
        sa_by_event_date, la_by_event_date = fetch_attendance_by_event_dates(
            staff_ids, date_from, date_to
        )
        dates_with_any_sa = {ed + datetime.timedelta(days=1) for ed in sa_by_event_date}

        location_cache = get_class_location_cache()
        location_searcher = location_cache.get("searcher")
        if location_searcher is None and location_cache.get("searcher_payload"):
            try:
                location_searcher = utils.LocationSearcher(
                    location_cache["searcher_payload"]
                )
            except Exception:
                location_searcher = None
        class_locations = list(
            models.ClassLocation.objects.only("name", "address").values(
                "name", "address"
            )
        )
        name_to_address = {loc["name"]: loc["address"] for loc in class_locations}
        address_to_name = {loc["address"]: loc["name"] for loc in class_locations}

        results = []
        d = date_from
        while d <= date_to:
            sa_records = sa_by_event_date.get(d, [])
            la_records = la_by_event_date.get(d, [])
            day_payload = _build_one_day_from_records(
                d,
                staff_list,
                sa_records,
                la_records,
                data_available=None,
                location_searcher=location_searcher,
                name_to_address=name_to_address,
                address_to_name=address_to_name,
                dates_with_any_sa=dates_with_any_sa,
            )
            results.append(day_payload)
            d += datetime.timedelta(days=1)

        payload = {
            "child_department_id": child_department_id,
            "child_department_name": dept_name,
            "results": results,
        }
        Cache.set(cache_key, payload, DEPARTMENT_CONFIRMATION_CACHE_TTL)
        return Response(payload, status=status.HTTP_200_OK)

    day_payload = _build_one_day_confirmation(dept, target_date, staff_list)
    payload = {
        "date": day_payload["date"],
        "child_department_id": child_department_id,
        "child_department_name": dept_name,
        "data_available": day_payload["data_available"],
        "total": day_payload["total"],
        "locations": day_payload["locations"],
        "by_pin_short": day_payload["by_pin_short"],
    }
    Cache.set(cache_key, payload, DEPARTMENT_CONFIRMATION_CACHE_TTL)
    return Response(payload, status=status.HTTP_200_OK)


@swagger_auto_schema(
    method="GET",
    operation_summary="Короткий список координатных обманщиков",
    operation_description=(
        "Возвращает короткий список сотрудников с punishable-паттернами по "
        "LessonAttendance: shared micro-point паттернами по датам и устойчивой "
        "персональной повторяемостью."
    ),
    tags=["Attendance & Statistics"],
    manual_parameters=[
        openapi.Parameter(
            "child_department_id",
            openapi.IN_QUERY,
            description="ID подразделения. Нужен, если не передан staff_pins.",
            type=openapi.TYPE_STRING,
            required=False,
        ),
        openapi.Parameter(
            "staff_pins",
            openapi.IN_QUERY,
            description=(
                "CSV список PIN. Поддерживает short-формат (25812) и wrapped "
                "формат (S25812S)."
            ),
            type=openapi.TYPE_STRING,
            required=False,
        ),
        openapi.Parameter(
            STAFF_PINS_HEADER_NAME,
            openapi.IN_HEADER,
            description="Fallback CSV список PIN, если query staff_pins не передан.",
            type=openapi.TYPE_STRING,
            required=False,
        ),
        openapi.Parameter(
            "date_from",
            openapi.IN_QUERY,
            description="Начало диапазона YYYY-MM-DD.",
            type=openapi.TYPE_STRING,
            required=True,
        ),
        openapi.Parameter(
            "date_to",
            openapi.IN_QUERY,
            description="Конец диапазона YYYY-MM-DD.",
            type=openapi.TYPE_STRING,
            required=True,
        ),
        openapi.Parameter(
            "include_medium",
            openapi.IN_QUERY,
            description="Если 1/true, возвращает также medium-кейсы person_repeat.",
            type=openapi.TYPE_BOOLEAN,
            required=False,
        ),
    ],
    responses={
        200: openapi.Response(
            description="Короткий список подозрительных сотрудников по датам и PIN.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "summary": openapi.Schema(type=openapi.TYPE_OBJECT),
                    "reasonLegend": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        additional_properties=openapi.Schema(type=openapi.TYPE_STRING),
                    ),
                    "datesByDate": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        additional_properties=openapi.Schema(type=openapi.TYPE_OBJECT),
                    ),
                    "usersByPin": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        additional_properties=openapi.Schema(type=openapi.TYPE_OBJECT),
                    ),
                },
            ),
        ),
        400: openapi.Response(
            description="Ошибка параметров: нет дат или нет источника выборки.",
        ),
        404: openapi.Response(description="Подразделение не найдено."),
    },
)
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def suspicious_location_patterns(request):
    date_from_str = str(request.query_params.get("date_from") or "").strip()
    date_to_str = str(request.query_params.get("date_to") or "").strip()
    child_department_id = str(
        request.query_params.get("child_department_id") or ""
    ).strip()
    query_staff_pins = _parse_staff_pins_csv(request.query_params.get("staff_pins"))
    header_staff_pins = _parse_staff_pins_header(
        request.headers.get(STAFF_PINS_HEADER_NAME)
    )
    staff_pins = query_staff_pins or header_staff_pins
    include_medium = _parse_query_bool(request.query_params.get("include_medium"))

    if not date_from_str or not date_to_str:
        return Response(
            {"error": "date_from and date_to are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        date_from = datetime.datetime.strptime(date_from_str, "%Y-%m-%d").date()
        date_to = datetime.datetime.strptime(date_to_str, "%Y-%m-%d").date()
    except ValueError:
        return Response(
            {"error": "Invalid date format, use YYYY-MM-DD"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if date_from > date_to:
        return Response(
            {"error": "date_from must be <= date_to"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if not child_department_id and not staff_pins:
        return Response(
            {"error": "Either child_department_id or staff_pins is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if staff_pins:
        staff_list = list(
            models.Staff.objects.filter(pin__in=staff_pins)
            .only("id", "pin", "name", "surname")
            .order_by("pin")
        )
    else:
        department = models.ChildDepartment.objects.filter(
            id=child_department_id
        ).first()
        if department is None:
            return Response(
                {"error": f"ChildDepartment {child_department_id} not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        staff_list = list(
            models.Staff.objects.filter(department_id=child_department_id)
            .only("id", "pin", "name", "surname")
            .order_by("pin")
        )

    cache_key = _build_suspicious_location_patterns_cache_key(
        date_from_str=date_from_str,
        date_to_str=date_to_str,
        child_department_id=child_department_id or None,
        staff_pins=staff_pins,
        include_medium=include_medium,
    )
    cached_payload = Cache.get(cache_key)
    if cached_payload is not None:
        return Response(cached_payload, status=status.HTTP_200_OK)

    if not staff_list:
        payload = {
            "summary": {
                "staffAnalyzed": 0,
                "staffFlagged": 0,
                "datesFlagged": 0,
            },
            "reasonLegend": dict(SUSPICIOUS_LOCATION_REASON_LEGEND),
            "datesByDate": {},
            "usersByPin": {},
        }
        Cache.set(cache_key, payload, SUSPICIOUS_LOCATION_PATTERNS_CACHE_TTL)
        return Response(payload, status=status.HTTP_200_OK)

    staff_by_id = {staff.id: staff for staff in staff_list}
    staff_ids = list(staff_by_id.keys())
    class_locations = list(
        models.ClassLocation.objects.filter(
            latitude__isnull=False,
            longitude__isnull=False,
        ).only("id", "name", "address", "latitude", "longitude", "acceptance_radius_m")
    )
    location_radii = get_class_location_cache().get("location_acceptance_radius_m", {})

    lesson_rows = list(
        models.LessonAttendance.exclude_report_invalid_days(
            models.LessonAttendance.objects.filter(
                staff_id__in=staff_ids,
                date_at__gte=date_from,
                date_at__lte=date_to,
                latitude__isnull=False,
                longitude__isnull=False,
            )
        )
        .values(
            "id",
            "staff_id",
            "date_at",
            "first_in",
            "latitude",
            "longitude",
        )
        .order_by("date_at", "staff_id", "first_in", "id")
    )

    lesson_rows_by_staff_day: dict[tuple[int, datetime.date], list[dict[str, Any]]] = (
        defaultdict(list)
    )
    for row in lesson_rows:
        lesson_rows_by_staff_day[(int(row["staff_id"]), row["date_at"])].append(row)

    anchors_by_staff: dict[int, list[dict[str, Any]]] = defaultdict(list)
    anchors_by_date: dict[datetime.date, list[dict[str, Any]]] = defaultdict(list)
    for (_staff_id, _date), rows_for_day in lesson_rows_by_staff_day.items():
        anchor = _build_day_anchor(rows_for_day)
        if anchor is None:
            continue
        anchors_by_staff[int(anchor["staff_id"])].append(anchor)
        anchors_by_date[anchor["date"]].append(anchor)

    person_profiles: dict[int, dict[str, Any]] = {}
    for staff_id, anchors in anchors_by_staff.items():
        anchors.sort(
            key=lambda anchor: (
                anchor["date"],
                _sort_datetime_value(anchor.get("first_in")),
                anchor["attendance_ids"][0] if anchor["attendance_ids"] else 0,
            )
        )
        profile = _build_person_repeat_profile(
            anchors,
            class_locations,
            location_radii,
        )
        if profile is not None:
            person_profiles[staff_id] = profile

    daily_exact_signals: list[dict[str, Any]] = []
    daily_near_signals: list[dict[str, Any]] = []
    for day, day_anchors in anchors_by_date.items():
        total_with_attendance = len({anchor["staff_id"] for anchor in day_anchors})
        if total_with_attendance == 0:
            continue
        min_group_count = max(2, get_confirmable_threshold(total_with_attendance))
        min_group_share = get_min_leader_share(total_with_attendance)

        exact_buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for anchor in day_anchors:
            exact_buckets[str(anchor["exact_signature"])].append(anchor)
        for signature, bucket in exact_buckets.items():
            distinct_staff_count = len({anchor["staff_id"] for anchor in bucket})
            share = distinct_staff_count / total_with_attendance
            if distinct_staff_count < min_group_count or share < min_group_share:
                continue

            exact_latitude, exact_longitude = map(float, signature.split("|"))
            location_context = _get_nearest_class_location_context(
                exact_latitude,
                exact_longitude,
                class_locations,
                location_radii,
            )
            exact_bucket_attendance_ids: dict[int, set[int]] = defaultdict(set)
            exact_bucket_staff_dates: dict[int, set[str]] = defaultdict(set)
            for anchor in bucket:
                staff_id = int(anchor["staff_id"])
                exact_bucket_attendance_ids[staff_id].update(anchor["attendance_ids"])
                exact_bucket_staff_dates[staff_id].add(day.isoformat())

            daily_exact_signals.append(
                {
                    "signal_type": "exact",
                    "exact_signature": signature,
                    "center_lat": exact_latitude,
                    "center_lon": exact_longitude,
                    "date_obj": day,
                    "date": day.isoformat(),
                    "staff_count": distinct_staff_count,
                    "staff_ids": sorted(exact_bucket_attendance_ids.keys()),
                    "attendance_ids_by_staff": exact_bucket_attendance_ids,
                    "staff_dates": exact_bucket_staff_dates,
                    "location_context": location_context,
                }
            )

        near_items = [
            {
                "lat": float(anchor["center_lat"]),
                "lon": float(anchor["center_lon"]),
                "sort_time": day,
                "sort_id": (
                    anchor["attendance_ids"][0] if anchor["attendance_ids"] else 0
                ),
                "anchor": anchor,
            }
            for anchor in day_anchors
        ]
        for near_cluster in _cluster_geo_items(
            near_items,
            SUSPICIOUS_LOCATION_GROUP_CLUSTER_RADIUS_M,
        ):
            cluster_anchors = [item["anchor"] for item in near_cluster["items"]]
            distinct_staff_count = len(
                {anchor["staff_id"] for anchor in cluster_anchors}
            )
            share = distinct_staff_count / total_with_attendance
            if distinct_staff_count < min_group_count or share < min_group_share:
                continue

            location_context = _get_nearest_class_location_context(
                near_cluster["center_lat"],
                near_cluster["center_lon"],
                class_locations,
                location_radii,
            )

            near_bucket_attendance_ids = defaultdict(set)
            near_bucket_staff_dates = defaultdict(set)
            for anchor in cluster_anchors:
                staff_id = int(anchor["staff_id"])
                near_bucket_attendance_ids[staff_id].update(anchor["attendance_ids"])
                near_bucket_staff_dates[staff_id].add(day.isoformat())

            daily_near_signals.append(
                {
                    "signal_type": "near",
                    "center_lat": near_cluster["center_lat"],
                    "center_lon": near_cluster["center_lon"],
                    "date_obj": day,
                    "date": day.isoformat(),
                    "staff_count": distinct_staff_count,
                    "staff_ids": sorted(near_bucket_attendance_ids.keys()),
                    "attendance_ids_by_staff": near_bucket_attendance_ids,
                    "staff_dates": near_bucket_staff_dates,
                    "location_context": location_context,
                }
            )

    exact_signals_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for signal in daily_exact_signals:
        exact_signals_map[str(signal["exact_signature"])].append(signal)
    aggregated_exact_patterns: list[dict[str, Any]] = []
    for signature, signals_by_signature in exact_signals_map.items():
        exact_latitude, exact_longitude = map(float, signature.split("|"))
        aggregated_exact_ids: dict[int, set[int]] = defaultdict(set)
        aggregated_exact_dates: dict[int, set[str]] = defaultdict(set)
        for signal in signals_by_signature:
            for staff_id, attendance_ids in signal["attendance_ids_by_staff"].items():
                aggregated_exact_ids[int(staff_id)].update(attendance_ids)
            for staff_id, dates_for_staff in signal["staff_dates"].items():
                aggregated_exact_dates[int(staff_id)].update(dates_for_staff)
        aggregated_exact_patterns.append(
            {
                "pattern_type": "group_exact",
                "center_lat": exact_latitude,
                "center_lon": exact_longitude,
                "location_context": signals_by_signature[0]["location_context"],
                "attendance_ids_by_staff": aggregated_exact_ids,
                "staff_dates": aggregated_exact_dates,
                "signals": sorted(
                    signals_by_signature,
                    key=lambda signal: (
                        signal["date"],
                        -signal["staff_count"],
                        round(float(signal["center_lat"]), 7),
                        round(float(signal["center_lon"]), 7),
                    ),
                ),
                "dates": sorted(
                    {
                        date_value
                        for signal in signals_by_signature
                        for date_value in [signal["date"]]
                    }
                ),
            }
        )

    near_signal_items = [
        {
            "lat": float(signal["center_lat"]),
            "lon": float(signal["center_lon"]),
            "sort_time": signal["date_obj"],
            "sort_id": index,
            "signal": signal,
        }
        for index, signal in enumerate(daily_near_signals)
    ]
    aggregated_near_patterns: list[dict[str, Any]] = []
    for near_pattern_cluster in _cluster_geo_items(
        near_signal_items,
        SUSPICIOUS_LOCATION_GROUP_CLUSTER_RADIUS_M,
    ):
        signals = [item["signal"] for item in near_pattern_cluster["items"]]
        aggregated_near_ids: dict[int, set[int]] = defaultdict(set)
        aggregated_near_dates: dict[int, set[str]] = defaultdict(set)
        for signal in signals:
            for staff_id, attendance_ids in signal["attendance_ids_by_staff"].items():
                aggregated_near_ids[int(staff_id)].update(attendance_ids)
            for staff_id, dates_for_staff in signal["staff_dates"].items():
                aggregated_near_dates[int(staff_id)].update(dates_for_staff)
        aggregated_near_patterns.append(
            {
                "pattern_type": "shared_point_near",
                "center_lat": near_pattern_cluster["center_lat"],
                "center_lon": near_pattern_cluster["center_lon"],
                "location_context": _get_nearest_class_location_context(
                    near_pattern_cluster["center_lat"],
                    near_pattern_cluster["center_lon"],
                    class_locations,
                    location_radii,
                ),
                "attendance_ids_by_staff": aggregated_near_ids,
                "staff_dates": aggregated_near_dates,
                "signals": sorted(
                    signals,
                    key=lambda signal: (
                        signal["date"],
                        -signal["staff_count"],
                        round(float(signal["center_lat"]), 7),
                        round(float(signal["center_lon"]), 7),
                    ),
                ),
                "dates": sorted({signal["date"] for signal in signals}),
            }
        )

    severity_rank = {"medium": 1, "high": 2, "critical": 3}
    candidates_by_staff: dict[int, list[dict[str, Any]]] = defaultdict(list)

    for pattern in aggregated_exact_patterns:
        for signal in pattern["signals"]:
            date_value = str(signal["date"])
            for staff_id in signal["staff_ids"]:
                dates_for_staff = sorted(
                    pattern["staff_dates"].get(int(staff_id), set())
                )
                group_days = len(dates_for_staff)
                severity = "critical" if group_days >= 2 else "high"
                reason_codes = {"shared_point"}
                if group_days >= 2:
                    reason_codes.add("multi_day_pattern")
                candidates_by_staff[int(staff_id)].append(
                    {
                        "pattern_type": "shared_point_exact",
                        "severity": severity,
                        "severity_rank": severity_rank[severity],
                        "group_days": group_days,
                        "staff_count": int(signal["staff_count"]),
                        "dates": [date_value],
                        "lat": round(float(signal["center_lat"]), 7),
                        "lon": round(float(signal["center_lon"]), 7),
                        "location_name": signal["location_context"].get(
                            "location_name"
                        ),
                        "reason": _sort_reason_codes(reason_codes),
                    }
                )

    for pattern in aggregated_near_patterns:
        for signal in pattern["signals"]:
            date_value = str(signal["date"])
            for staff_id in signal["staff_ids"]:
                dates_for_staff = sorted(
                    pattern["staff_dates"].get(int(staff_id), set())
                )
                group_days = len(dates_for_staff)
                severity = "critical" if group_days >= 2 else "high"
                reason_codes = {"shared_point"}
                if group_days >= 2:
                    reason_codes.add("multi_day_pattern")
                candidates_by_staff[int(staff_id)].append(
                    {
                        "pattern_type": "shared_point_near",
                        "severity": severity,
                        "severity_rank": severity_rank[severity],
                        "group_days": group_days,
                        "staff_count": int(signal["staff_count"]),
                        "dates": [date_value],
                        "lat": round(float(signal["center_lat"]), 7),
                        "lon": round(float(signal["center_lon"]), 7),
                        "location_name": signal["location_context"].get(
                            "location_name"
                        ),
                        "reason": _sort_reason_codes(reason_codes),
                    }
                )

    if include_medium:
        for staff_id, profile in person_profiles.items():
            if not profile["is_actionable"]:
                continue
            reason_codes = {"person_repeat"}
            if len(profile["dates"]) >= 2:
                reason_codes.add("multi_day_pattern")
            candidates_by_staff[staff_id].append(
                {
                    "pattern_type": "person_repeat",
                    "severity": "medium",
                    "severity_rank": severity_rank["medium"],
                    "group_days": 0,
                    "staff_count": 1,
                    "dates": list(profile["dates"]),
                    "lat": round(float(profile["center_lat"]), 7),
                    "lon": round(float(profile["center_lon"]), 7),
                    "location_name": profile["location_name"],
                    "reason": _sort_reason_codes(reason_codes),
                }
            )

    best_candidates_by_staff_date: dict[tuple[int, str], dict[str, Any]] = {}
    for staff_id, candidates in candidates_by_staff.items():
        if not candidates:
            continue
        repeat_pct = person_profiles.get(staff_id, {}).get("repeat_pct", 0.0)
        repeat_days = person_profiles.get(staff_id, {}).get("repeat_days", 0)
        for candidate in candidates:
            candidate_priority = _suspicious_candidate_priority(
                candidate,
                repeat_pct,
                repeat_days,
            )
            for date_value in candidate["dates"]:
                key = (staff_id, str(date_value))
                existing = best_candidates_by_staff_date.get(key)
                if existing is None or candidate_priority > existing["priority"]:
                    best_candidates_by_staff_date[key] = {
                        "candidate": candidate,
                        "priority": candidate_priority,
                    }

    date_groups: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    user_groups: dict[str, dict[str, Any]] = {}
    highest_severity_by_pin: dict[str, int] = {}

    for (staff_id, date_value), selected in sorted(
        best_candidates_by_staff_date.items()
    ):
        candidate = selected["candidate"]
        staff = staff_by_id[staff_id]
        pin_short = utils.pin_to_external_format(staff.pin)
        person_profile = person_profiles.get(staff_id)
        fio = f"{staff.surname} {staff.name}".strip()
        date_user_entry = {
            "fio": fio,
            "severity": candidate["severity"],
            "activeDays": person_profile["active_days"] if person_profile else 0,
            "repeatDays": person_profile["repeat_days"] if person_profile else 0,
            "repeatPct": person_profile["repeat_pct"] if person_profile else 0.0,
            "groupDays": candidate["group_days"],
            "lat": candidate["lat"],
            "lon": candidate["lon"],
            "locationName": candidate["location_name"],
            "reason": list(candidate["reason"]),
        }
        user_date_entry = {
            "severity": candidate["severity"],
            "groupDays": candidate["group_days"],
            "lat": candidate["lat"],
            "lon": candidate["lon"],
            "locationName": candidate["location_name"],
            "reason": list(candidate["reason"]),
        }
        date_groups[date_value][pin_short] = date_user_entry
        user_payload = user_groups.setdefault(
            pin_short,
            {
                "fio": fio,
                "activeDays": person_profile["active_days"] if person_profile else 0,
                "repeatDays": person_profile["repeat_days"] if person_profile else 0,
                "repeatPct": person_profile["repeat_pct"] if person_profile else 0.0,
                "highestSeverity": candidate["severity"],
                "datesByDate": {},
            },
        )
        user_payload["datesByDate"][date_value] = user_date_entry
        if severity_rank[candidate["severity"]] >= highest_severity_by_pin.get(
            pin_short, 0
        ):
            highest_severity_by_pin[pin_short] = severity_rank[candidate["severity"]]
            user_payload["highestSeverity"] = candidate["severity"]

    dates_payload = {
        date_value: {
            "usersByPin": {
                pin_short: date_groups[date_value][pin_short]
                for pin_short in sorted(date_groups[date_value])
            }
        }
        for date_value in sorted(date_groups)
    }
    users_payload = {
        pin_short: {
            "fio": user_groups[pin_short]["fio"],
            "activeDays": user_groups[pin_short]["activeDays"],
            "repeatDays": user_groups[pin_short]["repeatDays"],
            "repeatPct": user_groups[pin_short]["repeatPct"],
            "highestSeverity": user_groups[pin_short]["highestSeverity"],
            "datesByDate": {
                date_value: user_groups[pin_short]["datesByDate"][date_value]
                for date_value in sorted(user_groups[pin_short]["datesByDate"])
            },
        }
        for pin_short in sorted(user_groups)
    }

    payload = {
        "summary": {
            "staffAnalyzed": len(staff_list),
            "staffFlagged": len(users_payload),
            "datesFlagged": len(dates_payload),
        },
        "reasonLegend": dict(SUSPICIOUS_LOCATION_REASON_LEGEND),
        "datesByDate": dates_payload,
        "usersByPin": users_payload,
    }
    Cache.set(cache_key, payload, SUSPICIOUS_LOCATION_PATTERNS_CACHE_TTL)
    return Response(payload, status=status.HTTP_200_OK)


@swagger_auto_schema(
    method="GET",
    operation_summary="Список локаций для занятий",
    operation_description=(
        "**Без latitude/longitude:** все локации (id, name, address, latitude, longitude). "
        "Поле distance отсутствует.\n\n"
        "**С latitude и longitude:** фронт шлёт координаты, ответ — «в локации или нет». "
        "Только локации, где d ≤ R_loc: d — Haversine (user, pin) в м, R_loc из кэша/БД. "
        "Сортировка по d. В элементе: distance = d (м, 2 знака). Если ни одна не подходит → 404."
    ),
    tags=["Locations"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ (альтернатива JWT).",
        ),
        openapi.Parameter(
            name="latitude",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_NUMBER,
            format=openapi.FORMAT_FLOAT,
            required=False,
            description="Широта пользователя (WGS84). Для поиска по радиусу нужны оба: latitude и longitude.",
            example=43.246871,
        ),
        openapi.Parameter(
            name="longitude",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_NUMBER,
            format=openapi.FORMAT_FLOAT,
            required=False,
            description="Долгота пользователя (WGS84). Для поиска по радиусу нужны оба: latitude и longitude.",
            example=76.944923,
        ),
    ],
    responses={
        200: openapi.Response(
            description="locations. Без lat/lon: все, без distance. С lat/lon: d ≤ R_loc, distance = Haversine (м).",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                required=["locations"],
                properties={
                    "locations": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        description="При lat/lon: в радиусе (d ≤ R_loc), distance — Haversine в м. Без lat/lon: все локации.",
                        items=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            required=["id", "name", "address", "latitude", "longitude"],
                            properties={
                                "id": openapi.Schema(
                                    type=openapi.TYPE_INTEGER,
                                    description="ID локации",
                                ),
                                "name": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="Название",
                                ),
                                "address": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="Адрес",
                                ),
                                "latitude": openapi.Schema(
                                    type=openapi.TYPE_NUMBER,
                                    format=openapi.FORMAT_FLOAT,
                                    description="Широта точки локации",
                                ),
                                "longitude": openapi.Schema(
                                    type=openapi.TYPE_NUMBER,
                                    format=openapi.FORMAT_FLOAT,
                                    description="Долгота точки локации",
                                ),
                                "distance": openapi.Schema(
                                    type=openapi.TYPE_NUMBER,
                                    format=openapi.FORMAT_FLOAT,
                                    description=(
                                        "Только при lat/lon. Haversine(пользователь, pin) в метрах, 2 знака. "
                                        "Погрешность на практике — точность GPS (5–15 м у смартфона)."
                                    ),
                                ),
                            },
                        ),
                    ),
                },
            ),
        ),
        404: openapi.Response(
            description=(
                "При lat/lon: нет локаций с d ≤ R_loc — {message, detail}. "
                'Пустая БД по локациям — {error: "No locations available in database"}.'
            ),
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Системное: напр. «Ближайшая локация 85.2 м, превышен лимит 70 м».",
                    ),
                    "detail": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Для фронта: «Ничего не найдено».",
                    ),
                    "error": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="При пустой БД: «No locations available in database».",
                    ),
                },
            ),
        ),
        400: openapi.Response(
            description=(
                "Параметры отсутствуют или неверный формат. "
                "error, detail: «Широта обязательна», «Долгота обязательна», "
                "«Широта и долгота обязательны», «Invalid latitude or longitude format. Expected numbers.»"
            ),
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Код/текст ошибки",
                    ),
                    "detail": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение для пользователя",
                    ),
                },
            ),
        ),
        500: openapi.Response(
            description="Внутренняя ошибка сервера",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Описание ошибки",
                    ),
                },
            ),
        ),
    },
)
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def lesson_locations(request):
    """Список локаций: все либо только в приёмном радиусе R_loc.

    Без lat/lon: все локации (id, name, address, latitude, longitude).
    Поле distance не возвращается.

    С lat/lon: только локации, где d ≤ R_loc. d — Haversine (user, pin) в м,
    R_loc из кэша или ClassLocation.acceptance_radius_m. Сортировка по d.
    В ответе: distance = round(d, 2) (м). Погрешность — точность GPS (5–15 м).

    Args:
        request: GET; опционально latitude, longitude (WGS84, числа).

    Returns:
        Response: {locations: [...]}; при lat/lon и отсутствии подходящих — 404.
    """

    def _not_found(system_message: str):
        return Response(
            {"message": system_message, "detail": "Ничего не найдено"},
            status=status.HTTP_404_NOT_FOUND,
        )

    log_ll = logging.getLogger("monitoring_app.lesson_locations")
    log_ll_nf = logging.getLogger("monitoring_app.lesson_locations.not_found")

    try:
        latitude_param = request.GET.get("latitude")
        longitude_param = request.GET.get("longitude")

        def _is_empty(val):
            return val is None or (isinstance(val, str) and val.strip() == "")

        lat_empty = _is_empty(latitude_param)
        lon_empty = _is_empty(longitude_param)

        if latitude_param is None and longitude_param is None:
            pass
        elif lat_empty and lon_empty:
            log_ll.warning("MISSING both lat and lon")
            return Response(
                {
                    "error": "Широта и долгота обязательны",
                    "detail": "Широта и долгота обязательны",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        elif lat_empty:
            log_ll.warning("MISSING latitude lon=%s", longitude_param)
            return Response(
                {"error": "Широта обязательна", "detail": "Широта обязательна"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        elif lon_empty:
            log_ll.warning("MISSING longitude lat=%s", latitude_param)
            return Response(
                {"error": "Долгота обязательна", "detail": "Долгота обязательна"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if latitude_param is not None and longitude_param is not None:
            log_ll.info("request lat=%s lon=%s", latitude_param, longitude_param)
            try:
                latitude = float(latitude_param)
                longitude = float(longitude_param)
            except (ValueError, TypeError):
                log_ll.warning("INVALID lat=%s lon=%s", latitude_param, longitude_param)
                return Response(
                    {
                        "error": "Invalid latitude or longitude format. Expected numbers.",
                        "detail": "Неверный формат. Ожидаются числа.",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            all_locations = models.ClassLocation.objects.filter(
                latitude__isnull=False, longitude__isnull=False
            ).only("id", "name", "address", "latitude", "longitude")

            if not all_locations.exists():
                log_ll.warning("NO_LOCATIONS_IN_DB")
                return Response(
                    {"error": "No locations available in database"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            radii = get_class_location_cache().get("location_acceptance_radius_m", {})
            within = []
            min_overall = float("inf")
            nearest_loc = None

            for loc in all_locations:
                d = utils.calculate_distance_haversine(
                    latitude, longitude, loc.latitude, loc.longitude
                )
                acceptance_radius_m = radii.get(loc.id, DEFAULT_ACCEPTANCE_RADIUS_M)
                if d < min_overall:
                    min_overall = d
                    nearest_loc = loc
                if d <= acceptance_radius_m:
                    within.append((d, loc, acceptance_radius_m))

            within.sort(key=lambda x: x[0])

            if not within:
                nearest_acceptance_radius_m = (
                    radii.get(nearest_loc.id, DEFAULT_ACCEPTANCE_RADIUS_M)
                    if nearest_loc is not None
                    else DEFAULT_ACCEPTANCE_RADIUS_M
                )
                nearest_name = nearest_loc.name if nearest_loc else "N/A"
                loc_lat = nearest_loc.latitude if nearest_loc else 0
                loc_lon = nearest_loc.longitude if nearest_loc else 0
                loc_id = nearest_loc.id if nearest_loc else 0
                log_ll.warning(
                    "NOT_FOUND user(%.6f,%.6f) nearest_id=%d d=%.1fm R=%dm",
                    latitude,
                    longitude,
                    loc_id,
                    min_overall,
                    nearest_acceptance_radius_m,
                )
                log_ll_nf.warning(
                    "Haversine: d(user,loc)<=R => in_radius | "
                    "user(lat=%.6f,lon=%.6f) nearest=%s[id=%d](lat=%.6f,lon=%.6f) "
                    "d=%.1fm R=%dm => d>R NOT_FOUND",
                    latitude,
                    longitude,
                    nearest_name,
                    loc_id,
                    loc_lat,
                    loc_lon,
                    min_overall,
                    nearest_acceptance_radius_m,
                )
                return _not_found(
                    f"Ближайшая локация {min_overall:.1f} м, превышен лимит {nearest_acceptance_radius_m} м"
                )

            log_ll.info(
                "FOUND %d | %s",
                len(within),
                ", ".join(f"{loc.name}({d:.1f}m)" for d, loc, _ in within),
            )

            locations_data = [
                {
                    "id": loc.id,
                    "name": loc.name,
                    "address": loc.address,
                    "latitude": loc.latitude,
                    "longitude": loc.longitude,
                    "distance": round(d, 2),
                }
                for d, loc, _ in within
            ]

            return Response(
                {"locations": locations_data},
                status=status.HTTP_200_OK,
            )

        log_ll.info("request all locations")
        locations = models.ClassLocation.objects.only(
            "id", "name", "address", "latitude", "longitude"
        ).order_by("name")
        locations_data = [
            {
                "id": loc.id,
                "name": loc.name,
                "address": loc.address,
                "latitude": loc.latitude,
                "longitude": loc.longitude,
            }
            for loc in locations
        ]
        log_ll.info("returned %d locations", len(locations_data))

        return Response(
            {"locations": locations_data},
            status=status.HTTP_200_OK,
        )

    except Exception as e:
        log_ll.error(f"Critical error in lesson_locations: {str(e)}", exc_info=True)
        return Response(
            {"error": "A critical error occurred. Please try again later."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


_CLASSLOCATION_FIELDS = (
    "id",
    "name",
    "address",
    "latitude",
    "longitude",
    "acceptance_radius_m",
)

_classlocation_item_schema = openapi.Schema(
    type=openapi.TYPE_OBJECT,
    required=["name", "address", "latitude", "longitude"],
    properties={
        "name": openapi.Schema(type=openapi.TYPE_STRING, description="Название"),
        "address": openapi.Schema(type=openapi.TYPE_STRING, description="Адрес"),
        "latitude": openapi.Schema(
            type=openapi.TYPE_NUMBER, format=openapi.FORMAT_FLOAT, description="Широта"
        ),
        "longitude": openapi.Schema(
            type=openapi.TYPE_NUMBER, format=openapi.FORMAT_FLOAT, description="Долгота"
        ),
        "acceptance_radius_m": openapi.Schema(
            type=openapi.TYPE_INTEGER,
            nullable=True,
            description="Приёмный радиус (м), опционально",
        ),
    },
)

_classlocation_list_response = openapi.Response(
    description="Список локаций",
    schema=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            "results": openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "id": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "name": openapi.Schema(type=openapi.TYPE_STRING),
                        "address": openapi.Schema(type=openapi.TYPE_STRING),
                        "latitude": openapi.Schema(type=openapi.TYPE_NUMBER),
                        "longitude": openapi.Schema(type=openapi.TYPE_NUMBER),
                        "acceptance_radius_m": openapi.Schema(
                            type=openapi.TYPE_INTEGER, nullable=True
                        ),
                    },
                ),
            ),
        },
    ),
)


@swagger_auto_schema(
    method="get",
    operation_summary="Список локаций занятий (CRUD)",
    operation_description="Возвращает все локации: id, название, адрес, широта, долгота, приёмный радиус (м). Один запрос к БД.",
    tags=["ClassLocation"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ (альтернатива JWT).",
        ),
    ],
    responses={200: _classlocation_list_response},
)
@swagger_auto_schema(
    method="post",
    operation_summary="Создать одну или несколько локаций",
    operation_description="Тело: один объект или массив объектов. При массовом создании кэш локаций инвалидируется.",
    tags=["ClassLocation"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ (альтернатива JWT).",
        ),
    ],
    request_body=openapi.Schema(
        type=openapi.TYPE_ARRAY,
        description="Один объект или массив объектов: name, address, latitude, longitude, acceptance_radius_m (опционально).",
        items=_classlocation_item_schema,
    ),
    responses={
        201: openapi.Response(
            description="Создано",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "results": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(type=openapi.TYPE_OBJECT),
                    ),
                },
            ),
        ),
        400: openapi.Response(description="Ошибка валидации"),
    },
)
@swagger_auto_schema(
    method="delete",
    operation_summary="Массовое удаление локаций",
    operation_description="Query-параметр ids: список ID через запятую, например ?ids=1,2,3.",
    tags=["ClassLocation"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ (альтернатива JWT).",
        ),
        openapi.Parameter(
            name="ids",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            required=True,
            description="ID локаций через запятую (например 1,2,3)",
        ),
    ],
    responses={
        200: openapi.Response(description="Удалено"),
        400: openapi.Response(description="Не указаны ids"),
    },
)
@api_view(["GET", "POST", "DELETE"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def class_location_list_create(request):
    """GET: список (кэш 1 ч, один запрос .values()). POST/DELETE: инвалидация + прогрев списка и таска радиусов."""
    if request.method == "GET":
        data = Cache.get(CLASS_LOCATION_LIST_CACHE_KEY)
        if data is None:
            data = list(
                models.ClassLocation.objects.order_by("id").values(
                    "id",
                    "name",
                    "address",
                    "latitude",
                    "longitude",
                    "acceptance_radius_m",
                )
            )
            Cache.set(
                CLASS_LOCATION_LIST_CACHE_KEY,
                data,
                CLASS_LOCATION_LIST_CACHE_TTL,
            )
        return Response({"results": data}, status=status.HTTP_200_OK)

    if request.method == "POST":
        try:
            body = request.data
        except Exception:
            body = None
        if body is None:
            return Response(
                {"error": "Request body is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        is_list = isinstance(body, list)
        items = body if is_list else [body]
        serializer = serializers.ClassLocationSerializer(data=items, many=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        validated = serializer.validated_data
        if not validated or not isinstance(validated, (list, tuple)):
            return Response(
                {"error": "At least one item required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        with _db_atomic():
            created = []
            for d in validated:
                obj = models.ClassLocation.objects.create(
                    name=d["name"],
                    address=d["address"],
                    latitude=d["latitude"],
                    longitude=d["longitude"],
                    acceptance_radius_m=d.get("acceptance_radius_m"),
                )
                created.append(obj)
        if len(created) > 0:
            invalidate_class_location_cache_impl()
        result = [
            {
                "id": o.id,
                "name": o.name,
                "address": o.address,
                "latitude": o.latitude,
                "longitude": o.longitude,
                "acceptance_radius_m": getattr(o, "acceptance_radius_m", None),
            }
            for o in created
        ]
        return Response({"results": result}, status=status.HTTP_201_CREATED)

    if request.method == "DELETE":
        ids_param = request.query_params.get("ids") or request.query_params.get("id")
        if not ids_param:
            return Response(
                {"error": "Query parameter 'ids' is required (e.g. ids=1,2,3)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            id_list = [int(x.strip()) for x in ids_param.split(",") if x.strip()]
        except ValueError:
            return Response(
                {"error": "ids must be comma-separated integers"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not id_list:
            return Response(
                {"error": "At least one id required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        deleted, _ = models.ClassLocation.objects.filter(id__in=id_list).delete()
        return Response(
            {"deleted": deleted, "ids": id_list},
            status=status.HTTP_200_OK,
        )


@swagger_auto_schema(
    method="get",
    operation_summary="Одна локация по ID",
    tags=["ClassLocation"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ (альтернатива JWT).",
        ),
        openapi.Parameter(
            "id", openapi.IN_PATH, type=openapi.TYPE_INTEGER, description="ID локации"
        ),
    ],
    responses={
        200: _classlocation_list_response,
        404: openapi.Response(description="Не найдено"),
    },
)
@swagger_auto_schema(
    method="patch",
    operation_summary="Обновить одну локацию",
    tags=["ClassLocation"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ (альтернатива JWT).",
        ),
        openapi.Parameter(
            "id", openapi.IN_PATH, type=openapi.TYPE_INTEGER, description="ID локации"
        ),
    ],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            "name": openapi.Schema(type=openapi.TYPE_STRING),
            "address": openapi.Schema(type=openapi.TYPE_STRING),
            "latitude": openapi.Schema(type=openapi.TYPE_NUMBER),
            "longitude": openapi.Schema(type=openapi.TYPE_NUMBER),
            "acceptance_radius_m": openapi.Schema(
                type=openapi.TYPE_INTEGER, nullable=True
            ),
        },
    ),
    responses={
        200: _classlocation_list_response,
        404: openapi.Response(description="Не найдено"),
    },
)
@swagger_auto_schema(
    method="delete",
    operation_summary="Удалить одну локацию",
    tags=["ClassLocation"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ (альтернатива JWT).",
        ),
        openapi.Parameter(
            "id", openapi.IN_PATH, type=openapi.TYPE_INTEGER, description="ID локации"
        ),
    ],
    responses={
        200: openapi.Response(description="Удалено"),
        404: openapi.Response(description="Не найдено"),
    },
)
@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def class_location_detail(request, pk):
    """GET/PATCH/DELETE одной локации по pk. Кэш инвалидируется через signal при save/delete."""
    loc = get_object_or_404(
        models.ClassLocation.objects.only(*_CLASSLOCATION_FIELDS),
        pk=pk,
    )
    if request.method == "GET":
        return Response(
            {
                "id": loc.id,
                "name": loc.name,
                "address": loc.address,
                "latitude": loc.latitude,
                "longitude": loc.longitude,
                "acceptance_radius_m": loc.acceptance_radius_m,
            },
            status=status.HTTP_200_OK,
        )
    if request.method == "PATCH":
        serializer = serializers.ClassLocationSerializer(
            loc, data=request.data, partial=True
        )
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    if request.method == "DELETE":
        loc.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


_classlocation_bulk_update_schema = openapi.Schema(
    type=openapi.TYPE_ARRAY,
    items=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=["id"],
        properties={
            "id": openapi.Schema(type=openapi.TYPE_INTEGER, description="ID локации"),
            "name": openapi.Schema(type=openapi.TYPE_STRING),
            "address": openapi.Schema(type=openapi.TYPE_STRING),
            "latitude": openapi.Schema(type=openapi.TYPE_NUMBER),
            "longitude": openapi.Schema(type=openapi.TYPE_NUMBER),
            "acceptance_radius_m": openapi.Schema(
                type=openapi.TYPE_INTEGER, nullable=True
            ),
        },
    ),
)


@swagger_auto_schema(
    method="patch",
    operation_summary="Массовое обновление локаций",
    operation_description="Тело: массив объектов с обязательным полем id. После обновления кэш инвалидируется.",
    tags=["ClassLocation"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ (альтернатива JWT).",
        ),
    ],
    request_body=_classlocation_bulk_update_schema,
    responses={
        200: openapi.Response(
            description="Обновлено",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "updated": openapi.Schema(type=openapi.TYPE_INTEGER),
                    "results": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(type=openapi.TYPE_OBJECT),
                    ),
                },
            ),
        ),
        400: openapi.Response(description="Ошибка валидации"),
    },
)
@api_view(["PATCH"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def class_location_bulk_update(request):
    """Массовое обновление: bulk_update + инвалидация кэша."""
    try:
        body = request.data
    except Exception:
        body = None
    if body is None:
        return Response(
            {"error": "Body must be a non-empty array of objects with 'id'"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if not isinstance(body, list):
        body = [body] if isinstance(body, dict) and "id" in body else []
    if not body:
        return Response(
            {"error": "Body must be a non-empty array of objects with 'id'"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    ids = []
    for item in body:
        raw_id = item.get("id")
        if raw_id is None:
            return Response(
                {"error": "Each item must have integer 'id'"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            ids.append(int(raw_id))
        except (TypeError, ValueError):
            return Response(
                {"error": "Each item must have integer 'id'"},
                status=status.HTTP_400_BAD_REQUEST,
            )
    unique_ids = list(dict.fromkeys(ids))
    qs = list(
        models.ClassLocation.objects.only(*_CLASSLOCATION_FIELDS).filter(
            id__in=unique_ids
        )
    )
    found_ids = {o.id for o in qs}
    missing = [i for i in unique_ids if i not in found_ids]
    if missing:
        return Response(
            {"error": "Some ids not found", "missing_ids": missing},
            status=status.HTTP_404_NOT_FOUND,
        )
    by_id = {o.id: o for o in qs}
    update_fields = ["name", "address", "latitude", "longitude", "acceptance_radius_m"]
    for raw in body:
        obj = by_id.get(int(raw["id"]))
        if not obj:
            continue
        for f in update_fields:
            if f in raw:
                setattr(obj, f, raw[f])
    with _db_atomic():
        models.ClassLocation.objects.bulk_update(qs, update_fields)
    invalidate_class_location_cache_impl()
    results = [
        {
            "id": o.id,
            "name": o.name,
            "address": o.address,
            "latitude": o.latitude,
            "longitude": o.longitude,
            "acceptance_radius_m": o.acceptance_radius_m,
        }
        for o in qs
    ]
    return Response({"updated": len(qs), "results": results}, status=status.HTTP_200_OK)


@swagger_auto_schema(
    method="GET",
    operation_summary="Получить ID всех корневых (root) подразделений",
    operation_description="Возвращает список ID из ChildDepartment, где parent IS NULL.",
    tags=["Departments"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
    ],
    responses={
        200: openapi.Response(
            description="Ок",
            schema=openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="ID корневого подразделения (ChildDepartment).",
                ),
            ),
        ),
        404: openapi.Response(
            description="Не удалось найти корни",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={"error": openapi.Schema(type=openapi.TYPE_STRING)},
            ),
        ),
    },
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_parent_id(request):
    """
    Получить ID всех родительских департаментов.

    Этот метод возвращает список всех ID родительских департаментов.

    Возвращаемые данные:
    - Список ID родительских департаментов.

    Возможные ошибки:
    - 404: Департаменты не найдены.

    Пример ответа:
    - 200: [1, 2, 3, ...]
    - 404: {"error": "Не удалось найти департаменты."}
    """
    logger.info("Request received to get parent department IDs.")

    cache_key = "parent_department_ids"

    def fetch_parent_ids():
        roots = (
            models.ChildDepartment.objects.filter(parent__isnull=True)
            .order_by("id")
            .values_list("id", flat=True)
        )
        return [str(pk) for pk in roots]

    try:
        root_ids = get_cache(
            cache_key,
            query=fetch_parent_ids,
            timeout=30 * 60,
        )

        if not root_ids:
            return Response(
                {"error": "Корни не найдены"}, status=status.HTTP_404_NOT_FOUND
            )
        return Response(root_ids, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_404_NOT_FOUND)


def _get_breadcrumb_path(dept_id: str) -> list[dict]:
    """Возвращает путь от корня до отдела: [{id, name}, ...]."""
    try:
        dept = models.ChildDepartment.objects.get(id=dept_id)
    except models.ChildDepartment.DoesNotExist:
        return []
    path = []
    current = dept
    while current:
        path.append({"id": str(current.id), "name": current.name})
        current = current.parent
    path.reverse()
    return path


def _build_department_summary_data(parent_department_id: str):
    """Строит данные для department_summary. Используется в API и warmup."""

    def get_subtree_staff_count(dept_id: str) -> int:
        """Подсчёт сотрудников в поддереве через рекурсивный CTE (оптимизация для больших деревьев)."""
        with connection.cursor() as cursor:
            cursor.execute(
                """
                WITH RECURSIVE subtree AS (
                    SELECT id FROM monitoring_app_childdepartment WHERE id = %s
                    UNION ALL
                    SELECT cd.id FROM monitoring_app_childdepartment cd
                    JOIN subtree s ON cd.parent_id = s.id
                )
                SELECT COUNT(DISTINCT st.id) FROM monitoring_app_staff st
                INNER JOIN subtree s ON st.department_id = s.id
                """,
                [dept_id],
            )
            row = cursor.fetchone()
            return row[0] if row else 0

    parent_department = get_object_or_404(
        models.ChildDepartment, id=parent_department_id
    )
    total_staff_count = get_subtree_staff_count(parent_department.id)
    child_departments_data = models.ChildDepartment.objects.filter(
        parent=parent_department
    )
    child_departments_data_serialized = serializers.ChildDepartmentSerializer(
        child_departments_data, many=True
    ).data
    breadcrumb_path = _get_breadcrumb_path(parent_department_id)
    return {
        "name": parent_department.name,
        "date_of_creation": parent_department.date_of_creation,
        "child_departments": child_departments_data_serialized,
        "total_staff_count": total_staff_count,
        "breadcrumb_path": breadcrumb_path,
    }


@swagger_auto_schema(
    method="GET",
    operation_summary="Сводная информация о департаменте",
    operation_description="Метод для получения сводной информации о департаменте и его дочерних подразделениях с количеством сотрудников.",
    tags=["Departments"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
    ],
    responses={
        200: openapi.Response(
            description="Успешный запрос. Возвращается сводная информация о департаменте и его дочерних подразделениях.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "name": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Название департамента.",
                    ),
                    "date_of_creation": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        format="date-time",
                        description="Дата создания департамента.",
                    ),
                    "child_departments": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "child_id": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="ID дочернего подразделения.",
                                ),
                                "name": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="Название дочернего подразделения.",
                                ),
                                "date_of_creation": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    format="date-time",
                                    description="Дата создания дочернего подразделения.",
                                ),
                                "parent": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="ID родительского департамента.",
                                ),
                            },
                        ),
                        description="Список дочерних подразделений департамента.",
                    ),
                    "total_staff_count": openapi.Schema(
                        type=openapi.TYPE_INTEGER,
                        description="Общее количество сотрудников в департаменте и его дочерних подразделениях.",
                    ),
                },
            ),
        ),
        404: "Департамент не найден.",
    },
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def department_summary(request, parent_department_id):
    cache_key = f"department_summary_v2_{parent_department_id}"
    logger.info(
        f"Request received for department summary with ID {parent_department_id}"
    )

    if not models.ChildDepartment.objects.filter(id=parent_department_id).exists():
        logger.warning(f"Department with ID {parent_department_id} not found")
        return Response(
            status=status.HTTP_404_NOT_FOUND,
            data={"message": f"Department with ID {parent_department_id} not found"},
        )

    try:
        cached_data = get_cache(
            cache_key,
            query=lambda: _build_department_summary_data(parent_department_id),
            timeout=5 * 60,
            cache=Cache,
        )
        logger.info(f"Returning summary data for department ID {parent_department_id}")
        return Response(cached_data, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error while generating department summary: {str(e)}")
        return Response(data={"message": str(e)}, status=status.HTTP_404_NOT_FOUND)


def _fetch_root_departments_data():
    """
    Внутренняя функция для получения данных корневых департаментов.
    Может использоваться как в API endpoint, так и в preload.
    """
    all_departments = list(
        models.ChildDepartment.objects.only(
            "id", "parent_id", "name", "date_of_creation"
        ).values_list("id", "parent_id", "name", "date_of_creation")
    )

    dept_by_id = {}
    children_by_parent = {}
    root_ids = []

    for dept_id, parent_id, name, date_of_creation in all_departments:
        dept_by_id[dept_id] = {
            "id": dept_id,
            "name": name,
            "date_of_creation": date_of_creation,
        }
        if parent_id is None:
            root_ids.append(dept_id)
        else:
            children_by_parent.setdefault(parent_id, []).append(dept_id)

    if not root_ids:
        return {
            "departments": [],
            "total_staff_count": 0,
        }

    root_ids.sort()

    subtree_ids_by_root = {}
    for root_id in root_ids:
        visited = set()
        stack = [root_id]
        subtree_ids = []

        while stack:
            cur = stack.pop()
            if cur in visited:
                continue
            visited.add(cur)
            subtree_ids.append(cur)
            stack.extend(children_by_parent.get(cur, []))

        subtree_ids_by_root[root_id] = subtree_ids

    all_subtree_ids = []
    for subtree_ids in subtree_ids_by_root.values():
        all_subtree_ids.extend(subtree_ids)

    unique_subtree_ids = list(set(all_subtree_ids))

    staff_counts = (
        models.Staff.objects.filter(department_id__in=unique_subtree_ids)
        .values("department_id")
        .annotate(count=Count("id", distinct=True))
    )

    staff_count_by_dept = {
        item["department_id"]: item["count"] for item in staff_counts
    }

    all_child_dept_ids = []
    for children in children_by_parent.values():
        all_child_dept_ids.extend(children)

    child_depts_by_parent = {}
    if all_child_dept_ids:
        child_depts = models.ChildDepartment.objects.filter(
            id__in=all_child_dept_ids
        ).only("id", "name", "date_of_creation", "parent_id")

        for child in child_depts:
            parent_id = child.parent_id
            if parent_id:
                child_depts_by_parent.setdefault(parent_id, []).append(child)

    departments_data = []
    total_staff_count = 0

    for root_id in root_ids:
        dept = dept_by_id[root_id]
        subtree_ids = subtree_ids_by_root[root_id]
        dept_total = sum(staff_count_by_dept.get(dept_id, 0) for dept_id in subtree_ids)
        total_staff_count += dept_total

        has_children = bool(children_by_parent.get(root_id))

        child_depts = child_depts_by_parent.get(root_id, [])
        child_departments_serialized = [
            {
                "child_id": str(child.id),
                "name": child.name,
                "date_of_creation": child.date_of_creation,
                "parent": str(root_id),
            }
            for child in child_depts
        ]

        departments_data.append(
            {
                "child_id": str(root_id),
                "name": dept["name"],
                "date_of_creation": dept["date_of_creation"],
                "parent": "",
                "has_child_departments": has_children,
                "total_staff_count": dept_total,
                "child_departments": child_departments_serialized,
            }
        )

    return {
        "departments": departments_data,
        "total_staff_count": total_staff_count,
    }


@swagger_auto_schema(
    method="GET",
    operation_summary="Получить все корневые департаменты одним запросом",
    operation_description="endpoint для получения всех корневых департаментов с их сводной информацией одним запросом",
    tags=["Departments"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
    ],
    responses={
        200: openapi.Response(
            description="Успешный ответ",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "departments": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "child_id": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="ID корневого департамента",
                                ),
                                "name": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="Название департамента",
                                ),
                                "date_of_creation": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    format="date-time",
                                    description="Дата создания",
                                ),
                                "parent": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="ID родительского департамента (всегда пусто для корневых)",
                                ),
                                "has_child_departments": openapi.Schema(
                                    type=openapi.TYPE_BOOLEAN,
                                    description="Есть ли дочерние подразделения",
                                ),
                                "total_staff_count": openapi.Schema(
                                    type=openapi.TYPE_INTEGER,
                                    description="Общее количество сотрудников",
                                ),
                                "child_departments": openapi.Schema(
                                    type=openapi.TYPE_ARRAY,
                                    items=openapi.Schema(
                                        type=openapi.TYPE_OBJECT,
                                        properties={
                                            "child_id": openapi.Schema(
                                                type=openapi.TYPE_STRING
                                            ),
                                            "name": openapi.Schema(
                                                type=openapi.TYPE_STRING
                                            ),
                                            "date_of_creation": openapi.Schema(
                                                type=openapi.TYPE_STRING,
                                                format="date-time",
                                            ),
                                            "parent": openapi.Schema(
                                                type=openapi.TYPE_STRING
                                            ),
                                        },
                                    ),
                                    description="Список дочерних подразделений",
                                ),
                            },
                        ),
                    ),
                    "total_staff_count": openapi.Schema(
                        type=openapi.TYPE_INTEGER,
                        description="Общее количество сотрудников во всех корневых департаментах",
                    ),
                },
            ),
        ),
    },
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def root_departments_batch(request):
    """
    Получить все корневые департаменты одним оптимизированным запросом.
    Используется для быстрой загрузки главной страницы вместо множественных запросов.
    """
    cache_key = "root_departments_batch"
    logger.info("Request received for root departments batch")

    try:
        cached_data = get_cache(
            cache_key,
            query=_fetch_root_departments_data,
            timeout=15 * 60,
        )

        logger.info("Returning root departments batch data")
        return Response(cached_data, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error while generating root departments batch: {str(e)}")
        return Response(
            data={"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@swagger_auto_schema(
    method="get",
    operation_summary="Получить описание подотдела",
    operation_description="Получите подробную информацию о подотделе и его сотрудниках.",
    tags=["Departments"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
        openapi.Parameter(
            name="child_department_id",
            in_=openapi.IN_PATH,
            type=openapi.TYPE_INTEGER,
            description="ID подотдела",
            required=True,
        ),
    ],
    responses={
        200: openapi.Response(
            description="Сведения о подотделе и данные о персонале",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "child_department": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            "id": openapi.Schema(type=openapi.TYPE_INTEGER),
                            "name": openapi.Schema(type=openapi.TYPE_STRING),
                            "date_of_creation": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                format=openapi.FORMAT_DATETIME,
                            ),
                            "parent": openapi.Schema(type=openapi.TYPE_INTEGER),
                        },
                    ),
                    "staff_count": openapi.Schema(type=openapi.TYPE_INTEGER),
                    "staff_data": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        additional_properties=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "FIO": openapi.Schema(type=openapi.TYPE_STRING),
                                "date_of_creation": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    format=openapi.FORMAT_DATETIME,
                                ),
                                "avatar": openapi.Schema(type=openapi.TYPE_STRING),
                                "positions": openapi.Schema(
                                    type=openapi.TYPE_ARRAY,
                                    items=openapi.Schema(type=openapi.TYPE_STRING),
                                ),
                            },
                        ),
                    ),
                },
            ),
        ),
        404: "Not Found: Если подотдела не существует.",
    },
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def child_department_detail(request, child_department_id):
    """
    Получите подробную информацию о дочернем отделе вместе с его сотрудниками.

    Args:
    запрос: объект запроса.
    child_department_id (int): идентификатор дочернего отдела, который требуется получить.

    Returns:
    Ответ: ответ, содержащий сведения о дочернем отделе и данные о сотрудниках.

    Raises:
    Http404: Если дочерний отдел не существует.
    """
    logger.info(
        f"Request received for child department detail with ID {child_department_id}"
    )

    cache_key = f"child_department_detail_v2_{child_department_id}"

    def fetch_child_department_data():
        try:
            child_department = models.ChildDepartment.objects.get(
                id=child_department_id
            )
        except models.ChildDepartment.DoesNotExist:
            logger.warning(f"Child department with ID {child_department_id} not found")
            return None

        all_departments = [
            child_department
        ] + child_department.get_all_child_departments()
        staff_in_department = (
            models.Staff.objects.filter(department__in=all_departments)
            .select_related("department")
            .prefetch_related("positions")
        )

        staff_data = {}
        for staff_member in staff_in_department:
            if staff_member.surname == "Нет фамилии":
                fio = staff_member.name
            else:
                fio = f"{staff_member.surname} {staff_member.name}"
            staff_data[staff_member.pin] = {
                "FIO": fio,
                "date_of_creation": staff_member.date_of_creation,
                "avatar": (staff_member.avatar.url if staff_member.avatar else None),
                "positions": [p.name for p in staff_member.positions.all()],
            }

        sorted_staff_data = dict(
            sorted(staff_data.items(), key=lambda item: item[1]["FIO"])
        )
        breadcrumb_path = _get_breadcrumb_path(child_department_id)
        return {
            "child_department": serializers.ChildDepartmentSerializer(
                child_department
            ).data,
            "staff_count": staff_in_department.count(),
            "staff_data": sorted_staff_data,
            "breadcrumb_path": breadcrumb_path,
        }

    try:
        data = get_cache(
            cache_key,
            query=fetch_child_department_data,
            timeout=10 * 60,
        )

        if data is None:
            return Response(status=status.HTTP_404_NOT_FOUND)

        logger.info(
            f"Returning detailed data for child department ID {child_department_id}"
        )
        return Response(data, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error in child_department_detail: {str(e)}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method="GET",
    operation_summary="Получить информацию о сотруднике",
    operation_description=(
        "Возвращает подробную информацию о сотруднике за период: посещаемость по датам "
        "(first_in, last_out, effective_work_seconds — время в здании с учётом выходов, "
        "area_sequence — цепочка зон для карты перемещений), процент присутствия, бонус, "
        "тип контракта и зарплату. Данные кэшируются на 5 минут."
    ),
    tags=["Staff"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
        openapi.Parameter(
            name="staff_pin",
            in_=openapi.IN_PATH,
            type=openapi.TYPE_STRING,
            required=True,
            description="Уникальный идентификатор сотрудника (PIN)",
        ),
        openapi.Parameter(
            name="start_date",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            required=False,
            description="Дата начала периода для фильтрации данных о посещаемости (формат: YYYY-MM-DD)",
        ),
        openapi.Parameter(
            name="end_date",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            required=False,
            description="Дата окончания периода для фильтрации данных о посещаемости (формат: YYYY-MM-DD)",
        ),
    ],
    responses={
        200: openapi.Response(
            description="Данные сотрудника успешно получены",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "name": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Имя сотрудника",
                    ),
                    "surname": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Фамилия сотрудника",
                    ),
                    "positions": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(type=openapi.TYPE_STRING),
                        description="Список должностей сотрудника",
                    ),
                    "avatar": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        format=openapi.FORMAT_URI,
                        nullable=True,
                        description="URL аватара сотрудника",
                    ),
                    "department": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Отдел, к которому относится сотрудник",
                    ),
                    "department_id": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        description="Id отдела",
                    ),
                    "attendance": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        description="Данные о посещаемости по датам. Ключ - дата в формате DD-MM-YYYY",
                        additional_properties=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            description="Данные о посещаемости за конкретную дату",
                            properties={
                                "first_in": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    format=openapi.FORMAT_DATETIME,
                                    nullable=True,
                                    description="Время первого входа в формате ISO 8601",
                                ),
                                "last_out": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    format=openapi.FORMAT_DATETIME,
                                    nullable=True,
                                    description="Время последнего выхода в формате ISO 8601",
                                ),
                                "area_name_in": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    nullable=True,
                                    description="Название места первого входа",
                                ),
                                "area_name_out": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    nullable=True,
                                    description=(
                                        "Название места последнего выхода. "
                                        "Примечание: если значение равно 'Unknown', рекомендуется парсить его как null."
                                    ),
                                ),
                                "first_in_source": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    nullable=True,
                                    description="Источник данных о первом входе (staff_attendance или lesson_attendance)",
                                ),
                                "last_out_source": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    nullable=True,
                                    description="Источник данных о последнем выходе (staff_attendance или lesson_attendance)",
                                ),
                                "percent_day": openapi.Schema(
                                    type=openapi.TYPE_NUMBER,
                                    format=openapi.FORMAT_FLOAT,
                                    description="Процент отработанного времени за день",
                                ),
                                "total_minutes": openapi.Schema(
                                    type=openapi.TYPE_NUMBER,
                                    format=openapi.FORMAT_FLOAT,
                                    description="Отработанные минуты за день (с учётом effective_work_seconds при наличии)",
                                ),
                                "effective_work_seconds": openapi.Schema(
                                    type=openapi.TYPE_INTEGER,
                                    nullable=True,
                                    description="Секунды в здании по интервалам (без времени вне здания по турникетам выхода)",
                                ),
                                "area_sequence": openapi.Schema(
                                    type=openapi.TYPE_ARRAY,
                                    nullable=True,
                                    items=openapi.Schema(
                                        type=openapi.TYPE_OBJECT,
                                        properties={
                                            "t": openapi.Schema(
                                                type=openapi.TYPE_STRING,
                                                description="Время HH:MM",
                                            ),
                                            "area": openapi.Schema(
                                                type=openapi.TYPE_STRING,
                                                description="Название зоны",
                                            ),
                                        },
                                    ),
                                    description="Цепочка зон по времени для карты перемещений (только при границах из StaffAttendance)",
                                ),
                                "is_weekend": openapi.Schema(
                                    type=openapi.TYPE_BOOLEAN,
                                    description="Является ли день выходным",
                                ),
                                "is_remote_work": openapi.Schema(
                                    type=openapi.TYPE_BOOLEAN,
                                    description="Является ли работа удаленной",
                                ),
                                "is_absent_approved": openapi.Schema(
                                    type=openapi.TYPE_BOOLEAN,
                                    description="Утверждено ли отсутствие",
                                ),
                                "absent_reason": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    nullable=True,
                                    description="Причина отсутствия (если применимо)",
                                ),
                            },
                        ),
                    ),
                    "percent_for_period": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        format=openapi.FORMAT_FLOAT,
                        description="Общий процент работы за указанный период",
                    ),
                    "salary": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        format=openapi.FORMAT_FLOAT,
                        nullable=True,
                        description="Общая заработная плата сотрудника",
                    ),
                    "contract_type": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Тип контракта сотрудника",
                    ),
                    "bonus_percentage": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        description="Процент бонуса сотрудника",
                    ),
                },
            ),
        ),
        400: "Неверный запрос, дата начала не может быть позже даты окончания",
        404: "Сотрудник не найден",
    },
)
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def staff_detail(request, staff_pin):
    """Возвращает детальную информацию о сотруднике за указанный период.

    Включает посещаемость (first_in, last_out, effective_work_seconds, area_sequence),
    процент присутствия, бонус, тип контракта и зарплату. Данные кэшируются.

    Args:
        request: HttpRequest с query-параметрами start_date, end_date (YYYY-MM-DD).
        staff_pin: Уникальный идентификатор сотрудника (PIN), строка.

    Returns:
        Response с JSON-данными сотрудника (200), 400 при неверном диапазоне дат,
        404 если сотрудник не найден.
    """

    logger.info(f"Request received for staff details with PIN {staff_pin}")

    staff = get_cache(
        f"staff_{staff_pin}",
        query=lambda: fetch_staff_data(staff_pin),
        timeout=10 * 60,
    )

    if staff is None:
        logger.warning(f"Staff with PIN {staff_pin} not found")
        return Response(status=status.HTTP_404_NOT_FOUND)

    start_date, end_date = get_date_range(request)
    logger.debug(f"Retrieved date range: start_date={start_date}, end_date={end_date}")

    if start_date > end_date:
        logger.warning(
            f"Invalid date range: start_date {start_date} is greater than end_date {end_date}"
        )
        return Response(
            data={"error": "start_date cannot be greater than end_date"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    cache_key = (
        f"staff_detail_{LESSON_REPORT_CACHE_VERSION}_{staff_pin}_"
        f"{start_date}_{end_date}"
    )
    logger.debug(f"Generated cache key: {cache_key}")

    data = get_cache(
        cache_key,
        query=lambda: get_staff_detail(staff, start_date, end_date),
        timeout=300,
    )

    logger.info(f"Returning staff details for PIN {staff_pin}")
    return Response(data, status=status.HTTP_200_OK)


def fetch_staff_data(staff_pin):
    """Загружает объект сотрудника из БД по PIN с минимумом полей и department.

    Args:
        staff_pin: Уникальный идентификатор сотрудника (PIN), строка.

    Returns:
        Экземпляр models.Staff с полями id, pin, name, surname, avatar, department
        (select_related) или None, если сотрудник не найден.
    """
    try:
        return (
            models.Staff.objects.filter(pin=staff_pin)
            .select_related("department")
            .only(
                "id",
                "pin",
                "name",
                "surname",
                "avatar",
                "department_id",
                "department__id",
                "department__name",
            )
            .get()
        )
    except models.Staff.DoesNotExist:
        return None


def get_date_range(request):
    """Извлекает диапазон дат из query-параметров запроса.

    Если параметры не заданы, используется период: последние 7 дней до сегодня.
    Ожидаемые ключи: start_date, end_date в формате YYYY-MM-DD.

    Args:
        request: HttpRequest с query_params (DRF или Django).

    Returns:
        Кортеж (start_date, end_date) типа datetime.date.
    """
    end_date_str = request.query_params.get(
        "end_date", timezone.now().strftime("%Y-%m-%d")
    )
    start_date_str = request.query_params.get(
        "start_date", (timezone.now() - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
    )

    end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
    start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()

    return start_date, end_date


def get_staff_detail(staff, start_date, end_date):
    """Формирует полный словарь данных сотрудника за период для API.

    Объединяет StaffAttendance и LessonAttendance по датам, считает процент
    присутствия и бонус, подтягивает удалёнку, отсутствия, праздники и зарплату.
    Для посещаемости используется effective_work_seconds при наличии.

    Args:
        staff: Экземпляр models.Staff (должен иметь id, name, pin, department и т.д.).
        start_date: Начало периода, date.
        end_date: Конец периода, date.

    Returns:
        Словарь с ключами name, surname, positions, avatar, department, department_id,
        attendance (по датам), percent_for_period, bonus_percentage, contract_type, salary.
    """
    logger.info(f"Получение деталей сотрудника {staff.name} (PIN: {staff.pin})")
    logger.debug(f"Запрошенный диапазон дат: {start_date} до {end_date}")

    sa_by_event_date, la_by_event_date = fetch_attendance_by_event_dates(
        [staff.id], start_date, end_date
    )
    location_cache = get_class_location_cache()
    kd_tree = location_cache["kd_tree"]
    class_names = location_cache["class_names"]
    if kd_tree and class_names:
        logger.debug(f"KDTree initialized with {len(class_names)} locations")

    all_event_dates = sorted(
        set(sa_by_event_date.keys()) | set(la_by_event_date.keys())
    )
    combined_attendance = {}
    for event_date in all_event_dates:
        combined_attendance[event_date] = _merge_attendance_for_date(
            sa_by_event_date.get(event_date, []),
            la_by_event_date.get(event_date, []),
            kd_tree,
            class_names,
        )

    attendance_dates = list(sa_by_event_date.keys())
    lesson_dates = list(la_by_event_date.keys())

    logger.debug(f"Объединенные данные посещаемости: {combined_attendance}")

    holidays = models.PublicHoliday.objects.filter(
        date__range=[start_date, end_date]
    ).values_list("date", "is_working_day")
    logger.debug(f"Государственные праздники в периоде: {len(holidays)}")

    holiday_dict = dict(holidays)
    attendance_data = {}
    total_minutes_for_period = 0
    total_days_with_data = 0
    percent_for_period = 0

    remote_work_qs = models.RemoteWork.objects.filter(staff=staff).filter(
        Q(permanent_remote=True) | Q(start_date__lte=end_date, end_date__gte=start_date)
    )
    logger.debug(f"Получено периодов дистанционной работы: {remote_work_qs.count()}")

    absent_reason_qs = models.AbsentReason.objects.filter(
        staff=staff, start_date__lte=end_date, end_date__gte=start_date
    )
    logger.debug(f"Получено причин отсутствия: {absent_reason_qs.count()}")

    dates = []
    dates.extend(attendance_dates)
    dates.extend(lesson_dates)

    if remote_work_qs.exists():
        remote_dates = []
        for remote_work in remote_work_qs:
            if remote_work.permanent_remote:
                remote_dates.extend(attendance_dates)
            else:
                rw_start = (
                    remote_work.start_date
                    if remote_work.start_date is not None
                    else start_date
                )
                rw_end = (
                    remote_work.end_date
                    if remote_work.end_date is not None
                    else end_date
                )
                remote_start = max(rw_start, start_date)
                remote_end = min(rw_end, end_date)
                if remote_start <= remote_end:
                    remote_dates.extend(
                        [
                            remote_start + datetime.timedelta(days=x)
                            for x in range((remote_end - remote_start).days + 1)
                        ]
                    )
        dates.extend(remote_dates)
    else:
        remote_dates = []

    if absent_reason_qs.exists():
        absent_dates = []
        for absent_reason in absent_reason_qs:
            absent_start = max(absent_reason.start_date, start_date)
            absent_end = min(absent_reason.end_date, end_date)
            absent_dates.extend(
                [
                    absent_start + datetime.timedelta(days=x)
                    for x in range((absent_end - absent_start).days + 1)
                ]
            )
        dates.extend(absent_dates)
    else:
        absent_dates = []

    if not dates:
        logger.warning(
            "Нет данных о посещаемости, дистанционной работе или причинах отсутствия за указанный период."
        )
        staff_detail_data = {
            "name": staff.name,
            "surname": staff.surname if staff.surname != "Нет фамилии" else "",
            "positions": [position.name for position in staff.positions.all()],
            "avatar": (
                staff.avatar.url if staff.avatar else "/media/images/no-avatar.png"
            ),
            "department": staff.department.name if staff.department else "N/A",
            "department_id": staff.department.id if staff.department else "N/A",
            "attendance": {},
            "percent_for_period": 0.0,
            "contract_type": None,
            "salary": None,
        }
        return staff_detail_data

    min_date = max(min(dates), start_date)
    max_date = min(max(dates), end_date)
    logger.debug(f"Фактический диапазон дат с данными: {min_date} до {max_date}")

    start_date = min_date
    end_date = max_date

    date_set = set(dates)
    logger.debug(f"Общее количество уникальных дат с данными: {len(date_set)}")

    num_days = len(date_set)
    cost_per_day = 100 / num_days
    logger.debug(
        f"Количество дней с данными: {num_days}, стоимость дня: {cost_per_day}"
    )

    average_attendance = get_average_attendance_for_period(staff, start_date, end_date)
    logger.debug(f"Средняя посещаемость за период: {average_attendance}%")

    k_adj = 1.25

    if average_attendance <= 0:
        logger.error(
            f"Критическая ошибка: средняя посещаемость равна нулю или отрицательна ({average_attendance}). "
            f"Используется дефолтное значение 85.0% для расчета штрафного коэффициента."
        )
        average_attendance = 85.0

    penalty_rate = (100 / average_attendance) * k_adj
    logger.debug(
        f"Расчет штрафного коэффициента: penalty_rate = (100 / {average_attendance}) * {k_adj} = {penalty_rate:.4f}"
    )

    salary_qs = models.Salary.objects.filter(staff=staff).first()
    contract_type = salary_qs.contract_type if salary_qs else "full_time"
    total_minutes_expected_per_day = get_expected_minutes_per_day(contract_type)
    logger.debug(
        f"Тип контракта: {contract_type}, ожидаемые минуты в день: {total_minutes_expected_per_day}"
    )

    for event_date in sorted(date_set):
        logger.debug(f"Обработка даты: {event_date}")

        attendance = combined_attendance.get(event_date)
        (
            attendance_record,
            total_minutes_for_period,
            total_days_with_data,
            percent_for_period,
        ) = process_attendance(
            attendance,
            event_date,
            start_date,
            end_date,
            holiday_dict,
            total_minutes_expected_per_day,
            cost_per_day,
            penalty_rate,
            total_minutes_for_period,
            total_days_with_data,
            percent_for_period,
            remote_work_qs,
            absent_reason_qs,
        )

        if attendance_record:
            attendance_data[event_date.strftime("%d-%m-%Y")] = attendance_record
            logger.debug(
                f"Добавлена запись посещаемости для {event_date}: {attendance_record}"
            )

    if total_days_with_data > 0:
        percent_for_period /= total_days_with_data
        percent_for_period = max(percent_for_period, 0)
        logger.debug(f"Итоговый процент за период: {percent_for_period}")
    else:
        percent_for_period = 0.0
        logger.debug("Нет рабочих дней для расчета процента за период.")

    num_days = len(date_set)
    bonus_percentage = utils.get_bonus_percentage(num_days, percent_for_period)
    logger.info(
        f"Рассчитанный бонус: {bonus_percentage}% для {num_days} дней и {percent_for_period}% присутствия."
    )

    avatar_url = staff.avatar.url if staff.avatar else "/media/images/no-avatar.png"
    logger.debug(f"URL аватара: {avatar_url}")

    staff_detail_payload = {
        "name": staff.name,
        "surname": staff.surname if staff.surname != "Нет фамилии" else "",
        "positions": [position.name for position in staff.positions.all()],
        "avatar": avatar_url,
        "department": staff.department.name if staff.department else "N/A",
        "department_id": staff.department.id if staff.department else "N/A",
        "attendance": attendance_data,
        "percent_for_period": round(percent_for_period, 2),
        "bonus_percentage": bonus_percentage,
        "contract_type": salary_qs.contract_type if salary_qs else None,
        "salary": salary_qs.total_salary if salary_qs else None,
    }

    logger.info(
        f"Генерация деталей сотрудника завершена для {staff.name} (PIN: {staff.pin})"
    )
    return staff_detail_payload


def get_average_attendance_for_period(staff, start_date, end_date):
    """Считает средний процент присутствия за предыдущие 30 дней для KPI.

    Используется при расчёте штрафного коэффициента в get_staff_detail. При
    наличии effective_work_seconds берётся он, иначе (last_out - first_in).
    Норма — 8 часов в день.

    Args:
        staff: Экземпляр models.Staff.
        start_date: Начало текущего запрошенного периода, date.
        end_date: Конец текущего запрошенного периода, date.

    Returns:
        Число с плавающей точкой (процент 0–100+). Не менее 1.0 и не более
        разумного; при отсутствии данных возвращается 85.0.
    """
    logger.info(
        f"Calculating average attendance for staff {staff.name} (PIN: {staff.pin}) from {start_date} to {end_date}"
    )

    previous_start_date = start_date - datetime.timedelta(days=30)
    previous_end_date = end_date - datetime.timedelta(days=30)
    logger.debug(f"Previous period range: {previous_start_date} to {previous_end_date}")

    previous_attendance_qs = models.StaffAttendance.objects.filter(
        staff=staff,
        date_at__gte=previous_start_date + datetime.timedelta(days=1),
        date_at__lte=previous_end_date + datetime.timedelta(days=1),
    ).only("id", "date_at", "first_in", "last_out", "effective_work_seconds")
    logger.debug(
        f"Retrieved {previous_attendance_qs.count()} attendance records for previous period"
    )

    if not previous_attendance_qs.exists():
        logger.warning(
            "No attendance records found for the previous period. Returning default average attendance of 85.0%"
        )
        return 85.0

    total_minutes = 0
    total_days = 0

    for attendance in previous_attendance_qs:
        if attendance.effective_work_seconds is not None:
            minutes_present = attendance.effective_work_seconds / 60.0
        elif attendance.first_in and attendance.last_out:
            minutes_present = (
                attendance.last_out - attendance.first_in
            ).total_seconds() / 60
        else:
            minutes_present = 0
        if minutes_present > 0:
            total_minutes += minutes_present
            total_days += 1
            logger.debug(
                f"Processed attendance for {attendance.date_at}: {minutes_present} minutes present"
            )

    if total_days == 0:
        logger.warning(
            "No complete attendance days found for the previous period. Returning default average attendance of 85.0%"
        )
        return 85.0

    average_attendance = (total_minutes / (total_days * 8 * 60)) * 100

    if average_attendance <= 0:
        logger.warning(
            f"Calculated average attendance is {average_attendance}% (invalid). "
            f"Using default value of 85.0% for KPI calculation."
        )
        return 85.0

    if average_attendance < 1.0:
        logger.warning(
            f"Calculated average attendance is very low ({average_attendance}%). "
            f"Using minimum threshold of 1.0% for KPI calculation to prevent unrealistic penalties."
        )
        return 1.0

    logger.info(
        f"Calculated average attendance for previous period: {average_attendance}%"
    )
    return average_attendance


def get_expected_minutes_per_day(contract_type):
    """Возвращает норму рабочих минут в день по типу контракта.

    Args:
        contract_type: Строка типа контракта (например "part_time", "gph", "full_time").

    Returns:
        Целое число минут: 240 для part_time/gph, 480 для остальных.
    """
    if contract_type in ["part_time", "gph"]:
        return 4 * 60
    return 8 * 60


def process_attendance(
    attendance,
    event_date,
    start_date,
    end_date,
    holiday_dict,
    total_minutes_expected_per_day,
    cost_per_day,
    penalty_rate,
    total_minutes_for_period,
    total_days_with_data,
    percent_for_period,
    remote_work_qs,
    absent_reason_qs,
):
    """Обрабатывает посещаемость за одну дату и обновляет накопительные показатели периода.

    Учитывает выходные, праздники, удалёнку и утверждённые отсутствия. Для расчёта
    минут используется effective_work_seconds при наличии, иначе first_in/last_out
    (с учётом обеда через calculate_effective_minutes_with_lunch).

    Args:
        attendance: Словарь объединённой посещаемости за дату (first_in, last_out,
            effective_work_seconds, area_sequence и др.) или None.
        event_date: Дата события, date.
        start_date: Начало периода, date.
        end_date: Конец периода, date.
        holiday_dict: Словарь {date: is_working_day} по праздникам.
        total_minutes_expected_per_day: Норма минут в день (int).
        cost_per_day: Доля стоимости одного дня в процентах (float).
        penalty_rate: Штрафной коэффициент (float).
        total_minutes_for_period: Накопленные минуты за период (float).
        total_days_with_data: Количество дней с данными (int).
        percent_for_period: Накопленный процент за период (float).
        remote_work_qs: QuerySet периодов дистанционной работы сотрудника.
        absent_reason_qs: QuerySet причин отсутствия сотрудника.

    Returns:
        Кортеж (attendance_record, total_minutes_for_period, total_days_with_data,
        percent_for_period). attendance_record — словарь для ответа API или None.
    """
    logger.info(f"Обработка посещаемости за дату {event_date}")

    if not start_date <= event_date <= end_date:
        logger.warning(
            f"Дата события {event_date} вне указанного диапазона {start_date} до {end_date}"
        )
        return None, total_minutes_for_period, total_days_with_data, percent_for_period

    is_off_day = check_off_day(event_date, holiday_dict)
    logger.debug(f"Является ли выходным днем: {is_off_day} для даты {event_date}")

    absent_reason = absent_reason_qs.filter(
        start_date__lte=event_date, end_date__gte=event_date
    ).first()

    first_in = attendance.get("first_in") if attendance else None
    last_out = attendance.get("last_out") if attendance else None
    area_name_in = attendance.get("area_name_in") if attendance else None
    area_name_out = attendance.get("area_name_out") if attendance else None
    first_in_source = attendance.get("first_in_source") if attendance else None
    last_out_source = attendance.get("last_out_source") if attendance else None
    effective_work_seconds = (
        attendance.get("effective_work_seconds") if attendance else None
    )
    area_sequence = attendance.get("area_sequence") if attendance else None

    if is_off_day:
        if effective_work_seconds is not None:
            total_minutes_worked = effective_work_seconds / 60.0
            percent_day = (total_minutes_worked / total_minutes_expected_per_day) * 100
        elif first_in and last_out:
            total_minutes_worked = calculate_effective_minutes_with_lunch(
                first_in, last_out
            )
            percent_day = (total_minutes_worked / total_minutes_expected_per_day) * 100
            logger.info(
                f"Сотрудник работал в выходной день {event_date}. Данные отображаются, но не влияют на расчеты."
            )
        else:
            total_minutes_worked = 0
            percent_day = 0
            logger.info(
                f"Выходной день {event_date} без данных о посещаемости. Пропускаем."
            )

        attendance_record = {
            "first_in": (
                first_in.astimezone(timezone.get_current_timezone())
                if first_in
                else None
            ),
            "last_out": (
                last_out.astimezone(timezone.get_current_timezone())
                if last_out
                else None
            ),
            "area_name_in": area_name_in,
            "area_name_out": area_name_out,
            "first_in_source": first_in_source,
            "last_out_source": last_out_source,
            "percent_day": round(percent_day, 2),
            "total_minutes": round(total_minutes_worked, 2),
            "effective_work_seconds": effective_work_seconds,
            "area_sequence": area_sequence,
            "is_weekend": True,
            "is_remote_work": False,
            "is_absent_approved": False,
            "absent_reason": None,
        }
        return (
            attendance_record,
            total_minutes_for_period,
            total_days_with_data,
            percent_for_period,
        )

    is_remote_work = remote_work_qs.filter(
        Q(permanent_remote=True)
        | Q(start_date__lte=event_date, end_date__gte=event_date)
    ).exists()

    is_absent_approved = False
    absent_reason_display = None

    if is_remote_work:
        has_physical_attendance = effective_work_seconds is not None or (
            first_in is not None and last_out is not None
        )
        if has_physical_attendance:
            if effective_work_seconds is not None:
                total_minutes_worked = effective_work_seconds / 60.0
            else:
                total_minutes_worked = calculate_effective_minutes_with_lunch(
                    first_in, last_out
                )
            percent_day = (
                total_minutes_worked / total_minutes_expected_per_day * 100
                if total_minutes_expected_per_day
                else 100.0
            )
            total_minutes_for_period += total_minutes_worked
            total_days_with_data += 1
            percent_for_period += percent_day
            logger.info(
                f"{event_date} дистанционный день с явкой: {total_minutes_worked:.1f} мин, {percent_day:.1f}%."
            )
        else:
            percent_day = 100.0
            total_minutes_worked = total_minutes_expected_per_day
            total_minutes_for_period += total_minutes_worked
            total_days_with_data += 1
            percent_for_period += percent_day
            logger.info(
                f"{event_date} отмечен как день дистанционной работы (без явки)."
            )
    elif absent_reason:
        is_absent_approved = absent_reason.approved
        absent_reason_display = absent_reason.get_reason_display()
        if is_absent_approved:
            logger.info(
                f"{event_date} утвержденная причина отсутствия: {absent_reason_display}."
            )
            attendance_record = {
                "first_in": (
                    first_in.astimezone(timezone.get_current_timezone())
                    if first_in
                    else None
                ),
                "last_out": (
                    last_out.astimezone(timezone.get_current_timezone())
                    if last_out
                    else None
                ),
                "area_name_in": area_name_in,
                "area_name_out": area_name_out,
                "first_in_source": first_in_source,
                "last_out_source": last_out_source,
                "percent_day": 0,
                "total_minutes": 0,
                "effective_work_seconds": effective_work_seconds,
                "area_sequence": area_sequence,
                "is_weekend": False,
                "is_remote_work": False,
                "is_absent_approved": True,
                "absent_reason": absent_reason_display,
            }
            return (
                attendance_record,
                total_minutes_for_period,
                total_days_with_data,
                percent_for_period,
            )
        else:
            percent_day = 0
            total_minutes_worked = 0
            total_days_with_data += 1
            penalty = penalty_rate * cost_per_day
            percent_for_period -= penalty
            logger.warning(
                f"{event_date} неутвержденная причина отсутствия: {absent_reason_display}. Применяется штраф {penalty}%."
            )
    else:
        if effective_work_seconds is not None:
            total_minutes_worked = effective_work_seconds / 60.0
            percent_day = (total_minutes_worked / total_minutes_expected_per_day) * 100
        elif first_in and last_out:
            total_minutes_worked = calculate_effective_minutes_with_lunch(
                first_in, last_out
            )
            percent_day = (total_minutes_worked / total_minutes_expected_per_day) * 100
        else:
            percent_day = 0
            total_minutes_worked = 0

        if effective_work_seconds is not None or (first_in and last_out):
            total_minutes_for_period += total_minutes_worked
            total_days_with_data += 1
            percent_for_period += percent_day
            logger.debug(
                f"Отработано минут: {total_minutes_worked}, Процент дня: {percent_day}"
            )
        else:
            total_days_with_data += 1
            penalty = penalty_rate * cost_per_day
            percent_for_period -= penalty
            logger.warning(
                f"Нет записей о посещаемости за дату {event_date}. Применяется штраф {penalty}%. "
            )

    attendance_record = {
        "first_in": (
            first_in.astimezone(timezone.get_current_timezone()) if first_in else None
        ),
        "last_out": (
            last_out.astimezone(timezone.get_current_timezone()) if last_out else None
        ),
        "area_name_in": area_name_in,
        "area_name_out": area_name_out,
        "first_in_source": first_in_source,
        "last_out_source": last_out_source,
        "percent_day": round(percent_day, 2),
        "total_minutes": round(total_minutes_worked, 2),
        "effective_work_seconds": effective_work_seconds,
        "area_sequence": area_sequence,
        "is_weekend": is_off_day,
        "is_remote_work": is_remote_work,
        "is_absent_approved": is_absent_approved,
        "absent_reason": absent_reason_display,
    }
    logger.info(
        f"Обработана запись посещаемости за дату {event_date}: {attendance_record}"
    )

    return (
        attendance_record,
        total_minutes_for_period,
        total_days_with_data,
        percent_for_period,
    )


def check_off_day(event_date, holiday_dict):
    """
    Проверка, является ли дата выходным или праздничным днем.

    Args:
        event_date (datetime.date): Дата для проверки.
        holiday_dict (dict): Словарь с информацией о праздничных днях.

    Returns:
        bool: True, если день является выходным или праздничным, иначе False.
    """
    is_weekend = event_date.weekday() >= 5
    is_holiday = event_date in holiday_dict
    return (is_weekend and event_date not in holiday_dict) or (
        is_holiday and not holiday_dict[event_date]
    )


def update_percent_for_period(
    percent_for_period,
    percent_day,
    is_off_day,
    total_minutes_worked,
    cost_per_day,
    penalty_rate,
):
    """
    Обновление накопленного процента за период на основе ежедневной посещаемости.

    Args:
        percent_for_period (float): Текущий накопленный процент.
        percent_day (float): Процент присутствия за день.
        is_off_day (bool): Является ли день выходным.
        total_minutes_worked (float): Отработано минут за день.
        cost_per_day (float): Стоимость одного дня в процентах.
        penalty_rate (float): Штрафной коэффициент за отсутствие.

    Returns:
        float: Обновленный процент за период.
    """
    logger.debug(
        f"Updating percent for period. Initial: {percent_for_period}%, "
        f"Day percent: {percent_day}%, Is off day: {is_off_day}, "
        f"Total minutes worked: {total_minutes_worked}, "
        f"Cost per day: {cost_per_day}%, Penalty rate: {penalty_rate}%"
    )

    if is_off_day and total_minutes_worked > 0:
        percent_for_period += percent_day * 1.5
        logger.info(f"Off day with work. Increasing percent by {percent_day * 1.5}%.")
    elif not is_off_day and total_minutes_worked == 0:
        penalty = penalty_rate * cost_per_day
        percent_for_period -= penalty
        logger.warning(f"Workday with no work. Decreasing percent by {penalty}%.")
    else:
        percent_for_period += percent_day
        logger.info(f"Regular day. Adding {percent_day}% to the period percent.")

    logger.debug(f"Updated percent for period: {percent_for_period}%")
    return percent_for_period


@swagger_auto_schema(
    method="get",
    operation_summary="Статус задачи посещаемости",
    operation_description=(
        "Как использовать:\n"
        "1. Возьмите task_id из POST /api/lesson_attendance/ или POST /api/lesson_attendance/json/.\n"
        "2. Подставьте task_id в URL и выполните GET.\n\n"
        "Коды ответа:\n"
        "202 - задача в очереди.\n"
        "200 - задача выполнена, вернутся lesson_ids.\n"
        "500 - ошибка выполнения."
    ),
    tags=["Lesson Attendance"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
        openapi.Parameter(
            "task_id",
            openapi.IN_PATH,
            description="ID задачи из ответа 202",
            type=openapi.TYPE_STRING,
            required=True,
        ),
    ],
    responses={
        200: openapi.Response(
            description="Задача выполнена успешно или в процессе выполнения",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "status": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Статус задачи (Success, Pending, или другой)",
                    ),
                    "lesson_ids": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(type=openapi.TYPE_INTEGER),
                        description="Список ID созданных записей посещаемости (только при Success)",
                    ),
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение о статусе задачи",
                    ),
                },
            ),
        ),
        202: openapi.Response(
            description="Задача в очереди, ожидает выполнения",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "status": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Статус задачи (Pending)",
                    ),
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение о том, что задача в очереди",
                    ),
                },
            ),
        ),
        500: openapi.Response(
            description="Ошибка при выполнении задачи или проверке статуса",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "status": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Статус задачи (Failure)",
                    ),
                    "error": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Описание ошибки",
                    ),
                },
            ),
        ),
    },
)
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def check_lesson_task_status(request, task_id):
    """
    Проверка статуса задачи и получение lesson_id.
    """
    log_prefix = "[lesson_attendance]"
    ip_address = request.META.get("REMOTE_ADDR", "Неизвестный IP")

    try:
        task_result = AsyncResult(task_id)

        if task_result.state == "PENDING":
            lesson_attendance_logger.debug(
                "%s task_status PENDING task_id=%s ip=%s",
                log_prefix,
                task_id,
                ip_address,
            )
            return Response(
                {
                    "status": "Pending",
                    "message": "Задача в очереди, ожидайте завершения",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        elif task_result.state == "SUCCESS":
            result = task_result.result or {}
            success_records = result.get("success_records", [])
            error_records = result.get("error_records", [])
            lesson_attendance_logger.info(
                "%s task_status SUCCESS task_id=%s ip=%s created=%s failed=%s",
                log_prefix,
                task_id,
                ip_address,
                len(success_records),
                len(error_records),
            )
            if error_records:
                lesson_attendance_logger.warning(
                    "%s task_status partial_failures task_id=%s error_records=%s",
                    log_prefix,
                    task_id,
                    error_records,
                )
            return Response(
                {"status": "Success", "lesson_ids": success_records},
                status=status.HTTP_200_OK,
            )

        elif task_result.state == "FAILURE":
            lesson_attendance_logger.warning(
                "%s task_status FAILURE task_id=%s ip=%s error=%s",
                log_prefix,
                task_id,
                ip_address,
                str(task_result.info),
            )
            return Response(
                {"status": "Failure", "error": str(task_result.info)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        else:
            lesson_attendance_logger.info(
                "%s task_status state=%s task_id=%s ip=%s",
                log_prefix,
                task_result.state,
                task_id,
                ip_address,
            )
            return Response(
                {
                    "status": task_result.state,
                    "message": "Задача в процессе выполнения",
                },
                status=status.HTTP_200_OK,
            )

    except Exception as e:
        lesson_attendance_logger.exception(
            "%s task_status EXCEPTION task_id=%s ip=%s error=%s",
            log_prefix,
            task_id,
            ip_address,
            str(e),
        )
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _lesson_attendance_responses():
    """Общие ответы для создания записей посещаемости (multipart и JSON)."""
    return {
        202: openapi.Response(
            description="Задача принята в очередь. Идентификатор задачи — в task_id.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(type=openapi.TYPE_STRING),
                    "task_id": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Проверка: GET /api/lesson_attendance/task_status/{task_id}/",
                    ),
                },
            ),
        ),
        400: openapi.Response(
            description="Ошибка валидации.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                    "missing": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(type=openapi.TYPE_STRING),
                    ),
                },
            ),
        ),
        500: openapi.Response(
            description="Ошибка сервера.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={"error": openapi.Schema(type=openapi.TYPE_STRING)},
            ),
        ),
    }


LESSON_ATTENDANCE_RECORD_EXAMPLE = {
    "staff_pin": "T861T",
    "tutor_id": 101,
    "tutor": "Нео Андерсон",
    "first_in": "2026-03-16T10:00:00+05:00",
    "latitude": 43.2389,
    "longitude": 76.8897,
    "subject_name": "Матрица и Морбиус",
}
LESSON_ATTENDANCE_ARRAY_EXAMPLE_TEXT = json.dumps(
    [LESSON_ATTENDANCE_RECORD_EXAMPLE], ensure_ascii=False, indent=2
)
LESSON_ATTENDANCE_JSON_BODY_EXAMPLE = {
    "attendance_data": [LESSON_ATTENDANCE_RECORD_EXAMPLE],
    "image": "<base64>",
}
LESSON_ATTENDANCE_JSON_BODY_EXAMPLE_TEXT = json.dumps(
    LESSON_ATTENDANCE_JSON_BODY_EXAMPLE, ensure_ascii=False, indent=2
)


@swagger_auto_schema(
    method="post",
    auto_schema=FormOnlySwaggerAutoSchema,  # type: ignore[reportArgumentType]
    operation_summary="Создать записи посещаемости (multipart)",
    operation_description=(
        "Try it out (в Swagger):\n"
        "1. Нажмите Try it out.\n"
        "2. В поле attendance_data уже подставлен рабочий шаблон. "
        "Скопируйте его и замените staff_pin, tutor, first_in и координаты под ваш кейс.\n"
        "3. Если нужна фотофиксация, в поле image выберите файл через Choose File.\n"
        "4. Нажмите Execute.\n\n"
        "Обязательные поля записи: staff_pin, tutor_id, tutor, first_in, latitude, longitude.\n"
        "subject_name - опционально.\n\n"
        "Ответ: 202 + task_id.\n"
        "Проверка статуса: GET /api/lesson_attendance/task_status/{task_id}/."
    ),
    tags=["Lesson Attendance"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API-ключ (альтернатива JWT).",
        ),
        openapi.Parameter(
            name="attendance_data",
            in_=openapi.IN_FORM,
            type=openapi.TYPE_STRING,
            required=True,
            description=(
                "JSON-строка с массивом записей.\n"
                "Что обычно меняют: staff_pin, tutor, first_in, latitude, longitude.\n"
                "Готовый шаблон (скопируйте как есть):\n"
                f"{LESSON_ATTENDANCE_ARRAY_EXAMPLE_TEXT}"
            ),
            default=LESSON_ATTENDANCE_ARRAY_EXAMPLE_TEXT,
        ),
        openapi.Parameter(
            name="image",
            in_=openapi.IN_FORM,
            type=openapi.TYPE_FILE,
            required=False,
            description="Файл фотографии (JPG/PNG), опционально. Загружайте через Choose File.",
        ),
    ],
    request_body=no_body,
    consumes=["multipart/form-data"],
    responses=_lesson_attendance_responses(),
)
@api_view(["POST"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def create_lesson_attendance(request):
    """
    POST /api/lesson_attendance/ — создание записей посещаемости занятий (с фото).

    Как заполнять запрос
    --------------------
    Поддерживаются два типа запроса.

    **1. multipart/form-data** (форма с файлом, в т.ч. в Swagger «Try it out»):
    - Поле **attendance_data**: строка, содержащая JSON-массив объектов. Каждый объект — одна запись посещаемости.
    - Поле **image**: файл изображения (JPG/PNG), опционально.

    Обязательные поля в каждом объекте массива attendance_data:
    - staff_pin (str): PIN сотрудника из справочника Staff.
    - tutor_id (int): ID преподавателя (допускается 0).
    - tutor (str): ФИО преподавателя.
    - first_in (str): Время начала занятия, ISO 8601 с таймзоной, напр. "2024-10-06T14:24:24+05:00".
    - latitude (float): Широта (допускается 0).
    - longitude (float): Долгота (допускается 0).

    Необязательное поле: subject_name (str).

    Пример значения для attendance_data (одна запись):
        [
          {
            "staff_pin": "s00260",
            "tutor_id": 1,
            "tutor": "Иванов И.И.",
            "first_in": "2024-10-06T14:24:24+05:00",
            "latitude": 43.21,
            "longitude": 76.85
          }
        ]

    **2. application/json** (предпочтительно отправлять на POST /api/lesson_attendance/json/):
    - Тело:
      {
        "attendance_data": [ {...}, ... ],
        "image": "<base64-строка>" (опционально)
      }.
    - Структура объектов в attendance_data — та же, image — фото в Base64 без префикса data:... (опционально).

    Ответ
    -----
    - 202: в теле {"message": "Task accepted", "task_id": "<uuid>"}. Результат проверять в GET /api/lesson_attendance/task_status/<task_id>/.
    - 400: ошибка валидации (нет полей, неверный JSON и т.п.).
    - 500: ошибка сервера при постановке задачи в очередь.
    """
    ip_address = request.META.get("REMOTE_ADDR", "Неизвестный IP")
    domain = request.get_host()
    log_prefix = "[lesson_attendance]"

    lesson_attendance_logger.info(
        "%s POST create ip=%s host=%s content_type=%s",
        log_prefix,
        ip_address,
        domain,
        getattr(request, "content_type", None),
    )

    try:
        ct = (getattr(request, "content_type") or "") or ""
        if "multipart" in ct or "form-data" in ct:
            attendance_data_raw = request.POST.get(
                "attendance_data"
            ) or request.data.get("attendance_data")
            image_base64 = request.POST.get("image") or request.data.get("image")
        else:
            attendance_data_raw = request.data.get("attendance_data")
            image_base64 = request.data.get("image")
        has_file = bool(request.FILES.get("image"))

        if not attendance_data_raw:
            _data_keys = (
                list(request.data.keys()) if getattr(request, "data", None) else []
            )
            _post_keys = list(request.POST.keys()) if hasattr(request, "POST") else []
            lesson_attendance_logger.warning(
                "%s BAD_REQUEST attendance_data_missing ip=%s data_keys=%s post_keys=%s",
                log_prefix,
                ip_address,
                _data_keys,
                _post_keys,
            )
            return Response(
                {"error": "Attendance data is missing"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if isinstance(attendance_data_raw, list):
            attendance_data = attendance_data_raw
        elif isinstance(attendance_data_raw, (str, bytes)):
            attendance_data_raw = (
                attendance_data_raw.decode("utf-8")
                if isinstance(attendance_data_raw, bytes)
                else attendance_data_raw
            )
            try:
                attendance_data = json.loads(attendance_data_raw)
            except json.JSONDecodeError as e:
                lesson_attendance_logger.warning(
                    "%s BAD_REQUEST attendance_data_not_json ip=%s error=%s",
                    log_prefix,
                    ip_address,
                    str(e),
                )
                return Response(
                    {"error": "Invalid JSON in attendance_data"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            attendance_data = []

        if not isinstance(attendance_data, list):
            lesson_attendance_logger.warning(
                "%s BAD_REQUEST attendance_data_not_list ip=%s type=%s",
                log_prefix,
                ip_address,
                type(attendance_data).__name__,
            )
            return Response(
                {"error": "attendance_data must be a list"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def _missing_required(rec):
            missing = []
            if not rec.get("staff_pin"):
                missing.append("staff_pin")
            if "tutor_id" not in rec:
                missing.append("tutor_id")
            if not rec.get("tutor"):
                missing.append("tutor")
            if not rec.get("first_in"):
                missing.append("first_in")
            if "latitude" not in rec:
                missing.append("latitude")
            if "longitude" not in rec:
                missing.append("longitude")
            return missing

        for idx, record in enumerate(attendance_data):
            if not isinstance(record, dict):
                lesson_attendance_logger.warning(
                    "%s BAD_REQUEST record_not_dict ip=%s record_index=%s",
                    log_prefix,
                    ip_address,
                    idx,
                )
                return Response(
                    {"error": "Each attendance record must be an object"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            missing = _missing_required(record)
            if missing:
                lesson_attendance_logger.warning(
                    "%s BAD_REQUEST missing_required_fields ip=%s record_index=%s missing=%s staff_pin=%s tutor_id=%s",
                    log_prefix,
                    ip_address,
                    idx,
                    missing,
                    record.get("staff_pin"),
                    record.get("tutor_id"),
                )
                return Response(
                    {"error": "Missing required fields in record", "missing": missing},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        image_content = None
        image_name = None

        if has_file:
            staff_image = request.FILES["image"]
            image_content = staff_image.read()
        elif image_base64:
            try:
                image_content = base64.b64decode(image_base64)
            except Exception as e:
                lesson_attendance_logger.warning(
                    "%s BAD_REQUEST image_base64_decode_failed ip=%s error=%s",
                    log_prefix,
                    ip_address,
                    str(e),
                )
                return Response(
                    {"error": "Invalid Base64 image format"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        task = tasks.process_lesson_attendance_batch.apply_async(
            args=[attendance_data, image_name, image_content]
        )
        lesson_attendance_logger.info(
            "%s ACCEPTED task_id=%s ip=%s records_count=%s image_size_bytes=%s",
            log_prefix,
            task.id,
            ip_address,
            len(attendance_data),
            len(image_content) if image_content else 0,
        )

        return Response(
            {"message": "Task accepted", "task_id": task.id},
            status=status.HTTP_202_ACCEPTED,
        )

    except Exception as e:
        lesson_attendance_logger.exception(
            "%s EXCEPTION create ip=%s error=%s",
            log_prefix,
            ip_address,
            str(e),
        )
        return Response(
            {"error": "Error with creating job"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


_lesson_attendance_json_schema = openapi.Schema(
    type=openapi.TYPE_OBJECT,
    required=["attendance_data"],
    description="Тело запроса: массив записей посещаемости; фото в Base64 опционально.",
    example=LESSON_ATTENDANCE_JSON_BODY_EXAMPLE,
    properties={
        "attendance_data": openapi.Schema(
            type=openapi.TYPE_ARRAY,
            description="Массив записей посещаемости.",
            items=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                required=[
                    "staff_pin",
                    "tutor_id",
                    "tutor",
                    "first_in",
                    "latitude",
                    "longitude",
                ],
                properties={
                    "staff_pin": openapi.Schema(
                        type=openapi.TYPE_STRING, example="s00260"
                    ),
                    "tutor_id": openapi.Schema(type=openapi.TYPE_INTEGER, example=1),
                    "tutor": openapi.Schema(
                        type=openapi.TYPE_STRING, example="Иванов И.И."
                    ),
                    "first_in": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        format=openapi.FORMAT_DATETIME,
                        example="2024-10-06T14:24:24+05:00",
                    ),
                    "latitude": openapi.Schema(
                        type=openapi.TYPE_NUMBER, example=43.207674
                    ),
                    "longitude": openapi.Schema(
                        type=openapi.TYPE_NUMBER, example=76.851377
                    ),
                    "subject_name": openapi.Schema(
                        type=openapi.TYPE_STRING, example="Математика"
                    ),
                },
            ),
        ),
        "image": openapi.Schema(
            type=openapi.TYPE_STRING,
            description="Фото в Base64 (без префикса data:image/...;base64,). Опционально.",
        ),
    },
)


@swagger_auto_schema(
    method="post",
    operation_summary="Создать записи посещаемости (JSON)",
    operation_description=(
        "Формат запроса: application/json.\n"
        "Тело запроса:\n"
        f"{LESSON_ATTENDANCE_JSON_BODY_EXAMPLE_TEXT}\n"
        "image - опционально.\n\n"
        "Обязательные поля записи: staff_pin, tutor_id, tutor, first_in, latitude, longitude.\n"
        "subject_name - опционально.\n\n"
        "Ответ: 202 + task_id.\n"
        "Проверка статуса: GET /api/lesson_attendance/task_status/{task_id}/."
    ),
    tags=["Lesson Attendance"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API-ключ (альтернатива JWT).",
        ),
    ],
    request_body=_lesson_attendance_json_schema,
    consumes=["application/json"],
    responses=_lesson_attendance_responses(),
)
@api_view(["POST"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def create_lesson_attendance_json(request):
    return create_lesson_attendance(getattr(request, "_request", request))


@swagger_auto_schema(
    method="put",
    auto_schema=FormOnlySwaggerAutoSchema,  # type: ignore[reportArgumentType]
    operation_summary="Обновить запись посещаемости",
    operation_description=(
        "Try it out (в Swagger):\n"
        "1. Укажите id записи в path-параметре.\n"
        "2. Заполните только те поля, которые хотите изменить.\n"
        "3. Для фото используйте только поле image (кнопка Choose File).\n"
        "4. Нажмите Execute.\n\n"
        "Пример JSON (для Postman/curl):\n"
        "{\n"
        '  "last_out": "2026-03-16T11:40:00+05:00",\n'
        '  "latitude": 43.2389,\n'
        '  "longitude": 76.8897\n'
        "}"
    ),
    tags=["Lesson Attendance"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
        openapi.Parameter(
            "id",
            openapi.IN_PATH,
            description="ID записи для обновления",
            type=openapi.TYPE_INTEGER,
            required=True,
        ),
        openapi.Parameter(
            "first_in",
            openapi.IN_FORM,
            description="Начало занятия (ISO 8601). Пример: 2026-03-16T10:00:00+05:00",
            type=openapi.TYPE_STRING,
            required=False,
            default="2026-03-16T10:00:00+05:00",
        ),
        openapi.Parameter(
            "last_out",
            openapi.IN_FORM,
            description="Окончание занятия (ISO 8601). Пример: 2026-03-16T11:40:00+05:00",
            type=openapi.TYPE_STRING,
            required=False,
            default="2026-03-16T11:40:00+05:00",
        ),
        openapi.Parameter(
            "latitude",
            openapi.IN_FORM,
            description="Широта. Пример: 43.2389",
            type=openapi.TYPE_NUMBER,
            required=False,
            default=43.2389,
        ),
        openapi.Parameter(
            "longitude",
            openapi.IN_FORM,
            description="Долгота. Пример: 76.8897",
            type=openapi.TYPE_NUMBER,
            required=False,
            default=76.8897,
        ),
        openapi.Parameter(
            name="image",
            in_=openapi.IN_FORM,
            type=openapi.TYPE_FILE,
            required=False,
            description="Фото сотрудника (jpg/png). Загружайте через Choose File.",
        ),
    ],
    request_body=no_body,
    consumes=["multipart/form-data"],
    responses={
        200: openapi.Response(
            description="Запись успешно обновлена",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение об успешном обновлении записи",
                    ),
                    "lesson_id": openapi.Schema(
                        type=openapi.TYPE_INTEGER,
                        description="ID обновленной записи посещаемости",
                    ),
                },
            ),
        ),
        400: openapi.Response(
            description="Неверные данные в теле запроса",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
        404: openapi.Response(
            description="Запись с указанным ID не найдена",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
        500: openapi.Response(
            description="Ошибка сервера",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
    },
)
@api_view(["PUT"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def update_lesson_attendance(request, attendance_id=None, **kwargs):
    """
    Обновление записи посещаемости занятия.

    Args:
        id (int): ID записи для обновления.
        request (Request): HTTP запрос, содержащий данные для обновления записи.

    Все поля в теле запроса опциональны: можно изменить время начала/окончания,
    координаты и/или добавить/обновить фото (image: файл в multipart или Base64 в JSON).
    """
    ip_address = request.META.get("REMOTE_ADDR", "Неизвестный IP")
    log_prefix = "[lesson_attendance]"
    attendance_id = attendance_id if attendance_id is not None else kwargs.get("id")
    if attendance_id is None:
        return Response(
            {"error": "Attendance id is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        lesson_attendance = get_object_or_404(models.LessonAttendance, id=attendance_id)

        image_content = None
        if request.FILES.get("image"):
            image_content = request.FILES["image"].read()
        elif request.data.get("image"):
            try:
                image_content = base64.b64decode(request.data.get("image"))
            except Exception as e:
                lesson_attendance_logger.warning(
                    "%s PUT BAD_REQUEST image_base64_decode_failed id=%s error=%s",
                    log_prefix,
                    attendance_id,
                    str(e),
                )
                return Response(
                    {"error": "Invalid Base64 image format"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        first_in = request.data.get("first_in", lesson_attendance.first_in)
        last_out = request.data.get("last_out")
        latitude = request.data.get("latitude", lesson_attendance.latitude)
        longitude = request.data.get("longitude", lesson_attendance.longitude)

        if last_out is None:
            last_out = lesson_attendance.last_out

        update_fields = ["first_in", "last_out", "latitude", "longitude"]
        if image_content:
            staff_pin = lesson_attendance.staff.pin
            base_dir, file_path = utils.get_lesson_attendance_photo_path(staff_pin)
            os.makedirs(base_dir, exist_ok=True)
            try:
                with open(file_path, "wb") as destination:
                    destination.write(image_content)
            except OSError as e:
                lesson_attendance_logger.error(
                    "%s PUT image_save_failed id=%s path=%s error=%s",
                    log_prefix,
                    attendance_id,
                    file_path,
                    str(e),
                )
                return Response(
                    {"error": f"Image save failed: {e}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            lesson_attendance.staff_image_path = file_path
            update_fields.append("staff_image_path")

        lesson_attendance.first_in = first_in
        lesson_attendance.last_out = last_out
        lesson_attendance.latitude = latitude
        lesson_attendance.longitude = longitude
        lesson_attendance.save(update_fields=update_fields)

        lesson_attendance_logger.info(
            "%s PUT OK lesson_id=%s ip=%s last_out=%s has_photo=%s",
            log_prefix,
            lesson_attendance.id,
            ip_address,
            last_out,
            bool(image_content),
        )
        return Response(
            {
                "message": "LessonAttendance updated successfully",
                "lesson_id": lesson_attendance.id,
            },
            status=status.HTTP_200_OK,
        )

    except models.LessonAttendance.DoesNotExist:
        lesson_attendance_logger.warning(
            "%s PUT NOT_FOUND id=%s ip=%s",
            log_prefix,
            attendance_id,
            ip_address,
        )
        return Response(
            {"error": "LessonAttendance not found."}, status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        lesson_attendance_logger.exception(
            "%s PUT EXCEPTION id=%s ip=%s error=%s",
            log_prefix,
            attendance_id,
            ip_address,
            str(e),
        )
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


PHOTO_VERDICT_ACTION_MANUAL_CLEAN = "manual_clean"
PHOTO_VERDICT_ACTION_MANUAL_SUSPICIOUS = "manual_suspicious"
PHOTO_VERDICT_ACTION_MANUAL_RESET = "manual_reset"
PHOTO_VERDICT_ACTION_RESCAN = "rescan"
PHOTO_VERDICT_ACTIONS = {
    PHOTO_VERDICT_ACTION_MANUAL_CLEAN,
    PHOTO_VERDICT_ACTION_MANUAL_SUSPICIOUS,
    PHOTO_VERDICT_ACTION_MANUAL_RESET,
    PHOTO_VERDICT_ACTION_RESCAN,
}
PHOTO_VERDICT_MAX_IDS_PER_REQUEST = 2000
PHOTO_VERDICT_DEFAULT_LIMIT = 200
PHOTO_VERDICT_MAX_LIMIT = 1000
PHOTO_VERDICT_ONLY_FIELDS = (
    "id",
    "date_at",
    "first_in",
    "staff_image_path",
    "tutor",
    "tutor_id",
    "subject_name",
    "photo_spoof_status",
    "photo_spoof_score",
    "photo_spoof_tags",
    "photo_spoof_checked_at",
    "photo_spoof_model_version",
    "photo_trust_confirmed",
    "photo_manual_verdict",
    "photo_manual_comment",
    "photo_manual_by_id",
    "photo_manual_at",
    "staff__pin",
    "staff__surname",
    "staff__name",
    "staff__department__name",
    "photo_manual_by__username",
)


def _photo_verdict_choices_payload() -> dict[str, Any]:
    manual_reviewable_statuses = [
        models.LessonAttendance.PHOTO_SPOOF_STATUS_PENDING,
        models.LessonAttendance.PHOTO_SPOOF_STATUS_REVIEW,
        models.LessonAttendance.PHOTO_SPOOF_STATUS_ERROR,
    ]
    return {
        "photo_spoof_statuses": [
            {"value": value, "label": label}
            for value, label in models.LessonAttendance.PHOTO_SPOOF_STATUS_CHOICES
        ],
        "photo_manual_verdicts": [
            {"value": value, "label": label}
            for value, label in models.LessonAttendance.PHOTO_MANUAL_VERDICT_CHOICES
        ],
        "actions": [
            {
                "value": PHOTO_VERDICT_ACTION_MANUAL_CLEAN,
                "label": "Ручной вердикт: нормальное",
            },
            {
                "value": PHOTO_VERDICT_ACTION_MANUAL_SUSPICIOUS,
                "label": "Ручной вердикт: подозрительное",
            },
            {
                "value": PHOTO_VERDICT_ACTION_MANUAL_RESET,
                "label": "Сбросить ручной вердикт",
            },
            {
                "value": PHOTO_VERDICT_ACTION_RESCAN,
                "label": "Пересканировать автоматически",
            },
        ],
        "manual_reviewable_statuses": manual_reviewable_statuses,
    }


def _parse_positive_int(value: Any, *, default: int, minimum: int, maximum: int) -> int:
    try:
        numeric = int(value)
    except (TypeError, ValueError):
        return default
    numeric = max(minimum, numeric)
    return min(maximum, numeric)


def _parse_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    lowered = str(value).strip().lower()
    if lowered in {"1", "true", "yes", "y", "on"}:
        return True
    if lowered in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _parse_photo_record_ids(raw_ids: Any) -> list[int]:
    if raw_ids is None:
        return []

    raw_items: list[Any]
    if isinstance(raw_ids, (list, tuple, set)):
        raw_items = list(raw_ids)
    elif isinstance(raw_ids, str):
        raw_items = [part.strip() for part in raw_ids.split(",")]
    else:
        raw_items = [raw_ids]

    parsed: list[int] = []
    seen: set[int] = set()
    for item in raw_items:
        token = str(item).strip()
        if not token:
            continue
        if not token.isdigit():
            raise ValidationError(f"Некорректный id: {token}")
        value = int(token)
        if value <= 0 or value in seen:
            continue
        seen.add(value)
        parsed.append(value)
    if len(parsed) > PHOTO_VERDICT_MAX_IDS_PER_REQUEST:
        raise ValidationError(
            f"Слишком много id. Максимум за запрос: {PHOTO_VERDICT_MAX_IDS_PER_REQUEST}."
        )
    return parsed


def _extract_photo_ids_from_request(request, attendance_id: Optional[int]) -> list[int]:
    if attendance_id is not None:
        return [int(attendance_id)]
    payload = request.data if isinstance(request.data, dict) else {}
    if "ids" in payload:
        return _parse_photo_record_ids(payload.get("ids"))
    if "id" in payload:
        return _parse_photo_record_ids(payload.get("id"))
    if request.query_params.get("ids"):
        return _parse_photo_record_ids(request.query_params.get("ids"))
    if request.query_params.get("id"):
        return _parse_photo_record_ids(request.query_params.get("id"))
    return []


def _serialize_lesson_attendance_photo(
    record: models.LessonAttendance,
) -> dict[str, Any]:
    checked_at = (
        timezone.localtime(record.photo_spoof_checked_at).isoformat()
        if record.photo_spoof_checked_at
        else None
    )
    manual_at = (
        timezone.localtime(record.photo_manual_at).isoformat()
        if record.photo_manual_at
        else None
    )
    manual_by_username = ""
    manual_by = getattr(record, "photo_manual_by", None)
    if manual_by is not None:
        manual_by_username = str(getattr(manual_by, "username", "") or "")

    return {
        "id": record.id,
        "dateAt": record.date_at.isoformat(),
        "hasPhoto": bool(record.staff_image_path),
        "staffPin": record.staff.pin,
        "staffFullName": f"{record.staff.surname} {record.staff.name}",
        "department": (
            record.staff.department.name if record.staff.department else "Unknown"
        ),
        "photoUrl": record.image_url,
        "attendanceTime": timezone.localtime(record.first_in).isoformat(),
        "tutorInfo": record.tutor_info,
        "photoSpoofStatus": record.photo_spoof_status,
        "photoSpoofScore": record.photo_spoof_score,
        "photoSpoofTags": (
            record.photo_spoof_tags if isinstance(record.photo_spoof_tags, list) else []
        ),
        "photoSpoofCheckedAt": checked_at,
        "photoSpoofModelVersion": record.photo_spoof_model_version,
        "photoTrustConfirmed": record.photo_trust_confirmed,
        "photoManualVerdict": record.photo_manual_verdict,
        "photoManualComment": record.photo_manual_comment or "",
        "photoManualBy": record.photo_manual_by_id,
        "photoManualByUsername": manual_by_username,
        "photoManualAt": manual_at,
        "photoEffectiveStatus": record.photo_effective_status,
        "photoEffectiveTrustConfirmed": record.photo_effective_trust_confirmed,
        "photoCanSetManualVerdict": record.photo_can_set_manual_verdict,
    }


def _effective_status_filter(status_value: str) -> Q:
    lesson = models.LessonAttendance
    if status_value == lesson.PHOTO_SPOOF_STATUS_CLEAN:
        return cast(
            Q,
            Q(photo_manual_verdict=lesson.PHOTO_MANUAL_VERDICT_CLEAN)
            | (
                Q(photo_manual_verdict=lesson.PHOTO_MANUAL_VERDICT_NONE)
                & Q(photo_spoof_status=lesson.PHOTO_SPOOF_STATUS_CLEAN)
            ),
        )
    if status_value == lesson.PHOTO_SPOOF_STATUS_SUSPICIOUS:
        return cast(
            Q,
            Q(photo_manual_verdict=lesson.PHOTO_MANUAL_VERDICT_SUSPICIOUS)
            | (
                Q(photo_manual_verdict=lesson.PHOTO_MANUAL_VERDICT_NONE)
                & Q(photo_spoof_status=lesson.PHOTO_SPOOF_STATUS_SUSPICIOUS)
            ),
        )
    return cast(
        Q,
        Q(photo_manual_verdict=lesson.PHOTO_MANUAL_VERDICT_NONE)
        & Q(photo_spoof_status=status_value),
    )


def _sanitize_photo_group_name(name: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_\\-\\.]", "_", name)[:100]


def _invalidate_photo_cache_for_records(records: List[models.LessonAttendance]) -> None:
    if not records:
        return
    unique_dates = {record.date_at for record in records}
    for lesson_date in unique_dates:
        cache.delete(f"photos_for_{lesson_date}")


def _broadcast_photo_updates(records: List[models.LessonAttendance]) -> None:
    if not records:
        return
    channel_layer = get_channel_layer()
    if channel_layer is None:
        return
    version_ts = timezone.now().isoformat()
    grouped_ids: dict[str, list[int]] = {}
    for record in records:
        key = record.date_at.isoformat()
        grouped_ids.setdefault(key, []).append(record.id)

    for iso_date, raw_ids in grouped_ids.items():
        group_name = _sanitize_photo_group_name(f"photos_{iso_date}")
        unique_ids = list(dict.fromkeys(raw_ids))
        try:
            for start in range(0, len(unique_ids), 200):
                chunk = unique_ids[start : start + 200]
                payload = {
                    "type": "new_photo",
                    "attendance_ids": chunk,
                    "op": "updated",
                    "stateCode": "UPDATED_META",
                    "versionTs": version_ts,
                }
                if len(chunk) == 1:
                    payload["attendance_id"] = chunk[0]
                async_to_sync(channel_layer.group_send)(group_name, payload)
        except Exception:
            logger.exception(
                "Failed to broadcast lesson attendance photo bulk update ids=%s date=%s",
                unique_ids[:10],
                iso_date,
            )


@api_view(["GET", "POST", "PUT", "PATCH"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def lesson_attendance_photo_verdicts(request, attendance_id=None):
    """GET — список/детали PAD статусов; POST/PUT/PATCH — ручные и bulk-вердикты."""
    attendance_id = int(attendance_id) if attendance_id is not None else None
    base_qs = (
        models.LessonAttendance.objects.select_related(
            "staff__department", "photo_manual_by"
        )
        .only(*PHOTO_VERDICT_ONLY_FIELDS)
        .order_by("-first_in", "-id")
    )

    if request.method == "GET":
        ids = _extract_photo_ids_from_request(request, attendance_id)
        qs = base_qs
        if ids:
            qs = qs.filter(id__in=ids)
            records = list(qs)
            by_id = {record.id: record for record in records}
            ordered = [by_id[id_value] for id_value in ids if id_value in by_id]
            return Response(
                {
                    "choices": _photo_verdict_choices_payload(),
                    "count": len(ordered),
                    "limit": len(ordered),
                    "offset": 0,
                    "results": [
                        _serialize_lesson_attendance_photo(record) for record in ordered
                    ],
                },
                status=status.HTTP_200_OK,
            )

        date_str = request.query_params.get("date")
        if date_str:
            try:
                filter_date = datetime.date.fromisoformat(date_str)
            except ValueError:
                raise ValidationError("Неверный формат date. Используй YYYY-MM-DD.")
        else:
            filter_date = timezone.localdate()
        qs = qs.filter(date_at=filter_date)

        staff_pin = str(request.query_params.get("staff_pin") or "").strip().upper()
        if staff_pin:
            qs = qs.filter(staff__pin__iexact=staff_pin)

        auto_status = str(request.query_params.get("photo_spoof_status") or "").strip()
        if auto_status:
            allowed_auto_statuses = {
                value for value, _ in models.LessonAttendance.PHOTO_SPOOF_STATUS_CHOICES
            }
            if auto_status not in allowed_auto_statuses:
                raise ValidationError(f"Некорректный photo_spoof_status={auto_status}.")
            qs = qs.filter(photo_spoof_status=auto_status)

        manual_verdict = str(
            request.query_params.get("photo_manual_verdict") or ""
        ).strip()
        if manual_verdict:
            allowed_manual_verdicts = {
                value
                for value, _ in models.LessonAttendance.PHOTO_MANUAL_VERDICT_CHOICES
            }
            if manual_verdict not in allowed_manual_verdicts:
                raise ValidationError(
                    f"Некорректный photo_manual_verdict={manual_verdict}."
                )
            qs = qs.filter(photo_manual_verdict=manual_verdict)

        effective_status = str(
            request.query_params.get("photo_effective_status") or ""
        ).strip()
        if effective_status:
            allowed_effective_statuses = {
                value for value, _ in models.LessonAttendance.PHOTO_SPOOF_STATUS_CHOICES
            }
            if effective_status not in allowed_effective_statuses:
                raise ValidationError(
                    f"Некорректный photo_effective_status={effective_status}."
                )
            qs = qs.filter(_effective_status_filter(effective_status))

        has_photo_param = request.query_params.get("has_photo")
        if has_photo_param is not None:
            has_photo = _parse_bool(has_photo_param, default=True)
            if has_photo:
                qs = qs.filter(staff_image_path__isnull=False).exclude(
                    staff_image_path=""
                )
            else:
                qs = qs.filter(
                    Q(staff_image_path__isnull=True) | Q(staff_image_path="")
                )

        limit = _parse_positive_int(
            request.query_params.get("limit"),
            default=PHOTO_VERDICT_DEFAULT_LIMIT,
            minimum=1,
            maximum=PHOTO_VERDICT_MAX_LIMIT,
        )
        offset = _parse_positive_int(
            request.query_params.get("offset"),
            default=0,
            minimum=0,
            maximum=200000,
        )

        total = qs.count()
        records = list(qs[offset : offset + limit])
        return Response(
            {
                "choices": _photo_verdict_choices_payload(),
                "count": total,
                "limit": limit,
                "offset": offset,
                "results": [
                    _serialize_lesson_attendance_photo(record) for record in records
                ],
            },
            status=status.HTTP_200_OK,
        )

    ids = _extract_photo_ids_from_request(request, attendance_id)
    if not ids:
        raise ValidationError("Передай id или ids для изменения вердикта.")

    payload = request.data if isinstance(request.data, dict) else {}
    action = str(payload.get("action") or "").strip().lower()
    manual_verdict = str(payload.get("manual_verdict") or "").strip().lower()
    if not action and manual_verdict:
        if manual_verdict == models.LessonAttendance.PHOTO_MANUAL_VERDICT_CLEAN:
            action = PHOTO_VERDICT_ACTION_MANUAL_CLEAN
        elif manual_verdict == models.LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS:
            action = PHOTO_VERDICT_ACTION_MANUAL_SUSPICIOUS
        elif manual_verdict == models.LessonAttendance.PHOTO_MANUAL_VERDICT_NONE:
            action = PHOTO_VERDICT_ACTION_MANUAL_RESET

    if action not in PHOTO_VERDICT_ACTIONS:
        raise ValidationError(
            "Некорректный action. Используй one of: "
            f"{', '.join(sorted(PHOTO_VERDICT_ACTIONS))}."
        )

    records = list(base_qs.filter(id__in=ids))
    if not records:
        return Response(
            {"detail": "Записи не найдены."},
            status=status.HTTP_404_NOT_FOUND,
        )

    by_id = {record.id: record for record in records}
    ordered_records = [by_id[id_value] for id_value in ids if id_value in by_id]

    changed_records: list[models.LessonAttendance] = []
    skipped_ids: list[int] = []
    skipped_reasons: list[dict[str, Any]] = []
    error_items: list[dict[str, Any]] = []

    def _reason_manual_verdict_unavailable(record: models.LessonAttendance) -> str:
        if not bool(record.staff_image_path):
            return "no_photo"
        if (
            record.photo_manual_verdict
            != models.LessonAttendance.PHOTO_MANUAL_VERDICT_NONE
        ):
            return "verdict_already_set"
        if record.photo_spoof_status not in {
            models.LessonAttendance.PHOTO_SPOOF_STATUS_PENDING,
            models.LessonAttendance.PHOTO_SPOOF_STATUS_REVIEW,
            models.LessonAttendance.PHOTO_SPOOF_STATUS_ERROR,
        }:
            return "status_not_reviewable"
        return "unknown"

    if action in {
        PHOTO_VERDICT_ACTION_MANUAL_CLEAN,
        PHOTO_VERDICT_ACTION_MANUAL_SUSPICIOUS,
        PHOTO_VERDICT_ACTION_MANUAL_RESET,
    }:
        now_dt = timezone.now()
        actor = (
            request.user if getattr(request.user, "is_authenticated", False) else None
        )
        actor_id = actor.id if actor is not None else None
        manual_comment = str(payload.get("manual_comment") or "").strip()

        if action == PHOTO_VERDICT_ACTION_MANUAL_CLEAN:
            verdict = models.LessonAttendance.PHOTO_MANUAL_VERDICT_CLEAN
            manual_at = now_dt
            db_comment = manual_comment
        elif action == PHOTO_VERDICT_ACTION_MANUAL_SUSPICIOUS:
            verdict = models.LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS
            manual_at = now_dt
            db_comment = manual_comment
        else:
            verdict = models.LessonAttendance.PHOTO_MANUAL_VERDICT_NONE
            manual_at = None
            db_comment = ""
            actor_id = None
            actor = None

        updatable_records = ordered_records
        if action in {
            PHOTO_VERDICT_ACTION_MANUAL_CLEAN,
            PHOTO_VERDICT_ACTION_MANUAL_SUSPICIOUS,
        }:
            updatable_records = [
                record
                for record in ordered_records
                if record.photo_can_set_manual_verdict
            ]
            for record in ordered_records:
                if not record.photo_can_set_manual_verdict:
                    skipped_ids.append(record.id)
                    reason = _reason_manual_verdict_unavailable(record)
                    skipped_reasons.append({"id": record.id, "reason": reason})
                    photo_verdict_logger.warning(
                        "manual verdict unavailable attendance_id=%s reason=%s "
                        "photo_manual_verdict=%s photo_spoof_status=%s",
                        record.id,
                        reason,
                        record.photo_manual_verdict,
                        record.photo_spoof_status,
                    )

        updatable_ids = [record.id for record in updatable_records]
        if not updatable_ids:
            updatable_records = []

        models.LessonAttendance.objects.filter(id__in=updatable_ids).update(
            photo_manual_verdict=verdict,
            photo_manual_comment=db_comment,
            photo_manual_by_id=actor_id,
            photo_manual_at=manual_at,
        )
        for record in updatable_records:
            record.photo_manual_verdict = verdict
            record.photo_manual_comment = db_comment
            record.photo_manual_by_id = actor_id
            record.photo_manual_at = manual_at
            record.photo_manual_by = actor
        changed_records = updatable_records

    elif action == PHOTO_VERDICT_ACTION_RESCAN:
        from monitoring_app.photo_pad import MANUAL_NONE, check_photo, normalize_device

        force_manual = _parse_bool(payload.get("force_manual"), default=False)
        device = normalize_device(payload.get("device"))
        for record in ordered_records:
            image_path = record.staff_image_path
            if not image_path:
                error_items.append({"id": record.id, "error": "no_photo"})
                continue
            if not force_manual and record.photo_manual_verdict != MANUAL_NONE:
                skipped_ids.append(record.id)
                skipped_reasons.append(
                    {"id": record.id, "reason": "rescan_skipped_has_verdict"}
                )
                photo_verdict_logger.warning(
                    "rescan skipped attendance_id=%s reason=rescan_skipped_has_verdict "
                    "photo_manual_verdict=%s",
                    record.id,
                    record.photo_manual_verdict,
                )
                continue
            try:
                result = check_photo(image_path=image_path, device=device)
            except Exception as exc:
                logger.exception(
                    "lesson_attendance_photo_verdicts rescan failed id=%s path=%s",
                    record.id,
                    image_path,
                )
                error_items.append({"id": record.id, "error": str(exc)})
                continue

            update_kwargs = result.to_update_kwargs()
            models.LessonAttendance.objects.filter(id=record.id).update(**update_kwargs)
            record.photo_trust_confirmed = update_kwargs["photo_trust_confirmed"]
            record.photo_spoof_status = update_kwargs["photo_spoof_status"]
            record.photo_spoof_score = update_kwargs["photo_spoof_score"]
            record.photo_spoof_tags = update_kwargs["photo_spoof_tags"]
            record.photo_spoof_checked_at = update_kwargs["photo_spoof_checked_at"]
            record.photo_spoof_model_version = update_kwargs[
                "photo_spoof_model_version"
            ]
            changed_records.append(record)

    _invalidate_photo_cache_for_records(changed_records)
    _broadcast_photo_updates(changed_records)

    return Response(
        {
            "action": action,
            "updated_count": len(changed_records),
            "skipped_ids": skipped_ids,
            "skipped_reasons": skipped_reasons,
            "errors": error_items,
            "results": [
                _serialize_lesson_attendance_photo(record) for record in changed_records
            ],
            "choices": _photo_verdict_choices_payload(),
        },
        status=status.HTTP_200_OK,
    )


@swagger_auto_schema(
    method="get",
    operation_summary="Посещаемость сотрудников по отделу",
    operation_description="Получить данные о посещаемости сотрудников по ID подразделения и его дочерним подразделениям за указанный период.",
    tags=["Attendance & Statistics"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
        openapi.Parameter(
            "end_date",
            openapi.IN_QUERY,
            description="Конечная дата периода в формате YYYY-MM-DD",
            type=openapi.TYPE_STRING,
            required=True,
        ),
        openapi.Parameter(
            "start_date",
            openapi.IN_QUERY,
            description="Начальная дата периода в формате YYYY-MM-DD",
            type=openapi.TYPE_STRING,
            required=True,
        ),
        openapi.Parameter(
            "page",
            openapi.IN_QUERY,
            description="Номер страницы для пагинации",
            type=openapi.TYPE_INTEGER,
            required=False,
        ),
        openapi.Parameter(
            "page_size",
            openapi.IN_QUERY,
            description="Количество записей на странице (максимум 500)",
            type=openapi.TYPE_INTEGER,
            required=False,
        ),
    ],
)
@api_view(["GET"])
@permission_classes([AllowAny])
def staff_detail_by_department_id(request, department_id):
    """
    Получить данные о посещаемости сотрудников по ID подразделения.

    Этот эндпоинт возвращает данные о посещаемости сотрудников для указанного подразделения и его дочерних подразделений за указанный период.

    Параметры запроса:
    - end_date: Конечная дата периода в формате YYYY-MM-DD.
    - start_date: Начальная дата периода в формате YYYY-MM-DD.
    - page: Номер страницы для пагинации (по умолчанию 1).
    - page_size: Количество записей на странице (максимум 500).

    Возвращаемые данные:
    - count: Общее количество записей.
    - next: URL следующей страницы результатов.
    - previous: URL предыдущей страницы результатов.
    - results: Список посещаемости сотрудников, сгруппированных по датам и по каждому сотруднику.

    Пример ответа:
    {
        "count": 1,
        "next": null,
        "previous": null,

        "results": [
            {
                "2024-10-29": {
                    "department": "Test",
                    "attendance": [
                        {
                            "staff_fio": "Иванов Иван",
                            "first_in": "2024-10-29T08:00:00+05:00",
                            "last_out": "2024-10-29T17:00:00+05:00",
                            "area_name": "Улица примерочная 1",
                            "remote_work": false,
                            "absence_reason": null
                        },
                        {
                            "staff_fio": "Петров Петр",
                            "first_in": "2024-10-29T09:15:00+05:00",
                            "last_out": "2024-10-29T16:45:00+05:00",
                            "area_name": "Проспект примеров 7",
                            "remote_work": true,
                            "absence_reason": "business_trip"
                        }
                    ]
                }
            },
            ...
        ]
    }

    Возможные ошибки:
    - 400: Не указаны параметры начала или конца периода.
    - 404: Подразделение не найдено или данные о посещаемости не найдены.
    - 500: Внутренняя ошибка сервера.
    """
    logger.info(
        f"Request received for staff attendance by department ID {department_id}"
    )

    try:
        end_date_str = request.query_params.get("end_date")
        start_date_str = request.query_params.get("start_date")
        page = request.query_params.get("page", 1)

        if not end_date_str or not start_date_str:
            logger.warning("Missing startDate or endDate in request parameters")
            return Response(
                status=status.HTTP_400_BAD_REQUEST,
                data={"error": "Не указаны параметры начала или конца периода"},
            )

        try:
            start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
            logger.info(
                f"Parsed date range: start_date={start_date}, end_date={end_date}"
            )
        except ValueError as ve:
            logger.warning(f"Invalid date format {ve}")
            return Response(
                status=status.HTTP_400_BAD_REQUEST,
                data={"error": "Неверный формат даты. Используйте YYYY-MM-DD."},
            )

        if start_date > end_date:
            logger.warning("Start date is greater than end date")
            return Response(
                status=status.HTTP_400_BAD_REQUEST,
                data={"error": "Дата начала не может быть больше даты конца"},
            )

        try:
            department = models.ChildDepartment.objects.get(id=department_id)
            logger.info(f"Department found: {department.name} (ID: {department_id})")
        except models.ChildDepartment.DoesNotExist:
            logger.warning(f"Department with ID {department_id} not found")
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={"error": "Подразделение не найдено"},
            )

        def get_all_child_department_ids(department_id):
            """
            Uses a recursive CTE to get all child departments via an SQL query.

            Args:
                department_id (str): ID of the parent department.

            Returns:
                list: List of IDs of all child departments, including the given ID.
            """
            query = """
            WITH RECURSIVE childdepartment_cte AS (
                SELECT id, parent_id
                FROM monitoring_app_childdepartment
                WHERE id = %s
                UNION ALL
                SELECT cd.id, cd.parent_id
                FROM monitoring_app_childdepartment cd
                JOIN childdepartment_cte cte ON cd.parent_id = cte.id
            )
            SELECT id FROM childdepartment_cte;
            """
            with connection.cursor() as cursor:
                cursor.execute(query, [department_id])
                result = cursor.fetchall()
            return [row[0] for row in result]

        department_ids = get_all_child_department_ids(department_id)
        logger.info(f"Departments for ID {department_id}: {department_ids}")

        cache_key = (
            f"staff_detail_{LESSON_REPORT_CACHE_VERSION}_{department_id}_"
            f"{start_date_str}_{end_date_str}_page_{page}"
        )
        logger.info(f"Generated cache key: {cache_key}")

        def daterange(start_date, end_date):
            for n in range(int((end_date - start_date).days) + 1):
                yield start_date + datetime.timedelta(n)

        reason_display = dict(models.AbsentReason.ABSENT_REASON_CHOICES)

        def query():
            logger.info("Querying staff attendance data")

            staff_objects = models.Staff.objects.filter(
                department_id__in=department_ids
            ).select_related("department")

            staff_dict = {staff.id: staff for staff in staff_objects}
            staff_ids = list(staff_dict.keys())

            staff_attendance_qs = models.StaffAttendance.objects.filter(
                staff_id__in=staff_ids,
                date_at__range=(start_date, end_date),
            ).values("staff_id", "date_at", "first_in", "last_out", "area_name_in")

            lesson_attendance_qs = models.LessonAttendance.exclude_report_invalid_days(
                models.LessonAttendance.objects.filter(
                    staff_id__in=staff_ids,
                    date_at__range=(start_date, end_date),
                )
            ).values(
                "staff_id",
                "date_at",
                "first_in",
                "last_out",
                "latitude",
                "longitude",
            )

            absent_reasons_qs = (
                models.AbsentReason.objects.filter(
                    staff_id__in=staff_ids,
                    start_date__lte=end_date,
                    end_date__gte=start_date,
                )
                .select_related("staff")
                .values("staff_id", "reason", "start_date", "end_date")
            )
            logger.info(f"AbsentReason records fetched: {absent_reasons_qs.count()}")

            remote_works_qs = (
                models.RemoteWork.objects.filter(
                    Q(staff_id__in=staff_ids)
                    & (
                        Q(start_date__lte=end_date, end_date__gte=start_date)
                        | Q(permanent_remote=True)
                    )
                )
                .select_related("staff")
                .values("staff_id", "permanent_remote", "start_date", "end_date")
            )
            logger.debug(f"RemoteWork records fetched: {remote_works_qs.count()}")

            location_cache = get_class_location_cache()
            location_searcher = location_cache["searcher"]
            if location_searcher is None:
                location_searcher = utils.LocationSearcher(
                    location_cache["searcher_payload"] or []
                )

            staff_attendance_map = defaultdict(lambda: defaultdict(list))
            for sa in staff_attendance_qs:
                date_key = (sa["date_at"] - datetime.timedelta(days=1)).strftime(
                    "%Y-%m-%d"
                )
                staff_attendance_map[sa["staff_id"]][date_key].append(sa)

            lesson_attendance_map = defaultdict(lambda: defaultdict(list))
            for la in lesson_attendance_qs:
                date_key = la["date_at"].strftime("%Y-%m-%d")
                lesson_attendance_map[la["staff_id"]][date_key].append(la)

            absence_map = defaultdict(lambda: defaultdict(list))
            for ar in absent_reasons_qs:
                ar_start = max(ar["start_date"], start_date)
                ar_end = min(ar["end_date"], end_date)
                for single_date in daterange(ar_start, ar_end):
                    date_key = single_date.strftime("%Y-%m-%d")
                    display_reason = reason_display.get(ar["reason"], ar["reason"])
                    absence_map[ar["staff_id"]][date_key].append(display_reason)

            remote_work_map = defaultdict(lambda: defaultdict(bool))
            for rw in remote_works_qs:
                if rw["permanent_remote"]:
                    rw_start = start_date
                    rw_end = end_date
                else:
                    raw_start = rw.get("start_date")
                    raw_end = rw.get("end_date")
                    tmp_start = raw_start if raw_start is not None else start_date
                    tmp_end = raw_end if raw_end is not None else end_date
                    rw_start = max(tmp_start, start_date)
                    rw_end = min(tmp_end, end_date)
                if rw_start <= rw_end:
                    for single_date in daterange(rw_start, rw_end):
                        date_key = single_date.strftime("%Y-%m-%d")
                        remote_work_map[rw["staff_id"]][date_key] = True

            results = []

            for single_date in daterange(start_date, end_date):
                date_key = single_date.strftime("%Y-%m-%d")

                department_attendance_map = defaultdict(list)

                for staff_id, staff in staff_dict.items():
                    staff_fio = f"{staff.surname} {staff.name}"
                    department_name = (
                        staff.department.name
                        if staff.department
                        else "Unknown Department"
                    )

                    sa_records = staff_attendance_map.get(staff_id, {}).get(
                        date_key, []
                    )
                    la_records = lesson_attendance_map.get(staff_id, {}).get(
                        date_key, []
                    )

                    first_in = None
                    last_out = None
                    area_names = []

                    for sa in sa_records:
                        if sa["first_in"]:
                            sa_first_in = sa["first_in"].astimezone(
                                timezone.get_default_timezone()
                            )
                            if not first_in or sa_first_in < first_in:
                                first_in = sa_first_in
                        if sa["last_out"]:
                            sa_last_out = sa["last_out"].astimezone(
                                timezone.get_default_timezone()
                            )
                            if not last_out or sa_last_out > last_out:
                                last_out = sa_last_out
                        area_address = utils.resolve_area_address(
                            sa.get("area_name_in")
                        )
                        if area_address:
                            area_names.append(area_address)

                    for la in la_records:
                        if la["first_in"]:
                            la_first_in = la["first_in"].astimezone(
                                timezone.get_default_timezone()
                            )
                            if not first_in or la_first_in < first_in:
                                first_in = la_first_in
                        if la["last_out"]:
                            la_last_out = la["last_out"].astimezone(
                                timezone.get_default_timezone()
                            )
                            if not last_out or la_last_out > last_out:
                                last_out = la_last_out

                        closest_location_name = location_searcher.find_nearest(
                            la["latitude"], la["longitude"], radius=200
                        )
                        if closest_location_name != "Unknown Area":
                            area_names.append(closest_location_name)

                    area_name = area_names[0] if area_names else "Unknown Area"

                    remote_work = remote_work_map.get(staff_id, {}).get(date_key, False)
                    reasons = absence_map.get(staff_id, {}).get(date_key, [])
                    absence_reason = ", ".join(reasons) if reasons else None

                    attendance_entry = {
                        "staff_fio": staff_fio,
                        "first_in": first_in.isoformat() if first_in else None,
                        "last_out": last_out.isoformat() if last_out else None,
                        "area_name": area_name,
                        "remote_work": remote_work,
                        "absence_reason": absence_reason,
                    }

                    department_attendance_map[department_name].append(attendance_entry)

                for dept, attendance in department_attendance_map.items():
                    date_result = {
                        date_key: {"department": dept, "attendance": attendance}
                    }
                    results.append(date_result)

            paginator = StaffAttendancePagination()
            result_page = paginator.paginate_queryset(results, request)
            return paginator.get_paginated_response(result_page).data

        cached_data = get_cache(cache_key, query=query, timeout=1 * 60 * 60)
        logger.info("Returning cached or queried data")
        return Response(cached_data)

    except Exception as e:
        logger.error(f"Server error while processing request: {str(e)}")
        return Response(
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            data={"error": str(e)},
        )


@swagger_auto_schema(
    method="post",
    operation_summary="Зарегистрировать нового пользователя (доступно только для администратора)",
    operation_description="Регистрирует нового пользователя в системе. Разрешено только для администратора.",
    tags=["Authentication"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
    ],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=["username", "password"],
        properties={
            "username": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Желаемое имя для нового пользователя",
            ),
            "password": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Пароль для нового пользователя",
            ),
        },
    ),
    responses={
        201: openapi.Response(
            description="Created - Пользователь успешно зарегистрирован",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение о результате регистрации",
                    )
                },
            ),
        ),
        400: openapi.Response(
            description="Bad Request - Ошибка в запросе",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Описание ошибки запроса",
                    ),
                },
            ),
        ),
    },
)
@api_view(http_method_names=["POST"])
@permission_classes([IsAdminUser])
def user_register(request):
    """
    Регистрирует нового пользователя в системе. Разрешено только для администратора.

    Этот view ожидает запрос POST, содержащий в теле запроса следующие данные:
    - username (str): Желаемое имя для нового пользователя.
    - password (str): Пароль для нового пользователя.

    Возвращаемые данные:
    - status (int): HTTP статус код:
        - 201 Created: Если пользователь успешно зарегистрирован.
        - 400 Bad Request: Если имя пользователя или пароль отсутствуют, не соответствуют требованиям или имя пользователя уже занято.
    - message (str): Сообщение о результате регистрации.

    Возможные ошибки:
    - 400 Bad Request: Если имя пользователя или пароль отсутствуют, пароль не соответствует требованиям или имя пользователя уже занято.

    Исключения:
    - Стандартные исключения Django, если во время создания или сохранения пользователя возникают какие-либо ошибки.
    """

    logger.info("Received request to register a new user")

    username = request.data.get("username", None)
    password = request.data.get("password", None)

    if not username or not password:
        logger.warning("Username or password not provided")
        return Response(
            status=status.HTTP_400_BAD_REQUEST,
            data={"message": "Требуется юзернейм и пароль"},
        )

    if not utils.password_check(password):
        logger.warning("Password does not meet the requirements")
        return Response(
            status=status.HTTP_400_BAD_REQUEST,
            data={"message": "Пароль не прошел требования"},
        )

    try:
        user, created = User.objects.get_or_create(username=username)
        if not created:
            logger.warning(f"Username '{username}' is already taken")
            return Response(
                status=status.HTTP_400_BAD_REQUEST,
                data={"message": "Данный username уже занят"},
            )

        user.set_password(password)
        user.save()
        logger.info(f"User '{username}' successfully created")
        return Response(
            status=status.HTTP_201_CREATED,
            data={"message": "пользователь успешно создан"},
        )
    except Exception as e:
        logger.error(f"Error occurred while creating user '{username}': {str(e)}")
        return Response(
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            data={"message": str(e)},
        )


def logout_view(request):
    logout(request)
    return redirect("login_view")


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("uploadFile")

    return render(request, "login.html", context={})


@swagger_auto_schema(
    method="get",
    operation_summary="Запуск синхронизации посещаемости с внешним СКУД",
    operation_description=(
        "Запускает загрузку посещаемости сотрудников с внешнего сервера и сохранение в базу. "
        "Аутентификация: `X-API-KEY` или `Authorization: Bearer <access_token>`."
    ),
    tags=["Fetcher"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
        token_param_config,
        openapi.Parameter(
            name="days",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_INTEGER,
            required=False,
            description=(
                "На сколько дней назад брать данные (0..365). "
                "Если не передан, используется значение из settings.DAYS."
            ),
        ),
        openapi.Parameter(
            name="max_concurrent_requests",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_INTEGER,
            required=False,
            description=(
                "Максимум одновременных запросов к внешнему API (1..30, по умолчанию 6)."
            ),
        ),
    ],
    responses={
        200: openapi.Response(
            description="Синхронизация завершена без ошибок.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Статус выполнения.",
                    ),
                    "days": openapi.Schema(
                        type=openapi.TYPE_INTEGER,
                        description="Фактически использованное значение days.",
                    ),
                    "duration_seconds": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        format=openapi.FORMAT_FLOAT,
                        description="Время выполнения в секундах.",
                    ),
                    "duration_human_readable": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Человекочитаемая длительность.",
                    ),
                    "status": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Статус выполнения: success/partial_error/failed.",
                    ),
                    "fetch_summary": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            "source_date": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                description="Дата источника (день, который запрашивали во внешнем API).",
                            ),
                            "save_date": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                description="Дата, в которую сохранялись записи.",
                            ),
                            "total_pins": openapi.Schema(type=openapi.TYPE_INTEGER),
                            "successful_requests": openapi.Schema(
                                type=openapi.TYPE_INTEGER
                            ),
                            "failed_requests": openapi.Schema(
                                type=openapi.TYPE_INTEGER
                            ),
                            "pins_with_events": openapi.Schema(
                                type=openapi.TYPE_INTEGER
                            ),
                            "pins_without_events": openapi.Schema(
                                type=openapi.TYPE_INTEGER
                            ),
                            "created_records": openapi.Schema(
                                type=openapi.TYPE_INTEGER
                            ),
                            "updated_records": openapi.Schema(
                                type=openapi.TYPE_INTEGER
                            ),
                            "event_time_parse_errors": openapi.Schema(
                                type=openapi.TYPE_INTEGER
                            ),
                            "ambiguous_exit_candidates": openapi.Schema(
                                type=openapi.TYPE_INTEGER,
                                description="Количество неоднозначных кандидатов на выход (ambiguous exit devices).",
                            ),
                            "ambiguous_resolved_as_exit": openapi.Schema(
                                type=openapi.TYPE_INTEGER,
                                description="Количество неоднозначных событий, классифицированных как выход.",
                            ),
                            "ambiguous_resolved_as_transfer": openapi.Schema(
                                type=openapi.TYPE_INTEGER,
                                description="Количество неоднозначных событий, классифицированных как переход в пристройку.",
                            ),
                            "error_statuses": openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                description="Сводка кодов ошибок внешнего API, например {'401': 12, 'network_or_unknown': 2}.",
                            ),
                        },
                    ),
                },
            ),
        ),
        207: openapi.Response(
            description="Синхронизация завершена частично: были ошибки по части PIN-ов.",
        ),
        400: openapi.Response(
            description="Некорректные query-параметры.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
        403: "Forbidden: Если доступ запрещен (нет валидной авторизации).",
        429: "Too Many Requests: Fetcher уже выполняется.",
        500: "Internal Server Error: Внутренняя ошибка сервера.",
        502: "Bad Gateway: Все обращения к внешнему API завершились ошибкой.",
    },
)
@async_logic.async_drf_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
async def fetch_data_view(request):
    """
    Асинхронный обработчик запросов на получение данных о посещаемости.
    """
    function_name = "fetch_data_view"
    start_time = time.perf_counter()
    logger.info("%s: Request received", function_name)
    lock_key = "attendance_fetcher_run_lock"
    lock_ttl_seconds = int(getattr(settings, "FETCHER_LOCK_TTL_SECONDS", 30 * 60))
    lock_acquired = cache.add(lock_key, "running", timeout=lock_ttl_seconds)

    if not lock_acquired:
        logger.warning(
            "%s: rejected because another fetcher run is in progress", function_name
        )
        return Response(
            status=status.HTTP_429_TOO_MANY_REQUESTS,
            data={
                "status": "busy",
                "message": "Fetcher is already running. Try again later.",
            },
        )

    def parse_int_query_param(
        name: str,
        *,
        default: int | None,
        min_value: int | None = None,
        max_value: int | None = None,
    ) -> tuple[int | None, str | None]:
        raw_value = request.query_params.get(name)
        if raw_value in (None, ""):
            return default, None
        try:
            parsed_value = int(raw_value)
        except (TypeError, ValueError):
            return None, f"Параметр '{name}' должен быть целым числом."

        if min_value is not None and parsed_value < min_value:
            return (
                None,
                f"Параметр '{name}' должен быть не меньше {min_value}.",
            )
        if max_value is not None and parsed_value > max_value:
            return (
                None,
                f"Параметр '{name}' должен быть не больше {max_value}.",
            )
        return parsed_value, None

    try:
        has_api_key_header = bool(
            request.headers.get("X-API-KEY") or request.headers.get("x-api-key")
        )
        if (
            request.user.is_authenticated
            and not request.user.is_staff
            and not has_api_key_header
        ):
            logger.warning(
                "%s: forbidden for non-staff authenticated user id=%s",
                function_name,
                request.user.id,
            )
            return Response(
                status=status.HTTP_403_FORBIDDEN,
                data={"error": "Недостаточно прав для запуска fetcher."},
            )

        days, days_error = parse_int_query_param(
            "days",
            default=None,
            min_value=0,
            max_value=365,
        )
        if days_error:
            return Response(
                status=status.HTTP_400_BAD_REQUEST, data={"error": days_error}
            )

        max_concurrent_requests, concurrency_error = parse_int_query_param(
            "max_concurrent_requests",
            default=6,
            min_value=1,
            max_value=30,
        )
        if concurrency_error:
            return Response(
                status=status.HTTP_400_BAD_REQUEST,
                data={"error": concurrency_error},
            )
        if max_concurrent_requests is None:
            max_concurrent_requests = 6

        logger.info(
            "%s: Starting attendance fetch with params days=%s, max_concurrent_requests=%s",
            function_name,
            days,
            max_concurrent_requests,
        )

        fetcher = attendance_fetcher.AsyncAttendanceFetcher(
            max_concurrent_requests=max_concurrent_requests,
        )
        fetch_summary = await fetcher.get_all_attendance(days=days)
        raw_errors = fetch_summary.get("errors", [])

        error_statuses_counter: Counter[str] = Counter()
        if isinstance(raw_errors, list):
            for error in raw_errors:
                if not isinstance(error, dict):
                    error_statuses_counter["network_or_unknown"] += 1
                    continue
                status_value = error.get("status")
                if status_value is None:
                    error_statuses_counter["network_or_unknown"] += 1
                else:
                    error_statuses_counter[str(status_value)] += 1

        failed_requests = int(fetch_summary.get("failed_requests", 0))
        total_pins = int(fetch_summary.get("total_pins", 0))
        elapsed_seconds = time.perf_counter() - start_time
        duration_human_readable = utils.format_duration(elapsed_seconds)
        response_summary = {
            "source_date": fetch_summary.get("source_date"),
            "save_date": fetch_summary.get("save_date"),
            "total_pins": total_pins,
            "successful_requests": int(fetch_summary.get("successful_requests", 0)),
            "failed_requests": failed_requests,
            "pins_with_events": int(fetch_summary.get("pins_with_events", 0)),
            "pins_without_events": int(fetch_summary.get("pins_without_events", 0)),
            "created_records": int(fetch_summary.get("created_records", 0)),
            "updated_records": int(fetch_summary.get("updated_records", 0)),
            "event_time_parse_errors": int(
                fetch_summary.get("event_time_parse_errors", 0)
            ),
            "ambiguous_exit_candidates": int(
                fetch_summary.get("ambiguous_exit_candidates", 0)
            ),
            "ambiguous_resolved_as_exit": int(
                fetch_summary.get("ambiguous_resolved_as_exit", 0)
            ),
            "ambiguous_resolved_as_transfer": int(
                fetch_summary.get("ambiguous_resolved_as_transfer", 0)
            ),
        }
        if error_statuses_counter:
            response_summary["error_statuses"] = dict(error_statuses_counter)

        response_data = {
            "message": "Done",
            "status": "success",
            "days": fetch_summary.get("days", days),
            "duration_seconds": round(elapsed_seconds, 2),
            "duration_human_readable": duration_human_readable,
            "fetch_summary": response_summary,
        }

        if failed_requests == 0:
            response_status = status.HTTP_200_OK
            response_data["message"] = "Done"
        elif failed_requests < total_pins:
            response_status = status.HTTP_207_MULTI_STATUS
            response_data["status"] = "partial_error"
            response_data["message"] = "Done with errors."
        else:
            response_status = status.HTTP_502_BAD_GATEWAY
            response_data["status"] = "failed"
            response_data["message"] = "Fetcher failed."

        logger.info(
            "%s: Completed with status=%s in %.2f seconds (total_pins=%s, failed_requests=%s)",
            function_name,
            response_status,
            elapsed_seconds,
            total_pins,
            failed_requests,
        )

        return Response(status=response_status, data=response_data)

    except Exception as e:
        elapsed_seconds = time.perf_counter() - start_time
        duration_human_readable = utils.format_duration(elapsed_seconds)
        logger.error(
            f"{function_name}: Error occurred while fetching attendance data: {str(e)}",
            exc_info=True,
        )
        return Response(
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            data={
                "status": "failed",
                "error": "Ошибка при выполнении fetcher.",
                "duration_seconds": round(elapsed_seconds, 2),
                "duration_human_readable": duration_human_readable,
            },
        )
    finally:
        if lock_acquired:
            cache.delete(lock_key)


@swagger_auto_schema(
    method="get",
    operation_summary="Выгрузка посещаемости зданий по кафедрам",
    operation_description=(
        "Возвращает Excel-отчёт по посещаемости зданий в разрезе кафедр (ChildDepartment). "
        "По умолчанию формируется по последним 7 датам с данными. "
        "Если передать date_from/date_to, отчёт строится по датам с данными внутри указанного диапазона. "
        "Если date_to не указан, а date_from задан — date_to принимается равным текущей дате."
    ),
    tags=["Files & Downloads"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
        openapi.Parameter(
            name="date_from",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            format="date",
            required=False,
            description="Дата начала периода (YYYY-MM-DD).",
        ),
        openapi.Parameter(
            name="date_to",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            format="date",
            required=False,
            description="Дата конца периода (YYYY-MM-DD). Используется только с date_from.",
        ),
        openapi.Parameter(
            name="days_with_data",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_INTEGER,
            required=False,
            description="Количество последних дат с данными (по умолчанию 7, применяется без date_from/date_to).",
        ),
    ],
    responses={
        200: openapi.Response(
            description="Excel-файл отчёта.",
            examples={
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "Binary Excel data"
            },
        ),
        400: "Bad Request: invalid query parameters.",
        500: "Internal Server Error: failed to generate report.",
    },
)
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def download_building_attendance_report(request):
    try:
        params = building_attendance_report.parse_report_request_params(
            date_from_raw=request.query_params.get("date_from"),
            date_to_raw=request.query_params.get("date_to"),
            days_with_data_raw=request.query_params.get("days_with_data"),
        )
    except ValueError as exc:
        return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    try:
        report_result = (
            building_attendance_report.build_building_attendance_report_excel(
                date_from=params.date_from,
                date_to=params.date_to,
                days_with_data=params.days_with_data,
            )
        )
        filename = building_attendance_report.build_report_filename(
            report_result.selected_dates
        )

        response = HttpResponse(
            report_result.excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["X-Report-Dates-Count"] = str(len(report_result.selected_dates))
        return response
    except Exception as exc:
        logger.error(
            "Failed to generate building attendance report: %s",
            str(exc),
            exc_info=True,
        )
        return Response(
            {"error": "Failed to generate building attendance report."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@swagger_auto_schema(
    method="get",
    operation_summary="Generate Attendance Excel File",
    operation_description=(
        "Generates an Excel file containing attendance data for the specified department and its child departments. "
        "You must provide the department ID as a path parameter, and the report period via the `startDate` and `endDate` "
        "query parameters formatted as YYYY-MM-DD. If the provided endDate is greater than today, it will be capped to today's date. "
        "Authentication is required by providing a valid API key in the X-API-KEY header or by using other credentials."
    ),
    tags=["Files & Downloads"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=False,
            description="API ключ для аутентификации (альтернатива JWT токену).",
        ),
        openapi.Parameter(
            name="department_id",
            in_=openapi.IN_PATH,
            type=openapi.TYPE_INTEGER,
            required=True,
            description="The ID of the department for which the attendance report is to be generated.",
        ),
        openapi.Parameter(
            name="startDate",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            format="date",
            required=True,
            description="The start date of the attendance report period, formatted as YYYY-MM-DD.",
        ),
        openapi.Parameter(
            name="endDate",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            format="date",
            required=True,
            description="The end date of the attendance report period, formatted as YYYY-MM-DD.",
        ),
    ],
    responses={
        200: openapi.Response(
            description="Excel file generated successfully. Returns an Excel file stream.",
            examples={
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "Binary Excel data"
            },
        ),
        400: "Bad Request: Missing or invalid query parameters (startDate, endDate) or date format issues.",
        404: "Not Found: The specified department or associated staff records were not found.",
        500: "Internal Server Error: Failed to generate the Excel file due to a server error.",
    },
)
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def sent_excel(request, department_id):
    """
    Optimized function for generating and returning attendance Excel files.

    Args:
        request: HTTP request object
        department_id: ID of the department to generate report for

    Returns:
        HttpResponse with Excel file or error Response
    """
    start_time = datetime.datetime.now()
    logger.info(f"Starting Excel generation for department ID {department_id}")

    end_date_str = request.query_params.get("endDate")
    start_date_str = request.query_params.get("startDate")

    if not all([end_date_str, start_date_str]):
        logger.warning(
            f"Missing startDate or endDate in request for department ID {department_id}"
        )
        return Response(
            {"error": "Missing startDate or endDate"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()

        today = datetime.datetime.now().date()
        if end_date > today:
            logger.info(f"Capping end_date from {end_date} to {today}")
            end_date = today

        if start_date > end_date:
            logger.warning(f"Start date {start_date} is after end date {end_date}")
            return Response(
                {"error": "Start date cannot be after end date"},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except ValueError as e:
        logger.error(f"Invalid date format: {e}")
        return Response(
            {"error": f"Invalid date format: {e}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        department = models.ChildDepartment.objects.get(id=department_id)
        logger.info(f"Department found: {department.name}")
    except models.ChildDepartment.DoesNotExist:
        logger.warning(f"Department with ID {department_id} not found")
        return Response(
            {"error": "Department not found"},
            status=status.HTTP_404_NOT_FOUND,
        )

    all_departments = utils.get_all_child_departments(department)
    department_ids = [dept.id for dept in all_departments]
    logger.info(f"Found {len(department_ids)} departments in hierarchy")

    staff_list = models.Staff.objects.filter(
        department_id__in=department_ids
    ).select_related("department")
    if not staff_list.exists():
        logger.warning(
            f"No staff found for department ID {department_id} or its children"
        )
        return Response(
            {"error": "No staff found for this department"},
            status=status.HTTP_404_NOT_FOUND,
        )

    attendance_data = utils.collect_attendance_data(staff_list, start_date, end_date)

    try:
        excel_file = utils.generate_excel_file(
            attendance_data, department.name, start_date, end_date
        )

        processing_time = datetime.datetime.now() - start_time
        logger.info(
            f"Excel file generated in {utils.format_duration(processing_time.total_seconds())}"
        )

        response = HttpResponse(
            excel_file,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f"attachment; filename=Attendance_{department_id}_{start_date_str}_{end_date_str}.xlsx"
        )
        return response

    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}", exc_info=True)
        return Response(
            {"error": "Failed to generate Excel file: " + str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


class UploadFileView(View):
    """
    Класс представления для обработки действий по загрузке файлов.

    Отображает форму загрузки файла (upload_file.html)
    и обрабатывает POST-запросы для импорта данных из файла.
    """

    template_name = "upload_file.html"

    def get(self, request, *args, **kwargs):
        """
        Обрабатывает GET-запросы.

        Args:
            request (HttpRequest): Объект запроса.
            *args: Дополнительные позиционные аргументы.
            **kwargs: Дополнительные именованные аргументы.

        Returns:
            HttpResponse: Отрисовывает шаблон upload_file.html
            с контекстом, содержащим список всех категорий файлов (categories) и список родительских отделов (parent_departments).
        """
        logger.info("GET request received for file upload view")
        categories = models.FileCategory.objects.all()
        parent_departments = models.ParentDepartment.objects.exclude(id=1)
        context = {"categories": categories, "parent_departments": parent_departments}
        logger.debug(f"Rendering template with context: {context}")
        return render(request, self.template_name, context=context)

    def post(self, request, *args, **kwargs):
        """
        Обрабатывает POST-запросы.

        Args:
            request (HttpRequest): Объект запроса.
            *args: Дополнительные позиционные аргументы.
            **kwargs: Дополнительные именованные аргументы.

        Returns:
            HttpResponse: Возвращает редирект на страницу загрузки файла или
            рендеринг upload_file.html с соответствующим контекстом.

        Raises:
            Exception: Если произошла ошибка при обработке файла.
        """
        logger.info("POST request received for file upload view")
        file_path = request.FILES.get("file")
        category_slug = request.POST.get("category")
        parent_department_id = request.POST.get("parent_department")

        if file_path and category_slug:
            logger.debug(f"File received: {file_path.name}, Category: {category_slug}")
            try:
                if file_path.name.endswith(".xlsx"):
                    logger.info("Processing Excel file")
                    rows = self.handle_excel(file_path)
                    if category_slug == "delete_staff":
                        logger.info("Deleting staff based on Excel data")
                        self.delete_staff(request, rows, parent_department_id)
                    elif category_slug == "staff":
                        logger.info("Processing staff data from Excel")
                        self.process_staff(request, rows)
                    elif category_slug == "departments":
                        logger.info("Processing departments data from Excel")
                        self.process_departments(request, rows)
                    elif category_slug == "public_holidays":
                        logger.info("Processing public holidays data from Excel")
                        self.process_public_holidays(request, rows)
                    elif category_slug == "load_geo":
                        rows = rows[1:]
                        logger.info("Processing ClassLocation data from Excel")
                        self.process_class_locations(request, rows)
                    messages.success(
                        request, "Файл успешно обработан и данные обновлены."
                    )
                elif file_path.name.endswith(".zip") and category_slug == "photo":
                    logger.info("Processing ZIP file for photos")
                    self.handle_zip(request, file_path)
                    messages.success(request, "Фото успешно загружены.")
                else:
                    logger.warning("Invalid file format or category")
                    messages.error(request, "Неверный формат файла или категория.")
                    return render(request, self.template_name)

                return redirect("uploadFile")
            except Exception as error:
                logger.error(f"Error processing file: {str(error)}")
                messages.error(request, f"Ошибка при обработке файла: {str(error)}")
        else:
            logger.warning("File or category missing in the POST request")
            messages.error(
                request,
                "Проверьте правильность заполненных данных или неверный формат файла.",
            )

        return render(request, self.template_name)

    def handle_excel(self, file_path) -> List[ExcelRow]:
        """
        Обрабатывает загрузку и импорт данных из файла Excel.

        Args:
            file_path (File): Путь к загруженному файлу.
            category_slug (str): Категория файла для обработки.

        Raises:
            Exception: Если произошла ошибка при обработке файла Excel.
        """
        logger.info("Handling Excel file")
        try:
            with atomic_block():
                wb = load_workbook(file_path)
                ws = wb.active
                ws.delete_rows(1, 2)
                rows = list(ws.iter_rows())
                logger.debug(f"Rows before sorting: {[row[0].value for row in rows]}")

                rows.sort(
                    key=lambda row: (
                        not str(row[0].value).isdigit(),
                        str(row[0].value).zfill(10),
                    ),
                    reverse=False,
                )
                logger.debug(f"Rows after sorting: {[row[0].value for row in rows]}")
                logger.debug(f"Excel file processed, number of rows: {len(rows)}")
                return rows
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            raise

    def process_class_locations(self, request, rows):
        """
        Processes a list of Excel file rows to populate the ClassLocation model using bulk_create and bulk_update.

        This method processes rows containing class location data, extracting details such as
        name, address, latitude, and longitude. It then either creates new ClassLocation records
        or updates existing ones based on matching name and address. Records with missing or invalid
        data are skipped, and errors are logged.

        Args:
            request (HttpRequest): The request object.
            rows (list): A list of Excel rows, where each row contains data in the format
                [name, address, geo].

        Raises:
            ValueError: If the 'geo' column value is missing or invalid.

        Returns:
            None: Populates the ClassLocation model and sends success or error messages
            to the request user regarding records that were created, updated, or skipped due to errors.

        Logs:
            Logs details of processed rows, including created, updated, and skipped rows. If errors
            occur, they are logged and the user is notified.
        """
        with transaction.atomic():  # type: ignore[reportGeneralTypeIssues]
            to_create = []
            to_update = []
            existing_locations = {
                (loc.name, loc.address): loc
                for loc in models.ClassLocation.objects.only(
                    "id",
                    "name",
                    "address",
                    "latitude",
                    "longitude",
                    "acceptance_radius_m",
                )
            }

            error_count = 0
            error_details = []
            max_error_details = 10

            for index, row in enumerate(rows):
                try:
                    name = str(row[0].value or "").strip()
                    address = str(row[1].value or "").strip()
                    geo_data = str(row[2].value or "") if len(row) > 2 else ""
                    radius_val = row[3].value if len(row) > 3 else None

                    if not geo_data or geo_data.lower() == "none":
                        raise ValueError("Отсутствует значение в столбце 'geo'.")

                    latitude, longitude = utils.extract_coordinates(geo_data)

                    if not all([name, address, latitude, longitude]):
                        raise ValueError("Отсутствуют необходимые данные.")

                    try:
                        if latitude is None or str(latitude).strip() == "":
                            raise ValueError("Latitude is missing or empty")
                        if longitude is None or str(longitude).strip() == "":
                            raise ValueError("Longitude is missing or empty")

                        lat_str = (
                            latitude
                            if isinstance(latitude, (int, float))
                            else str(latitude).strip().replace(",", ".")
                        )
                        lon_str = (
                            longitude
                            if isinstance(longitude, (int, float))
                            else str(longitude).strip().replace(",", ".")
                        )

                        latitude = float(lat_str)
                        longitude = float(lon_str)
                    except (TypeError, ValueError) as e:
                        raise ValueError(f"Invalid coordinates: {e}")

                    acceptance_radius_m = None
                    if radius_val is not None and str(radius_val).strip() != "":
                        try:
                            acceptance_radius_m = int(float(str(radius_val).strip()))
                            if acceptance_radius_m <= 0:
                                acceptance_radius_m = None
                        except (TypeError, ValueError):
                            pass

                    if (name, address) in existing_locations:
                        location = existing_locations[(name, address)]
                        location.latitude = latitude
                        location.longitude = longitude
                        if acceptance_radius_m is not None:
                            location.acceptance_radius_m = acceptance_radius_m
                        to_update.append(location)
                    else:
                        loc_kw = dict(
                            name=name,
                            address=address,
                            latitude=latitude,
                            longitude=longitude,
                        )
                        if acceptance_radius_m is not None:
                            loc_kw["acceptance_radius_m"] = acceptance_radius_m
                        to_create.append(models.ClassLocation(**loc_kw))
                except Exception as e:
                    logger.error(f"Error processing row {index} for ClassLocation: {e}")
                    error_count += 1
                    if len(error_details) < max_error_details:
                        error_details.append(f"Строка {index + 2}: {e}")
                    continue

            if to_create:
                try:
                    models.ClassLocation.objects.bulk_create(to_create)
                    logger.info(f"Создано новых записей: {len(to_create)}")
                except Exception as e:
                    logger.error(f"Error during bulk_create: {e}")
                    messages.error(
                        request, "Не удалось создать новые записи ClassLocation."
                    )

            if to_update:
                try:
                    models.ClassLocation.objects.bulk_update(
                        to_update, ["latitude", "longitude", "acceptance_radius_m"]
                    )
                    logger.info(f"Обновлено существующих записей: {len(to_update)}")
                except Exception as e:
                    logger.error(f"Error during bulk_update: {e}")
                    messages.error(
                        request,
                        "Не удалось обновить существующие записи ClassLocation.",
                    )

            if to_create or to_update:
                try:
                    invalidate_class_location_cache_impl()
                except Exception as inv_err:
                    logger.warning(f"Cache invalidation after bulk ops: {inv_err}")

            success_message = f"Успешно добавлено {len(to_create)} новых записей и обновлено {len(to_update)} записей."
            if error_count > 0:
                success_message += f" Пропущено {error_count} записей из-за ошибок."
            messages.success(request, success_message)

            if error_details:
                error_message = (
                    "Некоторые записи были пропущены из-за ошибок:\n"
                    + "\n".join(error_details)
                )
                if error_count > max_error_details:
                    error_message += (
                        f"\n...и ещё {error_count - max_error_details} ошибок."
                    )
                messages.warning(request, error_message)

    def delete_staff(self, request, rows, parent_department_id):
        """
        Удаляет сотрудников дочерних отделов, отсутствующих в переданном списке PIN-кодов.

        Метод получает родительский отдел по `parent_department_id` и находит все связанные
        дочерние отделы. Затем проверяет, какие PIN-коды сотрудников из базы данных
        отсутствуют в списке, переданном в `rows`, и удаляет таких сотрудников.

        Args:
            request: HTTP-запрос для отправки сообщений об успешном или неудачном удалении.
            rows: Список строк с PIN-кодами сотрудников, которых нужно оставить.
            parent_department_id: ID родительского отдела для поиска связанных дочерних отделов.

        Exceptions:
            ValueError: Если не передан `parent_department_id` или не найдены дочерние отделы.
            models.ParentDepartment.DoesNotExist: Если родительский отдел с данным ID не найден.
            Exception: Любая другая ошибка, возникшая при удалении сотрудников.
        """
        logger.info(f"Deleting staff for parent department ID: {parent_department_id}")
        try:
            if not parent_department_id:
                raise ValueError("ID родительского отдела не был передан.")

            parent_department = models.ParentDepartment.objects.get(
                id=parent_department_id
            )

            child_departments = models.ChildDepartment.objects.filter(
                parent__name=parent_department.name
            )
            if not child_departments.exists():
                raise ValueError(
                    f"Для родительского отдела {parent_department.name} не найдены дочерние отделы."
                )

            pin_list_from_file = [row[0].value for row in rows if row[0].value]

            staff_in_db = models.Staff.objects.filter(department__in=child_departments)

            staff_to_delete = staff_in_db.exclude(pin__in=pin_list_from_file)

            deleted_count, _ = staff_to_delete.delete()
            logger.info(f"Deleted {deleted_count} staff members")
            messages.success(
                request, f"Успешно удалено {deleted_count} сотрудника(ов)."
            )

        except models.ParentDepartment.DoesNotExist:
            error_message = f"Родительский отдел с ID {parent_department_id} не найден."
            logger.error(error_message)
            messages.error(request, error_message)
        except ValueError as ve:
            logger.warning(f"ValueError during staff deletion: {str(ve)}")
            messages.error(request, str(ve))
        except Exception as e:
            logger.error(f"Unexpected error during staff deletion: {str(e)}")
            messages.error(
                request, f"Произошла ошибка при удалении сотрудников: {str(e)}"
            )

    def process_departments(self, request, rows):
        """
        Обрабатывает данные для категории "departments" из Excel файла.
        Args:
            rows (list): Список строк из файла Excel.
        Raises:
            Exception: Если произошла ошибка при обработке строки.
        """
        logger.info("Processing departments data")

        created_parent_departments = []
        created_child_departments = []

        try:
            for row in rows:
                parent_department_id_value = row[2].value
                parent_department_name = row[3].value
                child_department_name = row[1].value
                child_department_id_value = row[0].value

                parent_department_id = (
                    utils.normalize_id(str(parent_department_id_value).strip())
                    if parent_department_id_value
                    else None
                )
                child_department_id = (
                    utils.normalize_id(str(child_department_id_value).strip())
                    if child_department_id_value
                    else None
                )

                if not parent_department_id or not child_department_id:
                    logger.debug("Skipping row due to missing or invalid ID")
                    continue

                if parent_department_name:
                    (
                        parent_department,
                        parent_created,
                    ) = models.ParentDepartment.objects.get_or_create(
                        id=parent_department_id,
                        defaults={"name": parent_department_name},
                    )
                    if parent_created:
                        created_parent_departments.append(parent_department_name)
                        logger.info(
                            f"Created new parent department: {parent_department_name}"
                        )

                    (
                        parent_department_as_child,
                        child_created,
                    ) = models.ChildDepartment.objects.get_or_create(
                        id=parent_department.id,
                        defaults={"name": parent_department.name, "parent": None},
                    )
                else:
                    parent_department_as_child = models.ChildDepartment.objects.get(
                        id="1"
                    )

                (
                    _child_department,
                    child_created,
                ) = models.ChildDepartment.objects.get_or_create(
                    id=child_department_id,
                    defaults={
                        "name": child_department_name,
                        "parent": parent_department_as_child,
                    },
                )
                if child_created:
                    created_child_departments.append(child_department_name)
                    logger.info(
                        f"Created new child department: {child_department_name}"
                    )

            if created_parent_departments or created_child_departments:
                messages.success(
                    request,
                    f"Создано родительских отделов: {len(created_parent_departments)}, "
                    f"дочерних отделов: {len(created_child_departments)}.",
                )

        except Exception as error:
            logger.error(f"Error processing departments: {str(error)}")
            messages.error(request, f"Ошибка при обработке отдела: {str(error)}")

    def process_staff(self, request, rows):
        """
        Обрабатывает данные для категории "staff" из Excel файла.
        except Exception as error:
            messages.error(request, f"Ошибка при обработке отдела: {str(error)}")

        Args:
            rows (list): Список строк из файла Excel.

        Raises:
            Exception: Если произошла ошибка при обработке строки.
        """
        logger.info("Processing staff data")
        staff_instances = []
        departments_cache = {}

        try:
            for row in rows:
                pin = row[0].value
                name = row[1].value
                surname = row[2].value or "Нет фамилии"
                department_id = str(row[3].value) if row[3].value else None
                position_name = (
                    row[5].value or "Сотрудник" if len(row) > 5 else "Сотрудник"
                )

                position, _ = models.Position.objects.get_or_create(name=position_name)

                if department_id:
                    if department_id in departments_cache:
                        department = departments_cache[department_id]
                    else:
                        try:
                            department = models.ChildDepartment.objects.get(
                                id=department_id
                            )
                            departments_cache[department_id] = department
                        except models.ChildDepartment.DoesNotExist:
                            department = None
                else:
                    department = None

                staff_instance = models.Staff(
                    pin=pin,
                    name=name,
                    surname=surname,
                    department=department,
                )

                staff_instances.append((staff_instance, position))

            pin_list = [staff[0].pin for staff in staff_instances]
            existing_staff = models.Staff.objects.filter(pin__in=pin_list)
            existing_staff_dict = {staff.pin: staff for staff in existing_staff}

            staff_to_create = []
            staff_to_update = []

            for staff_instance, position in staff_instances:
                if staff_instance.pin in existing_staff_dict:
                    existing = existing_staff_dict[staff_instance.pin]
                    if staff_instance.name and staff_instance.name != existing.name:
                        existing.name = staff_instance.name
                    if (
                        staff_instance.surname
                        and staff_instance.surname != existing.surname
                    ):
                        existing.surname = staff_instance.surname
                    if (
                        staff_instance.department
                        and staff_instance.department != existing.department
                    ):
                        existing.department = staff_instance.department
                    if position.name and position.name != "Сотрудник":
                        if not existing.positions.filter(name=position.name).exists():
                            existing.positions.add(position)
                    staff_to_update.append(existing)
                else:
                    try:
                        staff_instance.save()
                        staff_instance.positions.add(position)
                        staff_to_create.append(staff_instance)
                    except IntegrityError:
                        logger.warning(
                            f"Duplicate entry for pin {staff_instance.pin}, skipping."
                        )
                        continue

            for staff in staff_to_update:
                staff.save()

            logger.info(f"Updated {len(staff_to_update)} staff members")
            logger.info(f"Created {len(staff_to_create)} new staff members")
            messages.success(
                request, f"Успешно обновлено {len(staff_to_update)} сотрудников."
            )
            messages.success(
                request, f"Успешно добавлено {len(staff_to_create)} новых сотрудников."
            )

        except Exception as e:
            logger.error(f"Error processing staff data: {str(e)}")
            messages.error(request, f"Ошибка при обработке сотрудников: {str(e)}")

    def process_public_holidays(self, request, rows):
        """
        Обрабатывает данные для категории "public_holidays" из Excel файла.

        Args:
            request (HttpRequest): Объект запроса.
            rows (list): Список строк из файла Excel.

        Raises:
            Exception: Если произошла ошибка при обработке строки.
        """
        logger.info("Processing public holidays data")

        created_holidays = 0
        updated_holidays = 0
        errors = []
        max_error_details = 10

        working_day_mapping = {
            "да": True,
            "нет": False,
            "yes": True,
            "no": False,
            "true": True,
            "false": False,
            "рабочий": True,
            "не рабочий": False,
        }

        with atomic_block():
            for index, row in enumerate(rows):
                try:
                    date_cell = row[0].value
                    name_cell = row[1].value
                    is_working_day_cell = row[2].value

                    if not date_cell or not name_cell:
                        raise ValueError(
                            "Отсутствуют обязательные поля 'Дата праздника' или 'Название праздника'."
                        )

                    if isinstance(date_cell, datetime.datetime) or isinstance(
                        date_cell, datetime.date
                    ):
                        date = (
                            date_cell.date()
                            if isinstance(date_cell, datetime.datetime)
                            else date_cell
                        )
                    else:
                        try:
                            date = datetime.datetime.strptime(
                                str(date_cell), "%d.%m.%Y"
                            ).date()
                        except ValueError:
                            try:
                                date = datetime.datetime.strptime(
                                    str(date_cell), "%Y-%m-%d"
                                ).date()
                            except ValueError:
                                raise ValueError(
                                    "Неверный формат даты. Ожидается DD.MM.YYYY или YYYY-MM-DD."
                                )

                    if isinstance(is_working_day_cell, bool):
                        is_working_day = is_working_day_cell
                    else:
                        is_working_day_str = str(is_working_day_cell).strip().lower()
                        is_working_day = working_day_mapping.get(is_working_day_str)
                        if is_working_day is None:
                            raise ValueError(
                                "Неверное значение в поле 'Рабочий день'. Ожидается 'Да' или 'Нет'."
                            )

                    _holiday, created = models.PublicHoliday.objects.update_or_create(
                        date=date,
                        defaults={
                            "name": name_cell.strip(),
                            "is_working_day": is_working_day,
                        },
                    )

                    if created:
                        created_holidays += 1
                    else:
                        updated_holidays += 1

                except Exception as e:
                    logger.error(
                        f"Error processing row {index + 2} for PublicHoliday: {e}"
                    )
                    errors.append(f"Строка {index + 2}: {e}")
                    if len(errors) >= max_error_details:
                        break
                    continue

        success_message = f"Успешно создано {created_holidays} праздников и обновлено {updated_holidays} праздников."
        messages.success(request, success_message)

        if errors:
            error_message = (
                "Некоторые записи не были обработаны из-за ошибок:\n"
                + "\n".join(errors)
            )
            if len(errors) > max_error_details:
                error_message += f"\n...и ещё {len(errors) - max_error_details} ошибок."
            messages.warning(request, error_message)

    def handle_zip(self, request, file_path):
        """
        Обрабатывает загрузку и импорт данных из ZIP архива для Staff.
        Args:
            file_path (File): Путь к загруженному файлу.

        Raises:
            Exception: Если произошла ошибка при обработке ZIP файла.
        """
        logger.info("Processing ZIP file for staff photos")
        try:
            with zipfile.ZipFile(file_path, "r") as zip_file:
                zip_file.extractall("/tmp")
                for filename in zip_file.namelist():
                    pin = os.path.splitext(filename)[0]
                    staff_member = models.Staff.objects.filter(pin=pin).first()
                    if staff_member:
                        with zip_file.open(filename) as file:
                            new_avatar = ContentFile(file.read())
                            new_avatar.name = filename
                            if new_avatar:
                                if staff_member.avatar:
                                    staff_member.avatar.delete(save=False)
                                staff_member.avatar.save(
                                    new_avatar.name, new_avatar, save=False
                                )
                                staff_member.save()
            logger.info("Staff photos updated successfully")
            messages.success(request, "Фотографии успешно обновлены.")
        except Exception as e:
            logger.error(f"Error processing ZIP file: {str(e)}")
            messages.error(
                request, f"Ошибка при обработке архива с фотографиями: {str(e)}"
            )


class APIKeyCheckView(APIView):
    """
    Проверка API ключа.

    Проверяет наличие и валидность переданного API ключа в заголовке запроса.
    Если ключ отсутствует или недействителен, возвращает соответствующее сообщение об ошибке.

    Методы:
        get: Проверяет API ключ и возвращает данные о его создании и статусе активности.

    Права доступа:
        AllowAny: Доступ открыт для всех пользователей, аутентификация не требуется.
    """

    permission_classes = [AllowAny]

    @swagger_auto_schema(
        operation_summary="Проверка API ключа",
        operation_description="Проверяет наличие и валидность переданного API ключа в заголовке запроса.",
        tags=["Authentication"],
        manual_parameters=[
            openapi.Parameter(
                name="X-API-KEY",
                in_=openapi.IN_HEADER,
                type=openapi.TYPE_STRING,
                required=True,
                description="API ключ для проверки.",
            )
        ],
        responses={
            200: openapi.Response(
                description="Данные о создании и статусе активности API ключа.",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "data": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "created_at": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    format="date-time",
                                    description="Дата и время создания ключа.",
                                ),
                                "is_active": openapi.Schema(
                                    type=openapi.TYPE_BOOLEAN,
                                    description="Статус активности ключа.",
                                ),
                            },
                        )
                    },
                ),
            ),
            400: openapi.Response(
                description="Некорректный запрос: отсутствует или недействителен API ключ.",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "message": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="Сообщение об ошибке.",
                        ),
                        "error": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="Описание ошибки.",
                        ),
                    },
                ),
            ),
        },
    )
    def get(self, request, *args, **kwargs):
        """
        Проверяет API ключ и возвращает данные о его создании и статусе активности.

        Аргументы:
            request (HttpRequest): Объект HTTP запроса.

        Возвращает:
            Response: Ответ с данными о создании и статусе активности API ключа, либо сообщение об ошибке.
        """
        logger.info("API Key check request received")

        api_key = request.headers.get("X-API-KEY")

        if not api_key:
            logger.warning("API Key is missing in the request")
            return Response(
                {"message": "API Key is missing"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            secret_key = utils.APIKeyUtility.get_secret_key()
            logger.debug("Secret key retrieved successfully")
            data = utils.APIKeyUtility.decrypt_data(
                api_key, secret_key, fields=("created_at", "is_active")
            )
            logger.info("API Key is valid")
            return Response({"data": data}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Invalid API Key: {str(e)}")
            return Response(
                {"message": "Invalid API Key", "error": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


def password_reset_request_view(request):
    if request.method == "POST":
        identifier = request.POST.get("identifier")
        ip_address = utils.get_client_ip(request)
        logger.error(str(ip_address))
        user = (
            User.objects.filter(username=identifier).first()
            or User.objects.filter(email=identifier).first()
        )

        user_timezone = utils.get_user_timezone(request)

        if user:
            last_request_time = models.PasswordResetRequestLog.get_last_request_time(
                user, ip_address
            )
            if not models.PasswordResetRequestLog.can_request_again(user, ip_address):
                if last_request_time:
                    next_possible_time = timezone.localtime(
                        last_request_time + timezone.timedelta(minutes=5),
                        user_timezone,
                    )
                    last_request_time_local = timezone.localtime(
                        last_request_time, user_timezone
                    )
                    messages.warning(
                        request,
                        f"Запрос уже был отправлен. Повторный запрос возможен в {next_possible_time.strftime('%H:%M:%S %Z')} ({next_possible_time.tzinfo}). Последний запрос был в {last_request_time_local.strftime('%H:%M:%S %Z')} ({last_request_time_local.tzinfo}).",
                    )
                else:
                    current_time_local = timezone.localtime(
                        timezone.now(), user_timezone
                    )
                    messages.warning(
                        request,
                        f"Запрос уже был отправлен. Повторный запрос возможен в ближайшее время. Последний запрос: неизвестен. Текущее время {current_time_local.strftime('%H:%M:%S %Z')} ({current_time_local.tzinfo}).",
                    )
            else:
                utils.send_password_reset_email(user, request)
                models.PasswordResetRequestLog.log_request(user, ip_address)
                current_time_local = timezone.localtime(timezone.now(), user_timezone)
                messages.success(
                    request,
                    f"Если пользователь существует, ссылка для сброса пароля была отправлена на его электронную почту. Последний запрос был в {current_time_local.strftime('%H:%M:%S %Z')} ({current_time_local.tzinfo}).",
                )
        else:
            messages.info(
                request,
                "Если пользователь существует, ссылка для сброса пароля была отправлена на его электронную почту.",
            )

        return redirect("password_reset_request")

    return render(request, "password_reset_request.html")


def password_reset_confirm_view(request, token):
    reset_token = get_object_or_404(models.PasswordResetToken, token=token)

    if not reset_token.is_valid():
        messages.error(request, "Этот токен для сброса пароля больше не действителен.")
        return redirect("password_reset_request")

    if request.method == "POST":
        new_password = request.POST.get("password")
        user = reset_token.user
        user.set_password(new_password)
        user.save()

        if models.PasswordResetToken.objects.mark_as_used(token):
            messages.success(
                request,
                "Пароль успешно сброшен. Вы можете войти в систему с новым паролем.",
            )
            return redirect("react_app")
        else:
            messages.error(request, "Ошибка при обновлении токена. Попробуйте снова.")
            return redirect("password_reset_confirm", token=token)

    return render(request, "password_reset_confirm.html", {"token": token})


def download_examples_zip(request):
    """
    Serve the 'examples.zip' file from the media directory to the user for download.

    This view checks for the existence of 'examples.zip' in the MEDIA_ROOT directory
    and serves it as a downloadable file if found. Logs an error and raises a 404
    error if the file is missing.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        FileResponse: A response object containing the file for download.

    Raises:
        Http404: If the file does not exist in the specified directory.
    """
    file_path = Path(settings.MEDIA_ROOT) / "examples.zip"

    if not file_path.exists():
        logger.error(f"File not found: {file_path}")
        raise Http404("Requested file does not exist.")

    try:
        response = FileResponse(file_path.open("rb"), content_type="application/zip")
        response["Content-Disposition"] = 'attachment; filename="examples.zip"'
        return response
    except Exception as e:
        logger.error(f"Error serving file {file_path}: {e}")
        raise Http404("An error occurred while serving the file.")


def serve_attendance_media(request, path):
    attendance_root = Path(settings.ATTENDANCE_ROOT).resolve()
    requested_path = (attendance_root / path).resolve(strict=False)

    try:
        requested_path.relative_to(attendance_root)
    except ValueError as exc:
        raise Http404("File not found") from exc

    if not requested_path.is_file():
        raise Http404("File not found")

    content_type, _ = mimetypes.guess_type(requested_path.name)
    response = FileResponse(
        requested_path.open("rb"),
        content_type=content_type or "application/octet-stream",
    )
    response["Cache-Control"] = "private, max-age=300"
    return response


@api_view(["GET"])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def face_lab_departments(request):
    """
    Child departments that have at least one staff member (Face Lab UI selects).
    """
    dept_ids = (
        models.Staff.objects.filter(department_id__isnull=False)
        .values_list("department_id", flat=True)
        .distinct()
    )
    rows = list(
        models.ChildDepartment.objects.filter(id__in=dept_ids)
        .order_by("name")
        .values("id", "name", "parent_id")
    )
    return Response(rows, status=status.HTTP_200_OK)


@api_view(["GET"])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def face_lab_staff_options(request):
    """
    Flat list of all staff with department (Face Lab «С эталоном»).
    One DB round-trip + cache instead of N requests per department.
    """
    cache_key = "face_lab_staff_options_v1"

    def fetch_rows():
        rows = []
        qs = (
            models.Staff.objects.filter(department_id__isnull=False)
            .select_related("department")
            .order_by("surname", "name", "pin")
        )
        for s in qs.iterator(chunk_size=800):
            dept = s.department
            dept_name = dept.name if dept else ""
            dept_id = s.department_id
            if s.surname == "Нет фамилии":
                fio = s.name
            else:
                fio = f"{s.surname} {s.name}"
            rows.append(
                {
                    "pin": s.pin,
                    "fio": fio,
                    "dept_id": dept_id,
                    "dept_name": dept_name,
                }
            )
        return rows

    data = get_cache(cache_key, query=fetch_rows, timeout=10 * 60)
    return Response(data, status=status.HTTP_200_OK)


@api_view(["POST"])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def face_lab_pad_test(request):
    """Run presentation-attack detection on an uploaded frame (authenticated testers)."""
    from monitoring_app.photo_pad import check_photo_bgr

    uploaded = request.FILES.get("image")
    if not uploaded:
        return Response(
            {"error": "image file is required."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if uploaded.size == 0:
        return Response(
            {"error": "Uploaded image is empty."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        img_bgr = ml.load_image_from_memory(uploaded)
    except ValidationError as ve:
        return Response(
            {"error": _drf_validation_error_text(ve)},
            status=status.HTTP_400_BAD_REQUEST,
        )

    result = check_photo_bgr(img_bgr)
    return Response(
        {
            "status": result.status,
            "trust_confirmed": result.trust_confirmed,
            "risk_score": result.risk_score,
            "tags": result.tags,
            "model_version": result.model_version,
            "elapsed_ms": result.elapsed_ms,
            "deepface_score": result.deepface_score,
            "device_score": result.device_score,
            "frame_score": result.frame_score,
            "quality_penalty": result.quality_penalty,
        },
        status=status.HTTP_200_OK,
    )


@swagger_auto_schema(
    method="post",
    auto_schema=FormOnlySwaggerAutoSchema,  # type: ignore[reportArgumentType]
    operation_summary="Верификация лица",
    operation_description="Верифицирует лицо сотрудника по PIN и изображению. Требует передачи заголовка X-API-KEY для просмотра в Swagger.",
    tags=["Face Recognition - Verify"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=True,
            description="API ключ для доступа к этому эндпоинту. Без этого ключа эндпоинт скрыт в Swagger.",
        ),
        openapi.Parameter(
            name="pin",
            in_=openapi.IN_FORM,
            type=openapi.TYPE_STRING,
            required=True,
            description="PIN сотрудника для верификации.",
        ),
        openapi.Parameter(
            name="image",
            in_=openapi.IN_FORM,
            type=openapi.TYPE_FILE,
            required=True,
            description="Изображение лица для верификации.",
        ),
    ],
    request_body=no_body,
    responses={
        200: openapi.Response(
            description="Результат верификации",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "verified": openapi.Schema(
                        type=openapi.TYPE_BOOLEAN,
                        description="Результат верификации (True/False).",
                    ),
                    "score": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        description="Оценка схожести (0–1), смесь max и среднего по лучшим прототипам.",
                    ),
                    "max_cosine": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        description="Максимальный косинус по прототипам галереи.",
                    ),
                    "trained_model_present": openapi.Schema(
                        type=openapi.TYPE_BOOLEAN,
                        description="Есть ли сохранённый per-staff .pt после обучения.",
                    ),
                    "gallery_templates": openapi.Schema(
                        type=openapi.TYPE_INTEGER,
                        description="Число прототипов (маска + аватар + embeddings.npy).",
                    ),
                    "threshold_used": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        description="Порог: строже при наличии .pt, мягче в fallback.",
                    ),
                    "verification_mode": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="embedding_gallery_strict | embedding_gallery_fallback",
                    ),
                },
            ),
        ),
        400: "Bad Request: Неверные данные запроса.",
        404: "Not Found: Сотрудник не найден.",
    },
)
@api_view(["POST"])
@permission_classes([AllowAny])
def verify_face(request):
    import numpy as np

    face_logger = logging.getLogger("django")
    face_logger.info("Received request to verify face.")

    staff_pin = request.data.get("pin")
    staff_image = request.FILES.get("image")

    if not staff_pin or not staff_image:
        face_logger.warning("PIN or image is missing in the request.")
        return Response(
            {"error": "PIN and image are required."}, status=status.HTTP_400_BAD_REQUEST
        )

    if staff_image.size == 0:
        face_logger.warning("Uploaded image is empty.")
        return Response(
            {"error": "Uploaded image is empty."}, status=status.HTTP_400_BAD_REQUEST
        )

    try:
        staff = models.Staff.objects.get(pin=staff_pin)
        face_logger.info(f"Staff with PIN {staff_pin} found.")
    except models.Staff.DoesNotExist:
        face_logger.error(f"Staff with PIN {staff_pin} does not exist.")
        return Response({"error": "Staff not found."}, status=status.HTTP_404_NOT_FOUND)

    try:
        new_image = ml.load_image_from_memory(staff_image)
        new_embedding = ml.create_face_encoding(new_image)
        if new_embedding is None:
            logger.warning("No face detected in the uploaded image.")
            return Response(
                {"error": "No face detected in the image."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        face_logger.info("Face detected, calculating gallery similarity.")
        probe = np.asarray(new_embedding, dtype=np.float64)
        try:
            verified, score, meta = ml.verify_staff_face_embedding_score(staff, probe)
        except ValueError as ve:
            face_logger.warning("verify_face: no gallery for PIN %s: %s", staff_pin, ve)
            return Response(
                {
                    "error": (
                        "Нет эталона лица: нужен файл аватара и/или запись маски "
                        "с эмбеддингом."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        face_logger.info(
            "Verification PIN %s: score=%s verified=%s mode=%s templates=%s",
            staff_pin,
            score,
            verified,
            meta.get("verification_mode"),
            meta.get("gallery_templates"),
        )
        return Response(
            {
                "verified": verified,
                "score": score,
                "max_cosine": meta["max_cosine"],
                "trained_model_present": meta["trained_model_present"],
                "gallery_templates": meta["gallery_templates"],
                "threshold_used": meta["threshold_used"],
                "verification_mode": meta["verification_mode"],
                "relaxed_match": bool(meta.get("relaxed_match")),
            },
            status=status.HTTP_200_OK,
        )

    except ValidationError as ve:
        face_logger.warning("verify_face validation: %s", _drf_validation_error_text(ve))
        return Response(
            {"error": _drf_validation_error_text(ve)},
            status=status.HTTP_400_BAD_REQUEST,
        )
    except Exception as e:
        face_logger.error(
            f"Error during face verification for PIN {staff_pin}: {str(e)}"
        )
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method="post",
    auto_schema=FormOnlySwaggerAutoSchema,  # type: ignore[reportArgumentType]
    operation_summary="Распознавание лиц",
    operation_description="Распознает лица сотрудников на изображении. Требует передачи заголовка X-API-KEY для просмотра в Swagger.",
    tags=["Face Recognition - Recognize"],
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=True,
            description="API ключ для доступа к этому эндпоинту. Без этого ключа эндпоинт скрыт в Swagger.",
        ),
        openapi.Parameter(
            name="image",
            in_=openapi.IN_FORM,
            type=openapi.TYPE_FILE,
            required=True,
            description="Изображение с лицами для распознавания (PNG, JPG, JPEG).",
        ),
    ],
    request_body=no_body,
    responses={
        200: openapi.Response(
            description="Результат распознавания",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "recognized_staff": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(type=openapi.TYPE_OBJECT),
                        description="Список распознанных сотрудников.",
                    ),
                    "unknown_faces": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(type=openapi.TYPE_OBJECT),
                        description="Список нераспознанных лиц.",
                    ),
                },
            ),
        ),
        400: "Bad Request: Неверные данные запроса.",
        404: "Not Found: Лица не распознаны.",
        500: "Internal Server Error: Ошибка при распознавании.",
    },
)
@api_view(["POST"])
@permission_classes([AllowAny])
def recognize_faces(request):
    recognize_logger = logging.getLogger("django")
    recognize_logger.info("Received request to recognize faces.")

    staff_image = request.FILES.get("image")

    if not staff_image:
        recognize_logger.warning("No image provided in the request.")
        return Response(
            {"error": "Image is required."}, status=status.HTTP_400_BAD_REQUEST
        )

    if not staff_image.name.lower().endswith((".png", ".jpg", ".jpeg")):
        recognize_logger.warning("Invalid image format provided.")
        return Response(
            {"error": "Invalid image format. Only PNG, JPG, and JPEG are allowed."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if staff_image.size == 0:
        recognize_logger.warning("Uploaded image is empty.")
        return Response(
            {"error": "Uploaded image is empty."}, status=status.HTTP_400_BAD_REQUEST
        )

    try:
        recognized_staff, unknown_faces = ml.recognize_faces_in_image(staff_image)

        if not recognized_staff and not unknown_faces:
            logger.info("No staff members or unknown faces recognized.")
            return Response(
                {"error": "No staff members recognized."},
                status=status.HTTP_404_NOT_FOUND,
            )

        recognize_logger.info(
            f"Recognition completed. Recognized staff: {len(recognized_staff)}, Unknown faces: {len(unknown_faces)}."
        )
        return Response(
            {"recognized_staff": recognized_staff, "unknown_faces": unknown_faces},
            status=status.HTTP_200_OK,
        )

    except ValidationError as ve:
        msg = _drf_validation_error_text(ve)
        recognize_logger.warning("Validation error during face recognition: %s", msg)
        return Response({"error": msg}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        recognize_logger.error(f"Unexpected error during face recognition: {str(e)}")
        return Response(
            {"error": f"Face recognition error: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


class AbsentReasonView(APIView):
    permission_classes = [permissions.IsAuthenticatedOrAPIKey]
    authentication_classes = [JWTAuthentication]

    def get_date_interval(self, request):
        """
        Извлекает интервал дат из параметров запроса или подставляет интервал за последнюю неделю.

        Параметры запроса:
          - start_date (str, опционально): Дата начала в формате "YYYY-MM-DD".
          - end_date (str, опционально): Дата окончания в формате "YYYY-MM-DD".

        Возвращает:
          tuple(datetime.date, datetime.date): Кортеж (query_start, query_end).

        Генерирует:
          ValueError: Если передан неверный формат даты.
        """
        today = datetime.datetime.today().date()
        default_start = today - datetime.timedelta(days=7)
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")

        if start_date_str:
            try:
                query_start = datetime.datetime.strptime(
                    start_date_str, "%Y-%m-%d"
                ).date()
            except ValueError as e:
                logger.error(f"Неверный формат start_date: {start_date_str}")
                raise ValueError(
                    "Неверный формат start_date. Ожидается YYYY-MM-DD."
                ) from e
        else:
            query_start = default_start

        if end_date_str:
            try:
                query_end = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError as e:
                logger.error(f"Неверный формат end_date: {end_date_str}")
                raise ValueError(
                    "Неверный формат end_date. Ожидается YYYY-MM-DD."
                ) from e
        else:
            query_end = today

        return query_start, query_end

    @swagger_auto_schema(
        operation_summary="Получение записей отсутствия",
        operation_description=(
            "Метод возвращает записи отсутствия с возможностью фильтрации по интервалу дат и сотруднику.\n\n"
            "Параметры запроса:\n"
            " - **start_date** (опционально): Дата начала (формат YYYY-MM-DD).\n"
            " - **end_date** (опционально): Дата окончания (формат YYYY-MM-DD).\n"
            " - **staffs_all** (опционально): Если true, возвращаются записи для всех сотрудников, сгруппированные по сотрудникам.\n"
            " - **staff_pin** (обязательно, если staffs_all не true): PIN сотрудника для фильтрации записей.\n"
            " - **download** (опционально): Если true, возвращается ZIP-архив документов.\n"
            "\nПри параметре **download=true** архив формируется с файлами, имена которых имеют формат:\n"
            "   `staff.pin_fio_absenceID.ext` (например: `001_Ivanov_Ivan_7.pdf`)."
        ),
        tags=["Absence"],
        manual_parameters=[
            openapi.Parameter(
                name="X-API-KEY",
                in_=openapi.IN_HEADER,
                type=openapi.TYPE_STRING,
                required=False,
                description="API ключ для аутентификации (альтернатива JWT токену).",
            ),
            openapi.Parameter(
                "start_date",
                openapi.IN_QUERY,
                description="Дата начала (YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "end_date",
                openapi.IN_QUERY,
                description="Дата окончания (YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "staffs_all",
                openapi.IN_QUERY,
                description="Если true, возвращаются записи для всех сотрудников, сгруппированные по сотрудникам",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "staff_pin",
                openapi.IN_QUERY,
                description="PIN сотрудника (обязательный, если staffs_all не true)",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "download",
                openapi.IN_QUERY,
                description="Если true, возвращается ZIP-архив документов",
                type=openapi.TYPE_STRING,
            ),
        ],
        responses={
            200: openapi.Response(
                description=(
                    "При успешном выполнении возвращается:\n"
                    " - Если download=true, ZIP-архив документов.\n"
                    " - Если staffs_all=true, сгруппированные данные по сотрудникам.\n"
                    " - Иначе – список записей отсутствия для указанного сотрудника."
                )
            ),
            400: openapi.Response(description="Ошибка в параметрах запроса"),
        },
    )
    def get(self, request, *args, **kwargs):
        """
        **GET** запрос для получения записей отсутствия.

        **Параметры запроса:**
          - **start_date** (опционально): Дата начала (формат YYYY-MM-DD).
          - **end_date** (опционально): Дата окончания (формат YYYY-MM-DD).
          - **staffs_all** (опционально): Если true, возвращаются записи для всех сотрудников.
          - **staff_pin** (обязательно, если staffs_all не true): PIN сотрудника.
          - **download** (опционально): Если true, возвращается ZIP-архив документов.

        **Ответ:**
          - **HTTP 200 OK**: JSON с записями отсутствия или сгруппированными данными по сотрудникам, либо ZIP-архив.
          - **HTTP 400 BAD REQUEST**: При отсутствии обязательных параметров или неверном формате даты.
        """
        staffs_all = request.query_params.get("staffs_all", "").lower() == "true"
        staff_pin = request.query_params.get("staff_pin")

        if not staffs_all and not staff_pin:
            logger.warning("Не передан обязательный параметр 'staff_pin'.")
            return Response(
                {"error": "Параметр 'staff_pin' обязателен."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            query_start, query_end = self.get_date_interval(request)
        except ValueError as e:
            logger.error(f"Ошибка при разборе интервала дат: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        absences = models.AbsentReason.objects.filter(
            end_date__gte=query_start,
            start_date__lte=query_end,
        )

        if not staffs_all:
            absences = absences.filter(staff__pin=staff_pin)

        download = request.query_params.get("download", "").lower() == "true"
        if download:
            in_memory = BytesIO()
            with zipfile.ZipFile(in_memory, "w", zipfile.ZIP_DEFLATED) as zf:
                for absence in absences:
                    if absence.document:
                        file_path = absence.document.path
                        if os.path.exists(file_path):
                            _, ext = os.path.splitext(file_path)
                            ext = ext.lower()
                            staff = absence.staff
                            fio = f"{staff.surname}_{staff.name}"
                            new_filename = f"{staff.pin}_{fio}_{absence.id}{ext}"
                            zf.write(file_path, arcname=new_filename)
            in_memory.seek(0)
            logger.info(
                f"Возвращается ZIP-архив документов за период {query_start} - {query_end}."
            )
            response = HttpResponse(
                in_memory.getvalue(), content_type="application/zip"
            )
            response["Content-Disposition"] = (
                f'attachment; filename="documents_{query_start}_{query_end}.zip"'
            )
            return response

        if not staffs_all:
            serializer_context = {"minimal_staff": True}
            serializer = serializers.AbsentReasonSerializer(
                absences, many=True, context=serializer_context
            )
            logger.info(
                f"Возвращается список записей отсутствия для staff_pin: {staff_pin}. Количество записей: {len(absences)}."
            )
            return Response(serializer.data, status=status.HTTP_200_OK)

        grouped = {}
        for absence in absences:
            staff = absence.staff
            key = staff.pin
            if key not in grouped:
                grouped[key] = {
                    "staff": {"pin": staff.pin, "fio": f"{staff.surname} {staff.name}"},
                    "absences": [],
                }
            absence_data = serializers.AbsentReasonSerializer(
                absence, context={"minimal_staff": True}
            ).data
            grouped[key]["absences"].append(absence_data)
        result = list(grouped.values())
        logger.info(
            f"Возвращается сгруппированный список записей отсутствия для всех сотрудников. Количество групп: {len(result)}."
        )
        return Response(result, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_summary="Создание записи отсутствия",
        operation_description=(
            "Метод создает новую запись отсутствия.\n\n"
            "Тело запроса должно содержать следующие поля:\n"
            " - **staff** (строка): PIN сотрудника.\n"
            " - **reason** (строка): Код причины отсутствия (например, `sick_leave`).\n"
            '   Если передано некорректное значение, по умолчанию устанавливается `other` (отображается как "Другая причина").\n'
            " - **start_date** (строка): Дата начала (формат YYYY-MM-DD).\n"
            " - **end_date** (строка): Дата окончания (формат YYYY-MM-DD).\n"
            " - **approved** (bool): Статус утверждения.\n"
            " - **document** (файл, опционально): Прикрепленный документ. Разрешенные форматы: pdf, jpg, jpeg, png."
        ),
        tags=["Absence"],
        request_body=serializers.AbsentReasonSerializer,
        responses={
            201: openapi.Response(description="Запись отсутствия успешно создана."),
            400: openapi.Response(description="Неверные входные данные"),
        },
    )
    def post(self, request, *args, **kwargs):
        """
        **POST** запрос для создания новой записи отсутствия.

        **Тело запроса:**
          - **staff** (строка): PIN сотрудника.
          - **reason** (строка): Код причины отсутствия (например, `sick_leave` или `other`).
          - **start_date** (строка): Дата начала (формат YYYY-MM-DD).
          - **end_date** (строка): Дата окончания (формат YYYY-MM-DD).
          - **approved** (bool): Статус утверждения.
          - **document** (файл, опционально): Прикрепленный документ. Допустимые расширения: pdf, jpg, jpeg, png.

        **Ответ:**
          - **HTTP 201 CREATED**: Сообщение об успешном создании записи.
          - **HTTP 400 BAD REQUEST**: Если входные данные неверны.
        """
        if not request.user.is_authenticated:
            logger.warning(
                "Пользователь не аутентифицирован для POST /api/absent_staff/"
            )
            return Response(
                {"error": "Требуется аутентификация"},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = serializers.AbsentReasonSerializer(data=request.data)

        if serializer.is_valid():
            try:
                instance = serializer.save()
                logger.info(
                    f"Запись отсутствия создана. ID: {instance.id}, "
                    f"Сотрудник: {instance.staff.pin if hasattr(instance, 'staff') and hasattr(instance.staff, 'pin') else 'N/A'}, "
                    f"Период: {instance.start_date} - {instance.end_date}"
                )
                return Response(
                    {"message": "Запись отсутствия успешно создана."},
                    status=status.HTTP_201_CREATED,
                )
            except Exception as e:
                logger.error(
                    f"Ошибка при сохранении записи отсутствия: {str(e)}", exc_info=True
                )
                return Response(
                    {"error": f"Ошибка при сохранении: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        else:
            logger.warning(f"Ошибка валидации данных: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
