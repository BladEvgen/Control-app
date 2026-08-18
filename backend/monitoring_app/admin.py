import logging
import math
import mimetypes
import os
from calendar import month_abbr, monthrange
from collections import defaultdict
from contextlib import AbstractContextManager
from datetime import date, datetime, time, timedelta
from functools import reduce
from operator import or_
from typing import Any, cast
from urllib.parse import quote

from django.apps import apps as django_apps
from django.conf import settings
from django.contrib import admin, messages
from django.contrib.admin import SimpleListFilter
from django.contrib.admin.models import LogEntry
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import logout
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Avg, Case, Count, F, IntegerField, Q, Value, When
from django.db.utils import DatabaseError, OperationalError
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.formats import date_format
from django.utils.html import escape, format_html, format_html_join
from django.utils.safestring import mark_safe
from django_admin_geomap import ModelAdmin
from monitoring_app import utils as monitoring_utils
from monitoring_app.group_match import childdepartment_pks_for_group_style_search
from monitoring_app.lesson_locations_conf import (
    ACCEPTANCE_R_CLUSTER,
    ACCEPTANCE_R_SAME_POINT,
    ACCEPTANCE_R_STANDALONE,
    CLASS_LOCATION_ACCEPTANCE_RADII_CACHE_KEY,
    CLASS_LOCATION_ACCEPTANCE_RADII_CACHE_TTL,
    CLUSTER_THRESHOLD_M,
    SAME_POINT_THRESHOLD_M,
)
from monitoring_app.models import (
    AbsentReason,
    APIKey,
    ChildDepartment,
    ClassLocation,
    FileCategory,
    LessonAttendance,
    ParentDepartment,
    PasswordResetRequestLog,
    PasswordResetToken,
    PerformanceBonusRule,
    Position,
    PublicHoliday,
    RemoteWork,
    Salary,
    Staff,
    StaffAttendance,
    StaffFaceMask,
    StaffFaceSample,
    UserProfile,
)
from monitoring_app.pad_admin_summary import (
    _is_auto_insufficient_input,
    format_lesson_attendance_antifraud_list_hint,
    format_lesson_attendance_antifraud_operator_panel,
    format_lesson_attendance_pad_technical_compact,
)
from monitoring_app.staff_face_ml import (
    allowed_augment_basename,
    allowed_ml_basenames,
    augment_dir_for_pin,
    count_augment_images,
    face_ml_list_badge,
    list_augment_basenames,
    render_staff_face_ml_table,
    staff_workspace_dir,
)
from monitoring_app.staff_ml_preview_viz import (
    build_npy_embeddings_preview_body,
    build_pt_checkpoint_preview_body,
)

logger = logging.getLogger("monitoring_app.admin")
_MARKER = object()


def _db_atomic() -> AbstractContextManager[None]:
    """Типизированная обёртка над transaction.atomic() для статического анализа."""
    return cast(AbstractContextManager[None], transaction.atomic())


def _to_local_datetime(value):
    if value is None:
        return None
    if timezone.is_aware(value):
        return timezone.localtime(value)
    return timezone.make_aware(value, timezone.get_current_timezone())


def _format_local_time(value) -> str | None:
    local_value = _to_local_datetime(value)
    if local_value is None:
        return None
    return local_value.strftime("%H:%M")


def _shift_month_start(current_month_start: date, months_back: int) -> date:
    year = current_month_start.year
    month = current_month_start.month - months_back
    while month <= 0:
        month += 12
        year -= 1
    return current_month_start.replace(year=year, month=month, day=1)


def _parse_admin_iso_date(raw) -> date | None:
    try:
        return datetime.strptime(str(raw or "").strip()[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _aware_day_bounds(start_day: date, end_day: date):
    tz = timezone.get_current_timezone()
    return (
        timezone.make_aware(datetime.combine(start_day, time.min), tz),
        timezone.make_aware(datetime.combine(end_day, time.max), tz),
    )


def _admin_theme_name() -> str:
    if django_apps.is_installed("unfold"):
        return "unfold"
    if django_apps.is_installed("grappelli"):
        return "grappelli"
    return "django"


def _academic_year_bounds(today: date) -> tuple[date, date]:
    start_year = today.year if today.month >= 9 else today.year - 1
    return date(start_year, 9, 1), date(start_year + 1, 7, 31)


def _excel_col_width(max_len: int, *, min_w: float, max_w: float) -> float:
    return min(max(max_len + 2.0, min_w), max_w)


def _radius_bbox(latitude: float, radius_m: int) -> tuple[float, float]:
    lat_margin = radius_m / 111_320
    cos_lat = max(math.cos(math.radians(latitude)), 0.01)
    lon_margin = radius_m / (111_320 * cos_lat)
    return lat_margin, lon_margin


def _admin_badge(label: str, *, background: str, color: str = "#fff"):
    return format_html(
        '<span style="display:inline-flex; align-items:center; gap:4px; padding:3px 8px; '
        'border-radius:999px; font-size:11px; font-weight:600; background:{}; color:{};">{}</span>',
        background,
        color,
        label,
    )


def _staff_attendance_history_legend_badge(
    label: str, *, background: str, color: str = "#fff"
):
    """Бейдж в блоке легенды истории посещаемости (отступы задаются в CSS)."""
    return format_html(
        '<span class="staff-attendance-history__badge" style="display:inline-flex; align-items:center; '
        "gap:4px; padding:5px 11px; border-radius:999px; font-size:11px; font-weight:600; "
        'background:{}; color:{};">{}</span>',
        background,
        color,
        label,
    )


# ===== Admin Site Configuration =====
class MonitoringAdminSite(admin.AdminSite):
    site_header = "Панель управления мониторинга"
    site_title = "Административная панель"
    index_title = "Управление системой мониторинга"

    def get_app_list(self, request, app_label=None):
        """
        Override to organize models into custom groups
        """
        if app_label:
            return super().get_app_list(request)

        app_dict = self._build_app_dict(request)

        groups = {
            "auth": {
                "name": "Авторизация и безопасность",
                "models": [
                    "PasswordResetToken",
                    "PasswordResetRequestLog",
                    "APIKey",
                    "UserProfile",
                ],
                "icon": "fa fa-lock",
            },
            "staff": {
                "name": "Персонал",
                "models": [
                    "Staff",
                    "Position",
                    "StaffFaceMask",
                    "Salary",
                    "AbsentReason",
                    "RemoteWork",
                ],
                "icon": "fa fa-users",
            },
            "department": {
                "name": "Организационная структура",
                "models": ["ParentDepartment", "ChildDepartment"],
                "icon": "fa fa-sitemap",
            },
            "attendance": {
                "name": "Учет посещаемости",
                "models": ["StaffAttendance", "LessonAttendance", "PublicHoliday"],
                "icon": "fa fa-calendar-check-o",
            },
            "location": {
                "name": "Локации и пространственные данные",
                "models": ["ClassLocation"],
                "icon": "fa fa-map-marker",
            },
            "configuration": {
                "name": "Настройки системы",
                "models": ["FileCategory", "PerformanceBonusRule"],
                "icon": "fa fa-cogs",
            },
        }

        grouped_apps = []
        all_models = {}
        for app_label, app_data in app_dict.items():
            models_list = app_data.get("models", [])
            if not models_list:
                continue
            for model_data in models_list:
                model_name = model_data.get("object_name")
                if model_name:
                    all_models[model_name] = model_data

        for group_id, group_info in groups.items():
            group_models = []
            for model_name in group_info["models"]:
                if model_name in all_models:
                    group_models.append(all_models[model_name])

            if group_models:
                grouped_apps.append(
                    {
                        "name": group_info["name"],
                        "app_label": group_id,
                        "app_url": "#",
                        "has_module_perms": True,
                        "models": sorted(group_models, key=lambda x: x.get("name", "")),
                        "icon": group_info["icon"],
                    }
                )

        used_models = set()
        for group_info in groups.values():
            used_models.update(group_info["models"])

        unused_models = [
            model_data
            for model_name, model_data in all_models.items()
            if model_name not in used_models
        ]

        if unused_models:
            grouped_apps.append(
                {
                    "name": "Другие",
                    "app_label": "other",
                    "app_url": "#",
                    "has_module_perms": True,
                    "models": sorted(unused_models, key=lambda x: x.get("name", "")),
                    "icon": "fa fa-list",
                }
            )

        if not grouped_apps:
            return super().get_app_list(request)

        return sorted(grouped_apps, key=lambda x: x["name"])

    def each_context(self, request):
        context = super().each_context(request)
        context.update(
            {
                "has_permission": request.user.is_active and request.user.is_staff,
            }
        )
        return context

    def app_index(self, request, app_label, extra_context=None):
        """При клике на «Система мониторинга» в хлебных крошках — на главную с дашбордом."""
        if app_label == "monitoring_app":
            return redirect(reverse(f"{self.name}:index"))
        return super().app_index(request, app_label, extra_context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("logout/", self.admin_view(self.admin_logout_view), name="logout"),
            path("dashboard/", self.admin_view(self.dashboard_view), name="dashboard"),
            path(
                "api/attendance-stats/",
                self.admin_view(self.attendance_stats_api),
                name="attendance-stats-api",
            ),
            path(
                "api/department-stats/",
                self.admin_view(self.department_stats_api),
                name="department-stats-api",
            ),
        ]
        return custom_urls + urls

    def admin_logout_view(self, request):
        """Logout с поддержкой GET (для ссылок, закладок). Django 5+ по умолчанию требует POST."""
        logout(request)
        return redirect("/admin/")

    @method_decorator(staff_member_required)
    def dashboard_view(self, request):
        logger.debug("MonitoringAdminSite.dashboard_view")
        context = {
            **self.each_context(request),
            "title": "Панель мониторинга",
        }

        context["staff_count"] = Staff.objects.count()
        context["today_attendance"] = StaffAttendance.objects.filter(
            date_at=_staff_attendance_db_date_for_calendar_work_day(
                timezone.now().date()
            )
        ).count()

        context["recent_logs"] = LogEntry.objects.select_related(
            "content_type", "user"
        )[:10]

        departments = ChildDepartment.objects.annotate(
            staff_count=Count("staff")
        ).order_by("-staff_count")[:5]

        context["departments"] = departments

        return TemplateResponse(request, "admin/dashboard.html", context)

    def attendance_stats_api(self, request):
        """API endpoint for attendance statistics"""
        days = int(request.GET.get("days", 30))
        start_date = timezone.now().date() - timedelta(days=days)

        attendance_data = (
            StaffAttendance.objects.filter(
                date_at__gte=_staff_attendance_db_date_for_calendar_work_day(start_date)
            )
            .values("date_at")
            .annotate(count=Count("id"))
            .order_by("date_at")
        )

        return JsonResponse(
            {
                "labels": [
                    str(item["date_at"] - timedelta(days=1)) for item in attendance_data
                ],
                "data": [item["count"] for item in attendance_data],
            }
        )

    def department_stats_api(self, request):
        """API endpoint for department statistics"""
        departments = ChildDepartment.objects.annotate(
            staff_count=Count("staff", distinct=True),
            avg_salary=Avg("staff__salary__net_salary"),
        ).values("name", "staff_count", "avg_salary")

        return JsonResponse(
            {
                "departments": list(departments),
            }
        )


admin_site = MonitoringAdminSite(name="monitoring_admin")


def _staff_attendance_db_date_for_calendar_work_day(work_day: date) -> date:
    """Дата поля date_at строки СКУД: на сутки позже календарного дня смены (день выгрузки)."""
    return work_day + timedelta(days=1)


def _format_staffattendance_effective_duration(record) -> str | None:
    """Длительность «в здании» как в StaffAttendance: сначала effective_work_seconds, иначе last_out − first_in."""
    total_seconds = None
    if getattr(record, "effective_work_seconds", None) is not None:
        total_seconds = int(record.effective_work_seconds)
    elif getattr(record, "first_in", None) and getattr(record, "last_out", None):
        total_seconds = int((record.last_out - record.first_in).total_seconds())
    if total_seconds is None:
        return None
    if total_seconds < 0:
        return "ошибка времени"
    if total_seconds > 86400:
        total_seconds = 86400
    hours, remainder = divmod(total_seconds, 3600)
    minutes = remainder // 60
    return f"{hours:d}ч {minutes:02d}м"


# ===== Common Filter Classes =====


class UsedFilter(admin.SimpleListFilter):
    title = "Статус использования"
    parameter_name = "used"

    def lookups(self, request, model_admin):
        return (
            ("yes", "Использован"),
            ("no", "Не использован"),
        )

    def queryset(self, request, queryset):
        if self.value() == "yes":
            return queryset.filter(_used=True)
        elif self.value() == "no":
            return queryset.filter(_used=False)
        return queryset


class DepartmentHierarchyFilter(SimpleListFilter):
    title = "Отдел"
    parameter_name = "department_hierarchy"

    def lookups(self, request, model_admin):
        cache_key = "department_hierarchy_lookups"
        lookup_list = cache.get(cache_key)

        if lookup_list is None:
            departments = ChildDepartment.objects.all().select_related("parent")
            hierarchy = self.build_hierarchy(departments)
            lookup_list = self.get_department_choices(hierarchy)
            cache.set(cache_key, lookup_list, 3600)

        return lookup_list

    def queryset(self, request, queryset):
        if self.value():
            department_ids = self.get_all_descendants(self.value())
            return queryset.filter(department__in=department_ids)
        return queryset

    def build_hierarchy(self, departments):
        hierarchy = defaultdict(list)
        for dept in departments:
            hierarchy[dept.parent_id].append(dept)
        return hierarchy

    def get_all_descendants(self, department_id):
        cache_key = f"department_descendants_{department_id}"
        descendants = cache.get(cache_key)

        if descendants is None:
            try:
                dept_id = int(department_id)
            except (ValueError, TypeError):
                dept_id = department_id

            queue = [dept_id]
            descendants = set(queue)
            while queue:
                current = queue.pop(0)
                children = ChildDepartment.objects.filter(
                    parent_id=current
                ).values_list("id", flat=True)
                queue.extend(children)
                descendants.update(children)
            cache.set(cache_key, descendants, 3600)

        return descendants

    def get_department_choices(self, hierarchy, parent_id=None, level=0):
        choices = []
        if parent_id is None:
            root_departments = hierarchy[None]
        else:
            root_departments = hierarchy.get(parent_id, [])

        for dept in root_departments:
            indent = "—" * level
            choices.append((dept.id, f"{indent} {dept.name}"))
            choices.extend(self.get_department_choices(hierarchy, dept.id, level + 1))

        return choices


class DateRangeFilter(admin.SimpleListFilter):
    """A filter for date ranges"""

    title = "Период"
    parameter_name = "date_range"

    def lookups(self, request, model_admin):
        return (
            ("last_14", "Последние 14 календарных дней"),
            ("last_30", "Последние 30 календарных дней"),
            ("today", "Сегодня"),
            ("yesterday", "Вчера"),
            ("this_week", "Эта неделя"),
            ("last_week", "Прошлая неделя"),
            ("this_month", "Этот месяц"),
            ("last_month", "Прошлый месяц"),
            ("this_quarter", "Этот квартал"),
            ("this_year", "Этот год"),
            ("all", "Все даты (медленно на больших таблицах)"),
        )

    def queryset(self, request, queryset):
        today = timezone.now().date()
        one = timedelta(days=1)
        is_skud = queryset.model == StaffAttendance

        if self.value() == "all":
            return queryset
        if self.value() == "last_14":
            end_wd = today
            start_wd = today - timedelta(days=13)
            if is_skud:
                return queryset.filter(
                    date_at__gte=_staff_attendance_db_date_for_calendar_work_day(
                        start_wd
                    ),
                    date_at__lte=_staff_attendance_db_date_for_calendar_work_day(
                        end_wd
                    ),
                )
            return queryset.filter(date_at__gte=start_wd, date_at__lte=end_wd)
        if self.value() == "last_30":
            end_wd = today
            start_wd = today - timedelta(days=29)
            if is_skud:
                return queryset.filter(
                    date_at__gte=_staff_attendance_db_date_for_calendar_work_day(
                        start_wd
                    ),
                    date_at__lte=_staff_attendance_db_date_for_calendar_work_day(
                        end_wd
                    ),
                )
            return queryset.filter(date_at__gte=start_wd, date_at__lte=end_wd)
        if self.value() == "today":
            row_day = (
                _staff_attendance_db_date_for_calendar_work_day(today)
                if is_skud
                else today
            )
            return queryset.filter(date_at=row_day)
        elif self.value() == "yesterday":
            row_day = today if is_skud else today - one
            return queryset.filter(date_at=row_day)
        elif self.value() == "this_week":
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            if is_skud:
                return queryset.filter(
                    date_at__gte=_staff_attendance_db_date_for_calendar_work_day(
                        week_start
                    ),
                    date_at__lte=_staff_attendance_db_date_for_calendar_work_day(
                        week_end
                    ),
                )
            return queryset.filter(date_at__gte=week_start, date_at__lte=week_end)
        elif self.value() == "last_week":
            week_start = today - timedelta(days=today.weekday() + 7)
            week_end = week_start + timedelta(days=6)
            if is_skud:
                return queryset.filter(
                    date_at__gte=_staff_attendance_db_date_for_calendar_work_day(
                        week_start
                    ),
                    date_at__lte=_staff_attendance_db_date_for_calendar_work_day(
                        week_end
                    ),
                )
            return queryset.filter(date_at__gte=week_start, date_at__lte=week_end)
        elif self.value() == "this_month":
            first = today.replace(day=1)
            last_dom = monthrange(today.year, today.month)[1]
            last = today.replace(day=last_dom)
            if is_skud:
                return queryset.filter(
                    date_at__gte=_staff_attendance_db_date_for_calendar_work_day(first),
                    date_at__lte=_staff_attendance_db_date_for_calendar_work_day(last),
                )
            return queryset.filter(date_at__year=today.year, date_at__month=today.month)
        elif self.value() == "last_month":
            first_this = today.replace(day=1)
            last_prev = first_this - one
            first_prev = last_prev.replace(day=1)
            last_prev_dom = monthrange(first_prev.year, first_prev.month)[1]
            last_prev = first_prev.replace(day=last_prev_dom)
            if is_skud:
                return queryset.filter(
                    date_at__gte=_staff_attendance_db_date_for_calendar_work_day(
                        first_prev
                    ),
                    date_at__lte=_staff_attendance_db_date_for_calendar_work_day(
                        last_prev
                    ),
                )
            last_month = today.month - 1 if today.month > 1 else 12
            year = today.year if today.month > 1 else today.year - 1
            return queryset.filter(date_at__year=year, date_at__month=last_month)
        elif self.value() == "this_quarter":
            quarter = (today.month - 1) // 3 + 1
            first_month = 3 * quarter - 2
            first_day = date(today.year, first_month, 1)
            last_month_of_q = first_month + 2
            last_dom = monthrange(today.year, last_month_of_q)[1]
            last_day = date(today.year, last_month_of_q, last_dom)
            if is_skud:
                return queryset.filter(
                    date_at__gte=_staff_attendance_db_date_for_calendar_work_day(
                        first_day
                    ),
                    date_at__lte=_staff_attendance_db_date_for_calendar_work_day(
                        last_day
                    ),
                )
            return queryset.filter(
                date_at__year=today.year,
                date_at__month__gte=first_month,
                date_at__month__lte=first_month + 2,
            )
        elif self.value() == "this_year":
            if is_skud:
                return queryset.filter(
                    date_at__gte=_staff_attendance_db_date_for_calendar_work_day(
                        date(today.year, 1, 1)
                    ),
                    date_at__lte=_staff_attendance_db_date_for_calendar_work_day(
                        date(today.year, 12, 31)
                    ),
                )
            return queryset.filter(date_at__year=today.year)
        return queryset


class StaffAttendanceRowZoneFilter(admin.SimpleListFilter):
    """Скрытие строк без зон — раньше задавалось GET exclude_unknown=yes."""

    title = "Строки без зоны (вход/выход)"
    parameter_name = "staff_row_zone"

    def lookups(self, request, model_admin):
        return (
            ("hide", "Скрывать пустые и Unknown (быстрее)"),
            ("show", "Показать все строки"),
        )

    def queryset(self, request, queryset):
        if queryset.model != StaffAttendance:
            return queryset
        if request.GET.get("exclude_unknown") == "no":
            return queryset
        if self.value() == "show":
            return queryset
        q_conditions = [
            Q(area_name_in__isnull=True),
            Q(area_name_out__isnull=True),
            Q(area_name_in="Unknown"),
            Q(area_name_out="Unknown"),
        ]
        return queryset.exclude(reduce(or_, q_conditions))


class StaffAttendanceSkudDataFilter(admin.SimpleListFilter):
    title = "Данные СКУД (время / секунды)"
    parameter_name = "staff_skud_data"

    def lookups(self, request, model_admin):
        return (
            ("has", "Есть first_in или effective_work_seconds"),
            ("empty", "Нет ни времени, ни секунд"),
        )

    def queryset(self, request, queryset):
        if queryset.model != StaffAttendance:
            return queryset
        if self.value() == "has":
            return queryset.filter(
                Q(first_in__isnull=False) | Q(effective_work_seconds__isnull=False)
            )
        if self.value() == "empty":
            return queryset.filter(
                first_in__isnull=True,
                effective_work_seconds__isnull=True,
            )
        return queryset


class AttendanceStatusFilter(admin.SimpleListFilter):
    title = "Статус присутствия"
    parameter_name = "attendance_status"

    def lookups(self, request, model_admin):
        return (
            ("present", "Присутствует"),
            ("absent", "Отсутствует"),
            ("late", "Опоздал"),
            ("left_early", "Ушел раньше"),
            ("remote", "Удаленно"),
            ("partial", "Неполный рабочий день"),
        )

    def queryset(self, request, queryset):
        today = timezone.now().date()
        skud_row_date = _staff_attendance_db_date_for_calendar_work_day(today)

        late_threshold_minutes = 15
        early_leave_threshold_minutes = 15
        minimum_workday_hours = 4
        standard_workday_hours = 8

        work_start = timezone.now().replace(hour=9, minute=0, second=0, microsecond=0)
        work_end = timezone.now().replace(hour=18, minute=0, second=0, microsecond=0)

        late_threshold = work_start + timedelta(minutes=late_threshold_minutes)
        early_leave_threshold = work_end - timedelta(
            minutes=early_leave_threshold_minutes
        )

        if self.value() == "present":
            return queryset.filter(
                Q(attendance__date_at=skud_row_date),
                Q(attendance__first_in__isnull=False),
                Q(attendance__absence_reason__isnull=True),
                Q(attendance__last_out__isnull=True)
                | Q(attendance__last_out__gte=early_leave_threshold),
            ).distinct()

        elif self.value() == "absent":
            staff_with_attendance = queryset.filter(
                attendance__date_at=skud_row_date,
                attendance__first_in__isnull=False,
            ).distinct()

            return queryset.filter(
                Q(id__in=queryset.exclude(id__in=staff_with_attendance))
                | Q(
                    attendance__date_at=skud_row_date,
                    attendance__absence_reason__isnull=False,
                )
            ).distinct()

        elif self.value() == "late":
            return queryset.filter(
                attendance__date_at=skud_row_date,
                attendance__first_in__gt=late_threshold,
            ).distinct()

        elif self.value() == "left_early":
            return queryset.filter(
                attendance__date_at=skud_row_date,
                attendance__last_out__isnull=False,
                attendance__last_out__lt=early_leave_threshold,
            ).distinct()

        elif self.value() == "remote":
            return queryset.filter(
                Q(remote_work__permanent_remote=True)
                | Q(
                    remote_work__start_date__lte=today, remote_work__end_date__gte=today
                )
            ).distinct()

        elif self.value() == "partial":
            return (
                queryset.filter(
                    attendance__date_at=skud_row_date,
                    attendance__first_in__isnull=False,
                    attendance__last_out__isnull=False,
                )
                .annotate(
                    workday_duration=F("attendance__last_out")
                    - F("attendance__first_in")
                )
                .filter(
                    workday_duration__gte=timedelta(hours=minimum_workday_hours),
                    workday_duration__lt=timedelta(hours=standard_workday_hours),
                )
                .distinct()
            )

        return queryset


# ===== AUTHENTICATION MODELS =====


@admin.register(PasswordResetToken, site=admin_site)
class PasswordResetTokenAdmin(admin.ModelAdmin):
    list_display = ("user", "created_at", "used", "is_valid", "expiration_time")
    list_filter = (UsedFilter, "created_at")
    search_fields = ("user__username", "user__email", "token")
    readonly_fields = (
        "user",
        "token",
        "created_at",
        "used",
        "is_valid",
        "expiration_time",
    )
    ordering = ("-created_at",)

    def is_valid(self, obj):
        return obj.is_valid()

    is_valid.boolean = True
    is_valid.short_description = "Действительный токен"

    def expiration_time(self, obj):
        if obj.is_valid():
            expiration = obj.created_at + timezone.timedelta(hours=1)
            time_left = expiration - timezone.now()
            hours = time_left.seconds // 3600
            minutes = (time_left.seconds % 3600) // 60

            if time_left.days < 0 or (
                time_left.days == 0 and hours == 0 and minutes == 0
            ):
                return format_html('<span style="color: red;">Истек</span>')

            return format_html(
                '<span style="color: green;">Действителен еще {} ч. {} мин.</span>',
                hours,
                minutes,
            )
        return format_html('<span style="color: red;">Истек</span>')

    expiration_time.short_description = "Время до истечения"

    fieldsets = (
        (
            None,
            {
                "fields": (
                    "user",
                    "token",
                    "created_at",
                    "used",
                    "is_valid",
                    "expiration_time",
                ),
                "classes": ("wide",),
            },
        ),
    )

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(PasswordResetRequestLog, site=admin_site)
class PasswordResetRequestLogAdmin(admin.ModelAdmin):
    list_display = (
        "user",
        "ip_address",
        "requested_at",
        "next_possible_request",
        "time_until_next",
    )
    list_filter = ("requested_at",)
    search_fields = ("user__username", "user__email", "ip_address")
    readonly_fields = (
        "user",
        "ip_address",
        "requested_at",
        "next_possible_request",
        "time_until_next",
    )
    ordering = ("-requested_at",)

    def next_possible_request(self, obj):
        return obj.requested_at + timezone.timedelta(minutes=5)

    next_possible_request.short_description = "Следующий возможный запрос"

    def time_until_next(self, obj):
        next_time = obj.requested_at + timezone.timedelta(minutes=5)
        time_left = next_time - timezone.now()

        if time_left.total_seconds() <= 0:
            return format_html(
                '<span style="color: green;">Доступно (можно запросить новый сброс пароля)</span>'
            )

        minutes = int(time_left.total_seconds() // 60)
        seconds = int(time_left.total_seconds() % 60)

        return format_html(
            '<span style="color: orange;">Ожидание {} мин. {} сек. до следующего запроса</span>',
            minutes,
            seconds,
        )

    time_until_next.short_description = "Статус доступности запроса"

    fieldsets = (
        (
            None,
            {
                "fields": (
                    "user",
                    "ip_address",
                    "requested_at",
                    "next_possible_request",
                    "time_until_next",
                ),
            },
        ),
    )

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(APIKey, site=admin_site)
class APIKeyAdmin(admin.ModelAdmin):
    list_display = (
        "key_name",
        "created_by",
        "created_at",
        "short_key",
        "is_active",
    )
    list_filter = ("created_at", "created_by", "is_active")
    list_editable = ("is_active",)
    search_fields = ("key_name", "created_by__username")
    ordering = ("-created_at", "key_name")
    readonly_fields = ("key", "created_at", "created_by")
    actions = ["deactivate_keys", "reactivate_keys", "generate_new_keys"]

    fieldsets = (
        (
            "Основная информация",
            {
                "fields": ("key_name", "key", "is_active"),
                "classes": ("wide",),
            },
        ),
        (
            "Дополнительная информация",
            {
                "fields": ("created_by", "created_at"),
                "classes": ("grp-collapse grp-closed",),
            },
        ),
    )

    def short_key(self, obj):
        return format_html(
            '<span class="copy-to-clipboard api-key" data-clipboard-text="{}">'
            '<span class="key-preview">{}</span>'
            '<span class="copy-icon">📋</span>'
            "</span>",
            obj.key,
            f"{obj.key[:8]}...",
        )

    short_key.short_description = "Ключ API"

    def save_model(self, request, obj, form, change):
        logger.debug(
            "APIKeyAdmin.save_model key_name=%s change=%s", obj.key_name, change
        )
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
        logger.info("APIKeyAdmin.save_model OK key_name=%s", obj.key_name)

    def deactivate_keys(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f"{updated} ключей деактивировано.")

    deactivate_keys.short_description = "Деактивировать выбранные ключи"

    def reactivate_keys(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f"{updated} ключей активировано.")

    reactivate_keys.short_description = "Активировать выбранные ключи"

    class Media:
        css = {"all": ("admin/css/apikey_admin.css",)}
        js = ("admin/js/clipboard.min.js", "admin/js/copy-to-clipboard.js")


@admin.register(UserProfile, site=admin_site)
class UserProfileAdmin(admin.ModelAdmin):
    list_display = (
        "user",
        "is_banned",
        "phonenumber",
        "last_login_ip",
        "last_login",
    )
    list_filter = ("is_banned",)
    search_fields = ("user__username", "user__email", "phonenumber", "last_login_ip")
    ordering = ("user__username",)
    list_display_links = ("user",)
    list_editable = ("is_banned",)
    actions = ["ban_users", "unban_users"]
    readonly_fields = ("last_login_ip",)

    fieldsets = (
        (
            "Основная информация",
            {
                "fields": ("user", "phonenumber"),
                "classes": ("wide",),
            },
        ),
        (
            "Безопасность",
            {
                "fields": ("is_banned",),
                "classes": ("wide",),
            },
        ),
        (
            "Информация о входе",
            {
                "fields": ("last_login_ip",),
                "classes": ("grp-collapse grp-closed",),
            },
        ),
    )

    def last_login(self, obj):
        return obj.user.last_login if obj.user.last_login else "Никогда"

    last_login.short_description = "Последний вход"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.select_related("user")
        return qs

    def ban_users(self, request, queryset):
        updated = queryset.update(is_banned=True)
        self.message_user(request, f"{updated} пользователей заблокировано.")

    ban_users.short_description = "Заблокировать выбранных пользователей"

    def unban_users(self, request, queryset):
        updated = queryset.update(is_banned=False)
        self.message_user(request, f"{updated} пользователей разблокировано.")

    unban_users.short_description = "Разблокировать выбранных пользователей"

    def has_delete_permission(self, request, obj=None):
        """Удаление профиля каскадно удаляет User — в админке отключено."""
        return False


# ===== STAFF AND DEPARTMENT MODELS =====


@admin.register(FileCategory, site=admin_site)
class FileCategoryAdmin(admin.ModelAdmin):
    list_display = (
        "name",
        "slug",
    )
    search_fields = ("name", "slug")
    prepopulated_fields = {"slug": ("name",)}
    ordering = ("name",)
    list_display_links = ("name",)


@admin.register(ParentDepartment, site=admin_site)
class ParentDepartmentAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "name",
        "date_of_creation",
    )
    search_fields = ("name", "id")
    ordering = ("name",)
    readonly_fields = ("date_of_creation",)
    list_display_links = ("id", "name")

    fieldsets = (
        (
            "Информация об отделе",
            {
                "fields": ("id", "name", "date_of_creation"),
                "classes": ("wide",),
            },
        ),
    )

    def get_readonly_fields(self, request, obj=None):
        ro = list(self.readonly_fields)
        if obj is not None:
            ro.extend(["id", "name"])
        return ro

    def get_fieldsets(self, request, obj=None):
        if obj is None:
            return (
                (
                    "Информация об отделе",
                    {"fields": ("id", "name"), "classes": ("wide",)},
                ),
            )
        return self.fieldsets


@admin.register(ChildDepartment, site=admin_site)
class ChildDepartmentAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "name",
        "parent",
        "date_of_creation",
        "staff_count",
        "avg_salary",
    )
    search_fields = ("name", "parent__name")
    ordering = ("name",)
    list_filter = ("parent",)
    readonly_fields = ("date_of_creation", "staff_count", "avg_salary")
    list_display_links = ("id", "name")

    fieldsets = (
        (
            "Информация об отделе",
            {
                "fields": ("id", "name", "parent", "date_of_creation"),
                "classes": ("wide",),
            },
        ),
        (
            "Статистика",
            {
                "fields": ("staff_count", "avg_salary"),
                "classes": ("grp-collapse grp-closed",),
            },
        ),
    )

    def get_readonly_fields(self, request, obj=None):
        """При создании записи можно задать id, name и parent; после сохранения — только чтение."""
        ro = list(self.readonly_fields)
        if obj is not None:
            ro.extend(["parent", "name", "id"])
        return ro

    def get_fieldsets(self, request, obj=None):
        """На форме добавления скрыта пустая статистика и авто-поле даты."""
        info = (
            "Информация об отделе",
            {
                "fields": ("id", "name", "parent"),
                "classes": ("wide",),
            },
        )
        if obj is None:
            return (info,)
        return self.fieldsets

    def staff_count(self, obj):
        return getattr(obj, "_staff_count", 0)

    staff_count.short_description = "Количество сотрудников"

    def avg_salary(self, obj):
        avg = getattr(obj, "_avg_salary", None)
        return f"{int(avg)} руб." if avg else "Н/Д"

    avg_salary.short_description = "Средняя зарплата"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.annotate(
            _staff_count=Count("staff", distinct=True),
            _avg_salary=Avg("staff__salaries__net_salary"),
        )
        return qs

    def get_search_results(self, request, queryset, search_term):
        """Дополняет icontains поиск совпадением по ``compact_group_match_key`` (Nitro ↔ control).

        Логика ключей и префиксов — в :mod:`monitoring_app.group_match` (doctest там же).
        """
        qs, use_distinct = super().get_search_results(request, queryset, search_term)
        term = (search_term or "").strip()
        if not term:
            return qs, use_distinct
        pks = childdepartment_pks_for_group_style_search(queryset, term)
        if not pks:
            return qs, use_distinct
        key_qs = queryset.filter(pk__in=pks)
        combined = qs | key_qs
        return combined, True

    save_on_top = True


@admin.register(Position, site=admin_site)
class PositionAdmin(admin.ModelAdmin):
    list_display = ("name", "rate", "staff_count")
    search_fields = ("name",)
    ordering = ("-rate", "name")
    list_editable = ("rate",)
    list_display_links = ("name",)
    save_on_top = True

    fieldsets = (
        (
            "Информация о должности",
            {
                "fields": ("name", "rate"),
                "classes": ("wide",),
            },
        ),
    )

    def staff_count(self, obj):
        return getattr(obj, "_staff_count", 0)

    staff_count.short_description = "Количество сотрудников"

    def get_queryset(self, request):
        return super().get_queryset(request).annotate(_staff_count=Count("staff"))


class SalaryInline(admin.TabularInline):
    model = Salary
    extra = 0
    fields = ("net_salary", "total_salary", "contract_type")
    readonly_fields = ("total_salary",)


class AbsentReasonInline(admin.TabularInline):
    model = AbsentReason
    extra = 0
    fields = ("reason", "start_date", "end_date", "approved", "document")
    readonly_fields = ("approved",)


class RemoteWorkInline(admin.TabularInline):
    model = RemoteWork
    extra = 0
    fields = ("permanent_remote", "start_date", "end_date")


@admin.register(Staff, site=admin_site)
class StaffAdmin(admin.ModelAdmin):
    ATTENDANCE_HISTORY_DAYS = 7
    ATTENDANCE_HISTORY_CACHE_TTL = 3600
    ATTENDANCE_HISTORY_CACHE_VERSION = "event_day_v3-1"
    change_list_template = "admin/change_list_filter_sidebar.html"
    list_display = (
        "pin",
        "full_name",
        "department",
        "display_positions",
        "needs_training_status",
        "face_recognition_ml_badge",
    )
    list_display_links = ("pin", "full_name")
    list_per_page = 50
    list_filter = (
        DepartmentHierarchyFilter,
        "positions",
        "needs_training",
        AttendanceStatusFilter,
    )
    search_fields = ("pin", "surname", "name", "department__name")
    filter_horizontal = ("positions",)
    actions = [
        "clear_avatars",
        "assign_position",
        "mark_needs_training_true",
        "export_staff_data",
    ]
    ordering = ("-pin", "-department", "surname", "name")
    save_on_top = True
    inlines = [SalaryInline, AbsentReasonInline, RemoteWorkInline]
    readonly_fields = (
        "pin",
        "avatar_thumbnail",
        "face_recognition_ml_panel",
        "attendance_history",
    )

    fieldsets = (
        (
            "Личная информация",
            {
                "fields": (("surname", "name"), "pin", "avatar", "avatar_thumbnail"),
                "classes": ("wide",),
            },
        ),
        (
            "Должность и отдел",
            {
                "fields": ("department", "positions"),
                "classes": ("wide",),
            },
        ),
        (
            "Машинное обучение (лица)",
            {
                "fields": ("needs_training", "face_recognition_ml_panel"),
                "description": "Файлы рядом с аватаром: embeddings.npy, model.pt, best_model.pt; "
                "аугментации — в каталоге AUGMENT_ROOT (см. строку «Аугментации» в таблице).",
                "classes": ("wide", "grp-collapse grp-open"),
            },
        ),
        (
            "История посещаемости",
            {
                "fields": ("attendance_history",),
                "description": "Семь завершённых дней до вчера (без сегодня). Праздники — из справочника "
                "(рабочий/выходной). Эффективное время — как в списке посещаемости.",
                "classes": ("wide",),
            },
        ),
    )

    @admin.display(description="ФИО", ordering=cast(Any, ("surname", "name")))
    def full_name(self, obj):
        return f"{obj.surname} {obj.name}"

    def avatar_thumbnail(self, obj):
        cache_key = f"avatar_thumbnail_{obj.pin}"
        cached_html = cache.get(cache_key)

        if cached_html:
            return format_html(cached_html)

        if obj.avatar:
            html = format_html(
                """
                <div style="display: flex; justify-content: center; align-items: center; height: 80px; width: 80px; overflow: hidden; border-radius: 50%;">
                    <img src="{}" style="height: 100%; width: 100%; object-fit: cover; display: block;"/>
                </div>
                """,
                obj.avatar.url,
            )
            cache.set(cache_key, html, timeout=86400)
            return html

        no_photo_html = format_html(
            """
            <div style="display: flex; justify-content: center; align-items: center; height: 80px; width: 80px; border-radius: 50%; background-color: #f0f0f0;">
                <span style="color: #999; font-style: italic; text-align: center;">Нет фото</span>
            </div>
            """
        )
        cache.set(cache_key, no_photo_html, timeout=86400)
        return no_photo_html

    avatar_thumbnail.short_description = "Фото"

    def needs_training_status(self, obj):
        if obj.needs_training:
            return format_html(
                '<span style="color: red;">Требуется обучение модели</span>'
            )
        return format_html('<span style="color: green;">Модель обучена</span>')

    needs_training_status.short_description = "Статус обучения модели"

    @admin.display(description="ML-файлы")
    def face_recognition_ml_badge(self, obj):
        """Changelist badge for embeddings, checkpoints, and augment count.

        Args:
            obj: ``Staff`` row.

        Returns:
            ``SafeString`` compact HTML badge.
        """
        return face_ml_list_badge(obj)

    @admin.display(description="Эмбеддинги / .pt / аугментации")
    def face_recognition_ml_panel(self, obj):
        """Read-only ML artifact table with download, preview, and augment gallery links.

        Args:
            obj: ``Staff`` instance (unsaved records show a short help message).

        Returns:
            ``SafeString`` HTML from ``render_staff_face_ml_table``.
        """
        if not obj.pk:
            return mark_safe(
                '<p class="help" style="margin:0;">Сохраните запись — появится таблица файлов и ссылки.</p>'
            )
        dl = reverse(
            "admin:monitoring_app_staff_ml_file",
            kwargs={"object_id": obj.pk},
        )
        pv = reverse(
            "admin:monitoring_app_staff_ml_preview",
            kwargs={"object_id": obj.pk},
        )
        gal = reverse(
            "admin:monitoring_app_staff_ml_augment_gallery",
            kwargs={"object_id": obj.pk},
        )
        return render_staff_face_ml_table(
            obj,
            file_download_url=dl,
            file_preview_url=pv,
            augment_gallery_url=gal,
        )

    def get_urls(self):
        """Register staff-scoped ML file, preview, and augment gallery routes.

        Returns:
            URL patterns with custom paths prepended before default ``ModelAdmin`` URLs.
        """
        urls = super().get_urls()
        info = self.model._meta.app_label, self.model._meta.model_name
        custom = [
            path(
                "<int:object_id>/ml-file/",
                self.admin_site.admin_view(self.staff_ml_file_download),
                name=f"{info[0]}_{info[1]}_ml_file",
            ),
            path(
                "<int:object_id>/ml-preview/",
                self.admin_site.admin_view(self.staff_ml_preview),
                name=f"{info[0]}_{info[1]}_ml_preview",
            ),
            path(
                "<int:object_id>/ml-augment/",
                self.admin_site.admin_view(self.staff_ml_augment_serve),
                name=f"{info[0]}_{info[1]}_ml_augment",
            ),
            path(
                "<int:object_id>/ml-augment-gallery/",
                self.admin_site.admin_view(self.staff_ml_augment_gallery),
                name=f"{info[0]}_{info[1]}_ml_augment_gallery",
            ),
        ]
        return custom + urls

    def staff_ml_file_download(self, request, object_id):
        """Stream a whitelisted ML file from the staff workspace as a download.

        Args:
            request: HTTP request; query ``f`` is the basename.
            object_id: ``Staff`` primary key.

        Returns:
            ``FileResponse`` with ``Content-Disposition: attachment``, or 400/404.
        """
        fname = (request.GET.get("f") or "").strip()
        staff = get_object_or_404(Staff, pk=object_id)
        if fname not in allowed_ml_basenames(staff.pin):
            return HttpResponse(
                "Недопустимое имя файла".encode("utf-8"),
                status=400,
                content_type="text/plain; charset=utf-8",
            )
        ws = staff_workspace_dir(staff)
        if ws is None:
            raise Http404("Нет каталога сотрудника")
        fp = (ws / fname).resolve()
        try:
            fp.relative_to(ws.resolve())
        except ValueError:
            raise Http404()
        if not fp.is_file():
            raise Http404()
        return FileResponse(open(fp, "rb"), as_attachment=True, filename=fname)

    def staff_ml_preview(self, request, object_id):
        """Return an HTML page with a human-friendly overview of ``.npy`` or ``.pt`` files.

        Args:
            request: HTTP request; query ``f`` is the whitelisted basename.
            object_id: ``Staff`` primary key.

        Returns:
            ``HttpResponse`` ``text/html`` or 400/404.
        """
        fname = (request.GET.get("f") or "").strip()
        staff = get_object_or_404(Staff, pk=object_id)
        if fname not in allowed_ml_basenames(staff.pin):
            return HttpResponse(
                "Недопустимое имя файла".encode("utf-8"),
                status=400,
                content_type="text/plain; charset=utf-8",
            )
        ws = staff_workspace_dir(staff)
        if ws is None:
            raise Http404("Нет каталога сотрудника")
        fp = (ws / fname).resolve()
        try:
            fp.relative_to(ws.resolve())
        except ValueError:
            raise Http404()
        if not fp.is_file():
            raise Http404()

        pin_s = escape(staff.pin)
        file_s = escape(fname)
        back = reverse("admin:monitoring_app_staff_change", args=[staff.pk])
        dl_href = (
            reverse(
                "admin:monitoring_app_staff_ml_file",
                kwargs={"object_id": staff.pk},
            )
            + "?f="
            + quote(fname, safe="")
        )

        if fname.endswith(".npy"):
            try:
                body = build_npy_embeddings_preview_body(fp, fname)
            except Exception as exc:
                body = (
                    f"<p>Не удалось прочитать .npy: <code>{escape(str(exc))}</code></p>"
                )
        else:
            body = build_pt_checkpoint_preview_body(fp, fname, dl_href)

        html = (
            '<!DOCTYPE html><html><head><meta charset="utf-8"/><title>'
            f"Обзор файла {file_s}</title>"
            "<style>body{font-family:system-ui,sans-serif;margin:20px;max-width:960px;line-height:1.45;}"
            "a{color:#2563eb;}</style></head><body>"
            f'<p><a href="{escape(back)}">← к карточке сотрудника ({pin_s})</a></p>'
            '<h1 style="font-size:20px;font-weight:600;color:#0f172a;">Обзор ML-файла</h1>'
            f'<p style="color:#64748b;font-size:13px;margin:-6px 0 18px 0;">{file_s}</p>{body}</body></html>'
        )
        return HttpResponse(
            html.encode("utf-8"),
            content_type="text/html; charset=utf-8",
        )

    def staff_ml_augment_serve(self, request, object_id):
        """Serve one augment image inline from ``AUGMENT_ROOT`` (basename allowlist).

        Args:
            request: HTTP request; query ``f`` is the image basename.
            object_id: ``Staff`` primary key.

        Returns:
            ``FileResponse`` with ``Content-Disposition: inline``, or 400/404.
        """
        fname = (request.GET.get("f") or "").strip()
        staff = get_object_or_404(Staff, pk=object_id)
        if not allowed_augment_basename(staff.pin, fname):
            return HttpResponse(
                "Недопустимое имя файла".encode("utf-8"),
                status=400,
                content_type="text/plain; charset=utf-8",
            )
        aug_dir = augment_dir_for_pin(staff.pin)
        fp = (aug_dir / fname).resolve()
        try:
            fp.relative_to(aug_dir.resolve())
        except ValueError:
            raise Http404()
        if not fp.is_file():
            raise Http404()
        guessed, _ = mimetypes.guess_type(fname)
        content_type = guessed or "application/octet-stream"
        resp = FileResponse(open(fp, "rb"), as_attachment=False, filename=fname)
        resp["Content-Type"] = content_type
        resp["Content-Disposition"] = f'inline; filename="{fname}"'
        return resp

    def staff_ml_augment_gallery(self, request, object_id):
        """Render a standalone HTML page listing augment thumbnails for one staff.

        Args:
            request: HTTP request (unused; kept for admin_view signature).
            object_id: ``Staff`` primary key.

        Returns:
            ``HttpResponse`` ``text/html`` with embedded image URLs to
            ``staff_ml_augment_serve``.
        """
        staff = get_object_or_404(Staff, pk=object_id)
        pin = staff.pin
        names = list_augment_basenames(pin)
        base = reverse(
            "admin:monitoring_app_staff_ml_augment",
            kwargs={"object_id": staff.pk},
        )
        back = reverse("admin:monitoring_app_staff_change", args=[staff.pk])
        tiles: list[str] = []
        for name in names:
            q = quote(name, safe="")
            src = f"{base}?f={q}"
            tiles.append(
                '<div style="display:inline-block;margin:8px;text-align:center;'
                'vertical-align:top;">'
                f'<a href="{escape(src)}" target="_blank" rel="noopener">'
                f'<img src="{escape(src)}" alt="" loading="lazy" '
                'style="width:140px;height:140px;object-fit:cover;'
                'border:1px solid #e2e8f0;border-radius:6px;background:#fff;"/></a>'
                f'<div style="font-size:10px;max-width:140px;word-break:break-all;'
                f'margin-top:6px;color:#64748b;">{escape(name)}</div></div>'
            )
        body = (
            f'<p style="color:#64748b;">Файлов в каталоге аугментаций: <strong>{len(names)}</strong></p>'
            + "".join(tiles)
        )
        if not names:
            body = '<p style="color:#b91c1c;">Нет подходящих изображений в каталоге аугментаций.</p>'
        html = (
            '<!DOCTYPE html><html><head><meta charset="utf-8"/><title>'
            f"Аугментации {escape(pin)}</title>"
            "<style>body{font-family:system-ui,sans-serif;margin:20px;background:#f8fafc;}"
            "a{color:#2563eb;}</style></head><body>"
            f'<p><a href="{escape(back)}">← к карточке сотрудника ({escape(pin)})</a></p>'
            f'<h1 style="font-size:18px;">Галерея аугментаций ({escape(pin)})</h1>{body}</body></html>'
        )
        return HttpResponse(
            html.encode("utf-8"),
            content_type="text/html; charset=utf-8",
        )

    @admin.display(description="Отдел", ordering="department__name")
    def department(self, obj):
        if not obj.department_id:
            return "—"
        url = reverse(
            f"admin:{ChildDepartment._meta.app_label}_"
            f"{ChildDepartment._meta.model_name}_change",
            args=[obj.department_id],
        )
        return format_html('<a href="{}">{}</a>', url, obj.department.name)

    def display_positions(self, obj):
        positions = obj.positions.all()
        return ", ".join(position.name for position in positions[:5])

    def attendance_today(self, obj):
        today = timezone.now().date()
        skud_row_date = _staff_attendance_db_date_for_calendar_work_day(today)
        cache_key = f"attendance_today_{obj.pin}_{today}"
        cached_result = cache.get(cache_key)
        if cached_result is not None:
            return format_html(cached_result)

        attendance = None
        if hasattr(obj, "attendance_today_list") and obj.attendance_today_list:
            attendance = obj.attendance_today_list[0]
        else:
            attendance = (
                StaffAttendance.objects.filter(staff=obj, date_at=skud_row_date)
                .only("id", "first_in", "last_out")
                .first()
            )

        if not attendance:
            remote = False
            if hasattr(obj, "remote_work"):
                remote_works = list(obj.remote_work.all())
                remote = any(
                    rw.permanent_remote
                    or (
                        rw.start_date
                        and rw.end_date
                        and rw.start_date <= today <= rw.end_date
                    )
                    for rw in remote_works
                )
            else:
                remote = (
                    RemoteWork.objects.filter(staff=obj, permanent_remote=True).exists()
                    or RemoteWork.objects.filter(
                        staff=obj, start_date__lte=today, end_date__gte=today
                    ).exists()
                )

            if remote:
                result = '<span style="color: blue;">🏠 Удаленно</span>'
                cache.set(cache_key, result, 300)
                return format_html(result)

            absence = None
            if hasattr(obj, "absences"):
                absences = list(obj.absences.all())
                absence = next(
                    (
                        a
                        for a in absences
                        if a.start_date <= today <= a.end_date and a.approved
                    ),
                    None,
                )
            else:
                absence = AbsentReason.objects.filter(
                    staff=obj, start_date__lte=today, end_date__gte=today, approved=True
                ).first()

            if absence:
                result = f'<span style="color: orange;">⚠️ Отсутствует: {absence.get_reason_display()}</span>'
                cache.set(cache_key, result, 300)
                return format_html(result)

            result = '<span style="color: red;">❌ Отсутствует</span>'
            cache.set(cache_key, result, 300)
            return format_html(result)

        if attendance.first_in and not attendance.last_out:
            result = '<span style="color: green;">✓ Присутствует</span>'
            cache.set(cache_key, result, 300)
            return format_html(result)

        if attendance.first_in and attendance.last_out:
            result = '<span style="color: purple;">↩️ Ушел</span>'
            cache.set(cache_key, result, 300)
            return format_html(result)

        result = '<span style="color: gray;">? Неизвестно</span>'
        cache.set(cache_key, result, 300)
        return format_html(result)

    attendance_today.short_description = "Присутствие сегодня"

    def _attendance_history_cache_key(self, staff):
        return (
            f"attendance_history_{self.ATTENDANCE_HISTORY_CACHE_VERSION}_"
            f"{staff.pin}_{timezone.localdate().isoformat()}"
        )

    def _staff_attendance_event_date(self, record):
        for value in (record.first_in, record.last_out):
            local_value = _to_local_datetime(value)
            if local_value is not None:
                return local_value.date()
        return record.date_at - timedelta(days=1)

    def _is_remote_on_date(self, current_date, remote_periods):
        return any(
            (start is None and end is None)
            or (start is not None and end is not None and start <= current_date <= end)
            for start, end in remote_periods
        )

    def _render_attendance_history_line(
        self,
        label,
        value,
        *,
        color="#334155",
        emphasized=False,
    ):
        font_weight = "600" if emphasized else "500"
        return format_html(
            '<div style="margin-top:4px; font-size:12px; color:{};">'
            '<span style="font-weight:{};">{}:</span> {}</div>',
            color,
            font_weight,
            label,
            value,
        )

    def _render_attendance_history_card(
        self,
        *,
        current_date,
        record,
        lessons,
        is_remote,
        holiday,
    ):
        is_weekend = current_date.weekday() >= 5
        holiday_non_working = bool(holiday) and not holiday.is_working_day
        holiday_working = bool(holiday) and holiday.is_working_day
        badges = []
        if record:
            badges.append(_admin_badge("СКУД", background="#0f766e"))
        if lessons:
            badges.append(_admin_badge(f"Занятий {len(lessons)}", background="#7c3aed"))
        if holiday_working:
            badges.append(_admin_badge("Праздник, рабочий", background="#ca8a04"))
        elif holiday_non_working:
            badges.append(_admin_badge("Праздник, выходной", background="#9333ea"))
        elif is_weekend:
            badges.append(_admin_badge("Выходной", background="#64748b"))
        elif is_remote and not record:
            badges.append(_admin_badge("Удаленно", background="#2563eb"))

        lines = []
        if holiday:
            lines.append(
                format_html(
                    '<div style="font-size:13px; font-weight:600; color:{};">{}</div>',
                    "#a16207" if holiday_working else "#7e22ce",
                    holiday.name,
                )
            )
            if holiday_working:
                lines.append(
                    format_html(
                        '<div style="font-size:11px; color:#a16207; margin-top:2px;">'
                        "По календарю — рабочий день (учитывайте СКУД/занятия).</div>"
                    )
                )
            else:
                lines.append(
                    format_html(
                        '<div style="font-size:11px; color:#7e22ce; margin-top:2px;">'
                        "Нерабочий праздничный день.</div>"
                    )
                )
        elif is_remote and not record and not lessons:
            lines.append(
                format_html(
                    '<div style="font-size:13px; font-weight:600; color:#1d4ed8;">Удаленная работа</div>'
                )
            )

        if record:
            first_in = _format_local_time(record.first_in)
            last_out = _format_local_time(record.last_out)
            lines.append(
                self._render_attendance_history_line(
                    "Вход",
                    first_in or "Нет входа",
                    color="#047857" if first_in else "#b91c1c",
                    emphasized=True,
                )
            )
            lines.append(
                self._render_attendance_history_line(
                    "Выход",
                    last_out or "Нет выхода",
                    color="#1d4ed8" if last_out else "#64748b",
                    emphasized=True,
                )
            )
            if record.absence_reason:
                lines.append(
                    self._render_attendance_history_line(
                        "Причина",
                        record.absence_reason.get_reason_display(),
                        color="#b45309",
                    )
                )
            eff_dur = _format_staffattendance_effective_duration(record)
            if eff_dur:
                lines.append(
                    self._render_attendance_history_line(
                        "Эффективно в здании",
                        eff_dur,
                        color="#0f766e",
                        emphasized=True,
                    )
                )
            if record.date_at != current_date:
                lines.append(
                    self._render_attendance_history_line(
                        "Выгрузка",
                        record.date_at.strftime("%d.%m.%Y"),
                        color="#64748b",
                    )
                )

        if lessons:
            for lesson in lessons[:2]:
                lesson_name = lesson.subject_name
                if len(lesson_name) > 18:
                    lesson_name = f"{lesson_name[:18]}..."
                lines.append(
                    self._render_attendance_history_line(
                        "Пара",
                        f"{lesson_name} {_format_local_time(lesson.first_in) or ''}".strip(),
                        color="#6d28d9",
                    )
                )
            if len(lessons) > 2:
                lines.append(
                    format_html(
                        '<div style="margin-top:4px; font-size:12px; color:#7c3aed;">'
                        "...и еще {} занятия</div>",
                        len(lessons) - 2,
                    )
                )

        if not lines:
            lines.append(
                format_html(
                    '<div style="font-size:13px; color:#94a3b8;">Нет данных по посещаемости</div>'
                )
            )

        card_background = "#ffffff"
        if holiday_working:
            card_background = "#fffbeb"
        elif holiday_non_working or (is_weekend and not holiday_working):
            card_background = "#f8fafc"
        elif record:
            card_background = "#f0fdf4"
        elif lessons:
            card_background = "#faf5ff"
        elif is_remote:
            card_background = "#eff6ff"

        badges_html = (
            format_html_join(
                "",
                "{}",
                ((badge,) for badge in badges),
            )
            if badges
            else ""
        )
        lines_html = format_html_join("", "{}", ((line,) for line in lines))
        source_flags = []
        if record:
            source_flags.append("staff")
        if lessons:
            source_flags.append("lesson")
        if is_remote:
            source_flags.append("remote")
        if holiday:
            source_flags.append(
                "holiday_working" if holiday_working else "holiday_nonworking"
            )
        elif is_weekend:
            source_flags.append("weekend")

        return format_html(
            '<div data-attendance-day="{}" data-attendance-source="{}" '
            'style="border:1px solid #e2e8f0; border-radius:10px; padding:12px; background:{}; '
            'box-shadow:0 1px 2px rgba(15, 23, 42, 0.06); min-height:140px;">'
            '<div style="display:flex; justify-content:space-between; gap:8px; align-items:flex-start;">'
            "<div>"
            '<div style="font-weight:700; color:#0f172a;">{}</div>'
            '<div style="font-size:12px; color:#64748b;">{}</div>'
            "</div>"
            '<div style="display:flex; gap:6px; flex-wrap:wrap; justify-content:flex-end;">{}</div>'
            "</div>"
            '<div style="margin-top:10px;">{}</div>'
            "</div>",
            current_date.isoformat(),
            " ".join(source_flags) or "none",
            card_background,
            current_date.strftime("%d.%m.%Y"),
            date_format(current_date, "l"),
            badges_html,
            lines_html,
        )

    def attendance_history(self, obj):
        cache_key = self._attendance_history_cache_key(obj)
        cached_html = cache.get(cache_key)
        if cached_html:
            return format_html(cached_html)

        # Без «сегодня»: строка СКУД за текущую смену обычно попадает в БД после ночной выгрузки
        # (date_at на следующий календарный день), карточка «сегодня» выглядела бы пустой/вводящей в заблуждение.
        today = timezone.localdate()
        end_date = today - timedelta(days=1)
        start_date = end_date - timedelta(days=self.ATTENDANCE_HISTORY_DAYS - 1)

        attendance_records = (
            StaffAttendance.objects.filter(
                staff=obj,
                date_at__gte=_staff_attendance_db_date_for_calendar_work_day(
                    start_date
                ),
                date_at__lte=_staff_attendance_db_date_for_calendar_work_day(end_date),
            )
            .select_related("absence_reason")
            .only(
                "id",
                "date_at",
                "first_in",
                "last_out",
                "effective_work_seconds",
                "absence_reason_id",
                "absence_reason__reason",
            )
            .order_by("date_at")
        )
        records_dict = {}
        for record in attendance_records:
            event_date = self._staff_attendance_event_date(record)
            if not start_date <= event_date <= end_date:
                continue
            existing_record = records_dict.get(event_date)
            if existing_record is None or (
                existing_record.first_in is None and record.first_in is not None
            ):
                records_dict[event_date] = record

        lesson_records = (
            LessonAttendance.exclude_report_invalid_days(
                LessonAttendance.objects.filter(
                    staff=obj, date_at__range=(start_date, end_date)
                )
            )
            .only("id", "date_at", "first_in", "last_out", "staff_id", "subject_name")
            .order_by("date_at", "first_in")
        )
        lessons_dict = defaultdict(list)
        for lesson in lesson_records:
            lessons_dict[lesson.date_at].append(lesson)

        remote_works = (
            RemoteWork.objects.filter(staff=obj)
            .only("start_date", "end_date", "permanent_remote")
            .order_by("start_date")
        )
        remote_periods = []
        for rw in remote_works:
            if rw.permanent_remote:
                remote_periods.append((None, None))
            elif rw.start_date and rw.end_date:
                remote_periods.append((rw.start_date, rw.end_date))

        holidays = PublicHoliday.objects.filter(
            date__range=(start_date, end_date)
        ).only(
            "date",
            "name",
            "is_working_day",
        )
        holidays_dict = {hol.date: hol for hol in holidays}

        cards = []
        current_date = start_date
        while current_date <= end_date:
            record = records_dict.get(current_date)
            lessons = lessons_dict.get(current_date, [])
            is_remote = self._is_remote_on_date(current_date, remote_periods)
            holiday = holidays_dict.get(current_date)
            cards.append(
                self._render_attendance_history_card(
                    current_date=current_date,
                    record=record,
                    lessons=lessons,
                    is_remote=is_remote,
                    holiday=holiday,
                )
            )
            current_date += timedelta(days=1)

        legend_badges = format_html_join(
            "",
            "{}",
            (
                (_staff_attendance_history_legend_badge("СКУД", background="#0f766e"),),
                (
                    _staff_attendance_history_legend_badge(
                        "Удалённое занятие", background="#7c3aed"
                    ),
                ),
                (
                    _staff_attendance_history_legend_badge(
                        "Удалённо", background="#2563eb"
                    ),
                ),
                (
                    _staff_attendance_history_legend_badge(
                        "Праздник, выходной", background="#9333ea"
                    ),
                ),
                (
                    _staff_attendance_history_legend_badge(
                        "Праздник, рабочий", background="#ca8a04"
                    ),
                ),
                (
                    _staff_attendance_history_legend_badge(
                        "Выходной (сб/вс)", background="#64748b"
                    ),
                ),
            ),
        )
        legend_block = format_html(
            '<div class="staff-attendance-history__legend">'
            '<div class="staff-attendance-history__legend-title">Условные обозначения</div>'
            '<div class="staff-attendance-history__legend-badges">{}</div>'
            "</div>",
            legend_badges,
        )
        help_block = format_html(
            '<div class="staff-attendance-history__help help">'
            '<p class="staff-attendance-history__help-p">'
            "<strong>Период.</strong> Показаны <strong>{}</strong> полных календарных дней "
            "<strong>до вчера включительно</strong>. Сегодняшний день не включён: строка СКУД за текущую смену "
            "обычно появляется после ночной выгрузки."
            "</p>"
            '<p class="staff-attendance-history__help-p">'
            "<strong>Карточки</strong> привязаны к <strong>дню смены</strong>. В БД у СКУД поле "
            "<code>date_at</code> обычно на сутки позже календарного дня смены."
            "</p>"
            '<p class="staff-attendance-history__help-p">'
            "<strong>Эффективно в здании</strong> — сумма интервалов по турникетам; если данных нет — "
            "оценка по времени первого входа и последнего выхода."
            "</p>"
            '<p class="staff-attendance-history__help-p">'
            "<strong>Праздники</strong> берутся из справочника «Праздничные дни» "
            "(признак «рабочий день»)."
            "</p>"
            "</div>",
            self.ATTENDANCE_HISTORY_DAYS,
        )
        cards_html = format_html_join("", "{}", ((card,) for card in cards))
        html = format_html(
            '<div class="staff-attendance-history">'
            "{}"
            "{}"
            '<div class="staff-attendance-history__grid">{}</div>'
            "</div>",
            legend_block,
            help_block,
            cards_html,
        )
        cache.set(cache_key, str(html), self.ATTENDANCE_HISTORY_CACHE_TTL)
        return html

    attendance_history.short_description = (
        f"История: {ATTENDANCE_HISTORY_DAYS} дней до вчера (СКУД, праздники)"
    )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.select_related("department").prefetch_related("positions")
        return qs

    display_positions.short_description = "Должности"

    def clear_avatars(self, request, queryset):
        queryset.update(avatar=None)
        for staff in queryset:
            cache_key = f"avatar_thumbnail_{staff.pin}"
            cache.delete(cache_key)

    clear_avatars.short_description = "Очистить фото выбранных сотрудников"

    def mark_needs_training_true(self, request, queryset):
        queryset.update(needs_training=True)
        self.message_user(
            request,
            "Статус 'Требуется обучение модели' был изменён на 'True' для выбранных сотрудников.",
        )

    mark_needs_training_true.short_description = (
        "Установить 'Требуется обучение модели' для выбранных сотрудников"
    )

    def export_staff_data(self, request, queryset):
        self.message_user(
            request, f"Данные {queryset.count()} сотрудников экспортированы."
        )

    export_staff_data.short_description = "Экспортировать данные сотрудников"

    class Media:
        css = {"all": ("admin/css/custom_admin.css",)}
        js = ("admin/js/staff_admin.js",)


@admin.register(StaffFaceSample, site=admin_site)
class StaffFaceSampleAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "staff",
        "angle",
        "source",
        "with_glasses",
        "is_active",
        "quality_passed",
        "created_at",
    )
    list_filter = ("source", "angle", "is_active", "is_trusted", "with_glasses")
    search_fields = ("staff__pin", "staff__name", "staff__surname")
    autocomplete_fields = ("staff",)
    readonly_fields = ("created_at", "updated_at")
    ordering = ("-created_at",)
    save_on_top = True


@admin.register(StaffFaceMask, site=admin_site)
class StaffFaceMaskAdmin(admin.ModelAdmin):
    list_display = (
        "staff",
        "staff_department",
        "created_at",
        "updated_at",
        "staff_avatar",
        "augmentation_status",
    )
    search_fields = ("staff__name", "staff__surname", "staff__pin")
    readonly_fields = (
        "created_at",
        "updated_at",
        "mask_encoding",
        "staff",
        "staff_avatar",
        "staff_face_ml_bridge",
        "augmented_images",
    )
    list_filter = (
        "created_at",
        "updated_at",
        monitoring_utils.HierarchicalDepartmentFilter,
    )
    ordering = (
        "staff__department",
        "-updated_at",
    )
    actions = ["regenerate_masks", "force_augmentation"]

    def staff_avatar(self, obj):
        if obj.staff.avatar:
            return format_html(
                '<img src="{}" width="100" height="100" style="border-radius: 50%;" />',
                obj.staff.avatar.url,
            )
        return "No Avatar"

    staff_avatar.short_description = "Аватар сотрудника"

    def augmentation_status(self, obj):
        count, exists = count_augment_images(obj.staff.pin)
        if not exists:
            return format_html(
                '<span style="color: red;">❌ Нет каталога аугментаций</span>'
            )
        if count == 0:
            return format_html(
                '<span style="color: red;">Нет файлов <code>{}_aug_*</code></span>',
                obj.staff.pin,
            )
        if count < 10:
            return format_html(
                '<span style="color: orange;">Мало аугментаций ({}/10 рекомендуется)</span>',
                count,
            )
        return format_html(
            '<span style="color: green;">{} аугментированных кадров</span>',
            count,
        )

    augmentation_status.short_description = "Статус аугментации"

    def augmented_images(self, obj):
        pin = obj.staff.pin
        cache_key = f"augmented_images_v2_{pin}"
        images_html = cache.get(cache_key)

        if images_html is None:
            aug_dir = augment_dir_for_pin(pin)
            if not aug_dir.is_dir():
                return "Нет каталога аугментаций"

            names: list[str] = []
            for name in os.listdir(aug_dir):
                low = name.lower()
                if not low.endswith((".jpg", ".jpeg", ".png", ".webp")):
                    continue
                if name.startswith(f"{pin}_aug_") or name.startswith(
                    f"{pin}_augmented_"
                ):
                    names.append(name)
            names.sort()

            snippet_parts: list[str] = []
            base_url = f"{str(settings.AUGMENT_URL).rstrip('/')}/{pin}/"
            for name in names[:48]:
                snippet = format_html(
                    '<div style="border: 1px solid #ddd; padding: 3px; border-radius: 3px;">'
                    '<img src="{}" width="80" height="80" style="object-fit: cover;" alt="" />'
                    "</div>",
                    f"{base_url}{name}",
                )
                snippet_parts.append(str(snippet))

            if not snippet_parts:
                return "Нет аугментированных изображений"

            more = ""
            if len(names) > 48:
                more = f'<p style="margin:8px 0 0 0;font-size:12px;color:#64748b;">Ещё {len(names) - 48} файлов…</p>'

            images_html = (
                '<div style="display: flex; flex-wrap: wrap; gap: 5px;">'
                + "".join(snippet_parts)
                + "</div>"
                + more
            )
            cache.set(cache_key, images_html, timeout=3600)

        return format_html(images_html)

    augmented_images.short_description = "Аугментированные фото"

    @admin.display(description="Эмбеддинги / .pt / аугментации (сотрудник)")
    def staff_face_ml_bridge(self, obj):
        """Reuse ``StaffAdmin`` ML table links on the ``StaffFaceMask`` change form.

        Args:
            obj: ``StaffFaceMask`` instance.

        Returns:
            ``SafeString`` HTML from ``render_staff_face_ml_table`` or a dash help line.
        """
        s = obj.staff
        if not s.pk:
            return mark_safe('<p class="help">—</p>')
        dl = reverse(
            "admin:monitoring_app_staff_ml_file",
            kwargs={"object_id": s.pk},
        )
        pv = reverse(
            "admin:monitoring_app_staff_ml_preview",
            kwargs={"object_id": s.pk},
        )
        gal = reverse(
            "admin:monitoring_app_staff_ml_augment_gallery",
            kwargs={"object_id": s.pk},
        )
        return render_staff_face_ml_table(
            s,
            file_download_url=dl,
            file_preview_url=pv,
            augment_gallery_url=gal,
        )

    def staff_department(self, obj):
        return obj.staff.department

    staff_department.short_description = "Отдел"
    staff_department.admin_order_field = "staff__department"

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related("staff", "staff__department")
            .prefetch_related("staff__positions")
        )

    fieldsets = (
        (
            "Информация о сотруднике",
            {
                "fields": (
                    "staff",
                    "staff_avatar",
                ),
                "classes": ("wide",),
            },
        ),
        (
            "Face-ML артефакты на диске",
            {
                "fields": ("staff_face_ml_bridge",),
                "description": "Те же файлы, что в карточке сотрудника: .npy, .pt, счёт аугментаций.",
                "classes": ("wide", "grp-collapse grp-open"),
            },
        ),
        (
            "Аугментация изображений",
            {
                "fields": ("augmented_images",),
                "classes": ("wide",),
            },
        ),
        (
            "Временные метки",
            {
                "fields": ("created_at", "updated_at"),
                "classes": ("grp-collapse grp-closed",),
            },
        ),
        (
            "Технические данные",
            {
                "fields": ("mask_encoding",),
                "classes": ("grp-collapse grp-closed",),
            },
        ),
    )

    def regenerate_masks(self, request, queryset):
        count = queryset.count()
        self.message_user(
            request, f"Запущена регенерация масок для {count} сотрудников."
        )

    regenerate_masks.short_description = (
        "Регенерировать маски для выбранных сотрудников"
    )

    def force_augmentation(self, request, queryset):
        count = queryset.count()
        self.message_user(request, f"Запущена аугментация для {count} сотрудников.")

    force_augmentation.short_description = (
        "Запустить аугментацию для выбранных сотрудников"
    )

    def has_add_permission(self, request):
        """Маска одна на сотрудника; создаётся пайплайном Face-ML, не вручную."""
        return False


# ===== ATTENDANCE MODELS =====


class CachedCountQuerySet:
    """
    Обёртка QuerySet с кэшированным count() (ключ по хэшу запроса).
    Paginator вызывает object_list.count() — используем стандартный Paginator.
    """

    def __init__(self, queryset):
        self._queryset = queryset
        self._count = None

    def count(self):
        if self._count is not None:
            return self._count
        import hashlib

        try:
            q = self._queryset.query
            qstr = str(q.where) + str(q.order_by)
            h = hashlib.md5(qstr.encode()).hexdigest()[:16]
            cache_key = f"staffatt_count_{h}"
            self._count = cache.get(cache_key)
            if self._count is None:
                self._count = self._queryset.count()
                cache.set(cache_key, self._count, 300)
        except (ValueError, TypeError):
            self._count = 0
        except (DatabaseError, OperationalError) as e:
            logger.warning(
                "CachedCountQuerySet: database error on count: %s",
                e,
                exc_info=True,
            )
            try:
                self._count = self._queryset.count()
            except Exception:
                self._count = 0
        except Exception as e:
            logger.warning(
                "CachedCountQuerySet: unexpected error on count: %s",
                e,
                exc_info=True,
            )
            try:
                self._count = self._queryset.count()
            except Exception:
                self._count = 0
        return self._count

    def __getitem__(self, key):
        return self._queryset[key]

    def __iter__(self):
        return iter(self._queryset)

    def __len__(self) -> int:
        return len(self._queryset)

    def __getattr__(self, name):
        return getattr(self._queryset, name)


@admin.register(StaffAttendance, site=admin_site)
class StaffAttendanceAdmin(admin.ModelAdmin):
    change_list_template = "admin/monitoring_app/staffattendance/change_list.html"
    date_hierarchy = "date_at"
    list_display = (
        "staff_display",
        "staff_department",
        "event_calendar_day",
        "date_at",
        "formatted_first_in",
        "formatted_last_out",
        "duration",
        "area_name_in",
        "area_name_out",
        "absence_reason_display",
    )
    list_filter = (
        StaffAttendanceRowZoneFilter,
        StaffAttendanceSkudDataFilter,
        DateRangeFilter,
        monitoring_utils.HierarchicalDepartmentFilter,
        "absence_reason",
    )
    search_fields = (
        "staff__pin",
        "staff__name",
        "staff__surname",
        "staff__department__name",
    )
    ordering = ("-date_at", "staff__department_id", "staff_id")
    list_per_page = 50
    show_full_result_count = False
    sortable_by = (
        "event_calendar_day",
        "date_at",
        "staff_department",
        "staff_display",
        "duration",
    )

    def get_paginator(
        self, request, queryset, per_page, orphans=0, allow_empty_first_page=True
    ):
        return Paginator(
            CachedCountQuerySet(queryset),
            per_page,
            orphans=orphans,
            allow_empty_first_page=allow_empty_first_page,
        )

    autocomplete_fields = ("absence_reason",)
    actions = ["export_attendance_data", "mark_as_absent"]
    actions_on_top = False
    actions_on_bottom = True
    list_max_show_all = 400
    save_on_top = True

    readonly_fields = (
        "staff",
        "date_at",
        "first_in",
        "last_out",
        "duration",
        "formatted_effective_work_seconds",
        "area_name_out",
        "area_name_in",
        "formatted_area_sequence",
        "formatted_effective_work_intervals",
    )

    fieldsets = (
        (
            "Основная информация",
            {
                "fields": (
                    "staff",
                    "date_at",
                    "first_in",
                    "last_out",
                    "duration",
                    "formatted_effective_work_seconds",
                ),
                "classes": ("wide",),
            },
        ),
        (
            "Местоположение",
            {
                "fields": (
                    "area_name_in",
                    "area_name_out",
                    "formatted_area_sequence",
                    "formatted_effective_work_intervals",
                ),
                "classes": ("wide",),
            },
        ),
        (
            "Статус отсутствия",
            {
                "fields": ("absence_reason",),
                "classes": ("wide",),
            },
        ),
    )

    def formatted_first_in(self, obj):
        if obj.first_in:
            local_time = timezone.localtime(obj.first_in)
            return local_time.strftime("%H:%M:%S")
        return "-"

    formatted_first_in.short_description = "Вход"
    formatted_first_in.admin_order_field = "first_in"

    def formatted_last_out(self, obj):
        if obj.last_out:
            local_time = timezone.localtime(obj.last_out)
            return local_time.strftime("%H:%M:%S")
        return "-"

    formatted_last_out.short_description = "Выход"
    formatted_last_out.admin_order_field = "last_out"

    def duration(self, obj):
        total_seconds = None
        if obj.effective_work_seconds is not None:
            total_seconds = obj.effective_work_seconds
        elif obj.first_in and obj.last_out:
            total_seconds = int((obj.last_out - obj.first_in).total_seconds())
        if total_seconds is not None:
            if total_seconds < 0:
                return format_html('<span style="color: red;">Ошибка времени</span>')
            if total_seconds > 86400:
                total_seconds = 86400
            hours, remainder = divmod(total_seconds, 3600)
            minutes = remainder // 60
            time_str = f"{hours:02d}:{minutes:02d}"
            if hours < 1:
                return format_html('<span style="color: red;">{}</span>', time_str)
            if hours < 2:
                return format_html('<span style="color: orange;">{}</span>', time_str)
            return format_html('<span style="color: green;">{}</span>', time_str)
        return "-"

    duration.short_description = "Продолжительность (эффективная)"
    duration.admin_order_field = "effective_work_seconds"

    @admin.display(description="Сотрудник", ordering="staff_id")
    def staff_display(self, obj):
        url = reverse(
            f"admin:{Staff._meta.app_label}_{Staff._meta.model_name}_change",
            args=[obj.staff_id],
        )
        return format_html('<a href="{}">{}</a>', url, obj.staff)

    @admin.display(
        description="Уважит. причина",
        ordering="absence_reason__reason",
    )
    def absence_reason_display(self, obj):
        if not obj.absence_reason_id:
            return "—"
        url = reverse(
            f"admin:{AbsentReason._meta.app_label}_"
            f"{AbsentReason._meta.model_name}_change",
            args=[obj.absence_reason_id],
        )
        return format_html('<a href="{}">{}</a>', url, obj.absence_reason)

    def formatted_effective_work_seconds(self, obj):
        """Отображает effective_work_seconds в виде «N сек (X ч Y мин)»."""
        if obj is None or getattr(obj, "effective_work_seconds", None) is None:
            return format_html("<span style='color: #999;'>—</span>")
        sec = obj.effective_work_seconds
        hours, remainder = divmod(sec, 3600)
        minutes = remainder // 60
        return format_html(
            "{} сек <span style='color: #666;'>({} ч {} мин)</span>",
            sec,
            hours,
            minutes,
        )

    formatted_effective_work_seconds.short_description = (
        "Эффективное время в здании (сек)"
    )

    def formatted_area_sequence(self, obj):
        """Рендерит цепочку зон: №, Время, Зона, Устройство (devSn). Выход подсвечивается."""
        if obj is None:
            return format_html("<span style='color: #999;'>—</span>")
        seq = getattr(obj, "area_sequence", None)
        if not seq or not isinstance(seq, list):
            return format_html("<span style='color: #999;'>Нет данных</span>")
        rows = []
        for i, item in enumerate(seq[:50], 1):
            if not isinstance(item, dict):
                continue
            t = item.get("t") or ""
            area = item.get("area") or ""
            dev_sn = item.get("devSn") or "—"
            is_exit = item.get("is_exit") == "1"
            is_bridge_transfer = item.get("exit_resolution") == "bridge_transfer"
            row_bg = ""
            if is_exit:
                row_bg = "background: #e0f2fe;"
            elif is_bridge_transfer:
                row_bg = "background: #fef9c3;"
            elif i % 2 == 0:
                row_bg = "background: #f8fafc;"
            dev_cell = format_html(
                "<span style='font-family: monospace; font-size: 12px; color: #475569;'>{}</span>",
                dev_sn,
            )
            if is_exit:
                dev_cell = format_html(
                    "<span style='font-family: monospace; font-size: 12px; color: #0284c7;' title=\"Турникет выхода\">{} ✓</span>",
                    dev_sn,
                )
            elif is_bridge_transfer:
                dev_cell = format_html(
                    "<span style='font-family: monospace; font-size: 12px; color: #a16207;' title=\"Переход в пристройку\">{} ↔</span>",
                    dev_sn,
                )
            rows.append(
                format_html(
                    "<tr style='{}'>"
                    "<td style='padding: 8px 14px; text-align: right; min-width: 2.5em; font-weight: 500;'>{}</td>"
                    "<td style='padding: 8px 14px; min-width: 4em; font-variant-numeric: tabular-nums; font-size: 14px;'>{}</td>"
                    "<td style='padding: 8px 14px; min-width: 12em;'>{}</td>"
                    "<td style='padding: 8px 14px; min-width: 14em;'>{}</td></tr>",
                    row_bg,
                    i,
                    t,
                    area,
                    dev_cell,
                )
            )
        head = format_html(
            "<thead><tr style='background: #334155; color: #f1f5f9;'>"
            "<th style='text-align: right; padding: 10px 14px; font-weight: 600; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>№</th>"
            "<th style='text-align: left; padding: 10px 14px; font-weight: 600; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Время</th>"
            "<th style='text-align: left; padding: 10px 14px; font-weight: 600; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Зона</th>"
            "<th style='text-align: left; padding: 10px 14px; font-weight: 600; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Устройство / Статус</th></tr></thead>"
        )
        tail = (
            format_html(
                "<tr style='background: #f1f5f9;'><td colspan='4' style='padding: 8px 14px; color: #64748b; font-size: 12px;'>"
                "… и ещё {} пунктов</td></tr>",
                len(seq) - 50,
            )
            if len(seq) > 50
            else ""
        )
        return format_html(
            "<div style='max-height: 400px; overflow: auto; border: 1px solid #cbd5e1; border-radius: 8px; "
            "box-shadow: 0 1px 3px rgba(0,0,0,0.08); min-width: 420px;'>"
            "<table style='border-collapse: collapse; font-size: 13px; width: 100%; min-width: 400px;'>{}{}{}</table></div>",
            head,
            format_html("".join(rows)),
            tail,
        )

    formatted_area_sequence.short_description = "Цепочка зон (карта перемещений)"

    def formatted_effective_work_intervals(self, obj):
        """Рендерит интервалы «в здании» (effective_work_intervals) в читаемом виде."""

        if obj is None:
            return format_html("<span style='color: #999;'>—</span>")
        intervals = getattr(obj, "effective_work_intervals", None)
        if not intervals or not isinstance(intervals, list):
            return format_html("<span style='color: #999;'>Нет данных</span>")
        lines = []
        for item in intervals[:20]:
            if not isinstance(item, dict):
                continue
            start_s = item.get("start") or ""
            end_s = item.get("end") or ""
            try:
                start_dt = datetime.fromisoformat(start_s.replace("Z", "+00:00"))
                end_dt = datetime.fromisoformat(end_s.replace("Z", "+00:00"))
                start_local = timezone.localtime(start_dt)
                end_local = timezone.localtime(end_dt)
                start_str = start_local.strftime("%H:%M")
                end_str = end_local.strftime("%H:%M")
                delta_sec = int((end_dt - start_dt).total_seconds())
                mins = delta_sec // 60
                lines.append(
                    format_html(
                        "<div style='padding: 2px 0;'>{} — {} ({} мин)</div>",
                        start_str,
                        end_str,
                        mins,
                    )
                )
            except (ValueError, TypeError, AttributeError):
                lines.append(
                    format_html(
                        "<div style='padding: 2px 0; color: #999;'>{} — {}</div>",
                        start_s[:19] if start_s else "?",
                        end_s[:19] if end_s else "?",
                    )
                )
        if len(intervals) > 20:
            lines.append(
                format_html(
                    "<div style='color: #666; padding-top: 4px;'>… и ещё {} интервалов</div>",
                    len(intervals) - 20,
                )
            )
        return format_html(
            "<div style='max-height: 220px; overflow-y: auto; font-size: 13px;'>{}</div>",
            format_html("".join(lines)),
        )

    formatted_effective_work_intervals.short_description = (
        "Интервалы «в здании» (для объединения с LA)"
    )

    def staff_info(self, obj):
        if not obj.staff:
            return "Нет данных о сотруднике"

        cache_key = f"staff_info_{obj.staff.pin}_{obj.id}"
        cached_html = cache.get(cache_key)
        if cached_html:
            return format_html(cached_html)

        positions = list(obj.staff.positions.all()[:3])
        positions_str = (
            ", ".join(p.name for p in positions) if positions else "Не указаны"
        )

        avatar_html = self.staff_avatar(obj)

        html = f"""
        <div style="display: flex; align-items: center; gap: 20px;">
            <div style="flex-shrink: 0;">
                {avatar_html}
            </div>
            <div>
                <h3 style="margin: 0;">{obj.staff.surname} {obj.staff.name}</h3>
                <p style="margin: 5px 0;">PIN: {obj.staff.pin}</p>
                <p style="margin: 5px 0;">Отдел: {obj.staff.department.name if obj.staff.department else "Не указан"}</p>
                <p style="margin: 5px 0;">Должности: {positions_str}</p>
            </div>
        </div>
        """
        cache.set(cache_key, html, 3600)
        return format_html(html)

    staff_info.short_description = "Информация о сотруднике"

    def staff_avatar(self, obj):
        if not obj.staff:
            return format_html(
                '<div style="width: 100px; height: 100px; border-radius: 50%; background-color: #f0f0f0; display: flex; justify-content: center; align-items: center;">'
                '<span style="color: #999;">Нет фото</span>'
                "</div>"
            )

        cache_key = f"staff_avatar_url_{obj.staff.pin}"
        avatar_url = cache.get(cache_key)

        if avatar_url is None:
            if obj.staff.avatar:
                avatar_url = obj.staff.avatar.url
            else:
                avatar_url = ""
            cache.set(cache_key, avatar_url, 86400)

        if avatar_url:
            return format_html(
                '<img src="{}" width="100" height="100" style="border-radius: 50%; object-fit: cover;" />',
                avatar_url,
            )
        return format_html(
            '<div style="width: 100px; height: 100px; border-radius: 50%; background-color: #f0f0f0; display: flex; justify-content: center; align-items: center;">'
            '<span style="color: #999;">Нет фото</span>'
            "</div>"
        )

    def staff_department(self, obj):
        return obj.staff.department.name if obj.staff.department else "N/A"

    staff_department.short_description = "Отдел"
    staff_department.admin_order_field = "staff__department_id"

    def event_calendar_day(self, obj):
        if obj.date_at:
            return (obj.date_at - timedelta(days=1)).strftime("%d.%m.%Y")
        return "—"

    event_calendar_day.short_description = "День смены"
    event_calendar_day.admin_order_field = "date_at"

    def get_form(self, request, obj=None, change=False, **kwargs):
        form = super().get_form(request, obj, change=change, **kwargs)
        date_field = form.base_fields.get("date_at")
        if date_field is not None:
            date_field.help_text = (
                "Дата строки выгрузки в БД (как правило на календарный день позже дня смены). "
                "Календарный день явки — колонка «День смены» или дата локального first_in."
            )
        return form

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.select_related(
            "staff",
            "staff__department",
            "absence_reason",
            "absence_reason__staff",
        )
        qs = qs.defer("staff__avatar")
        if getattr(request, "staffattendance_changelist", False):
            if "date_range" not in request.GET:
                today = timezone.now().date()
                start_wd = today - timedelta(days=13)
                qs = qs.filter(
                    date_at__gte=_staff_attendance_db_date_for_calendar_work_day(
                        start_wd
                    ),
                    date_at__lte=_staff_attendance_db_date_for_calendar_work_day(today),
                )
            qs = qs.only(
                "id",
                "staff_id",
                "date_at",
                "first_in",
                "last_out",
                "effective_work_seconds",
                "area_name_in",
                "area_name_out",
                "absence_reason_id",
                "staff__id",
                "staff__pin",
                "staff__name",
                "staff__surname",
                "staff__department_id",
                "staff__department__id",
                "staff__department__name",
                "absence_reason__id",
                "absence_reason__reason",
                "absence_reason__start_date",
                "absence_reason__end_date",
                "absence_reason__staff_id",
                "absence_reason__staff__id",
                "absence_reason__staff__name",
                "absence_reason__staff__surname",
            )

        return qs

    def area_name_in(self, obj):
        value = obj.area_name_in
        if not value or value == "Unknown":
            return format_html('<span style="color: #999;">N/A</span>')
        return value

    def area_name_out(self, obj):
        value = obj.area_name_out
        if not value or value == "Unknown":
            return format_html('<span style="color: #999;">N/A</span>')
        return value

    def changelist_view(self, request, extra_context=None):
        request.staffattendance_changelist = True
        try:
            return super().changelist_view(request, extra_context=extra_context)
        finally:
            request.staffattendance_changelist = False

    def export_attendance_data(self, request, queryset):
        count = queryset.count()
        self.message_user(
            request, f"Экспортированы данные о посещаемости для {count} записей."
        )

    export_attendance_data.short_description = "Экспортировать данные о посещаемости"

    def mark_as_absent(self, request, queryset):
        count = queryset.count()
        self.message_user(request, f"{count} записей отмечены как отсутствие.")

    mark_as_absent.short_description = "Отметить как отсутствие"

    class Media:
        css = {"all": ("admin/css/custom_admin.css",)}
        js = ("admin/js/attendance_admin.js",)


class PhotoEffectiveStatusFilter(SimpleListFilter):
    title = "Итог фото"
    parameter_name = "photo_effective_status"

    def lookups(self, request, model_admin):
        _ = request
        _ = model_admin
        return (
            (LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN, "Нормальное"),
            ("insufficient_input_review", "Недостаточно данных"),
            (LessonAttendance.PHOTO_SPOOF_STATUS_REVIEW, "На проверку"),
            (LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS, "Подозрительное"),
            (LessonAttendance.PHOTO_SPOOF_STATUS_PENDING, "Ожидает проверки"),
            (LessonAttendance.PHOTO_SPOOF_STATUS_ERROR, "Ошибка проверки"),
        )

    def queryset(self, request, queryset):
        _ = request
        value = self.value()
        if value is None:
            return queryset

        if value == LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN:
            return queryset.filter(
                Q(photo_manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_CLEAN)
                | (
                    Q(photo_manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE)
                    & Q(photo_spoof_status=LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN)
                )
            )

        if value == LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS:
            return queryset.filter(
                Q(photo_manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS)
                | (
                    Q(photo_manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE)
                    & Q(
                        photo_spoof_status=LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS
                    )
                )
            )

        if value == "insufficient_input_review":
            return queryset.filter(
                photo_manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
                photo_spoof_status=LessonAttendance.PHOTO_SPOOF_STATUS_REVIEW,
                photo_spoof_tags__contains=[
                    "pad_rule:presentation_insufficient_input_review"
                ],
            )

        if value == LessonAttendance.PHOTO_SPOOF_STATUS_REVIEW:
            return queryset.filter(
                photo_manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
                photo_spoof_status=LessonAttendance.PHOTO_SPOOF_STATUS_REVIEW,
            ).exclude(
                photo_spoof_tags__contains=[
                    "pad_rule:presentation_insufficient_input_review"
                ]
            )

        return queryset.filter(
            photo_manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
            photo_spoof_status=value,
        )


class LessonAttendanceAdmin(ModelAdmin):
    change_list_template = "admin/lessonattendance_change_list.html"
    geomap_field_longitude = "id_longitude"
    geomap_field_latitude = "id_latitude"
    geomap_show_map_on_list = False
    geomap_item_zoom = "14"
    geomap_height = "450px"
    geomap_default_zoom = "16"
    geomap_autozoom = "15.9"

    readonly_fields = (
        "latitude",
        "longitude",
        "first_in",
        "last_out",
        "staff",
        "subject_name",
        "tutor",
        "tutor_id",
        "date_at",
        "photo_preview",
        "location_map",
        "lesson_duration",
        "formatted_duration_seconds",
        "photo_effective_status_badge",
        "photo_pad_scan_status_line",
        "photo_pad_operator_readout",
        "photo_spoof_checked_at",
        "photo_spoof_status_badge",
        "photo_trust_line",
        "photo_spoof_model_version_display",
        "photo_pad_technical_metrics",
        "photo_manual_verdict_badge",
        "photo_manual_by",
        "photo_manual_at",
    )
    date_hierarchy = "date_at"
    actions = [
        "export_lesson_data",
        "cleanup_old_photos",
        "mark_photo_manual_clean",
        "mark_photo_manual_suspicious",
        "reset_photo_manual_verdict",
        "rescan_selected_photos",
    ]
    show_full_result_count = False
    list_select_related = ("staff", "staff__department", "photo_manual_by")

    fieldsets = (
        (
            "Основная информация",
            {
                "fields": (
                    (
                        "first_in",
                        "last_out",
                        "lesson_duration",
                        "formatted_duration_seconds",
                    ),
                    "photo_preview",
                    "date_at",
                ),
                "classes": ("wide",),
            },
        ),
        (
            "Антифрод и проверка фото",
            {
                "description": (
                    "Сверху: итог, причина и ключевые сигналы. «Недостаточно данных» означает слабый кадр, а не подозрение."
                ),
                "fields": (
                    "photo_effective_status_badge",
                    "photo_pad_scan_status_line",
                    "photo_pad_operator_readout",
                ),
                "classes": ("wide",),
            },
        ),
        (
            "Ручное решение",
            {
                "fields": (
                    ("photo_manual_verdict", "photo_manual_verdict_badge"),
                    ("photo_manual_by", "photo_manual_at"),
                ),
                "classes": ("wide",),
            },
        ),
        (
            "Авто-проверка фото: время, версия и числа",
            {
                "description": (
                    "Сворачиваемый блок со справочными числами последнего авто-анализа."
                ),
                "fields": (
                    ("photo_spoof_status_badge", "photo_trust_line"),
                    "photo_pad_technical_metrics",
                ),
                "classes": ("wide", "grp-collapse", "grp-closed"),
            },
        ),
        (
            "Комментарий ручной проверки",
            {
                "fields": ("photo_manual_comment",),
                "classes": ("wide", "grp-collapse", "grp-closed"),
            },
        ),
        (
            "Местоположение",
            {
                "fields": (
                    ("latitude", "longitude"),
                    "location_map",
                ),
                "classes": ("wide",),
            },
        ),
        (
            "Преподаватель",
            {
                "fields": (
                    "staff",
                    "subject_name",
                    "tutor_id",
                    "tutor",
                ),
                "classes": ("grp-collapse grp-closed",),
            },
        ),
    )

    list_display = (
        "staff",
        "tutor",
        "subject_name",
        "formatted_first_in",
        "formatted_last_out",
        "lesson_duration",
        "date_at",
        "has_photo",
        "photo_effective_status_badge",
        "photo_antifraud_list_compact",
    )
    list_per_page = 50
    list_display_links = ("staff", "photo_effective_status_badge")

    list_filter = (
        PhotoEffectiveStatusFilter,
        DateRangeFilter,
        monitoring_utils.HierarchicalDepartmentFilter,
        "photo_manual_verdict",
    )

    search_fields = (
        "staff__pin",
        "staff__name",
        "staff__surname",
        "subject_name",
        "tutor",
        "tutor_id",
    )

    ordering = ("-date_at", "-first_in")

    @staticmethod
    def _is_deletable_attendance_photo_path(file_path: str) -> bool:
        if not file_path or file_path == "/static/media/images/no-avatar.png":
            return False
        normalized = os.path.abspath(str(file_path))
        attendance_root = os.path.abspath(str(settings.ATTENDANCE_ROOT))
        media_control_root = os.path.abspath(
            os.path.join(str(settings.MEDIA_ROOT), "control_image")
        )
        allowed_roots = (attendance_root, media_control_root)
        return any(
            normalized == root or normalized.startswith(f"{root}{os.sep}")
            for root in allowed_roots
        )

    @classmethod
    def _collect_candidate_photo_paths(cls, queryset):
        raw_paths = (
            queryset.exclude(staff_image_path__isnull=True)
            .exclude(staff_image_path="")
            .values_list("staff_image_path", flat=True)
        )
        candidate_paths = set()
        for stored_file_path in raw_paths:
            if cls._is_deletable_attendance_photo_path(stored_file_path):
                candidate_paths.add(os.path.abspath(str(stored_file_path)))
        return candidate_paths

    @staticmethod
    def _delete_orphaned_photo_paths(candidate_paths, deleted_ids):
        if not candidate_paths:
            return 0
        referenced_elsewhere = set(
            LessonAttendance.objects.filter(staff_image_path__in=candidate_paths)
            .exclude(id__in=deleted_ids)
            .values_list("staff_image_path", flat=True)
        )
        deleted_files = 0
        for file_path in candidate_paths:
            if file_path in referenced_elsewhere:
                continue
            if not os.path.isfile(file_path):
                continue
            try:
                os.remove(file_path)
                deleted_files += 1
            except OSError as exc:
                logger.warning(
                    "Failed to delete lesson attendance photo file path=%s error=%s",
                    file_path,
                    exc,
                )
        return deleted_files

    def delete_model(self, request, obj):
        deleted_ids = [obj.id] if obj.id is not None else []
        candidate_paths = set()
        if obj.staff_image_path and self._is_deletable_attendance_photo_path(
            obj.staff_image_path
        ):
            candidate_paths.add(os.path.abspath(str(obj.staff_image_path)))
        super().delete_model(request, obj)
        deleted_files = self._delete_orphaned_photo_paths(candidate_paths, deleted_ids)
        if deleted_files:
            self.message_user(
                request, f"Удалено файлов фотографий с диска: {deleted_files}."
            )

    def delete_queryset(self, request, queryset):
        deleted_ids = list(queryset.values_list("id", flat=True))
        candidate_paths = self._collect_candidate_photo_paths(queryset)
        super().delete_queryset(request, queryset)
        deleted_files = self._delete_orphaned_photo_paths(candidate_paths, deleted_ids)
        if deleted_files:
            self.message_user(
                request, f"Удалено файлов фотографий с диска: {deleted_files}."
            )

    def formatted_duration_seconds(self, obj):
        """Показывает duration_seconds в виде «N сек (X ч Y мин)»."""
        if obj is None:
            return format_html("<span style='color: #999;'>—</span>")
        sec = getattr(obj, "duration_seconds", None)
        if sec is None:
            return format_html("<span style='color: #999;'>—</span>")
        hours, remainder = divmod(sec, 3600)
        minutes = remainder // 60
        return format_html(
            "{} сек <span style='color: #666;'>({} ч {} мин)</span>",
            sec,
            hours,
            minutes,
        )

    formatted_duration_seconds.short_description = "Длительность занятия (сек)"

    def lesson_duration(self, obj):
        total_seconds = None
        if getattr(obj, "duration_seconds", None) is not None:
            total_seconds = obj.duration_seconds
        elif obj.first_in and obj.last_out:
            total_seconds = int((obj.last_out - obj.first_in).total_seconds())
        if total_seconds is not None:

            if total_seconds < 0:
                return format_html('<span style="color: red;">Ошибка времени</span>')

            if total_seconds > 86400:
                total_seconds = 86400

            hours, remainder = divmod(total_seconds, 3600)
            minutes = remainder // 60

            time_str = f"{hours:02d}:{minutes:02d}"

            if hours < 1:
                return format_html('<span style="color: red;">{}</span>', time_str)
            elif hours < 2:
                return format_html('<span style="color: orange;">{}</span>', time_str)
            else:
                return format_html('<span style="color: green;">{}</span>', time_str)
        return "-"

    lesson_duration.short_description = "Продолжительность"

    def formatted_first_in(self, obj):
        if obj.first_in:
            local_time = timezone.localtime(obj.first_in)
            return format_html(
                '<span title="{}">{}</span>',
                local_time.strftime("%d.%m.%Y"),
                local_time.strftime("%H:%M:%S"),
            )
        return "-"

    formatted_first_in.short_description = "Время начала"

    def formatted_last_out(self, obj):
        if obj.last_out:
            local_time = timezone.localtime(obj.last_out)
            return format_html(
                '<span title="{}">{}</span>',
                local_time.strftime("%d.%m.%Y"),
                local_time.strftime("%H:%M:%S"),
            )
        return format_html('<span style="color: blue;">Продолжается</span>')

    formatted_last_out.short_description = "Время окончания"

    def location_map(self, obj):
        if obj.latitude and obj.longitude:
            return format_html(
                '<div id="lesson-map" data-lat="{}" data-lng="{}" style="width: 100%; height: 300px;"></div>',
                obj.latitude,
                obj.longitude,
            )
        return "Координаты не указаны"

    location_map.short_description = "Карта местоположения"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.select_related("staff", "staff__department").only(
            "id",
            "staff__id",
            "staff__surname",
            "staff__name",
            "staff__pin",
            "staff__avatar",
            "staff__department__name",
            "subject_name",
            "tutor_id",
            "tutor",
            "first_in",
            "last_out",
            "date_at",
            "staff_image_path",
            "latitude",
            "longitude",
            "photo_spoof_status",
            "photo_spoof_score",
            "photo_spoof_tags",
            "photo_spoof_checked_at",
            "photo_spoof_model_version",
            "photo_trust_confirmed",
            "photo_manual_verdict",
            "photo_manual_comment",
            "photo_manual_by",
            "photo_manual_at",
        )

        photo_expired = request.GET.get("photo_expired")
        if photo_expired == "yes":
            thirty_days_ago = timezone.now().date() - timezone.timedelta(days=31)
            qs = qs.filter(date_at__lt=thirty_days_ago)
        elif photo_expired == "no":
            thirty_days_ago = timezone.now().date() - timezone.timedelta(days=31)
            qs = qs.filter(date_at__gte=thirty_days_ago)
        return qs

    def has_photo(self, obj):
        return (
            obj.staff_image_path
            and obj.staff_image_path != "/static/media/images/no-avatar.png"
        )

    has_photo.boolean = True
    has_photo.short_description = "Фотография"

    @staticmethod
    def _format_photo_status_badge(status_value: str, source: str):
        status_map = {
            LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN: ("#2e7d32", "Нормальное"),
            LessonAttendance.PHOTO_SPOOF_STATUS_REVIEW: ("#f57c00", "На проверку"),
            LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS: (
                "#c62828",
                "Подозрительное",
            ),
            "insufficient_input_review": ("#8a6d1f", "Недостаточно данных"),
            LessonAttendance.PHOTO_SPOOF_STATUS_PENDING: (
                "#6d6d6d",
                "Ожидает проверки",
            ),
            LessonAttendance.PHOTO_SPOOF_STATUS_ERROR: ("#616161", "Ошибка"),
        }
        color, label = status_map.get(
            status_value, ("#616161", status_value or "Неизвестно")
        )
        return format_html(
            "<span style='color:{}; font-weight:600;'>{}</span><br><small style='color:#666;'>{}</small>",
            color,
            label,
            source,
        )

    def photo_effective_status_badge(self, obj):
        if obj.photo_manual_verdict == LessonAttendance.PHOTO_MANUAL_VERDICT_CLEAN:
            return self._format_photo_status_badge(
                LessonAttendance.PHOTO_SPOOF_STATUS_CLEAN, "ручное"
            )
        if obj.photo_manual_verdict == LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS:
            return self._format_photo_status_badge(
                LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS, "ручное"
            )
        auto_status = (
            "insufficient_input_review"
            if _is_auto_insufficient_input(obj)
            else obj.photo_spoof_status
        )
        return self._format_photo_status_badge(auto_status, "авто")

    photo_effective_status_badge.short_description = "Фото-вердикт"

    def photo_pad_scan_status_line(self, obj):
        """Clarify whether ``check_photo`` has already persisted vs still pending."""
        if obj is None:
            return format_html("")
        if (
            not obj.staff_image_path
            or obj.staff_image_path == "/static/media/images/no-avatar.png"
        ):
            return format_html(
                "<p class='la-pad-muted'>Нет файла фото на диске — автоматическая проверка фото не запускалась.</p>"
            )
        la = LessonAttendance
        pending = obj.photo_spoof_status == la.PHOTO_SPOOF_STATUS_PENDING
        if obj.photo_spoof_checked_at is None and pending:
            return format_html(
                "<p class='la-pad-warn'><strong>Ожидает автоматической проверки</strong> "
                "(результат ещё не записан).</p>"
            )
        if obj.photo_spoof_checked_at is None:
            return format_html(
                "<p class='la-pad-muted'>Время последнего авто-скана в БД не заполнено.</p>"
            )
        local = timezone.localtime(obj.photo_spoof_checked_at).strftime(
            "%d.%m.%Y %H:%M:%S"
        )
        ver = (obj.photo_spoof_model_version or "").strip() or "—"
        st_label = (
            "Недостаточно данных"
            if _is_auto_insufficient_input(obj)
            else obj.get_photo_spoof_status_display()
        )
        return format_html(
            "<p class='la-pad-p'><strong>Последний авто-разбор</strong>: {} · модель {} · "
            "статус <strong>{}</strong>.</p>",
            local,
            ver,
            st_label,
        )

    photo_pad_scan_status_line.short_description = "Состояние авто-проверки фото"

    def photo_manual_verdict_badge(self, obj):
        verdict_map = {
            LessonAttendance.PHOTO_MANUAL_VERDICT_NONE: (
                "#6b7280",
                "Нет ручного вердикта",
            ),
            LessonAttendance.PHOTO_MANUAL_VERDICT_CLEAN: ("#2e7d32", "Нормальное"),
            LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS: (
                "#ad1457",
                "Подозрительное (ручное)",
            ),
        }
        color, label = verdict_map.get(
            obj.photo_manual_verdict,
            ("#6b7280", obj.get_photo_manual_verdict_display()),
        )
        return format_html(
            "<span style='color:{}; font-weight:600;'>{}</span>",
            color,
            label,
        )

    photo_manual_verdict_badge.short_description = "Ручной вердикт"

    def photo_spoof_status_badge(self, obj):
        auto_status = (
            "insufficient_input_review"
            if _is_auto_insufficient_input(obj)
            else obj.photo_spoof_status
        )
        return self._format_photo_status_badge(auto_status, "авто")

    photo_spoof_status_badge.short_description = "Авто-статус"

    def photo_pad_operator_readout(self, obj):
        """Render the short operator-facing PAD explanation block."""
        return format_lesson_attendance_antifraud_operator_panel(obj)

    photo_pad_operator_readout.short_description = "Пояснение проверки фото"

    def photo_pad_technical_metrics(self, obj):
        """Secondary numeric PAD table (collapsed fieldset by default)."""
        return format_lesson_attendance_pad_technical_compact(obj)

    photo_pad_technical_metrics.short_description = "Служебные метрики проверки"

    def photo_trust_line(self, obj):
        """Tri-state live-trust from the last automatic PAD scan."""
        la = LessonAttendance
        if obj.photo_manual_verdict != la.PHOTO_MANUAL_VERDICT_NONE:
            return format_html(
                "<span class='la-pad-muted'>Не применяется при ручном вердикте "
                "(ниже — значение последнего авто-скана).</span>"
            )
        val = obj.photo_trust_confirmed
        if val is True:
            return format_html("<strong>да</strong> (авто)")
        if val is False:
            return format_html("<strong>нет</strong> (авто)")
        return format_html("<span class='la-pad-muted'>не определено (авто)</span>")

    photo_trust_line.short_description = "Проверка фото, авто"

    def photo_spoof_model_version_display(self, obj):
        """PAD model version string stored on the row."""
        ver = (obj.photo_spoof_model_version or "").strip()
        if not ver:
            return format_html("<span class='la-pad-muted'>—</span>")
        return format_html("<code>{}</code>", ver)

    photo_spoof_model_version_display.short_description = "Версия модели проверки"

    def photo_antifraud_list_compact(self, obj):
        """One-line effective PAD hint for the changelist."""
        return format_lesson_attendance_antifraud_list_hint(obj)

    photo_antifraud_list_compact.short_description = "Проверка фото (кратко)"

    @staticmethod
    def _manual_verdict_update(
        request,
        queryset,
        verdict: str,
        comment: str,
    ) -> int:
        return queryset.update(
            photo_manual_verdict=verdict,
            photo_manual_comment=comment,
            photo_manual_by=request.user,
            photo_manual_at=timezone.now(),
        )

    def mark_photo_manual_clean(self, request, queryset):
        updated = self._manual_verdict_update(
            request=request,
            queryset=queryset,
            verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_CLEAN,
            comment="Manual clean via admin action",
        )
        self.message_user(
            request, f"Ручной вердикт «Нормальное» установлен: {updated}."
        )

    mark_photo_manual_clean.short_description = "Отметить как нормальное (manual)"

    def mark_photo_manual_suspicious(self, request, queryset):
        updated = self._manual_verdict_update(
            request=request,
            queryset=queryset,
            verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS,
            comment="Manual suspicious via admin action",
        )
        self.message_user(
            request,
            f"Ручной вердикт «Подозрительное (ручное)» установлен: {updated}.",
        )

    mark_photo_manual_suspicious.short_description = (
        "Отметить как подозрительное (manual)"
    )

    def reset_photo_manual_verdict(self, request, queryset):
        updated = queryset.update(
            photo_manual_verdict=LessonAttendance.PHOTO_MANUAL_VERDICT_NONE,
            photo_manual_comment="",
            photo_manual_by=None,
            photo_manual_at=None,
        )
        self.message_user(request, f"Ручной вердикт сброшен: {updated}.")

    reset_photo_manual_verdict.short_description = "Сбросить ручной вердикт фото"

    @staticmethod
    def _format_lesson_attendance_pad_rescan_message(
        *,
        selected_n: int,
        queued_n: int,
        skipped: dict[str, int],
    ) -> str:
        """Build a short admin message after a PAD rescan request."""

        parts: list[str] = []
        if queued_n > 0:
            parts.append(
                f"Выбрано {selected_n}, повторная проверка начата для {queued_n} "
                f"записей с фото (сброшены авто-разбор и ручной вердикт)."
            )
        else:
            parts.append(
                f"Выбрано записей: {selected_n}. Повторная проверка не начата: нет строк с путём к фото."
            )
        skip_bits: list[str] = []
        if skipped.get("no_photo", 0):
            skip_bits.append(f"нет пути к фото — {skipped['no_photo']}")
        if skipped.get("not_found", 0):
            skip_bits.append(f"не найдено в БД — {skipped['not_found']}")
        if skip_bits:
            parts.append("Пропущено: " + "; ".join(skip_bits) + ".")
        return " ".join(parts)

    def rescan_selected_photos(self, request, queryset):
        """Reset PAD/manual fields for selected rows with photos and start a full rescan.

        Only changelist-selected primary keys are touched. Every selected row that
        has a non-empty ``staff_image_path`` is cleared synchronously (automatic PAD
        outputs and manual verdict), then the same ids are submitted for a
        background ``check_photo`` run. Rows without a stored path are skipped.
        """
        from monitoring_app.tasks import (
            prepare_lesson_attendance_admin_pad_full_rescan,
            rescan_lesson_attendance_photo_ids,
        )

        selected_ids = list(queryset.values_list("id", flat=True))
        if not selected_ids:
            self.message_user(
                request,
                "Не выбраны записи для перескана.",
                level=messages.WARNING,
            )
            return

        photo_ids, skipped = prepare_lesson_attendance_admin_pad_full_rescan(
            selected_ids
        )

        if not photo_ids:
            msg = self._format_lesson_attendance_pad_rescan_message(
                selected_n=len(selected_ids),
                queued_n=0,
                skipped=skipped,
            )
            self.message_user(request, msg, level=messages.WARNING)
            return

        rescan_task = cast(Any, rescan_lesson_attendance_photo_ids)
        pad_device = getattr(settings, "PHOTO_PAD_DEVICE", "auto")
        pad_batch = getattr(settings, "PHOTO_PAD_HOURLY_BATCH_SIZE", 100)
        selected_count = len(selected_ids)
        queued_count = len(photo_ids)

        def enqueue_pad_rescan_after_commit() -> None:
            """Run async rescan after DB commit so PAD fields are visible."""
            try:
                task = rescan_task.delay(
                    attendance_ids=photo_ids,
                    device=pad_device,
                    force_manual=True,
                    batch_size=pad_batch,
                    auto_eligible_only=False,
                )
                logger.info(
                    "PAD admin: rescan task_id=%s lesson_ids=%s",
                    task.id,
                    photo_ids,
                )
            except Exception as exc:
                logger.exception(
                    "PAD admin rescan queue failed ids=%s error=%s",
                    selected_ids[:20],
                    exc,
                )
                self.message_user(
                    request,
                    "Не удалось запустить перескан. Повторите позже или обратитесь к администратору.",
                    level=messages.ERROR,
                )
                return

            msg = self._format_lesson_attendance_pad_rescan_message(
                selected_n=selected_count,
                queued_n=queued_count,
                skipped=skipped,
            )
            self.message_user(request, msg, level=messages.SUCCESS)

        transaction.on_commit(enqueue_pad_rescan_after_commit)

    rescan_selected_photos.short_description = "Перепроверить фото по выбранным"

    def closest_location(self, obj):
        if obj.latitude is None or obj.longitude is None:
            return "N/A"

        cache_key = f"closest_location_{obj.id}_{obj.latitude:.4f}_{obj.longitude:.4f}"
        cached_result = cache.get(cache_key)
        if cached_result:
            return format_html(cached_result)

        locations = cache.get("lesson_admin_closest_locations")
        if locations is None:
            locations = list(
                ClassLocation.objects.filter(
                    latitude__isnull=False, longitude__isnull=False
                ).only("id", "name", "address", "latitude", "longitude")
            )
            cache.set("lesson_admin_closest_locations", locations, 300)

        radius = 300
        obj_lat, obj_lon = obj.latitude, obj.longitude
        closest = None
        min_distance = float("inf")

        for loc in locations:
            lat_diff = abs(loc.latitude - obj_lat) * 111320
            lon_diff = abs(loc.longitude - obj_lon) * 111320 * abs(obj_lat / 90)
            distance = (lat_diff**2 + lon_diff**2) ** 0.5

            if distance < min_distance and distance <= radius:
                min_distance = distance
                closest = loc

        if closest:
            result = format_html(
                '<span style="color: green;">{}</span><br><small>{}</small>',
                closest.name,
                closest.address,
            )
            cache.set(cache_key, result, 86400)
            return result
        else:
            result = format_html('<span style="color: red;">Неизвестно</span>')
            cache.set(cache_key, result, 86400)
            return result

    closest_location.short_description = "Ближайшая локация"

    def photo_preview(self, obj):
        if (
            obj.staff_image_path
            and obj.staff_image_path != "/static/media/images/no-avatar.png"
        ):
            return format_html(
                """
                <div style="text-align: center;">
                    <div style="display: inline-block; border: 1px solid #ddd; padding: 5px; border-radius: 5px;">
                        <img src="{}" style="max-width: 200px; max-height: 200px; object-fit: contain;"/>
                    </div>
                </div>
                """,
                obj.image_url,
            )
        return format_html(
            """
            <div style="text-align: center;">
                <div style="display: inline-block; width: 200px; height: 200px; border-radius: 5px;
                      background-color: #f0f0f0; display: flex; justify-content: center; align-items: center;">
                    <span style="color: #999; font-style: italic;">Нет фото</span>
                </div>
            </div>
            """
        )

    photo_preview.short_description = "Фото"

    def export_lesson_data(self, request, queryset):
        count = queryset.count()
        self.message_user(request, f"Экспортированы данные о {count} занятиях.")

    export_lesson_data.short_description = "Экспортировать данные о занятиях"

    def cleanup_old_photos(self, request, queryset):
        count = queryset.count()
        self.message_user(request, f"Удалены старые фотографии для {count} занятий.")

    cleanup_old_photos.short_description = "Удалить старые фотографии"

    def has_add_permission(self, request):
        """Строки создаёт учёт посещаемости, не ручной ввод в админке."""
        return False

    def save_model(self, request, obj, form, change):
        if change and "photo_manual_verdict" in form.changed_data:
            if obj.photo_manual_verdict == LessonAttendance.PHOTO_MANUAL_VERDICT_NONE:
                obj.photo_manual_by = None
                obj.photo_manual_at = None
            else:
                obj.photo_manual_by = request.user
                obj.photo_manual_at = timezone.now()
        super().save_model(request, obj, form, change)

    class Media:
        css = {"all": ("admin/css/custom_admin.css",)}
        js = ("admin/js/lesson_admin.js", "admin/js/leaflet.js")


admin_site.register(LessonAttendance, LessonAttendanceAdmin)


class ClassLocationAdmin(ModelAdmin):
    ATTENDANCE_STATS_MONTHS = 6
    ATTENDANCE_STATS_CACHE_TTL = 3600
    ATTENDANCE_STATS_CACHE_VERSION = "geo_v3"
    ATTENDANCE_PERIOD_CACHE_VERSION = "attendance_period_v2"
    ATTENDANCE_PERIOD_DEFAULT_MONTHS = 6
    geomap_field_longitude = "longitude"
    geomap_field_latitude = "latitude"
    geomap_show_map_on_list = True
    geomap_item_zoom = "14"
    geomap_height = "450px"
    geomap_default_zoom = "16"
    geomap_autozoom = "15.9"

    readonly_fields = ("created_at", "updated_at", "attendance_stats")
    save_on_top = True

    class AttendancePeriodFilter(SimpleListFilter):
        title = "Период посещаемости"
        parameter_name = "attendance_period"
        template = "admin/attendance_period_filter.html"

        def __init__(self, request, params, model, model_admin):
            raw_from = params.pop("attendance_from", None)
            raw_to = params.pop("attendance_to", None)
            self.date_from = (raw_from[-1] if raw_from else "") or ""
            self.date_to = (raw_to[-1] if raw_to else "") or ""
            super().__init__(request, params, model, model_admin)
            self.admin_theme = _admin_theme_name()

        def lookups(self, request, model_admin):
            _ = request
            _ = model_admin
            return (
                ("6m", "6 месяцев"),
                ("academic", "Учебный год (1 сен — 31 июл)"),
                ("this_year", "Этот год"),
                ("this_month", "Этот месяц"),
                ("30", "30 дней"),
                ("90", "90 дней"),
                ("180", "180 дней"),
                ("365", "365 дней"),
            )

        def queryset(self, request, queryset):
            _ = request
            return queryset

        def choices(self, changelist):
            custom = bool(
                _parse_admin_iso_date(self.date_from)
                or _parse_admin_iso_date(self.date_to)
            )
            yield {
                "selected": self.value() is None and not custom,
                "query_string": changelist.get_query_string(
                    remove=[
                        self.parameter_name,
                        "attendance_from",
                        "attendance_to",
                    ]
                ),
                "display": "Все",
            }
            if custom:
                yield {
                    "selected": True,
                    "query_string": changelist.get_query_string(),
                    "display": f"{self.date_from or '…'} — {self.date_to or '…'}",
                }
            for lookup, title in self.lookup_choices:
                yield {
                    "selected": not custom and self.value() == str(lookup),
                    "query_string": changelist.get_query_string(
                        {self.parameter_name: lookup},
                        remove=["attendance_from", "attendance_to"],
                    ),
                    "display": title,
                }

    class AttendanceVolumeFilter(SimpleListFilter):
        title = "Активность локации"
        parameter_name = "attendance_volume"

        def lookups(self, request, model_admin):
            _ = request
            _ = model_admin
            return (
                ("zero", "Нет посещений"),
                ("low", "1-9 посещений"),
                ("medium", "10-49 посещений"),
                ("high", "50+ посещений"),
            )

        def queryset(self, request, queryset):
            _ = request
            value = self.value()
            if value == "zero":
                return queryset.filter(_attendance_hits_period=0)
            if value == "low":
                return queryset.filter(
                    _attendance_hits_period__gte=1, _attendance_hits_period__lte=9
                )
            if value == "medium":
                return queryset.filter(
                    _attendance_hits_period__gte=10, _attendance_hits_period__lte=49
                )
            if value == "high":
                return queryset.filter(_attendance_hits_period__gte=50)
            return queryset

    fieldsets = (
        (
            "Основная информация",
            {
                "fields": ("name", "address"),
                "classes": ("wide",),
            },
        ),
        (
            "Координаты",
            {
                "fields": (("latitude", "longitude"), "acceptance_radius_m"),
                "classes": ("wide",),
                "description": "Введите координаты вручную или выберите на карте ниже. Приёмный радиус (м): если задан — используется вместо вычисленного по соседям. Подберите по карте: круг должен охватывать здание/двор. 20–30 м — кабинет, 50–100 м — здание/двор. Пусто = по умолчанию (60–80 по соседям).",
            },
        ),
        (
            "Статистика",
            {
                "fields": ("attendance_stats",),
                "classes": ("wide",),
            },
        ),
        (
            "Системная информация",
            {
                "fields": ("created_at", "updated_at"),
                "classes": ("grp-collapse grp-closed",),
            },
        ),
    )

    list_display = (
        "name",
        "attendance_hits_period",
        "attendance_activity_status",
        "address",
        "formatted_latitude",
        "formatted_longitude",
        "acceptance_radius_m",
        "created_at",
    )
    list_filter = (
        AttendancePeriodFilter,
        AttendanceVolumeFilter,
        "created_at",
    )
    search_fields = ("name", "address")
    actions = ["export_for_upload", "export_attendance"]

    def export_for_upload(self, request, queryset):
        """Экспорт в Excel (формат загрузки). Выберите записи или нажмите «Выбрать все»."""
        content = monitoring_utils.export_class_locations_to_excel(queryset)
        filename = "class_locations_export.xlsx"
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{quote(filename)}"'
        return response

    export_for_upload.short_description = "Экспорт в Excel для загрузки"

    def export_attendance(self, request, queryset):
        """Посещаемость за фильтр периода: LessonAttendance → ближайшая локация в радиусе."""
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.styles.numbers import BUILTIN_FORMATS
        from openpyxl.worksheet.table import Table, TableStyleInfo

        period_start, period_end, period_label = self._resolve_attendance_period_window(
            request
        )
        all_locations = list(
            ClassLocation.objects.filter(
                latitude__isnull=False,
                longitude__isnull=False,
            ).only("id", "latitude", "longitude", "acceptance_radius_m")
        )
        counts = self._get_location_attendance_period_counts(
            all_locations,
            period_start=period_start,
            now=period_end,
            period_label=period_label,
        )
        start_s = timezone.localtime(period_start).strftime("%Y-%m-%d")
        end_s = timezone.localtime(period_end).strftime("%Y-%m-%d")
        start_label = timezone.localtime(period_start).strftime("%d.%m.%Y")
        end_label = timezone.localtime(period_end).strftime("%d.%m.%Y")
        headers = ("Название", "Адрес", "Дата создания", "Посещений")
        max_name, max_addr = len(headers[0]), len(headers[1])
        rows = []
        for loc in queryset.only("id", "name", "address", "created_at"):
            name = loc.name or ""
            address = loc.address or ""
            max_name = max(max_name, len(name))
            max_addr = max(max_addr, len(address))
            created = _to_local_datetime(loc.created_at)
            rows.append(
                (
                    name,
                    address,
                    created.replace(tzinfo=None) if created else None,
                    counts.get(loc.pk, 0),
                )
            )
        rows.sort(key=lambda row: (-row[3], row[0].casefold()))

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1D4ED8")
        wb = Workbook()
        ws = wb.active
        ws.title = "Посещаемость"
        ws.append([f"Период: {start_label} — {end_label}"])
        ws.merge_cells("A1:D1")
        ws["A1"].font = Font(bold=True, size=12)
        ws.append(list(headers))
        for row in rows:
            ws.append(row)
            ws.cell(row=ws.max_row, column=3).number_format = BUILTIN_FORMATS[22]
        ws.column_dimensions["A"].width = _excel_col_width(max_name, min_w=16, max_w=48)
        ws.column_dimensions["B"].width = _excel_col_width(max_addr, min_w=18, max_w=56)
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 14
        ws.freeze_panes = "A3"
        last_row = 2 + max(len(rows), 1)
        if rows:
            table = Table(displayName="Attendance", ref=f"A2:D{2 + len(rows)}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
            )
            ws.add_table(table)
        else:
            ws.auto_filter.ref = f"A2:D{last_row}"
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )

        out = io.BytesIO()
        wb.save(out)
        filename = f"class_locations_attendance_{start_s}_{end_s}.xlsx"
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{quote(filename)}"'
        return response

    export_attendance.short_description = "Выгрузить посещаемость за период"

    def _default_attendance_period_window(self):
        now = timezone.now()
        current_month_start = timezone.localdate().replace(day=1)
        start_date = _shift_month_start(
            current_month_start,
            self.ATTENDANCE_PERIOD_DEFAULT_MONTHS - 1,
        )
        period_start, _ = _aware_day_bounds(start_date, start_date)
        return (
            period_start,
            now,
            f"month_window_{self.ATTENDANCE_PERIOD_DEFAULT_MONTHS}",
        )

    def _resolve_attendance_period_window(self, request):
        today = timezone.localdate()
        now = timezone.now()
        from_date = _parse_admin_iso_date(request.GET.get("attendance_from"))
        to_date = _parse_admin_iso_date(request.GET.get("attendance_to"))
        if from_date or to_date:
            if from_date is None:
                from_date = to_date
            if to_date is None:
                to_date = today
            if from_date > to_date:
                from_date, to_date = to_date, from_date
            period_start, period_end = _aware_day_bounds(from_date, to_date)
            return (
                period_start,
                period_end,
                f"custom_{from_date.isoformat()}_{to_date.isoformat()}",
            )

        raw_value = request.GET.get("attendance_period")
        if raw_value in {"6m", "6"}:
            return self._default_attendance_period_window()
        if raw_value == "academic":
            from_date, to_date = _academic_year_bounds(today)
            period_start, period_end = _aware_day_bounds(from_date, to_date)
            return period_start, period_end, f"academic_{from_date.year}"
        if raw_value == "this_year":
            period_start, period_end = _aware_day_bounds(date(today.year, 1, 1), today)
            return period_start, period_end, f"year_{today.year}"
        if raw_value == "this_month":
            last_day = date(
                today.year, today.month, monthrange(today.year, today.month)[1]
            )
            period_start, period_end = _aware_day_bounds(today.replace(day=1), last_day)
            return period_start, period_end, f"month_{today.year}_{today.month:02d}"
        if not raw_value:
            return self._default_attendance_period_window()
        try:
            days = int(raw_value)
        except (TypeError, ValueError):
            return self._default_attendance_period_window()
        days = max(1, days)
        return now - timedelta(days=days), now, f"days_{days}"

    def _distance_to_location_m(self, location_meta, lesson) -> float:
        return monitoring_utils.calculate_distance_haversine(
            location_meta["latitude"],
            location_meta["longitude"],
            lesson.latitude,
            lesson.longitude,
        )

    def _get_location_attendance_period_counts(
        self, locations, *, period_start, now, period_label: str
    ):
        locations = [
            location
            for location in locations
            if getattr(location, "pk", None) is not None
            and getattr(location, "latitude", None) is not None
            and getattr(location, "longitude", None) is not None
        ]
        if not locations:
            return {}

        cache_key = (
            f"class_location_attendance_period_counts_{self.ATTENDANCE_PERIOD_CACHE_VERSION}_"
            f"{LessonAttendance.REPORT_FILTER_CACHE_VERSION}_{period_label}_"
            f"{timezone.localdate().isoformat()}"
        )
        cached = cache.get(cache_key)
        if cached is not None:
            return cached

        radii = self._get_acceptance_radii(locations)
        location_meta = []
        max_lat_margin = 0.0
        max_lon_margin = 0.0
        for location in locations:
            radius_m = max(1, monitoring_utils.get_location_radius(location, radii))
            lat_margin, lon_margin = _radius_bbox(location.latitude, radius_m)
            max_lat_margin = max(max_lat_margin, lat_margin)
            max_lon_margin = max(max_lon_margin, lon_margin)
            location_meta.append(
                {
                    "id": location.pk,
                    "latitude": location.latitude,
                    "longitude": location.longitude,
                    "radius_m": radius_m,
                    "lat_margin": lat_margin,
                    "lon_margin": lon_margin,
                }
            )

        candidate_lessons = LessonAttendance.exclude_report_invalid_days(
            LessonAttendance.objects.filter(
                first_in__gte=period_start,
                first_in__lte=now,
                latitude__gte=min(item["latitude"] for item in location_meta)
                - max_lat_margin,
                latitude__lte=max(item["latitude"] for item in location_meta)
                + max_lat_margin,
                longitude__gte=min(item["longitude"] for item in location_meta)
                - max_lon_margin,
                longitude__lte=max(item["longitude"] for item in location_meta)
                + max_lon_margin,
            )
        ).only("id", "latitude", "longitude")

        counts = {location.pk: 0 for location in locations}
        for lesson in candidate_lessons.iterator(chunk_size=1000):
            nearest_location_id = None
            nearest_distance = float("inf")
            for item in location_meta:
                if abs(lesson.latitude - item["latitude"]) > item["lat_margin"]:
                    continue
                if abs(lesson.longitude - item["longitude"]) > item["lon_margin"]:
                    continue
                distance_m = self._distance_to_location_m(item, lesson)
                if distance_m > item["radius_m"]:
                    continue
                if distance_m < nearest_distance:
                    nearest_distance = distance_m
                    nearest_location_id = item["id"]
            if nearest_location_id is not None:
                counts[nearest_location_id] += 1

        cache.set(cache_key, counts, timeout=self.ATTENDANCE_STATS_CACHE_TTL)
        return counts

    def _attendance_query_active(self, request) -> bool:
        get = request.GET
        return any(
            get.get(key)
            for key in (
                "attendance_period",
                "attendance_from",
                "attendance_to",
                "attendance_volume",
            )
        )

    def _should_attach_attendance_counts(self, request) -> bool:
        resolver_name = getattr(
            getattr(request, "resolver_match", None), "url_name", ""
        )
        if resolver_name and not resolver_name.endswith("_changelist"):
            return False
        return self._attendance_query_active(request)

    def get_queryset(self, request):
        queryset = (
            super()
            .get_queryset(request)
            .only(
                "id",
                "name",
                "address",
                "latitude",
                "longitude",
                "acceptance_radius_m",
                "created_at",
                "updated_at",
            )
        )
        if not self._should_attach_attendance_counts(request):
            return queryset.annotate(
                _attendance_hits_period=Value(0, output_field=IntegerField())
            )
        period_start, period_end, period_label = self._resolve_attendance_period_window(
            request
        )
        locations = list(queryset)
        counts = self._get_location_attendance_period_counts(
            locations,
            period_start=period_start,
            now=period_end,
            period_label=period_label,
        )
        if not counts:
            return queryset.annotate(
                _attendance_hits_period=Value(0, output_field=IntegerField())
            )
        count_cases = [
            When(pk=location_id, then=Value(hit_count))
            for location_id, hit_count in counts.items()
        ]
        return queryset.annotate(
            _attendance_hits_period=Case(
                *count_cases,
                default=Value(0),
                output_field=IntegerField(),
            )
        )

    def _get_acceptance_radii(self, locations=None):
        radii = cache.get(CLASS_LOCATION_ACCEPTANCE_RADII_CACHE_KEY)
        if radii is not None:
            return radii
        if locations is None:
            locations = list(
                ClassLocation.objects.filter(
                    latitude__isnull=False,
                    longitude__isnull=False,
                ).only("id", "latitude", "longitude", "acceptance_radius_m")
            )
        else:
            locations = [
                location
                for location in locations
                if getattr(location, "latitude", None) is not None
                and getattr(location, "longitude", None) is not None
            ]
        if not locations:
            return {}
        radii = monitoring_utils.compute_class_location_acceptance_radii(
            locations,
            r_same_point=ACCEPTANCE_R_SAME_POINT,
            r_cluster=ACCEPTANCE_R_CLUSTER,
            r_standalone=ACCEPTANCE_R_STANDALONE,
            same_point_threshold=SAME_POINT_THRESHOLD_M,
            cluster_threshold=CLUSTER_THRESHOLD_M,
        )
        cache.set(
            CLASS_LOCATION_ACCEPTANCE_RADII_CACHE_KEY,
            radii,
            CLASS_LOCATION_ACCEPTANCE_RADII_CACHE_TTL,
        )
        return radii

    def attendance_hits_period(self, obj):
        return getattr(obj, "_attendance_hits_period", 0)

    attendance_hits_period.short_description = "Посещений за период"
    attendance_hits_period.admin_order_field = "_attendance_hits_period"

    def attendance_activity_status(self, obj):
        hits = getattr(obj, "_attendance_hits_period", 0)
        if hits == 0:
            return _admin_badge("Пустая", background="#94a3b8")
        if hits < 10:
            return _admin_badge("Низкая", background="#b45309")
        if hits < 50:
            return _admin_badge("Средняя", background="#2563eb")
        return _admin_badge("Высокая", background="#0f766e")

    attendance_activity_status.short_description = "Активность"

    def _attendance_stats_cache_key(self, obj):
        return (
            f"attendance_stats_{self.ATTENDANCE_STATS_CACHE_VERSION}_"
            f"{LessonAttendance.REPORT_FILTER_CACHE_VERSION}_{obj.pk}_"
            f"{timezone.localdate().isoformat()}"
        )

    def _attendance_stats_month_order(self):
        current_month_start = timezone.localdate().replace(day=1)
        return [
            _shift_month_start(current_month_start, month_back)
            for month_back in range(self.ATTENDANCE_STATS_MONTHS - 1, -1, -1)
        ]

    def attendance_stats(self, obj):
        if (
            obj is None
            or obj.pk is None
            or obj.latitude is None
            or obj.longitude is None
        ):
            return format_html(
                '<div style="padding: 20px; color: #666;">'
                "Сохраните локацию с координатами для отображения статистики посещаемости."
                "</div>"
            )
        cache_key = self._attendance_stats_cache_key(obj)
        cached = cache.get(cache_key)
        if cached is not None:
            return format_html(cached)

        now = timezone.now()
        months_order = self._attendance_stats_month_order()
        stats_start = timezone.make_aware(
            datetime.combine(months_order[0], time.min),
            timezone.get_current_timezone(),
        )
        radii = self._get_acceptance_radii()
        analysis_radius_m = max(1, monitoring_utils.get_location_radius(obj, radii))
        lat_margin, lon_margin = _radius_bbox(obj.latitude, analysis_radius_m)

        candidate_lessons = LessonAttendance.exclude_report_invalid_days(
            LessonAttendance.objects.filter(
                latitude__gte=obj.latitude - lat_margin,
                latitude__lte=obj.latitude + lat_margin,
                longitude__gte=obj.longitude - lon_margin,
                longitude__lte=obj.longitude + lon_margin,
                first_in__gte=stats_start,
                first_in__lte=now,
            )
        ).only("id", "latitude", "longitude", "first_in")

        counts_by_ym = defaultdict(int)
        total_hits = 0
        location_meta = {
            "latitude": obj.latitude,
            "longitude": obj.longitude,
            "radius_m": analysis_radius_m,
        }
        for lesson in candidate_lessons.iterator(chunk_size=500):
            distance_m = self._distance_to_location_m(location_meta, lesson)
            if distance_m > analysis_radius_m:
                continue
            local_first_in = _to_local_datetime(lesson.first_in)
            if local_first_in is None:
                continue
            counts_by_ym[(local_first_in.year, local_first_in.month)] += 1
            total_hits += 1

        months_data = [
            counts_by_ym.get((month.year, month.month), 0) for month in months_order
        ]
        month_names = [
            date_format(month, "M") or month_abbr[month.month] for month in months_order
        ]
        max_count = max(months_data) if any(months_data) else 1

        columns = []
        for month_name, count in zip(month_names, months_data):
            height_percent = (count / max_count * 100) if max_count > 0 else 0
            bar_height = max(10, int(height_percent * 1.8)) if count else 10
            columns.append(
                format_html(
                    '<div style="flex:1; min-width:64px; display:flex; flex-direction:column; align-items:center; gap:6px;">'
                    '<div style="font-size:12px; font-weight:600; color:#475569;">{}</div>'
                    '<div style="width:100%; max-width:56px; height:180px; display:flex; align-items:flex-end;">'
                    '<div style="width:100%; border-radius:10px 10px 4px 4px; background:linear-gradient(180deg, #22c55e 0%, #15803d 100%); height:{}px;"></div>'
                    "</div>"
                    '<div style="font-size:12px; color:#0f172a;">{}</div>'
                    "</div>",
                    month_name,
                    bar_height,
                    count,
                )
            )

        columns_html = format_html_join("", "{}", ((column,) for column in columns))
        empty_state_html = ""
        if total_hits == 0:
            empty_state_html = format_html(
                '<div style="margin-top:12px; padding:12px; border-radius:10px; background:#f8fafc; color:#64748b; font-size:13px;">'
                "За выбранный период рядом с этой локацией не найдено ни одной подтвержденной отметки."
                "</div>"
            )

        html = format_html(
            '<div style="padding:20px; border-radius:14px; background:linear-gradient(180deg, #ffffff 0%, #f8fafc 100%); border:1px solid #e2e8f0;">'
            '<div style="display:flex; flex-wrap:wrap; justify-content:space-between; gap:10px; align-items:flex-start;">'
            "<div>"
            '<div style="font-size:16px; font-weight:700; color:#0f172a;">Статистика посещаемости</div>'
            '<div style="font-size:13px; color:#64748b;">Последние {} месяцев, аналитический радиус {} м.</div>'
            "</div>"
            '<div style="display:flex; gap:8px; flex-wrap:wrap;">'
            "{}{}"
            "</div>"
            "</div>"
            '<div style="margin-top:18px; display:flex; gap:12px; align-items:flex-end;">{}</div>'
            "{}"
            "</div>",
            self.ATTENDANCE_STATS_MONTHS,
            analysis_radius_m,
            _admin_badge(f"Всего отметок {total_hits}", background="#0f766e"),
            _admin_badge(
                f"Максимум за месяц {max(months_data) if months_data else 0}",
                background="#1d4ed8",
            ),
            columns_html,
            empty_state_html,
        )
        cache.set(cache_key, str(html), timeout=self.ATTENDANCE_STATS_CACHE_TTL)
        return html

    attendance_stats.short_description = "Статистика посещаемости"

    def formatted_latitude(self, obj):
        return f"{obj.latitude:.6f}"

    formatted_latitude.short_description = "Широта"

    def formatted_longitude(self, obj):
        return f"{obj.longitude:.6f}"

    formatted_longitude.short_description = "Долгота"

    def add_view(self, request, form_url="", extra_context=None):
        logger.debug("ClassLocationAdmin.add_view GET path=%s", request.path)
        try:
            response = super().add_view(request, form_url, extra_context)
            logger.info(
                "ClassLocationAdmin.add_view OK path=%s status=%s",
                request.path,
                getattr(response, "status_code", "N/A"),
            )
            return response
        except Exception as e:
            logger.exception(
                "ClassLocationAdmin.add_view ERROR path=%s error=%s",
                request.path,
                e,
            )
            raise

    def save_model(self, request, obj, form, change):
        logger.debug(
            "ClassLocationAdmin.save_model obj=%s change=%s",
            getattr(obj, "name", obj.pk),
            change,
        )
        try:
            with _db_atomic():
                super().save_model(request, obj, form, change)
                try:
                    from monitoring_app.signals import (
                        invalidate_class_location_cache_impl,
                    )

                    invalidate_class_location_cache_impl()
                except Exception as inv_err:
                    logger.warning(
                        "ClassLocationAdmin.save_model cache invalidation: %s", inv_err
                    )
            logger.info(
                "ClassLocationAdmin.save_model OK id=%s name=%s",
                obj.pk,
                getattr(obj, "name", ""),
            )
        except Exception as e:
            logger.exception(
                "ClassLocationAdmin.save_model ERROR name=%s error=%s",
                getattr(obj, "name", "?"),
                e,
            )
            raise

    def change_view(self, request, object_id, form_url="", extra_context=None):
        logger.debug("ClassLocationAdmin.change_view object_id=%s", object_id)
        response = super().change_view(request, object_id, form_url, extra_context)
        item = self.get_queryset(request).filter(pk=object_id).first()
        if (
            item
            and getattr(item, "geomap_longitude", None)
            and getattr(item, "geomap_latitude", None)
        ):
            radii = self._get_acceptance_radii()
            ctx = getattr(response, "context_data", None)
            if ctx is not None:
                ctx["geomap_draw_radius_circles"] = True
                ctx["geomap_radius_by_id"] = {
                    item.pk: monitoring_utils.get_location_radius(item, radii)
                }
                ctx["geomap_color_by_id"] = {item.pk: 0}
        return response

    def changelist_view(self, request, extra_context=None):
        logger.debug("ClassLocationAdmin.changelist_view")
        try:
            response = super().changelist_view(request, extra_context=extra_context)
        except Exception as e:
            logger.exception("ClassLocationAdmin.changelist_view ERROR error=%s", e)
            raise
        if not self.geomap_show_map_on_list:
            return response
        ctx = getattr(response, "context_data", None)
        if ctx is None or not isinstance(ctx, dict):
            return response
        cl = ctx.get("cl")

        if ctx.get("geomap_items", _MARKER) is _MARKER:
            ctx.update(self.set_common(request, {}))
            ctx.update(
                {
                    "geomap_items": (
                        cl.queryset
                        if cl
                        else ClassLocation.objects.filter(
                            latitude__isnull=False, longitude__isnull=False
                        )
                    )
                }
            )
        qs = (
            cl.queryset
            if cl
            else ClassLocation.objects.filter(
                latitude__isnull=False, longitude__isnull=False
            )
        )
        locs = [
            o
            for o in qs
            if getattr(o, "latitude", None) is not None
            and getattr(o, "longitude", None) is not None
        ]
        if not locs:
            ctx.setdefault("geomap_radius_by_id", {})
            ctx.setdefault("geomap_color_by_id", {})
            return response
        radii = self._get_acceptance_radii(locs)
        loc_ids_str = "_".join(str(o.id) for o in sorted(locs, key=lambda x: x.id))
        colors_cache_key = f"class_location_neighbor_colors_{loc_ids_str}"
        geomap_color_by_id = cache.get(colors_cache_key)
        if geomap_color_by_id is None:
            geomap_color_by_id = monitoring_utils.compute_neighbor_color_index(locs, 30)
            cache.set(colors_cache_key, geomap_color_by_id, timeout=300)
        ctx.update(
            {
                "geomap_draw_radius_circles": True,
                "geomap_radius_by_id": {
                    o.id: monitoring_utils.get_location_radius(o, radii) for o in locs
                },
                "geomap_color_by_id": geomap_color_by_id,
            }
        )
        return response

    class Media:
        css = {"all": ("admin/css/custom_admin.css",)}
        js = ("admin/js/location_admin.js", "admin/js/leaflet.js")


admin_site.register(ClassLocation, ClassLocationAdmin)


# ===== SALARY AND BENEFITS MODELS =====


@admin.register(Salary, site=admin_site)
class SalaryAdmin(admin.ModelAdmin):
    list_display = (
        "staff",
        "staff_department",
        "net_salary",
        "total_salary",
        "contract_type",
        "tax_amount",
    )
    search_fields = ("staff__surname", "staff__name", "staff__pin")
    list_filter = ("staff__department", "contract_type")
    autocomplete_fields = ("staff",)
    readonly_fields = ("total_salary", "tax_amount")
    save_on_top = True
    actions = ["calculate_bonuses", "export_salary_report"]

    fieldsets = (
        (
            "Основная информация",
            {
                "fields": ("staff", "net_salary", "total_salary", "contract_type"),
                "classes": ("wide",),
            },
        ),
    )

    def staff_department(self, obj):
        return obj.staff.department.name if obj.staff.department else "N/A"

    staff_department.short_description = "Отдел"
    staff_department.admin_order_field = "staff__department__name"

    def tax_amount(self, obj):
        tax_rate = 13
        tax = float(obj.net_salary) * (tax_rate / 100)
        return f"{tax:.2f} тг."

    tax_amount.short_description = "Сумма налога"

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related("staff", "staff__department")
            .prefetch_related("staff__positions")
        )

    def calculate_bonuses(self, request, queryset):
        count = queryset.count()
        self.message_user(request, f"Рассчитаны бонусы для {count} сотрудников.")

    calculate_bonuses.short_description = "Рассчитать бонусы для выбранных сотрудников"

    def export_salary_report(self, request, queryset):
        count = queryset.count()
        self.message_user(
            request, f"Экспортирован отчет по зарплате для {count} сотрудников."
        )

    export_salary_report.short_description = "Экспортировать отчет по зарплате"


@admin.register(PublicHoliday, site=admin_site)
class PublicHolidayAdmin(admin.ModelAdmin):
    class HolidayPeriodFilter(SimpleListFilter):
        title = "Период"
        parameter_name = "holiday_period"

        def lookups(self, request, model_admin):
            _ = request
            _ = model_admin
            return (
                ("today", "Сегодня"),
                ("week", "Ближайшие 7 дней"),
                ("month", "Ближайшие 30 дней"),
                ("future", "Будущие"),
                ("past", "Прошедшие"),
            )

        def queryset(self, request, queryset):
            _ = request
            today = timezone.localdate()
            value = self.value()
            if value == "today":
                return queryset.filter(date=today)
            if value == "week":
                return queryset.filter(
                    date__gte=today, date__lte=today + timedelta(days=7)
                )
            if value == "month":
                return queryset.filter(
                    date__gte=today, date__lte=today + timedelta(days=30)
                )
            if value == "future":
                return queryset.filter(date__gte=today)
            if value == "past":
                return queryset.filter(date__lt=today)
            return queryset

    class HolidayYearFilter(SimpleListFilter):
        title = "Год"
        parameter_name = "holiday_year"

        def lookups(self, request, model_admin):
            years = model_admin.get_queryset(request).dates("date", "year")
            return [(str(item.year), str(item.year)) for item in years]

        def queryset(self, request, queryset):
            _ = request
            value = self.value()
            if not value:
                return queryset
            return queryset.filter(date__year=value)

    class HolidayMonthFilter(SimpleListFilter):
        title = "Месяц"
        parameter_name = "holiday_month"

        def lookups(self, request, model_admin):
            _ = request
            _ = model_admin
            return [(str(month), month_abbr[month]) for month in range(1, 13)]

        def queryset(self, request, queryset):
            _ = request
            value = self.value()
            if not value:
                return queryset
            return queryset.filter(date__month=value)

    list_display = ("date", "weekday_name", "name", "is_working_day", "days_until")
    list_display_links = ("date", "name")
    list_editable = ("is_working_day",)
    list_filter = (
        HolidayPeriodFilter,
        HolidayYearFilter,
        HolidayMonthFilter,
        "is_working_day",
    )
    search_fields = ("name", "=date")
    ordering = ("date",)
    date_hierarchy = "date"
    list_per_page = 50
    save_on_top = True
    save_as = True
    actions = [
        "mark_as_working",
        "mark_as_non_working",
        "copy_selected_to_next_year",
    ]

    fieldsets = (
        (
            "Информация о празднике",
            {
                "fields": ("date", "name", "is_working_day"),
                "classes": ("wide",),
            },
        ),
        (
            "Календарный контекст",
            {
                "fields": ("weekday_name", "days_until"),
                "classes": ("wide",),
            },
        ),
    )
    readonly_fields = ("weekday_name", "days_until")

    def get_ordering(self, request):
        if request.GET.get("holiday_period") == "past":
            return ("-date",)
        return ("date",)

    def get_changeform_initial_data(self, request):
        initial: dict[str, Any] = dict(super().get_changeform_initial_data(request))
        if "date" not in initial:
            initial["date"] = timezone.localdate()
        if "is_working_day" not in initial:
            initial["is_working_day"] = False
        return initial

    def weekday_name(self, obj):
        if obj is None or not getattr(obj, "date", None):
            return "—"
        return date_format(obj.date, "l")

    weekday_name.short_description = "День недели"
    weekday_name.admin_order_field = "date"

    def days_until(self, obj):
        if obj is None or not getattr(obj, "date", None):
            return "—"
        today = timezone.localdate()
        days = (obj.date - today).days

        if days < 0:
            return format_html(
                '<span style="color:#64748b;">{} дн. назад</span>',
                abs(days),
            )
        if days == 0:
            return format_html(
                '<span style="color:#047857; font-weight:600;">Сегодня</span>'
            )
        if days <= 7:
            return format_html(
                '<span style="color:#b45309; font-weight:600;">через {} дн.</span>',
                days,
            )
        if days <= 30:
            return format_html('<span style="color:#1d4ed8;">через {} дн.</span>', days)
        return f"через {days} дн."

    days_until.short_description = "Дней до праздника"
    days_until.admin_order_field = "date"

    def mark_as_working(self, request, queryset):
        updated = queryset.update(is_working_day=True)
        self.message_user(request, f"{updated} праздников отмечены как рабочие дни.")

    mark_as_working.short_description = "Отметить как рабочие дни"

    def mark_as_non_working(self, request, queryset):
        updated = queryset.update(is_working_day=False)
        self.message_user(request, f"{updated} праздников отмечены как нерабочие дни.")

    mark_as_non_working.short_description = "Отметить как нерабочие дни"

    def copy_selected_to_next_year(self, request, queryset):
        created = 0
        skipped = 0
        for holiday in queryset:
            try:
                next_date = holiday.date.replace(year=holiday.date.year + 1)
            except ValueError:
                # 29 Feb fallback for non-leap year.
                next_date = holiday.date.replace(year=holiday.date.year + 1, day=28)

            _, is_created = PublicHoliday.objects.get_or_create(
                date=next_date,
                defaults={
                    "name": holiday.name,
                    "is_working_day": holiday.is_working_day,
                },
            )
            if is_created:
                created += 1
            else:
                skipped += 1

        if created and skipped:
            self.message_user(
                request,
                f"Скопировано в следующий год: {created}; уже существовали: {skipped}.",
                level=messages.WARNING,
            )
            return
        if created:
            self.message_user(request, f"Скопировано в следующий год: {created}.")
            return
        self.message_user(
            request,
            "Новые записи не созданы: все выбранные даты на следующий год уже существуют.",
            level=messages.INFO,
        )

    copy_selected_to_next_year.short_description = (
        "Скопировать выбранные праздники на следующий год"
    )


@admin.register(AbsentReason, site=admin_site)
class AbsentReasonAdmin(admin.ModelAdmin):
    list_display = (
        "staff",
        "reason",
        "start_date",
        "end_date",
        "duration_days",
        "approved",
        "has_document",
    )
    list_filter = (
        monitoring_utils.HierarchicalDepartmentFilter,
        "reason",
        "approved",
        "start_date",
    )
    search_fields = (
        "staff__pin",
        "staff__surname",
        "staff__name",
        "staff__department__name",
        "reason",
    )
    autocomplete_fields = ("staff",)
    readonly_fields = ("duration_days",)
    save_on_top = True
    actions = ["approve_selected", "reject_selected"]

    fieldsets = (
        (
            "Информация об отсутствии",
            {
                "fields": (
                    "staff",
                    "reason",
                    ("start_date", "end_date"),
                    "document",
                    "approved",
                ),
                "classes": ("wide",),
            },
        ),
    )

    def duration_days(self, obj):
        if obj.start_date and obj.end_date:
            days = (obj.end_date - obj.start_date).days + 1
            return days
        return "Н/Д"

    duration_days.short_description = "Продолжительность (дней)"

    def has_document(self, obj):
        return bool(obj.document)

    has_document.boolean = True
    has_document.short_description = "Документ"

    def document_preview(self, obj):
        if obj.document:
            file_url = obj.document.url
            file_name = os.path.basename(file_url)
            extension = file_name.split(".")[-1].lower()

            if extension in ["jpg", "jpeg", "png", "gif"]:
                return format_html(
                    '<img src="{}" style="max-width: 300px; max-height: 300px;" />',
                    file_url,
                )
            elif extension == "pdf":
                return format_html(
                    '<a href="{}" target="_blank" class="button">Просмотреть PDF</a>',
                    file_url,
                )
            else:
                return format_html(
                    '<a href="{}" target="_blank" class="button">Скачать файл</a>',
                    file_url,
                )
        return "Документ не прикреплен"

    document_preview.short_description = "Предпросмотр документа"

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related("staff", "staff__department")
            .prefetch_related("staff__positions")
        )

    def approve_selected(self, request, queryset):
        updated = queryset.update(approved=True)
        self.message_user(request, f"{updated} причин отсутствия одобрено.")

    approve_selected.short_description = "Одобрить выбранные причины отсутствия"

    def reject_selected(self, request, queryset):
        updated = queryset.update(approved=False)
        self.message_user(request, f"{updated} причин отсутствия отклонено.")

    reject_selected.short_description = "Отклонить выбранные причины отсутствия"


@admin.register(RemoteWork, site=admin_site)
class RemoteWorkAdmin(admin.ModelAdmin):
    list_display = (
        "staff",
        "permanent_remote",
        "start_date",
        "end_date",
        "duration_days",
        "get_remote_status",
    )
    list_filter = ("permanent_remote", "start_date", "end_date")
    search_fields = ("staff__surname", "staff__name", "staff__pin")
    autocomplete_fields = ("staff",)
    actions = ["extend_remote_work", "terminate_remote_work"]
    save_on_top = True

    fieldsets = (
        (
            "Удаленная работа",
            {
                "fields": ("staff", "permanent_remote", ("start_date", "end_date")),
                "classes": ("wide",),
            },
        ),
    )

    def duration_days(self, obj):
        if obj.permanent_remote:
            return "Постоянно"

        if obj.start_date and obj.end_date:
            days = (obj.end_date - obj.start_date).days + 1
            return f"{days} дн."
        return "Н/Д"

    duration_days.short_description = "Продолжительность"

    def get_remote_status(self, obj):
        status = obj.get_remote_status()

        if "Постоянно" in status:
            return format_html('<span style="color: green;">{}</span>', status)
        elif "Активно" in status:
            return format_html('<span style="color: blue;">{}</span>', status)
        elif "Завершено" in status:
            return format_html('<span style="color: gray;">{}</span>', status)
        elif "Ожидает" in status:
            return format_html('<span style="color: orange;">{}</span>', status)

        return status

    get_remote_status.short_description = "Статус удаленной работы"

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related("staff", "staff__department")
            .prefetch_related("staff__positions")
        )

    def save_model(self, request, obj, form, change):
        try:
            obj.full_clean()
            super().save_model(request, obj, form, change)
        except ValidationError as e:
            form.add_error(None, e)

    def extend_remote_work(self, request, queryset):
        count = queryset.count()
        self.message_user(
            request, f"Период удаленной работы продлен для {count} сотрудников."
        )

    extend_remote_work.short_description = "Продлить период удаленной работы"

    def terminate_remote_work(self, request, queryset):
        today = timezone.now().date()
        count = queryset.filter(end_date__gt=today).update(end_date=today)
        self.message_user(
            request, f"Удаленная работа завершена для {count} сотрудников."
        )

    terminate_remote_work.short_description = "Завершить удаленную работу"


@admin.register(PerformanceBonusRule, site=admin_site)
class PerformanceBonusRuleAdmin(admin.ModelAdmin):
    list_display = (
        "min_days",
        "max_days",
        "min_attendance_percent",
        "max_attendance_percent",
        "bonus_percentage",
        "rule_description",
    )
    list_filter = (
        "min_days",
        "max_days",
        "min_attendance_percent",
        "max_attendance_percent",
    )
    search_fields = ("bonus_percentage", "min_days", "max_days")
    ordering = ("min_days", "max_days", "min_attendance_percent")
    save_on_top = True

    fieldsets = (
        (
            "Критерии бонуса",
            {
                "fields": (
                    ("min_days", "max_days"),
                    ("min_attendance_percent", "max_attendance_percent"),
                    "bonus_percentage",
                ),
                "classes": ("wide",),
            },
        ),
    )

    def rule_description(self, obj):
        return f"За {obj.min_days}-{obj.max_days} дней с посещаемостью {obj.min_attendance_percent}%-{obj.max_attendance_percent}% бонус {obj.bonus_percentage}%"

    rule_description.short_description = "Описание правила"


standard_registry = getattr(admin.site, "_registry", {})
custom_registry = getattr(admin_site, "_registry", {})
for model, admin_class in standard_registry.items():
    if model not in custom_registry:
        admin_site.register(model, type(admin_class))

admin.site = admin_site
