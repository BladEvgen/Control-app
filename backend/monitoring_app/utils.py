import datetime
import hashlib
import json
import logging
import math
import os
import re
from collections import Counter, defaultdict
from difflib import get_close_matches
from functools import lru_cache
from typing import (
    Any,
    Dict,
    Iterable,
    List,
    Mapping,
    Sequence,
    Set,
    Tuple,
    TypedDict,
    cast,
)

import numpy as np
import pandas as pd
import pytz
from cryptography.fernet import Fernet
from django.conf import settings
from django.contrib.admin import SimpleListFilter
from django.core.mail import send_mail
from django.db.models import Count
from django.urls import reverse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from monitoring_app import models
from monitoring_app.cache_conf import get_cache
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sklearn.neighbors import BallTree, KDTree

DAYS = settings.DAYS

logger = logging.getLogger("django")


def get_lesson_attendance_photo_path(staff_pin: str):
    """
    Возвращает (base_dir, full_file_path) для сохранения фото посещаемости.
    base_dir нужно создать (makedirs); full_file_path — куда писать файл.
    """
    date_path = timezone.now().strftime("%Y-%m-%d")
    timestamp = int(timezone.now().timestamp())
    if settings.DEBUG:
        base_dir = os.path.join(
            settings.MEDIA_ROOT, "control_image", staff_pin, date_path
        )
    else:
        base_dir = os.path.join(settings.ATTENDANCE_ROOT, staff_pin, date_path)
    filename = f"{staff_pin}_{timestamp}.jpg"
    return base_dir, os.path.join(base_dir, filename)


def merge_work_intervals_to_total_seconds(
    intervals: List[Tuple[datetime.datetime, datetime.datetime]],
) -> int:
    """Объединяет перекрывающиеся интервалы и возвращает суммарную длительность в секундах.

    Интервалы сортируются по началу; пересекающиеся или смежные объединяются в один.
    Подходит для расчёта эффективного времени по SA и LA без двойного учёта.

    Args:
        intervals: Список кортежей (start, end) — timezone-aware datetime.

    Returns:
        Сумма длительностей объединённых интервалов в секундах (int). 0 при пустом списке.
    """
    if not intervals:
        return 0
    sorted_intervals = sorted(intervals, key=lambda x: (x[0], x[1]))
    merged: List[Tuple[datetime.datetime, datetime.datetime]] = [
        (sorted_intervals[0][0], sorted_intervals[0][1])
    ]
    for start, end in sorted_intervals[1:]:
        if start <= merged[-1][1]:
            merged[-1] = (merged[-1][0], max(merged[-1][1], end))
        else:
            merged.append((start, end))
    return sum(int((e - s).total_seconds()) for s, e in merged)


ARCFACE_MODEL = None

CANONICAL_ADDRESSES: dict[str, str] = {
    "abilai": "Проспект Абылай хана, 51/53",
    "torekulova": "Улица Торекулова, 71",
    "karasai": "Улица Карасай батыра, 75",
}
ATTENDANCE_BUILDING_CODES = ("abilai", "karasai", "torekulova")
ATTENDANCE_BUILDING_LABELS: dict[str, str] = {
    "abilai": "Абылай-хана",
    "karasai": "Карасай",
    "torekulova": "Торекулова",
}

_PUNCT = re.compile(r"[\"'’`.,:;!?\(\)\[\]{}_/\\]+")
_WS = re.compile(r"\s+")


def _norm(s: str) -> str:
    if not s:
        return ""
    s = s.strip().lower()
    s = s.replace("ё", "е")
    s = _PUNCT.sub(" ", s)
    s = _WS.sub(" ", s)
    return s


ALIASES: dict[str, str] = {
    "abilai": "abilai",
    "абылайхана": "abilai",
    "абылай хана": "abilai",
    "абылай-хана": "abilai",
    "абылай хана турникет": "abilai",
    "абылай-хана турникет": "abilai",
    "абылай": "abilai",
    "абылайхана турникет": "abilai",
    "вход абылайхана": "abilai",
    "цос": "abilai",
    "выход цос": "abilai",
    "военные 3 этаж": "abilai",
    "лифтовые с 1 по 7": "abilai",
    "torekulova": "torekulova",
    "торекулова": "torekulova",
    "торекулова турникет": "torekulova",
    "торекулов": "torekulova",
    "торекулва": "torekulova",
    "торекулва турникет": "torekulova",
    "karasai": "karasai",
    "карасай": "karasai",
    "карасай батыр": "karasai",
    "карасай батыра": "karasai",
    "карасай-батыра": "karasai",
    "карасай батыра турникет": "karasai",
    "карасай-батыра турникет": "karasai",
}

_RX_ABILAI = re.compile(
    r"(абылай[\s\-]*хана|абылайхана|цос|военные|\bвход\b|\bвыход\b|\bлифт\w*)",
    re.IGNORECASE,
)

_RX_TOREKULOVA = re.compile(r"торекулов\w*", re.IGNORECASE)

_RX_KARASAI = re.compile(r"карасай", re.IGNORECASE)

_RX_LIFT = re.compile(r"\bлифт[\s\-]*\d+\b|\bлифты?\b|\bлифт\w*\b", re.IGNORECASE)


def is_lift_terminal(area_name: str | None) -> bool:
    """Возвращает True, если зона/терминал — лифт (лифт 1, лифт 8, лифтовые и т.п.)."""
    if not area_name or not area_name.strip():
        return False
    return bool(_RX_LIFT.search(area_name.strip()))


@lru_cache(maxsize=4096)
def resolve_area_family(area_name: str | None) -> str | None:
    """Возвращает внутренний ключ корпуса: abilai, torekulova или karasai."""
    if not area_name:
        return None

    n = _norm(area_name)

    if n in ALIASES:
        return ALIASES[n]

    if _RX_TOREKULOVA.search(n):
        return "torekulova"
    if _RX_KARASAI.search(n):
        return "karasai"
    if _RX_ABILAI.search(n) or _RX_LIFT.search(n):
        return "abilai"

    return _fuzzy_family(n)


@lru_cache(maxsize=8192)
def pin_to_external_format(pin: str | None) -> str:
    """Приводит PIN к формату внешней системы (убирает обёртку S и T).

    Используется для сопоставления сотрудников с системой оценок, где PIN
    хранятся без префикса/суффикса (например, 9614 вместо S9614S).
    Результат кэшируется (lru_cache) для повторных вызовов с тем же pin.

    Args:
        pin: PIN сотрудника в формате системы контроля (может быть None).

    Returns:
        PIN без обёртки: S9614S → 9614, T861T → 861. Если обёртки нет
        или pin пустой — возвращается исходная строка или пустая строка.
    """
    if not pin:
        return ""
    if len(pin) >= 2 and pin[0] == "S" and pin[-1] == "S":
        return pin[1:-1]
    if len(pin) >= 2 and pin[0] == "T" and pin[-1] == "T":
        return pin[1:-1]
    return pin


KEYWORDS = {
    "abilai": ["абылай", "абылайхана", "цос", "военные", "вход", "выход", "лифт"],
    "torekulova": ["торекулова", "торекулов", "турекулова", "торекулв"],
    "karasai": ["карасай", "карасайбатыра", "карасай батыр"],
}


def _fuzzy_family(n: str) -> str | None:
    tokens = n.split()
    joined = " ".join(tokens)
    for fam, hints in KEYWORDS.items():
        if get_close_matches(joined, hints, n=1, cutoff=0.86):
            return fam
        if any(get_close_matches(t, hints, n=1, cutoff=0.88) for t in tokens):
            return fam
    return None


@lru_cache(maxsize=4096)
def resolve_area_address(area_name: str | None) -> str | None:
    """Преобразует произвольное имя зоны к каноническому адресу.

    Преобразует произвольное имя зоны (любой регистр/формат) к одному из:
    - 'Проспект Абылай хана, 51/53'
    - 'Улица Торекулова, 71'
    - 'Улица Карасай батыра, 75'

    Args:
        area_name (str | None): Имя зоны для преобразования.

    Returns:
        str | None: Канонический адрес или None, если распознать нельзя.
    """
    family = resolve_area_family(area_name)
    return CANONICAL_ADDRESSES[family] if family else None


def get_client_ip(request):
    """Получает IP адрес клиента из запроса с учетом прокси.

    Args:
        request: HTTP запрос Django.

    Returns:
        str: IP адрес клиента.
    """
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        ip = x_forwarded_for.split(",")[0].strip()
    else:
        ip = request.META.get("REMOTE_ADDR")
    return ip


def format_duration(duration_seconds):
    """Преобразует длительность в секундах в читаемый формат.

    Args:
        duration_seconds (float): Длительность в секундах.

    Returns:
        str: Отформатированная строка с длительностью (секунды, минуты, часы).
    """
    if duration_seconds < 60:
        return f"{duration_seconds:.2f} seconds"
    elif duration_seconds < 3600:
        minutes = duration_seconds // 60
        seconds = duration_seconds % 60
        return f"{minutes:.0f} minute(s) {seconds:.2f} seconds"
    else:
        hours = duration_seconds // 3600
        minutes = (duration_seconds % 3600) // 60
        seconds = duration_seconds % 60
        return f"{hours:.0f} hour(s) {minutes:.0f} minute(s) {seconds:.2f} seconds"


class HierarchicalDepartmentFilter(SimpleListFilter):
    """Иерархия от головного отдела к листьям; dept_name_q сужает список (GET-параметр)."""

    title = _("Department")
    parameter_name = "staff_department"
    _lookups_cache_ttl = 600

    def _lookups_cache_key(self, request) -> str:
        from hashlib import md5

        q = (request.GET.get("dept_name_q") or "").strip()
        if not q:
            return "hierarchical_dept_filter_tree_v2"
        h = md5(q.encode("utf-8")).hexdigest()[:16]
        return f"hierarchical_dept_filter_tree_v2_q_{h}"

    def lookups(self, request, model_admin):
        from django.core.cache import cache

        cache_key = self._lookups_cache_key(request)
        lookup_list = cache.get(cache_key)
        if lookup_list is not None:
            return lookup_list

        base = models.ChildDepartment.objects.only("id", "name", "parent_id")
        dept_q = (request.GET.get("dept_name_q") or "").strip()
        if dept_q:
            match_ids = set(
                base.filter(name__icontains=dept_q).values_list("id", flat=True)
            )
            if not match_ids:
                lookup_list = []
                cache.set(cache_key, lookup_list, self._lookups_cache_ttl)
                return lookup_list
            parent_by_id = dict(
                models.ChildDepartment.objects.values_list("id", "parent_id")
            )
            needed = set(match_ids)
            for mid in list(match_ids):
                pid = parent_by_id.get(mid)
                while pid:
                    needed.add(pid)
                    pid = parent_by_id.get(pid)
            qs = base.filter(id__in=needed)
        else:
            qs = base

        by_parent = defaultdict(list)
        for d in qs:
            by_parent[d.parent_id].append(d)
        for children in by_parent.values():
            children.sort(key=lambda x: (x.name or "").lower())

        lookup_list = self._walk_tree(by_parent, None, 0)
        cache.set(cache_key, lookup_list, self._lookups_cache_ttl)
        return lookup_list

    def _walk_tree(self, by_parent, parent_id, level: int):
        choices = []
        for dept in by_parent.get(parent_id, []):
            indent = "—" * level
            choices.append((dept.id, f"{indent} {dept.name}"))
            choices.extend(self._walk_tree(by_parent, dept.id, level + 1))
        return choices

    def queryset(self, request, queryset):
        if self.value():
            cache_key = f"dept_descendants_{self.value()}"
            from django.core.cache import cache

            descendant_ids = cache.get(cache_key)
            if descendant_ids is None:
                try:
                    department = models.ChildDepartment.objects.only("id").get(
                        pk=self.value()
                    )
                except models.ChildDepartment.DoesNotExist:
                    return queryset
                descendant_ids = self.get_all_descendant_ids(department.id)
                cache.set(cache_key, descendant_ids, self._lookups_cache_ttl)
            return queryset.filter(staff__department__in=descendant_ids)
        return queryset

    def get_all_descendant_ids(self, department_id):
        descendant_ids = {department_id}
        queue = [department_id]
        while queue:
            current_id = queue.pop(0)
            children = list(
                models.ChildDepartment.objects.filter(parent_id=current_id).values_list(
                    "id", flat=True
                )
            )
            queue.extend(children)
            descendant_ids.update(children)
        return descendant_ids


class APIKeyUtility:
    _secret_key = None

    @staticmethod
    def get_secret_key():
        if APIKeyUtility._secret_key is None:
            if not settings.SECRET_API:
                secret_key = Fernet.generate_key().decode("utf-8")

                dotenv = settings.DOTENV_PATH

                with open(dotenv, mode="ab") as f:
                    f.write(f"""\nSECRET_API={secret_key}\n""".encode("utf-8"))

                APIKeyUtility._secret_key = secret_key
            else:
                APIKeyUtility._secret_key = settings.SECRET_API

        return APIKeyUtility._secret_key

    @staticmethod
    def encrypt_data(data, secret_key):
        f = Fernet(secret_key.encode())
        encrypted_data = f.encrypt(json.dumps(data).encode())
        return encrypted_data.decode()

    @staticmethod
    def decrypt_data(encrypted_data, secret_key, fields=("is_active",)):
        f = Fernet(secret_key.encode())
        decrypted_data = f.decrypt(encrypted_data.encode())
        data = json.loads(decrypted_data.decode())

        return {
            field: (
                data.get("is_activate")
                if field == "is_active" and "is_active" not in data
                else data.get(field)
            )
            for field in fields
        }

    @staticmethod
    def generate_api_key(key_name, created_by):
        secret_key = APIKeyUtility.get_secret_key()
        data = {
            "key_name": key_name,
            "created_by": created_by.username,
            "created_at": timezone.now().isoformat(),
            "is_active": True,
        }
        encrypted_data = APIKeyUtility.encrypt_data(data, secret_key)
        return encrypted_data, secret_key


def password_check(password: str) -> bool:
    """
    Проверяет, соответствует ли пароль требованиям сложности системы.

    Эта функция проверяет строку пароля на основе следующих критериев:

        - Минимальная длина 8 символов.
        - Содержит хотя бы одну заглавную букву (A-Z)
        - Содержит хотя бы одну строчную букву (a-z)
        - Содержит хотя бы одну цифру (0-9)
        - Содержит хотя бы один специальный символ из следующего набора: #?!@$%^&*-

    Args:
        пароль (str): строка пароля, которую необходимо проверить.

    Returns:
        bool: True, если пароль соответствует всем требованиям сложности, в противном случае — False
    """
    return bool(
        re.match(
            r"^(?=.*?[A-Z])(?=.*?[a-z])(?=.*?[0-9])(?=.*?[#?!@$%^&*-]).{8,}$", password
        )
    )


def fetch_data(url: str) -> Dict[str, Any]:
    import requests

    try:
        response = requests.get(url, timeout=20)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        logger.error(f"Error fetching data: {e}")
        return {}


def send_password_reset_email(user, request):
    """
    Send a password reset email to the user with branding and design consistent with the website footer.

    Args:
        user: The user object who requested password reset
        request: The request object to build the reset URL

    Returns:
        bool: True if email was sent successfully, False otherwise
    """
    try:
        token = models.PasswordResetToken.generate_token(user)

        site_name = getattr(settings, "SITE_NAME", "KRMU")

        reset_scheme = "https" if request.is_secure() else request.scheme
        reset_link = f"{reset_scheme}://{request.get_host()}{reverse('password_reset_confirm', args=[token])}"

        expiry_time = timezone.now() + datetime.timedelta(hours=1)
        expiry_time_str = expiry_time.strftime("%H:%M %d.%m.%Y")

        current_year = timezone.now().year

        user_display_name = getattr(user, "first_name", user.username) or user.username

        subject = f"Сброс пароля на сайте {site_name}"

        html_message = format_html(
            """
            <!DOCTYPE html>
            <html lang="ru">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>Сброс пароля</title>
            </head>
            <body
                style="
                    margin: 0;
                    padding: 0;
                    font-family: 'Segoe UI', 'Roboto', 'Helvetica Neue', 'Arial', sans-serif;
                    background-color: #f7f9fc;
                    color: #333333;
                "
            >
                <div
                    style="
                        max-width: 600px;
                        margin: 20px auto;
                        padding: 30px;
                        border-radius: 12px;
                        background-color: #ffffff;
                        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.05);
                    "
                >
                    <div style="text-align: center; margin-bottom: 30px;">
                        <h1 style="color: #2563EB; font-size: 24px; margin: 0 0 5px 0;">Сброс пароля</h1>
                        <p style="color: #6B7280; font-size: 16px; margin: 0;">Инструкция по восстановлению доступа</p>
                    </div>

                    <div
                        style="padding: 20px; background-color: #F3F4F6; border-radius: 8px; margin-bottom: 25px;"
                    >
                        <p style="color: #4B5563; line-height: 1.6; margin: 0 0 15px 0;">
                            Здравствуйте, <strong>{user_name}</strong>!
                        </p>
                        <p style="color: #4B5563; line-height: 1.6; margin: 0 0 15px 0;">
                            Мы получили запрос на сброс пароля для вашего аккаунта.
                            Если это были вы, используйте кнопку ниже для
                            создания нового пароля.
                        </p>
                    </div>

                    <div style="text-align: center; margin-bottom: 30px;">
                        <a
                            href="{reset_link}"
                            style="
                                display: inline-block;
                                padding: 14px 32px;
                                color: #ffffff;
                                background-color: #2563EB;
                                text-decoration: none;
                                border-radius: 6px;
                                font-size: 16px;
                                font-weight: 600;
                                transition: background-color 0.2s ease;
                            "
                        >
                            Сбросить пароль
                        </a>
                    </div>

                    <div
                        style="
                            border-left: 4px solid #FCD34D;
                            padding: 12px 15px;
                            background-color: #FFFBEB;
                            margin-bottom: 25px;
                            border-radius: 0 6px 6px 0;
                        "
                    >
                        <p style="color: #92400E; font-size: 14px; line-height: 1.5; margin: 0;">
                            <strong>Важно:</strong> Ссылка действительна до <strong>{expiry_time}</strong>.<br>
                            Если вы не запрашивали сброс пароля, пожалуйста,
                            игнорируйте это письмо или обратитесь в службу
                            поддержки.
                        </p>
                    </div>

                    <div
                        style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #E5E7EB; text-align: center;"
                    >
                        <div style="display: inline-block; margin: 0 15px 15px 0;">
                            <a href="{site_url}" style="color: #4B5563; text-decoration: none; font-size: 14px;">
                                Home
                            </a>
                        </div>
                        <div style="display: inline-block; margin: 0 15px 15px 0;">
                            <a href="https://new.krmu.edu.kz" style="color: #4B5563; text-decoration: none; font-size: 14px;">
                                KRMU
                            </a>
                        </div>
                        <div style="display: inline-block; margin: 0 0 15px 0;">
                            <a
                                href="https://new.krmu.edu.kz/О_нас/Об_университете/"
                                style="color: #4B5563; text-decoration: none; font-size: 14px;"
                            >
                                About Us
                            </a>
                        </div>
                    </div>
                </div>
            </body>
            </html>
            """,
            user_name=user_display_name,
            reset_link=reset_link,
            expiry_time=expiry_time_str,
            site_name=site_name,
            current_year=current_year,
            site_url=f"{reset_scheme}://{request.get_host()}/",
        )

        plain_message = (
            f"Сброс пароля на сайте {site_name}\n"
            f"=============================================\n\n"
            f"Здравствуйте, {user_display_name}!\n\n"
            "Мы получили запрос на сброс пароля для вашего аккаунта.\n\n"
            f"Для создания нового пароля перейдите по следующей ссылке:\n{reset_link}\n\n"
            f"Важно: Ссылка действительна до {expiry_time_str}.\n\n"
            "Если вы не запрашивали сброс пароля, пожалуйста, игнорируйте это письмо\n"
            "или обратитесь в службу поддержки.\n\n"
            "---\n"
            "Home: https://krmu.edu.kz\n"
            "KRMU: https://new.krmu.edu.kz\n"
            "About Us: https://new.krmu.edu.kz/О_нас/Об_университете/\n\n"
            f"© {current_year} {site_name}, Inc."
        )

        success = send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )

        if success:
            logger.info(f"Password reset email sent for user ID: {user.id}")

            if hasattr(models, "SecurityAuditLog"):
                models.SecurityAuditLog.objects.create(
                    user=user,
                    action_type="password_reset_request",
                    ip_address=request.META.get("REMOTE_ADDR", ""),
                    user_agent=request.META.get("HTTP_USER_AGENT", ""),
                )

            return True
        else:
            logger.error(f"Failed to send password reset email to user ID: {user.id}")
            return False

    except Exception as e:
        logger.error(f"Failed to send password reset email: {str(e)}")
        return False


def get_user_timezone(request):
    user_timezone = request.session.get("timezone")
    if not user_timezone:
        user_timezone = settings.TIME_ZONE
    return pytz.timezone(user_timezone)


def normalize_id(department_id):
    """
    Нормализует ID отдела, удаляя ведущие нули, если ID состоит только из цифр.
    Если ID содержит буквы, он остаётся без изменений.

    Args:
        department_id (str): ID отдела.

    Returns:
        str: Нормализованный ID.
    """
    if department_id.isdigit():
        return str(int(department_id))
    return department_id


def transliterate(name):
    slovar = {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "yo",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "ts",
        "ч": "ch",
        "ш": "sh",
        "щ": "sch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
        " ": " ",
        "-": "-",
        ".": ".",
        ",": ",",
        "!": "!",
        "?": "?",
        ":": ":",
    }

    name = name.lower()
    translit = []
    for letter in name:
        translit.append(slovar.get(letter, letter))

    return "".join(translit)


def clean_address(address):
    """
    Очищает адрес, удаляя префиксы ('Улица', 'Проспект', и т.д.),
    скрытые символы и нормализует пробелы.

    Args:
        address (str): Исходный адрес.

    Returns:
        str: Очищенный и нормализованный адрес.
    """
    if address:
        address = address.replace("\u200b", "")
        address = re.sub(
            r"^(улица|проспект|переулок|бульвар|территория|микрорайон)\s+",
            "",
            address,
            flags=re.IGNORECASE,
        )
        address = re.sub(r"\s+", " ", address).strip().lower()
        return address
    return address


def generate_map_data(
    locations, date_at, search_staff_attendance=True, filter_empty=False
):
    """
    Генерирует данные по локациям, включая посещения сотрудников и занятия.

    Args:
        locations (QuerySet): Локации из модели ClassLocation.
        date_at (date): Дата для фильтрации данных (дата события).
        search_staff_attendance (bool): Если True, включает данные из StaffAttendance и LessonAttendance.
        filter_empty (bool): Если True, исключает локации с нулевым количеством посещений.

    Returns:
        list: Список словарей с данными по локациям, готовых для отображения на карте.
    """
    staff_by_address = defaultdict(int)
    lesson_attendance_by_address = defaultdict(int)

    if search_staff_attendance:
        try:
            data_insert_date = date_at + datetime.timedelta(days=1)
            logger.info(
                f"Начинаем обработку StaffAttendance для даты вставки: {data_insert_date} (дата события: {date_at})"
            )

            staff_attendances = (
                models.StaffAttendance.objects.filter(
                    date_at=data_insert_date, first_in__isnull=False
                )
                .values("area_name_in")
                .annotate(count=Count("id"))
            )

            if not staff_attendances:
                logger.warning(
                    f"Нет записей StaffAttendance для даты вставки {data_insert_date}"
                )
            else:
                for attendance in staff_attendances:
                    area_name_in = attendance.get("area_name_in")
                    if not area_name_in:
                        logger.warning(
                            f"Найдена запись StaffAttendance без area_name_in для даты вставки {data_insert_date}"
                        )
                        continue
                    address = resolve_area_address(area_name_in)
                    if address:
                        comparison_address = clean_address(address)
                        matched_location = next(
                            (
                                loc
                                for loc in locations
                                if clean_address(loc.address) == comparison_address
                            ),
                            None,
                        )
                        if matched_location:
                            original_address = matched_location.address.strip()
                            staff_by_address[original_address] += attendance.get(
                                "count", 0
                            )
                        else:
                            logger.warning(
                                f"Оригинальный адрес для '{area_name_in}' не найден в ClassLocation"
                            )
                    else:
                        logger.warning(
                            f"Название зоны '{area_name_in}' не найдено в AREA_ADDRESS_MAPPING"
                        )

                logger.info(f"Обработано StaffAttendance: {dict(staff_by_address)}")

            staff_with_attendance_qs = models.StaffAttendance.objects.filter(
                date_at=data_insert_date, first_in__isnull=False
            ).values_list("staff_id", flat=True)
            staff_with_attendance = list(staff_with_attendance_qs)
            staff_count = len(staff_with_attendance)
            logger.info(f"Количество сотрудников с посещением: {staff_count}")

            logger.info(f"Начинаем обработку LessonAttendance для даты: {date_at}")

            lesson_attendances_qs = models.LessonAttendance.exclude_report_invalid_days(
                models.LessonAttendance.objects.filter(date_at=date_at)
            ).exclude(staff_id__in=staff_with_attendance)

            lesson_count = lesson_attendances_qs.count()
            logger.info(f"Количество LessonAttendance для обработки: {lesson_count}")

            if lesson_count > 0:
                lesson_attendances_list = list(
                    lesson_attendances_qs.values_list(
                        "id", "latitude", "longitude", flat=False
                    )
                )

                class_locations = list(
                    models.ClassLocation.objects.only(
                        "id", "name", "latitude", "longitude"
                    )
                )
                if not class_locations:
                    logger.warning("Нет записей ClassLocation.")
                    return []

                class_coords = [
                    (loc.latitude, loc.longitude) for loc in class_locations
                ]

                kd_tree = KDTree(class_coords, metric="euclidean")
                logger.info("KDTree успешно построен.")

                nearest_addresses = []
                for lesson_id, lesson_lat, lesson_lon in lesson_attendances_list:
                    if lesson_lat is None or lesson_lon is None:
                        logger.warning(
                            f"LessonAttendance {lesson_id} не имеет координат"
                        )
                        continue

                    k_candidates = min(5, len(class_locations))
                    _distances_degrees, candidate_indices = kd_tree.query(
                        [[lesson_lat, lesson_lon]], k=k_candidates
                    )

                    candidate_list = []
                    if hasattr(candidate_indices, "flatten"):
                        candidate_list = candidate_indices.flatten().tolist()
                    elif (
                        hasattr(candidate_indices, "__len__")
                        and len(candidate_indices) > 0
                    ):
                        if (
                            hasattr(candidate_indices[0], "__len__")
                            and len(candidate_indices[0]) > 0
                        ):
                            candidate_list = [int(idx) for idx in candidate_indices[0]]
                        else:
                            candidate_list = [int(candidate_indices[0])]

                    nearest_location = None
                    min_distance = float("inf")
                    for idx in candidate_list:
                        if 0 <= idx < len(class_locations):
                            candidate_loc = class_locations[idx]
                            distance = calculate_distance_haversine(
                                lesson_lat,
                                lesson_lon,
                                candidate_loc.latitude,
                                candidate_loc.longitude,
                            )
                            if distance < min_distance:
                                min_distance = distance
                                nearest_location = candidate_loc

                    if nearest_location is None:
                        for loc in class_locations:
                            distance = calculate_distance_haversine(
                                lesson_lat,
                                lesson_lon,
                                loc.latitude,
                                loc.longitude,
                            )
                            if distance < min_distance:
                                min_distance = distance
                                nearest_location = loc

                    if nearest_location:
                        nearest_addresses.append(nearest_location.address.strip())
                    else:
                        logger.warning(
                            f"Не найдена ближайшая локация для LessonAttendance {lesson_id}"
                        )

                logger.info("KDTree запрос с точным расчетом завершен.")

                address_counts = Counter(nearest_addresses)
                lesson_attendance_by_address = defaultdict(int, address_counts)

                logger.info(
                    f"Обработано LessonAttendance: {dict(lesson_attendance_by_address)}"
                )
            else:
                logger.info("Нет записей LessonAttendance для обработки.")
        except Exception as e:
            logger.error(
                f"Ошибка при обработке данных посещений: {str(e)}", exc_info=True
            )
            raise

    try:
        aggregated_data = defaultdict(int)
        for address, count in staff_by_address.items():
            aggregated_data[address] += count
        for address, count in lesson_attendance_by_address.items():
            aggregated_data[address] += count
        logger.info(f"Агрегированные данные: {dict(aggregated_data)}")
    except Exception as e:
        logger.error(f"Ошибка при агрегации данных: {str(e)}", exc_info=True)
        raise

    result_list = []
    try:
        for loc in locations:
            original_address = loc.address.strip()
            location_data = {
                "name": loc.name,
                "address": original_address,
                "lat": loc.latitude,
                "lng": loc.longitude,
            }
            if search_staff_attendance:
                employees_count = aggregated_data.get(original_address, 0)
                if employees_count > 0:
                    location_data["employees"] = employees_count
                    if filter_empty and employees_count <= 1:
                        continue
                else:
                    continue
            result_list.append(location_data)
        logger.info(f"Сформирован список результатов с {len(result_list)} локациями.")
    except Exception as e:
        logger.error(
            f"Ошибка при формировании списка результатов: {str(e)}", exc_info=True
        )
        raise

    try:
        main_location = next(
            (
                item
                for item in result_list
                if re.search(
                    r"абылай\s*хана",
                    re.sub(r"[\"\'.,]", "", item["address"].lower()).strip(),
                )
            ),
            None,
        )

        if main_location:
            result_list.remove(main_location)
            result_list.insert(0, main_location)
            logger.info("Основная локация перемещена в начало списка.")
        else:
            result_list.sort(key=lambda x: x["name"])
            logger.info("Основная локация не найдена. Список отсортирован по имени.")
    except Exception as e:
        logger.error(
            f"Ошибка при сортировке списка результатов: {str(e)}", exc_info=True
        )

    return result_list


class LocationSearcher:
    """Класс для поиска ближайших локаций с использованием KDTree.

    Использует KDTree для быстрого поиска кандидатов и формулу Haversine
    для точного расчета расстояния.
    """

    def __init__(self, locations):
        """Инициализирует LocationSearcher со списком локаций.

        Args:
            locations (list): Список словарей с ключами `latitude`, `longitude`, `name`.
        """
        self.locations = locations
        self.location_coords = np.asarray(
            [(float(loc["latitude"]), float(loc["longitude"])) for loc in locations],
            dtype=float,
        )
        self.kd_tree = KDTree(self.location_coords, metric="euclidean")
        self.names = [loc["name"] for loc in locations]

    def _pick_nearest_candidate(
        self,
        lat: float,
        lon: float,
        candidate_indices,
        *,
        radius: float = 200,
    ):
        """Возвращает ближайшую локацию по индексам кандидатов.

        Args:
            lat: Широта искомой точки.
            lon: Долгота искомой точки.
            candidate_indices: Индексы кандидатов из KDTree.
            radius: Радиус поиска в метрах.

        Returns:
            dict | None: Payload ближайшей локации.

        Notes:
            Complexity: O(k), где ``k`` — число кандидатов в радиусе KDTree.
        """
        nearest_candidate = None
        min_distance = float("inf")

        for idx in candidate_indices:
            idx_int = int(idx)
            if 0 <= idx_int < len(self.locations):
                candidate = self.locations[idx_int]
                distance = calculate_distance_haversine(
                    lat,
                    lon,
                    float(candidate["latitude"]),
                    float(candidate["longitude"]),
                )
                if distance < min_distance and distance <= radius:
                    min_distance = distance
                    nearest_candidate = candidate

        return nearest_candidate

    def _find_nearest_candidate(self, lat, lon, radius=200):
        """Возвращает ближайшую локацию payload в заданном радиусе."""
        if lat is None or lon is None:
            return None
        if not self.locations:
            return None

        meters_to_degrees = radius / 111000
        candidate_indices = self.kd_tree.query_radius(
            [[lat, lon]], r=meters_to_degrees
        )[0]

        if len(candidate_indices) == 0:
            return None

        return self._pick_nearest_candidate(
            float(lat),
            float(lon),
            candidate_indices,
            radius=radius,
        )

    def find_nearest(self, lat, lon, radius=200):
        """Находит ближайшую локацию в заданном радиусе с точным расчетом расстояния.

        Использует KDTree для быстрого поиска кандидатов, затем пересчитывает
        точное расстояние через формулу Haversine для выбора ближайшей локации.

        Args:
            lat (float): Широта искомой точки.
            lon (float): Долгота искомой точки.
            radius (float): Радиус поиска в метрах.

        Returns:
            str: Название ближайшей локации или "Unknown Area".
        """
        nearest_candidate = self._find_nearest_candidate(lat, lon, radius=radius)
        if nearest_candidate is None:
            return "Unknown Area"
        return str(nearest_candidate.get("name") or "Unknown Area")

    def find_nearest_location(self, lat, lon, radius=200):
        """Возвращает payload ближайшей локации или None."""
        return self._find_nearest_candidate(lat, lon, radius=radius)

    def find_nearest_locations_bulk(
        self,
        coordinates: Sequence[tuple[float, float]],
        *,
        radius: float = 200,
    ) -> list[dict[str, Any] | None]:
        """Находит ближайшие локации сразу для набора координат.

        Args:
            coordinates: Список координат ``(lat, lon)``.
            radius: Радиус поиска в метрах.

        Returns:
            Список payload-локаций в том же порядке, что и ``coordinates``.

        Notes:
            Complexity: O(n log L + e), где ``n`` — число точек,
            ``L`` — число локаций, ``e`` — число candidate edges.
        """
        if not coordinates or not self.locations:
            return []

        meters_to_degrees = radius / 111000
        coords_array = np.asarray(coordinates, dtype=float)
        candidate_indices_list = self.kd_tree.query_radius(
            coords_array,
            r=meters_to_degrees,
        )

        return [
            self._pick_nearest_candidate(
                float(lat),
                float(lon),
                candidate_indices,
                radius=radius,
            )
            for (lat, lon), candidate_indices in zip(
                coordinates,
                candidate_indices_list,
            )
        ]


R_EARTH_M = 6_371_000


def calculate_distance_haversine(lat1, lon1, lat2, lon2):
    """Расстояние между двумя точками по формуле Haversine (большой круг на сфере).

    Подходит для WGS84 (lat/lon) при d < 1 км. Точность для смартфона (5–15 м) достаточна.

    Args:
        lat1, lon1: широта и долгота первой точки, градусы.
        lat2, lon2: широта и долгота второй точки, градусы.

    Returns:
        float: расстояние в метрах.
    """
    earth_radius_m = R_EARTH_M
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = (
        math.sin(delta_phi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    distance = earth_radius_m * c
    return distance


def _convert_to_local_with_tz(
    dt: datetime.date | datetime.datetime | None,
    local_tz: datetime.tzinfo,
) -> datetime.datetime | None:
    """Convert date/datetime into local timezone using a pre-fetched tz.

    Args:
        dt: Source date or datetime.
        local_tz: Target timezone object.

    Returns:
        Localized datetime or ``None`` when input is ``None``.

    Notes:
        Complexity: O(1).
    """
    if dt is None:
        return None
    if isinstance(dt, datetime.date) and not isinstance(dt, datetime.datetime):
        dt = datetime.datetime.combine(dt, datetime.time(0, 0, 0))
        dt = timezone.make_aware(dt, local_tz)
    return timezone.localtime(dt, local_tz)


def compute_class_location_acceptance_radii(
    locations,
    r_same_point=60,
    r_cluster=80,
    r_standalone=65,
    same_point_threshold=5,
    cluster_threshold=30,
):
    """Приёмный радиус R_loc (м) для каждой локации. Логика:

    1) Если у объекта задано acceptance_radius_m (БД) и > 0 — использовать его.
    2) Иначе — по соседству (min_d до ближайшей другой локации, Haversine):
       — min_d < same_point_threshold: одна точка, несколько организаций → r_same_point
       — same_point_threshold ≤ min_d < cluster_threshold: кластер/двор → r_cluster
       — min_d ≥ cluster_threshold: отдельно стоящая → r_standalone

    Кэшируется в Redis. Классификация по соседям учитывает двор/несколько пинов.

    Args:
        locations: объекты с .id, .latitude, .longitude; опционально .acceptance_radius_m
        r_same_point, r_cluster, r_standalone: радиусы в метрах
        same_point_threshold, cluster_threshold: пороги до ближайшей локации (м)

    Returns:
        dict[int, int]: {location_id: R в метрах}
    """
    locs = [
        o
        for o in locations
        if getattr(o, "latitude", None) is not None
        and getattr(o, "longitude", None) is not None
    ]
    out = {}
    for loc in locs:
        override = getattr(loc, "acceptance_radius_m", None)
        if override is not None and override > 0:
            out[loc.id] = int(override)
            continue
        min_d = float("inf")
        for other in locs:
            if getattr(other, "id", None) == getattr(loc, "id", None):
                continue
            d = calculate_distance_haversine(
                loc.latitude, loc.longitude, other.latitude, other.longitude
            )
            if d < min_d:
                min_d = d
        if min_d < same_point_threshold:
            radius_m = r_same_point
        elif min_d < cluster_threshold:
            radius_m = r_cluster
        else:
            radius_m = r_standalone
        out[loc.id] = radius_m
    return out


def get_location_radius(loc, radii_dict=None):
    """Радиус R (м) для локации: acceptance_radius_m или radii_dict или DEFAULT."""
    override = getattr(loc, "acceptance_radius_m", None)
    if override is not None and override > 0:
        return int(override)
    if radii_dict and getattr(loc, "id", None) in radii_dict:
        return int(radii_dict[loc.id])
    from monitoring_app.lesson_locations_conf import DEFAULT_ACCEPTANCE_RADIUS_M

    return DEFAULT_ACCEPTANCE_RADIUS_M


def compute_neighbor_color_index(locations, neighbor_threshold_m=30):
    """Индексы цветов для различения соседних локаций на карте.

    Сосед = расстояние < neighbor_threshold_m. В каждом кластере соседей
    раздаёт 0,1,2,... чтобы отличать друг от друга. Одинокие — 0.

    Returns:
        dict[int, int]: {location_id: 0..4}
    """
    thr = neighbor_threshold_m
    locs = [
        o
        for o in locations
        if getattr(o, "latitude", None) is not None
        and getattr(o, "longitude", None) is not None
    ]
    neighbors = {o.id: [] for o in locs}
    for i, a in enumerate(locs):
        for b in locs[i + 1 :]:
            d = calculate_distance_haversine(
                a.latitude, a.longitude, b.latitude, b.longitude
            )
            if d < thr:
                neighbors[a.id].append(b.id)
                neighbors[b.id].append(a.id)
    palette_size = 5
    out = {}
    for loc in locs:
        used = {out[n] for n in neighbors[loc.id] if n in out}
        c = 0
        while c in used:
            c += 1
        out[loc.id] = c % palette_size
    return out


def is_within_radius(lat1, lon1, lat2, lon2, radius=200):
    """Проверяет, находится ли точка в заданном радиусе от другой точки.

    Args:
        lat1 (float): Широта первой точки в градусах.
        lon1 (float): Долгота первой точки в градусах.
        lat2 (float): Широта второй точки в градусах.
        lon2 (float): Долгота второй точки в градусах.
        radius (float): Радиус в метрах. По умолчанию 200.

    Returns:
        bool: True если расстояние меньше или равно радиусу.
    """
    distance = calculate_distance_haversine(lat1, lon1, lat2, lon2)
    return distance <= radius


EXCEL_ATTENDANCE_CACHE_VERSION = "excel_alerts_v3"
EXCEL_CLASS_LOCATION_RESOLVE_RADIUS_M = 200
EXCEL_GPS_SPOOF_RADIUS_M = 2
EXCEL_GPS_SPOOF_MIN_DAYS = 3
EXCEL_GPS_SPOOF_MIN_SHARE = 0.5

EXCEL_ALERT_FACEID = "faceid"
EXCEL_ALERT_GPS_SPOOF = "gps_spoof"
EXCEL_ALERT_FACEID_GPS = "faceid_gps"
EXCEL_ALERT_NOTE_TEXT = {
    EXCEL_ALERT_FACEID: "Подозрение на обман FaceID",
    EXCEL_ALERT_GPS_SPOOF: "Подозрение на подмену локации",
    EXCEL_ALERT_FACEID_GPS: "Подозрение на обман FaceID и подмену локации",
}

ExcelAlertKey = tuple[int, str]
ExcelInvalidLessonDayKey = tuple[int, datetime.date]
ExcelLocationPayload = dict[str, object]
ExcelResolutionCache = dict[tuple[float, float], ExcelLocationPayload | None]


class ExcelGeoCluster(TypedDict):
    items: list[Mapping[str, Any]]
    center_lat: float
    center_lon: float


class ExcelDayAnchor(TypedDict):
    staff_id: int
    class_location_id: int
    date: datetime.date
    center_lat: float
    center_lon: float


class ExcelNormalizedLessonRow(TypedDict):
    id: int
    staff_id: int
    date_at: datetime.date
    date_key: str
    first_in: datetime.datetime | None
    last_out: datetime.datetime | None
    first_in_local: datetime.datetime | None
    last_out_local: datetime.datetime | None
    latitude: float | None
    longitude: float | None
    photo_spoof_status: str | None
    photo_manual_verdict: str | None
    class_location_id: int | None
    location_name: str


def _coerce_int_or_none(value: object) -> int | None:
    """Convert a loosely typed value into ``int`` when possible.

    Args:
        value: Source value that may represent an integer.

    Returns:
        Parsed integer or ``None`` when conversion is not possible.

    Notes:
        Complexity: O(1).
    """
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        try:
            return int(value)
        except ValueError:
            return None
    return None


def _excel_sort_token(value: object | None) -> tuple[bool, str]:
    """Build a stable sortable token for optional values.

    Args:
        value: Any sortable-like value or ``None``.

    Returns:
        Tuple where ``None`` values are always sorted last.

    Notes:
        Complexity: O(1).
    """
    return (value is None, str(value))


def _cluster_geo_items_for_excel(
    items: Sequence[Mapping[str, Any]],
    *,
    radius_m: float,
    lat_key: str = "latitude",
    lon_key: str = "longitude",
) -> list[ExcelGeoCluster]:
    """Cluster geo items by radius using BallTree and connected components.

    Args:
        items: Sequence of mapping-like geo records.
        radius_m: Clustering radius in meters.
        lat_key: Key used to read latitude from each item.
        lon_key: Key used to read longitude from each item.

    Returns:
        List of clusters with original items and centroid coordinates.

    Notes:
        Complexity: O(n^2) for very small buckets (``n <= 8``) to avoid tree
        overhead; otherwise O(n log n + e), where ``e`` is the number of
        candidate neighbor edges returned by BallTree within the radius.
    """
    if not items:
        return []
    if len(items) == 1:
        item = items[0]
        return [
            {
                "items": [item],
                "center_lat": float(item[lat_key]),
                "center_lon": float(item[lon_key]),
            }
        ]

    coords_deg = np.array(
        [[float(item[lat_key]), float(item[lon_key])] for item in items],
        dtype=float,
    )
    if len(items) <= 8:
        clusters = []
        visited = set()
        for start_idx in range(len(items)):
            if start_idx in visited:
                continue
            queue = [start_idx]
            component_indices = []
            while queue:
                idx = queue.pop()
                if idx in visited:
                    continue
                visited.add(idx)
                component_indices.append(idx)
                base_lat = coords_deg[idx][0]
                base_lon = coords_deg[idx][1]
                for other_idx in range(len(items)):
                    if other_idx == idx or other_idx in visited:
                        continue
                    distance = calculate_distance_haversine(
                        base_lat,
                        base_lon,
                        float(coords_deg[other_idx][0]),
                        float(coords_deg[other_idx][1]),
                    )
                    if distance <= radius_m:
                        queue.append(other_idx)

            cluster_items = [items[idx] for idx in component_indices]
            cluster_coords = coords_deg[component_indices]
            clusters.append(
                {
                    "items": cluster_items,
                    "center_lat": float(cluster_coords[:, 0].mean()),
                    "center_lon": float(cluster_coords[:, 1].mean()),
                }
            )
        return clusters

    coords_rad = np.radians(coords_deg)
    tree = BallTree(coords_rad, metric="haversine")
    radius_rad = radius_m / R_EARTH_M
    neighbor_indices = tree.query_radius(coords_rad, r=radius_rad)

    clusters = []
    visited = set()
    for start_idx in range(len(items)):
        if start_idx in visited:
            continue
        queue = [start_idx]
        component_indices = []
        while queue:
            idx = queue.pop()
            if idx in visited:
                continue
            visited.add(idx)
            component_indices.append(idx)
            base_lat = coords_deg[idx][0]
            base_lon = coords_deg[idx][1]
            for other_idx in neighbor_indices[idx]:
                other_idx = int(other_idx)
                if other_idx == idx or other_idx in visited:
                    continue
                distance = calculate_distance_haversine(
                    base_lat,
                    base_lon,
                    float(coords_deg[other_idx][0]),
                    float(coords_deg[other_idx][1]),
                )
                if distance <= radius_m:
                    queue.append(other_idx)

        cluster_items = [items[idx] for idx in component_indices]
        cluster_coords = coords_deg[component_indices]
        clusters.append(
            {
                "items": cluster_items,
                "center_lat": float(cluster_coords[:, 0].mean()),
                "center_lon": float(cluster_coords[:, 1].mean()),
            }
        )
    return clusters


def _pick_excel_dominant_cluster(
    items: Sequence[Mapping[str, Any]],
    *,
    radius_m: float,
    lat_key: str = "latitude",
    lon_key: str = "longitude",
) -> ExcelGeoCluster | None:
    """Pick the densest geo cluster for one day.

    Args:
        items: Sequence of geo items for one day.
        radius_m: Clustering radius in meters.
        lat_key: Key used to read latitude.
        lon_key: Key used to read longitude.

    Returns:
        Most representative cluster or ``None`` when input is empty.

    Notes:
        Complexity: O(n log n + e) for clustering plus O(c log c) for sorting
        ``c`` produced clusters.
    """
    clusters = _cluster_geo_items_for_excel(
        items,
        radius_m=radius_m,
        lat_key=lat_key,
        lon_key=lon_key,
    )
    if not clusters:
        return None
    return sorted(
        clusters,
        key=lambda cluster: (
            -len(cluster["items"]),
            min(_excel_sort_token(item.get("sort_time")) for item in cluster["items"]),
            min(int(item.get("sort_id", 0)) for item in cluster["items"]),
        ),
    )[0]


def _build_excel_day_anchor(
    day_records: Sequence[ExcelNormalizedLessonRow],
) -> ExcelDayAnchor | None:
    """Build one daily anchor point for a staff/location/day bucket.

    Args:
        day_records: Normalized lesson rows for one
            ``(staff_id, class_location_id, date)`` bucket.

    Returns:
        Daily anchor with centroid coordinates or ``None`` when there is no
        usable geo data.

    Notes:
        Complexity: O(m log m + e), where ``m`` is the number of rows in the
        day bucket.
    """
    if not day_records:
        return None

    items = []
    for record in day_records:
        latitude = record.get("latitude")
        longitude = record.get("longitude")
        record_id = record.get("id")
        if latitude is None or longitude is None or record_id is None:
            continue
        items.append(
            {
                "latitude": float(latitude),
                "longitude": float(longitude),
                "sort_time": record.get("first_in"),
                "sort_id": int(record_id),
                "record": record,
            }
        )
    if not items:
        return None

    dominant_cluster = _pick_excel_dominant_cluster(
        items,
        radius_m=EXCEL_GPS_SPOOF_RADIUS_M,
    )
    if dominant_cluster is None:
        return None

    base_record = dominant_cluster["items"][0]["record"]
    class_location_id = base_record.get("class_location_id")
    if class_location_id is None:
        return None
    return {
        "staff_id": int(base_record["staff_id"]),
        "class_location_id": int(class_location_id),
        "date": base_record["date_at"],
        "center_lat": dominant_cluster["center_lat"],
        "center_lon": dominant_cluster["center_lon"],
    }


def _is_faceid_suspicious_for_excel(
    photo_spoof_status: str | None,
    photo_manual_verdict: str | None,
) -> bool:
    """Resolve whether FaceID should be marked suspicious in Excel.

    Args:
        photo_spoof_status: Automatic spoof status from ``LessonAttendance``.
        photo_manual_verdict: Manual override verdict from ``LessonAttendance``.

    Returns:
        ``True`` when the effective verdict is suspicious.

    Notes:
        Complexity: O(1).
    """
    if photo_manual_verdict == models.LessonAttendance.PHOTO_MANUAL_VERDICT_SUSPICIOUS:
        return True
    if photo_manual_verdict == models.LessonAttendance.PHOTO_MANUAL_VERDICT_CLEAN:
        return False
    return photo_spoof_status == models.LessonAttendance.PHOTO_SPOOF_STATUS_SUSPICIOUS


def _build_excel_alert_code(*, has_faceid: bool, has_gps_spoof: bool) -> str:
    """Build the final Excel alert code for one cell.

    Args:
        has_faceid: Whether FaceID alert is active for the day.
        has_gps_spoof: Whether GPS spoof alert is active for the day.

    Returns:
        Alert code string used by the Excel renderer.

    Notes:
        Complexity: O(1).
    """
    if has_faceid and has_gps_spoof:
        return EXCEL_ALERT_FACEID_GPS
    if has_faceid:
        return EXCEL_ALERT_FACEID
    if has_gps_spoof:
        return EXCEL_ALERT_GPS_SPOOF
    return ""


def _append_excel_alert_note(status_info: str | None, alert_code: str) -> str:
    """Append a readable alert note to the attendance cell text.

    Args:
        status_info: Existing attendance text for the cell.
        alert_code: Alert code produced for the day.

    Returns:
        Human-readable cell text with alert note appended once.

    Notes:
        Complexity: O(len(status_info)).
    """
    note_text = EXCEL_ALERT_NOTE_TEXT.get(alert_code)
    if not note_text:
        return status_info or ""
    if not status_info:
        return f"({note_text})"
    if note_text in status_info:
        return status_info
    return f"{status_info}\n({note_text})"


def _normalize_excel_lesson_rows(
    raw_lesson_rows: Sequence[Mapping[str, Any]],
    location_searcher: "LocationSearcher | None",
    *,
    local_tz: datetime.tzinfo,
) -> list[ExcelNormalizedLessonRow]:
    """Normalize raw lesson rows for Excel processing.

    Args:
        raw_lesson_rows: Raw DB rows fetched via ``values(...)``.
        location_searcher: Initialized location resolver or ``None``.
        local_tz: Timezone object reused for local datetime conversion.

    Returns:
        Normalized lesson rows enriched with local datetimes and resolved class
        location metadata.

    Notes:
        Complexity: O(n + u log L + e), where ``u`` is the number of unique
        rounded coordinate pairs and ``L`` is the number of class locations.
    """
    resolution_cache: ExcelResolutionCache = {}
    normalized_rows: list[ExcelNormalizedLessonRow] = []
    coordinates_to_resolve: list[tuple[float, float]] = []

    if location_searcher is not None:
        for row in raw_lesson_rows:
            latitude = row.get("latitude")
            longitude = row.get("longitude")
            if latitude is None or longitude is None:
                continue
            cache_key = (round(float(latitude), 7), round(float(longitude), 7))
            if cache_key in resolution_cache:
                continue
            resolution_cache[cache_key] = None
            coordinates_to_resolve.append(cache_key)

        bulk_locations = location_searcher.find_nearest_locations_bulk(
            coordinates_to_resolve,
            radius=EXCEL_CLASS_LOCATION_RESOLVE_RADIUS_M,
        )
        for cache_key, location_payload in zip(coordinates_to_resolve, bulk_locations):
            resolution_cache[cache_key] = location_payload

    for row in raw_lesson_rows:
        staff_id = int(row["staff_id"])
        date_at = row["date_at"]
        first_in = row.get("first_in")
        last_out = row.get("last_out")
        latitude = row.get("latitude")
        longitude = row.get("longitude")
        location_payload = None
        if latitude is not None and longitude is not None:
            cache_key = (round(float(latitude), 7), round(float(longitude), 7))
            location_payload = resolution_cache.get(cache_key)

        class_location_id = None
        location_name = "Неизвестная локация"
        location_id_raw = location_payload.get("id") if location_payload else None
        class_location_id = _coerce_int_or_none(location_id_raw)
        if class_location_id is not None:
            location_name = str(location_payload.get("name") or location_name)

        normalized_rows.append(
            {
                "id": int(row["id"]),
                "staff_id": staff_id,
                "date_at": date_at,
                "date_key": date_at.isoformat(),
                "first_in": first_in,
                "last_out": last_out,
                "first_in_local": (
                    _convert_to_local_with_tz(first_in, local_tz) if first_in else None
                ),
                "last_out_local": (
                    _convert_to_local_with_tz(last_out, local_tz) if last_out else None
                ),
                "latitude": float(latitude) if latitude is not None else None,
                "longitude": float(longitude) if longitude is not None else None,
                "photo_spoof_status": row.get("photo_spoof_status"),
                "photo_manual_verdict": row.get("photo_manual_verdict"),
                "class_location_id": class_location_id,
                "location_name": location_name,
            }
        )

    return normalized_rows


def _detect_excel_faceid_alerts(
    raw_lesson_rows: Sequence[Mapping[str, Any]],
) -> tuple[Set[ExcelAlertKey], Set[ExcelInvalidLessonDayKey]]:
    """Detect FaceID alerts and report-invalid lesson days.

    Args:
        raw_lesson_rows: Normalized lesson rows for the selected period.

    Returns:
        Tuple of:
        - alert keys ``(staff_id, iso_date)`` used by Excel coloring;
        - invalid day keys ``(staff_id, date)`` used to exclude lesson-based
          attendance for days rejected by FaceID.

    Notes:
        Complexity: O(n), where ``n`` is the number of normalized lesson rows.
    """
    flagged_dates: Set[ExcelAlertKey] = set()
    flagged_days: Set[ExcelInvalidLessonDayKey] = set()
    for row in raw_lesson_rows:
        if _is_faceid_suspicious_for_excel(
            row.get("photo_spoof_status"),
            row.get("photo_manual_verdict"),
        ):
            staff_id = int(row["staff_id"])
            date_at = row["date_at"]
            flagged_dates.add((staff_id, row.get("date_key") or date_at.isoformat()))
            flagged_days.add((staff_id, date_at))
    return flagged_dates, flagged_days


def _detect_excel_gps_spoof_alerts(
    normalized_lesson_rows: Sequence[ExcelNormalizedLessonRow],
) -> Set[ExcelAlertKey]:
    """Detect repeated micro-point usage for each staff independently.

    GPS spoof detection is intentionally scoped only to one staff member inside
    one ``ClassLocation``. Other staff members using the same point do not
    affect the decision.

    Args:
        normalized_lesson_rows: Normalized lesson rows for the selected period.

    Returns:
        Set of alert keys ``(staff_id, iso_date)`` to color in Excel.

    Notes:
        Complexity: O(r + sum(m_i log m_i + e_i) + sum(a_j log a_j + g_j)),
        where ``r`` is the number of rows, ``m_i`` is the size of each
        staff/location/day bucket, ``a_j`` is the number of day anchors in each
        staff/location bucket, and ``e_i``/``g_j`` are the BallTree neighbor
        edge counts for those buckets.
    """
    if not normalized_lesson_rows:
        return set()

    rows_by_staff_location_day: dict[
        tuple[int, int, datetime.date],
        list[ExcelNormalizedLessonRow],
    ] = defaultdict(list)
    for row in normalized_lesson_rows:
        class_location_id = row.get("class_location_id")
        if class_location_id is None:
            continue

        key = (
            int(row["staff_id"]),
            int(class_location_id),
            row["date_at"],
        )
        rows_by_staff_location_day[key].append(row)

    anchors_by_staff_location: dict[
        tuple[int, int],
        list[ExcelDayAnchor],
    ] = defaultdict(list)
    for (
        staff_id,
        class_location_id,
        _date_at,
    ), day_records in rows_by_staff_location_day.items():
        day_anchor = _build_excel_day_anchor(day_records)
        if day_anchor is None:
            continue
        anchors_by_staff_location[(staff_id, class_location_id)].append(day_anchor)

    flagged_dates: Set[ExcelAlertKey] = set()
    for (staff_id, _class_location_id), anchors in anchors_by_staff_location.items():
        active_days = len({anchor["date"] for anchor in anchors})
        if active_days < EXCEL_GPS_SPOOF_MIN_DAYS:
            continue

        anchor_items = [
            {
                "latitude": float(anchor["center_lat"]),
                "longitude": float(anchor["center_lon"]),
                "sort_time": anchor["date"],
                "sort_id": index,
                "date": anchor["date"],
            }
            for index, anchor in enumerate(anchors)
        ]
        clusters = _cluster_geo_items_for_excel(
            anchor_items,
            radius_m=EXCEL_GPS_SPOOF_RADIUS_M,
        )
        for cluster in clusters:
            repeat_dates = {item["date"] for item in cluster["items"]}
            repeat_days = len(repeat_dates)
            if repeat_days < EXCEL_GPS_SPOOF_MIN_DAYS:
                continue
            if (repeat_days / active_days) < EXCEL_GPS_SPOOF_MIN_SHARE:
                continue
            for date_value in repeat_dates:
                flagged_dates.add((staff_id, date_value.isoformat()))

    return flagged_dates


def extract_coordinates(geo_data):
    """
    Extracts latitude and longitude from a geo data string.

    Supports formats:
    - latitude,longitude (e.g. "43.254926,76.929225") — для ручного редактирования
    - longitude%2Clatitude (e.g. "76.929225%2C43.254926")

    Returns:
        tuple: (latitude, longitude) or (None, None) if invalid.
    """
    if not geo_data or not isinstance(geo_data, str):
        return (None, None)
    geo_data = geo_data.strip()
    match = re.search(r"(-?\d+\.?\d*)\s*[,;]\s*(-?\d+\.?\d*)", geo_data)
    if match:
        a, b = float(match.group(1)), float(match.group(2))
        if -90 <= a <= 90 and -180 <= b <= 180:
            return (a, b)
        if -90 <= b <= 90 and -180 <= a <= 180:
            return (b, a)
        return (a, b)
    match = re.search(r"(-?\d+\.?\d*)%2C(-?\d+\.?\d*)", geo_data)
    if match:
        lon, lat = float(match.group(1)), float(match.group(2))
        return (lat, lon)
    return (None, None)


def export_class_locations_to_excel(queryset=None) -> bytes:
    """
    Экспортирует ClassLocation в Excel в формате загрузки (load_geo).
    Столбцы: name, address, geo (latitude,longitude), acceptance_radius_m.
    Можно отредактировать и загрузить обратно через upload_file.
    queryset: если передан — экспортируются только выбранные, иначе все.
    """
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Локации"

    ws.append(["Экспорт локаций для редактирования"])
    ws.append(["Формат: name, address, latitude,longitude, acceptance_radius_m (м)"])
    ws.append(["name", "address", "geo", "acceptance_radius_m"])

    if queryset is not None and queryset.exists():
        locs = queryset.only(
            "name", "address", "latitude", "longitude", "acceptance_radius_m"
        ).order_by("name")
    else:
        locs = models.ClassLocation.objects.only(
            "name", "address", "latitude", "longitude", "acceptance_radius_m"
        ).order_by("name")

    for loc in locs:
        geo = ""
        if loc.latitude is not None and loc.longitude is not None:
            geo = f"{loc.latitude},{loc.longitude}"
        radius = loc.acceptance_radius_m if loc.acceptance_radius_m is not None else ""
        ws.append([loc.name or "", loc.address or "", geo, radius])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def get_bonus_percentage(num_days, percent_for_period):
    """
    Определяет бонус на основе количества рабочих дней и итогового процента посещаемости.

    Если итоговый процент больше 100, по-прежнему осуществляется поиск правила.
    Если подходящего правила не найдено, возвращается 0.

    Аргументы:
        num_days (int): Количество уникальных рабочих дней в периоде.
        percent_for_period (float): Итоговый процент присутствия за период.

    Возвращает:
        float: Вычисленный бонус в процентах.
    """
    rule = models.PerformanceBonusRule.objects.filter(
        min_days__lte=num_days,
        max_days__gte=num_days,
        min_attendance_percent__lte=percent_for_period,
        max_attendance_percent__gte=percent_for_period,
    ).first()

    if rule:
        return float(rule.bonus_percentage)

    return 0.0


def get_all_child_departments(department):
    """
    Recursively get all child departments of a given department.

    Args:
        department: The parent department

    Returns:
        List of departments including the parent and all children
    """
    result = [department]
    children = models.ChildDepartment.objects.filter(parent=department)

    for child in children:
        result.extend(get_all_child_departments(child))

    return result


def collect_attendance_data(
    staff_list: Iterable[models.Staff],
    start_date: datetime.date,
    end_date: datetime.date,
) -> list[list[str]]:
    """Collect attendance data for the Excel export.

    Args:
        staff_list: Iterable of ``Staff`` objects or a queryset.
        start_date: Start date for attendance data.
        end_date: End date for attendance data.

    Returns:
        Flat list of Excel-ready rows:
        ``[fio, department, date_display, status_info, meta, alert_code]``.

    Notes:
        Complexity: Dominated by ``_collect_attendance_data_impl`` on a cold
        cache; O(1) on a warm cache aside from cache backend overhead.
    """
    if hasattr(staff_list, "select_related"):
        staff_list = list(staff_list.select_related("department"))
    else:
        staff_list = list(staff_list)

    if not staff_list:
        return []

    def generate_cache_key() -> str:
        start_str = start_date.strftime("%Y-%m-%d")
        end_str = end_date.strftime("%Y-%m-%d")

        dept_ids = sorted(
            set(
                staff.department_id
                for staff in staff_list
                if staff.department_id is not None
            )
        )
        if dept_ids:
            dept_str = hashlib.sha1(
                "|".join(map(str, dept_ids)).encode("utf-8")
            ).hexdigest()[:16]
        else:
            dept_str = "no_dept"

        staff_ids = sorted(str(staff.id) for staff in staff_list)
        staff_hash = hashlib.sha1("|".join(staff_ids).encode("utf-8")).hexdigest()[:16]

        cache_version = models.LessonAttendance.REPORT_FILTER_CACHE_VERSION
        return (
            f"attendance_data_{EXCEL_ATTENDANCE_CACHE_VERSION}_{cache_version}_"
            f"{start_str}_to_{end_str}_"
            f"dept_{dept_str}_staff_{staff_hash}"
        )

    cache_key = generate_cache_key()
    cached_results = get_cache(
        cache_key,
        query=lambda: _collect_attendance_data_impl(staff_list, start_date, end_date),
        timeout=6 * 60 * 60,
    )

    return cached_results or []


def _collect_attendance_data_impl(
    staff_list: Sequence[models.Staff],
    start_date: datetime.date,
    end_date: datetime.date,
) -> list[list[str]]:
    """Build Excel attendance rows from StaffAttendance and LessonAttendance.

    Args:
        staff_list: Materialized staff sequence with department relation loaded.
        start_date: Start date of the requested period.
        end_date: End date of the requested period.

    Returns:
        Flat list of Excel-ready rows:
        ``[fio, department, date_display, status_info, meta, alert_code]``.

    Notes:
        Complexity: O(s * d + sa + la + rw + ar + gps), where ``s`` is staff
        count, ``d`` is number of calendar days in the period, ``sa`` is the
        number of ``StaffAttendance`` rows, ``la`` is the number of
        ``LessonAttendance`` rows, ``rw`` is the number of remote-work spans,
        ``ar`` is the number of absence spans, and ``gps`` is the clustering
        work described in ``_detect_excel_gps_spoof_alerts``.
    """
    from django.db.models import Q

    date_range = [
        start_date + datetime.timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
    ]

    staff_list = list(staff_list)
    staff_ids = [s.id for s in staff_list]
    local_tz = timezone.get_current_timezone()
    logger.info(
        f"Collecting attendance data from {start_date} to {end_date} for {len(staff_list)} staff members"
    )

    holidays = (
        get_cache(
            "public_holidays",
            query=lambda: {
                holiday.date: holiday.is_working_day
                for holiday in models.PublicHoliday.objects.filter(
                    date__range=[start_date, end_date]
                )
            },
            timeout=10 * 60,
        )
        or {}
    )

    location_data = list(
        models.ClassLocation.objects.filter(
            latitude__isnull=False,
            longitude__isnull=False,
        ).values("id", "name", "address", "latitude", "longitude")
    )
    location_searcher = LocationSearcher(location_data) if location_data else None
    if location_searcher is not None:
        logger.info(f"LocationSearcher initialized with {len(location_data)} locations")

    attendance_qs = models.StaffAttendance.objects.filter(
        staff_id__in=staff_ids,
        date_at__range=[
            start_date + datetime.timedelta(days=1),
            end_date + datetime.timedelta(days=1),
        ],
    ).values(
        "staff_id",
        "date_at",
        "first_in",
        "last_out",
        "area_name_in",
        "area_name_out",
        "effective_work_seconds",
        "effective_work_intervals",
    )

    raw_lesson_rows = list(
        models.LessonAttendance.objects.filter(
            staff_id__in=staff_ids,
            date_at__range=[start_date, end_date],
        )
        .values(
            "id",
            "staff_id",
            "date_at",
            "first_in",
            "last_out",
            "latitude",
            "longitude",
            "photo_spoof_status",
            "photo_manual_verdict",
        )
        .order_by("date_at", "staff_id", "first_in", "id")
        .iterator(chunk_size=2000)
    )
    normalized_lesson_rows = _normalize_excel_lesson_rows(
        raw_lesson_rows,
        location_searcher,
        local_tz=local_tz,
    )
    faceid_alert_dates, invalid_lesson_days = _detect_excel_faceid_alerts(
        normalized_lesson_rows
    )
    gps_spoof_alert_dates = _detect_excel_gps_spoof_alerts(normalized_lesson_rows)

    remote_work_qs = models.RemoteWork.objects.filter(
        Q(staff_id__in=staff_ids)
        & (
            Q(start_date__lte=end_date, end_date__gte=start_date)
            | Q(permanent_remote=True)
        )
    ).values("id", "staff_id", "permanent_remote", "start_date", "end_date")

    absence_qs = models.AbsentReason.objects.filter(
        staff_id__in=staff_ids, start_date__lte=end_date, end_date__gte=start_date
    ).values("staff_id", "start_date", "end_date", "reason", "approved")

    attendance_map = defaultdict(lambda: defaultdict(dict))

    for att in attendance_qs.iterator(chunk_size=2000):
        first_in = att.get("first_in")
        last_out = att.get("last_out")
        date_at = att.get("date_at")
        area_name_in = att.get("area_name_in")
        area_name_out = att.get("area_name_out")
        first_in_local_full = (
            _convert_to_local_with_tz(first_in, local_tz) if first_in else None
        )
        last_out_local_full = (
            _convert_to_local_with_tz(last_out, local_tz) if last_out else None
        )

        if first_in_local_full:
            local_date = first_in_local_full
        else:
            local_date = _convert_to_local_with_tz(date_at, local_tz)
            if local_date is not None:
                local_date = local_date - datetime.timedelta(days=1)
        if local_date is None:
            continue
        date_key = local_date.strftime("%Y-%m-%d")
        staff_id = int(att["staff_id"])

        use_first_in = first_in_local_full and not is_lift_terminal(area_name_in)
        use_last_out = last_out_local_full and not is_lift_terminal(area_name_out)
        first_in_local = first_in_local_full if use_first_in else None
        last_out_local = last_out_local_full if use_last_out else None
        elevator_first_in = (
            first_in_local_full if first_in_local_full and not use_first_in else None
        )
        elevator_last_out = (
            last_out_local_full if last_out_local_full and not use_last_out else None
        )

        if staff_id not in attendance_map[date_key]:
            area_name = (
                (area_name_in if (area_name_in and use_first_in) else None)
                or (area_name_out if (area_name_out and use_last_out) else None)
                or "Неизвестная локация"
            )
            rec = {
                "first_in": first_in_local,
                "last_out": last_out_local,
                "area_name": area_name,
                "source": "staff_attendance",
                "first_in_source": "staff_attendance",
                "last_out_source": "staff_attendance",
                "effective_work_seconds": att.get("effective_work_seconds"),
                "effective_work_intervals": att.get("effective_work_intervals"),
                "la_intervals": [],
            }
            if elevator_first_in is not None or elevator_last_out is not None:
                rec["elevator_first_in"] = elevator_first_in
                rec["elevator_last_out"] = elevator_last_out
            attendance_map[date_key][staff_id] = rec
        else:
            current_rec = attendance_map[date_key][staff_id]
            if use_first_in and (
                not current_rec["first_in"]
                or first_in_local_full < current_rec["first_in"]
            ):
                current_rec["first_in"] = first_in_local_full
                current_rec["source"] = (
                    "staff_attendance"
                    if current_rec.get("source") == "lesson_attendance"
                    else "mixed"
                )
                current_rec["first_in_source"] = "staff_attendance"
                if area_name_in:
                    current_rec["area_name"] = area_name_in

            if use_last_out and (
                not current_rec["last_out"]
                or last_out_local_full > current_rec["last_out"]
            ):
                current_rec["last_out"] = last_out_local_full
                current_rec["source"] = (
                    "staff_attendance"
                    if current_rec.get("source") == "lesson_attendance"
                    else "mixed"
                )
                current_rec["last_out_source"] = "staff_attendance"
                if area_name_out:
                    current_rec["area_name"] = area_name_out
            if elevator_first_in is not None or elevator_last_out is not None:
                current_rec["elevator_first_in"] = elevator_first_in
                current_rec["elevator_last_out"] = elevator_last_out
            if att.get("effective_work_seconds") is not None:
                current_rec["effective_work_seconds"] = att["effective_work_seconds"]
            if "la_intervals" not in current_rec:
                current_rec["la_intervals"] = []

    for lesson_att in normalized_lesson_rows:
        if (lesson_att["staff_id"], lesson_att["date_at"]) in invalid_lesson_days:
            continue

        local_date = lesson_att["first_in_local"] or _convert_to_local_with_tz(
            lesson_att["date_at"],
            local_tz,
        )
        if local_date is None:
            continue
        date_key = local_date.strftime("%Y-%m-%d")
        staff_id = lesson_att["staff_id"]

        first_in_local = lesson_att["first_in_local"]
        last_out_local = lesson_att["last_out_local"]
        location_name = lesson_att["location_name"]

        if staff_id not in attendance_map[date_key]:
            la_intervals = []
            if first_in_local and last_out_local and last_out_local > first_in_local:
                la_intervals = [(first_in_local, last_out_local)]
            attendance_map[date_key][staff_id] = {
                "first_in": first_in_local,
                "last_out": last_out_local,
                "area_name": location_name,
                "source": "lesson_attendance",
                "first_in_source": "lesson_attendance",
                "last_out_source": "lesson_attendance",
                "effective_work_seconds": None,
                "effective_work_intervals": None,
                "la_intervals": la_intervals,
            }
        else:
            current_rec = attendance_map[date_key][staff_id]
            if "la_intervals" not in current_rec:
                current_rec["la_intervals"] = []
            if first_in_local and last_out_local and last_out_local > first_in_local:
                current_rec["la_intervals"].append((first_in_local, last_out_local))

            if first_in_local and (
                not current_rec["first_in"] or first_in_local < current_rec["first_in"]
            ):
                current_rec["first_in"] = first_in_local
                current_rec["source"] = (
                    "lesson_attendance"
                    if current_rec.get("source") == "staff_attendance"
                    else "mixed"
                )
                current_rec["first_in_source"] = "lesson_attendance"
                if current_rec.get("source") == "lesson_attendance":
                    current_rec["area_name"] = location_name

            if last_out_local and (
                not current_rec["last_out"] or last_out_local > current_rec["last_out"]
            ):
                current_rec["last_out"] = last_out_local
                current_rec["source"] = (
                    "lesson_attendance"
                    if current_rec.get("source") == "staff_attendance"
                    else "mixed"
                )
                current_rec["last_out_source"] = "lesson_attendance"
                if current_rec.get("source") == "lesson_attendance":
                    current_rec["area_name"] = location_name

    for date_key in attendance_map:
        for staff_id in attendance_map[date_key]:
            rec = attendance_map[date_key][staff_id]
            intervals: List[Tuple[datetime.datetime, datetime.datetime]] = []
            for raw in rec.get("effective_work_intervals") or []:
                try:
                    s = raw.get("start") and datetime.datetime.fromisoformat(
                        raw["start"].replace("Z", "+00:00")
                    )
                    e = raw.get("end") and datetime.datetime.fromisoformat(
                        raw["end"].replace("Z", "+00:00")
                    )
                    if s is not None and e is not None and e > s:
                        intervals.append((s, e))
                except (ValueError, TypeError, AttributeError):
                    continue
            for start, end in rec.get("la_intervals") or []:
                if start and end and end > start:
                    intervals.append((start, end))
            total_sec = merge_work_intervals_to_total_seconds(intervals)
            rec["effective_work_seconds"] = total_sec if total_sec > 0 else None
            rec.pop("la_intervals", None)
            rec.pop("effective_work_intervals", None)

    remote_work_map = defaultdict(set)
    for rw in remote_work_qs.iterator(chunk_size=1000):
        staff_id = int(rw["staff_id"])

        if rw["permanent_remote"]:
            rw_start = start_date
            rw_end = end_date
        else:
            rw_start_raw = rw.get("start_date")
            rw_end_raw = rw.get("end_date")
            if rw_start_raw is None or rw_end_raw is None:
                logger.warning(
                    "RemoteWork id=%s for staff_id=%s skipped because start/end dates are missing",
                    rw.get("id"),
                    staff_id,
                )
                continue
            rw_start = max(rw_start_raw, start_date)
            rw_end = min(rw_end_raw, end_date)

        if rw_start > rw_end:
            continue

        date_keys = [
            (rw_start + datetime.timedelta(days=i)).strftime("%Y-%m-%d")
            for i in range((rw_end - rw_start).days + 1)
        ]

        for date_key in date_keys:
            remote_work_map[date_key].add(staff_id)

    absence_map = defaultdict(lambda: defaultdict(list))
    absence_reason_display_map = {
        value: label for value, label in models.AbsentReason.ABSENT_REASON_CHOICES
    }
    for absence in absence_qs.iterator(chunk_size=1000):
        staff_id = int(absence["staff_id"])
        abs_start = max(absence["start_date"], start_date)
        abs_end = min(absence["end_date"], end_date)

        for i in range((abs_end - abs_start).days + 1):
            current = abs_start + datetime.timedelta(days=i)
            date_key = current.strftime("%Y-%m-%d")
            absence_map[date_key][staff_id].append(
                {
                    "reason": absence_reason_display_map.get(
                        absence["reason"],
                        str(absence["reason"]),
                    ),
                    "approved": bool(absence["approved"]),
                }
            )
    results = []
    staff_entries = [
        (
            staff.id,
            f"{staff.surname} {staff.name}",
            staff.department.name if staff.department else "N/A",
        )
        for staff in staff_list
    ]

    date_info = [
        (
            d.strftime("%Y-%m-%d"),
            d.strftime("%d.%m.%Y"),
            (d.weekday() >= 5) or (d in holidays and not holidays.get(d, False)),
        )
        for d in date_range
    ]

    for date_str, date_display, is_off_day in date_info:
        date_attendance = attendance_map.get(date_str, {})
        date_remote = remote_work_map.get(date_str, set())
        date_absence = absence_map.get(date_str, {})

        for staff_id, staff_fio, department_name in staff_entries:
            alert_key = (staff_id, date_str)
            alert_code = _build_excel_alert_code(
                has_faceid=alert_key in faceid_alert_dates,
                has_gps_spoof=alert_key in gps_spoof_alert_dates,
            )

            att_data = date_attendance.get(staff_id)
            has_attendance = att_data is not None
            is_remote = staff_id in date_remote
            has_absence = staff_id in date_absence

            attendance_info = None
            if has_attendance:
                first_in = att_data["first_in"]
                last_out = att_data["last_out"]
                effective_work_seconds = att_data.get("effective_work_seconds")

                if first_in or last_out:
                    raw_area = att_data.get("area_name", "Неизвестная локация")
                    display_area = resolve_area_address(raw_area) or raw_area
                    t_in = (
                        (first_in or last_out).strftime("%H:%M:%S")
                        if (first_in or last_out)
                        else "—"
                    )
                    t_out = (
                        (last_out or first_in).strftime("%H:%M:%S")
                        if (last_out or first_in)
                        else "—"
                    )
                    attendance_info = f"{t_in} - {t_out}\n({display_area})"
                    if effective_work_seconds is not None:
                        mins = int(effective_work_seconds // 60)
                        if mins >= 60:
                            h, m = divmod(mins, 60)
                            attendance_info += f"\n{h} ч {m} мин"
                        else:
                            attendance_info += f"\n{mins} мин"

            if is_off_day:
                if attendance_info:
                    status_info = attendance_info
                    meta = "holiday_with_attendance"
                else:
                    status_info = "Выходной"
                    meta = "holiday"
            else:
                first_in = att_data["first_in"] if att_data else None
                last_out = att_data["last_out"] if att_data else None
                has_elevator_times = att_data and (
                    att_data.get("elevator_first_in") is not None
                    or att_data.get("elevator_last_out") is not None
                )
                is_elevator_only = (
                    has_attendance
                    and first_in is None
                    and last_out is None
                    and not has_absence
                    and has_elevator_times
                )
                if is_elevator_only:
                    ei = att_data.get("elevator_first_in") if att_data else None
                    eo = att_data.get("elevator_last_out") if att_data else None
                    raw_area = att_data.get("area_name") if att_data else None
                    lift_addr = resolve_area_address(raw_area) if raw_area else None
                    addr_suffix = f"\n({lift_addr})" if lift_addr else ""
                    if ei is not None or eo is not None:
                        t_in = ei.strftime("%H:%M:%S") if ei else "—"
                        t_out = eo.strftime("%H:%M:%S") if eo else "—"
                        status_info = f"{t_in} - {t_out}\nЛИФТ{addr_suffix}"
                    else:
                        status_info = f"ЛИФТ{addr_suffix}"
                    meta = "elevator_only"
                elif has_attendance and is_remote:
                    status_info = (
                        f"Удаленная работа + Присутствие\n{attendance_info}"
                        if attendance_info
                        else "Удаленная работа"
                    )
                    meta = "remote_work"
                elif has_attendance and has_absence:
                    absence_info = date_absence[staff_id][0]
                    reason = absence_info["reason"]
                    approved = absence_info["approved"]

                    if approved:
                        status_info = (
                            f"{reason} + Присутствие\n{attendance_info}"
                            if attendance_info
                            else reason
                        )
                        meta = "absence_reason_approved"
                    else:
                        status_info = (
                            f"Неутв: {reason} + Присутствие\n{attendance_info}"
                            if attendance_info
                            else f"Не одобрено: {reason}"
                        )
                        meta = "absence_reason"
                elif attendance_info:
                    status_info = attendance_info
                    meta = "workday"
                elif is_remote:
                    status_info = "Удаленная работа"
                    meta = "remote_work"
                elif has_absence:
                    absence_info = date_absence[staff_id][0]
                    status_info = absence_info["reason"]
                    meta = (
                        "absence_reason_approved"
                        if absence_info["approved"]
                        else "absence_reason"
                    )
                else:
                    status_info = "Отсутствие"
                    meta = "absence"

            status_info = _append_excel_alert_note(status_info, alert_code)
            results.append(
                [
                    staff_fio,
                    department_name,
                    date_display,
                    status_info,
                    meta,
                    alert_code,
                ]
            )

    logger.info(
        f"Collected {len(results)} attendance records with combined status information"
    )
    return results


def generate_excel_file(
    attendance_data, department_name, user_start_date, user_end_date
):
    """
    Generate an Excel file from attendance data with improved formatting and filtering.

    Args:
        attendance_data: List of attendance records
            [ФИО, Отдел, 'DD.MM.YYYY', Посещаемость, meta, alert_code].
            The data has been obtained from the DB and may be for only a subset of the
            user-requested period.
        department_name: Name of the department (str).
        user_start_date: The user-requested start date (datetime.date).
        user_end_date: The user-requested end date (datetime.date).

    Returns:
        Bytes data of the Excel file.
    """
    import io

    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    logger.info("Generating refined Excel file...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет посещаемости"

    title_font = Font(name="Arial", size=16, bold=True)
    subtitle_font = Font(name="Arial", size=12, bold=True)
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10, color="000000")

    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(
        start_color="0070C0", end_color="0070C0", fill_type="solid"
    )
    fill_holiday = PatternFill(
        start_color="F59E0B", end_color="F59E0B", fill_type="solid"
    )
    fill_holiday_work = PatternFill(
        start_color="34D399", end_color="34D399", fill_type="solid"
    )
    fill_remote = PatternFill(
        start_color="38BDF8", end_color="38BDF8", fill_type="solid"
    )
    fill_approved = PatternFill(
        start_color="A78BFA", end_color="A78BFA", fill_type="solid"
    )
    fill_not_approved = PatternFill(
        start_color="FB7185", end_color="FB7185", fill_type="solid"
    )
    fill_elevator = PatternFill(
        start_color="9CA3AF", end_color="9CA3AF", fill_type="solid"
    )
    fill_faceid = PatternFill(
        start_color="DC2626", end_color="DC2626", fill_type="solid"
    )
    fill_gps_spoof = PatternFill(
        start_color="1D4ED8", end_color="1D4ED8", fill_type="solid"
    )
    fill_faceid_gps = PatternFill(
        start_color="7E22CE", end_color="7E22CE", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:E1")
    title_cell = ws.cell(
        row=1, column=1, value=f"Отчет посещаемости: {department_name}"
    )
    title_cell.font = title_font
    title_cell.alignment = center_wrap

    ws.merge_cells("A2:E2")
    subtitle_cell = ws.cell(
        row=2,
        column=1,
        value=f"Период: {user_start_date.strftime('%d.%m.%Y')} - {user_end_date.strftime('%d.%m.%Y')}",
    )
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = center_wrap

    legend_row = 4
    legend_title = ws.cell(row=legend_row, column=1, value="Легенда:")
    legend_title.font = subtitle_font
    legend_title.alignment = center_wrap

    legends = [
        ("Выходной день", fill_holiday),
        ("Активность в выходной день", fill_holiday_work),
        ("Удаленный формат", fill_remote),
        ("Согласованная причина отсутствия", fill_approved),
        ("Несогласованное отсутствие / отсутствие", fill_not_approved),
        ("Только событие лифта", fill_elevator),
        ("Попытка обмана FaceID", fill_faceid),
        ("Подозрение на подмену локации", fill_gps_spoof),
        ("FaceID + подмена локации", fill_faceid_gps),
    ]
    for i, (legend_text, legend_fill) in enumerate(legends, start=1):
        row_idx = legend_row + i
        color_cell = ws.cell(row=row_idx, column=1, value="")
        color_cell.fill = legend_fill
        color_cell.border = thin_border
        color_cell.alignment = center_wrap
        desc_cell = ws.cell(row=row_idx, column=2, value=legend_text)
        desc_cell.font = data_font
        desc_cell.border = thin_border
        desc_cell.alignment = center_wrap

    data_start_row = legend_row + len(legends) + 2

    columns_index = pd.Index(
        ["ФИО", "Отдел", "Дата", "Посещаемость", "meta", "alert_code"],
        dtype="object",
    )
    df = pd.DataFrame(attendance_data, columns=columns_index)
    df = cast(pd.DataFrame, df)
    df["date_obj"] = pd.to_datetime(df["Дата"], format="%d.%m.%Y")
    df = df[
        (df["date_obj"] >= pd.to_datetime(user_start_date))
        & (df["date_obj"] <= pd.to_datetime(user_end_date))
    ]
    if df.empty:
        logger.warning("No attendance data for the selected date range.")

    sort_method = getattr(df, "sort_values")
    df = cast(pd.DataFrame, sort_method(by="date_obj", ascending=False))

    unique_staff = df[["ФИО", "Отдел"]].drop_duplicates()
    sort_method_staff = getattr(unique_staff, "sort_values")
    unique_staff = cast(
        pd.DataFrame,
        sort_method_staff(by=["Отдел", "ФИО"], ascending=[True, True]),
    )

    unique_dates = df["Дата"].unique()

    headers = ["ФИО", "Отдел"] + list(unique_dates)
    for col_idx, header_val in enumerate(headers, start=1):
        cell = ws.cell(row=data_start_row, column=col_idx, value=header_val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_wrap
        cell.border = thin_border

    attendance_lookup = {}
    meta_lookup = {}
    alert_lookup = {}
    for row in df.itertuples(index=False):
        key = (row.ФИО, row.Отдел, row.Дата)
        attendance_lookup[key] = row.Посещаемость
        meta_lookup[key] = row.meta
        alert_lookup[key] = row.alert_code

    public_holidays = (
        get_cache(
            "public_holidays_for_excel",
            query=lambda: {
                holiday.date.strftime("%d.%m.%Y"): holiday.is_working_day
                for holiday in models.PublicHoliday.objects.filter(
                    date__range=[user_start_date, user_end_date]
                )
            },
            timeout=10 * 60,
        )
        or {}
    )

    row_idx = data_start_row + 1
    for __, staff_row in unique_staff.iterrows():
        fio_value = staff_row["ФИО"]
        dept_value = staff_row["Отдел"]
        fio = "" if bool(pd.isna(fio_value)) else str(fio_value)
        dept = "" if bool(pd.isna(dept_value)) else str(dept_value)

        fio_cell = ws.cell(row=row_idx, column=1, value=fio)
        fio_cell.font = data_font
        fio_cell.alignment = center_wrap
        fio_cell.border = thin_border

        dept_cell = ws.cell(row=row_idx, column=2, value=dept)
        dept_cell.font = data_font
        dept_cell.alignment = center_wrap
        dept_cell.border = thin_border

        for col_offset, date_str in enumerate(unique_dates, start=3):
            key = (fio, dept, date_str)
            value = attendance_lookup.get(key, "")
            meta = meta_lookup.get(key, "")
            alert_code = alert_lookup.get(key, "")
            data_cell = ws.cell(row=row_idx, column=col_offset, value=value)
            data_cell.font = data_font
            data_cell.alignment = center_wrap
            data_cell.border = thin_border

            is_working_holiday = (
                date_str in public_holidays and public_holidays[date_str]
            )

            if alert_code == EXCEL_ALERT_FACEID_GPS:
                data_cell.fill = fill_faceid_gps
                data_cell.font = Font(name="Arial", size=10, color="FFFFFF")
            elif alert_code == EXCEL_ALERT_FACEID:
                data_cell.fill = fill_faceid
                data_cell.font = Font(name="Arial", size=10, color="FFFFFF")
            elif alert_code == EXCEL_ALERT_GPS_SPOOF:
                data_cell.fill = fill_gps_spoof
                data_cell.font = Font(name="Arial", size=10, color="FFFFFF")
            elif meta == "holiday":
                if not is_working_holiday:
                    data_cell.fill = fill_holiday
            elif meta == "holiday_with_attendance":
                if not is_working_holiday:
                    data_cell.fill = fill_holiday_work
            elif meta == "remote_work":
                data_cell.fill = fill_remote
            elif meta == "absence_reason_approved":
                data_cell.fill = fill_approved
            elif meta in ["absence", "absence_reason"]:
                data_cell.fill = fill_not_approved
                data_cell.font = Font(name="Arial", size=10, color="FFFFFF")
            elif meta == "elevator_only":
                data_cell.fill = fill_elevator
        row_idx += 1

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for r_idx in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=r_idx, column=col_idx).value
            if cell_val and isinstance(cell_val, str):
                max_length = max(max_length, len(cell_val))
        ws.column_dimensions[col_letter].width = max_length + 2

    for r_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for c_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=r_idx, column=c_idx).value
            if cell_val and isinstance(cell_val, str):
                max_lines = max(max_lines, cell_val.count("\n") + 1)
        ws.row_dimensions[r_idx].height = 15 * max_lines

    excel_data = io.BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    logger.info(
        "Excel file generation completed with proper timezone conversion, filtering, and sorting."
    )
    return excel_data.getvalue()


def convert_to_local(dt):
    return _convert_to_local_with_tz(dt, timezone.get_current_timezone())
