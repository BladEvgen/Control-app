"""Выгрузка подозрительных LessonAttendance в XLSX и ZIP с фото.

Команда за один проход по данным формирует отчёт (отдел, ФИО, student_id/tutor_id,
дата, время, локация) и архив с фотографиями. Локация определяется по координатам
через LocationSearcher (Haversine). Пути к файлам фото разрешаются пакетно по
уникальным значениям staff_image_path, без повторных проходов по директориям.
"""

from __future__ import annotations

import datetime
import os
import zipfile
from pathlib import Path
from typing import Optional

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font

from monitoring_app import models, utils

CIS_DATE_FORMAT = "%d.%m.%Y"


class _StubLocationSearcher:
    """Заглушка: всегда возвращает «Unknown Area» (нет локаций для Haversine)."""

    def find_nearest(self, lat, lon, radius=200):
        return "Unknown Area"


class Command(BaseCommand):
    """Экспорт записей LessonAttendance с подозрительным фото в XLSX и ZIP.

    Формирует отчёт (отдел, ФИО, ID без S/T, дата СНГ, время локальное, локация)
    и рядом архив с фото: папки по дням (YYYY-MM-DD), имена файлов —
    staff_pin_ISO-время (двоеточия заменены на минус).
    """

    help = (
        "Export staff with suspicious LessonAttendance: FIO, pin (no S/T), "
        "department, date, location (from coordinates), time. Uses LocationSearcher (Haversine)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--date-from",
            type=str,
            dest="date_from",
            help="Start date YYYY-MM-DD (inclusive).",
        )
        parser.add_argument(
            "--date-to",
            type=str,
            dest="date_to",
            help="End date YYYY-MM-DD (inclusive).",
        )
        parser.add_argument(
            "--output",
            "-o",
            type=str,
            default="",
            dest="output",
            help="Output XLSX path. If omitted, writes to out/suspicious_attendance_<date>.xlsx.",
        )
        parser.add_argument(
            "--radius",
            type=int,
            default=200,
            dest="radius",
            help="Radius in meters for LocationSearcher.find_nearest (default: 200).",
        )

    def handle(self, *args, **options):
        """Запуск экспорта: фильтрация по датам, сбор строк и фото, запись XLSX и ZIP."""
        date_from, date_to = self._parse_dates(
            options.get("date_from"), options.get("date_to")
        )

        qs = (
            models.LessonAttendance.objects.filter(
                models.LessonAttendance.PHOTO_SUSPICIOUS_FOR_REPORTS_Q
            )
            .select_related("staff", "staff__department")
            .order_by("date_at", "first_in")
        )
        if date_from is not None:
            qs = qs.filter(date_at__gte=date_from)
        if date_to is not None:
            qs = qs.filter(date_at__lte=date_to)

        location_searcher = self._build_location_searcher()
        radius = int(options.get("radius") or 200)

        rows = []
        archive_records = []  # (rec, staff_pin_short) для архива с фото
        for rec in qs:
            staff = rec.staff
            fio = self._fio(staff)
            staff_pin = utils.pin_to_external_format(staff.pin) if staff.pin else ""
            department = (
                staff.department.name if staff.department else ""
            )
            date_str = rec.date_at.strftime(CIS_DATE_FORMAT) if rec.date_at else ""
            if rec.latitude is not None and rec.longitude is not None:
                location = location_searcher.find_nearest(
                    rec.latitude, rec.longitude, radius=radius
                )
            else:
                location = ""
            if rec.first_in:
                local_first_in = timezone.localtime(rec.first_in)
                time_str = local_first_in.strftime("%H:%M:%S")
            else:
                time_str = ""
            rows.append({
                "department": department,
                "fio": fio,
                "student_id_tutor_id": staff_pin,
                "date": date_str,
                "time": time_str,
                "location": location,
            })
            if rec.staff_image_path and str(rec.staff_image_path).strip():
                archive_records.append((rec, staff_pin or "unknown"))

        output_path = (options.get("output") or "").strip()
        if not output_path:
            output_path = self._default_output_path()
        path = Path(output_path)
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")
        path.parent.mkdir(parents=True, exist_ok=True)
        self._write_xlsx(rows, path)
        zip_path = path.with_suffix(".zip")
        added = self._write_archive(archive_records, zip_path)
        self.stdout.write(
            self.style.SUCCESS(
                f"Exported {len(rows)} rows to {path} (UTF-8), archive {zip_path} ({added} photos)"
            )
        )

    def _parse_dates(self, date_from_raw, date_to_raw):
        """Парсит и валидирует даты из аргументов команды.

        Args:
            date_from_raw: Строка начала периода (YYYY-MM-DD) или None.
            date_to_raw: Строка конца периода (YYYY-MM-DD) или None.

        Returns:
            Кортеж (date_from, date_to); каждый элемент — date или None.

        Raises:
            CommandError: Неверный формат даты или date_from > date_to.
        """
        date_from = None
        date_to = None
        if date_from_raw:
            try:
                date_from = datetime.datetime.strptime(
                    date_from_raw.strip(), "%Y-%m-%d"
                ).date()
            except ValueError as e:
                raise CommandError(
                    f"Invalid --date-from, use YYYY-MM-DD: {e}"
                ) from e
        if date_to_raw:
            try:
                date_to = datetime.datetime.strptime(
                    date_to_raw.strip(), "%Y-%m-%d"
                ).date()
            except ValueError as e:
                raise CommandError(
                    f"Invalid --date-to, use YYYY-MM-DD: {e}"
                ) from e
        if date_from is not None and date_to is not None and date_from > date_to:
            raise CommandError("date_from must be <= date_to")
        return date_from, date_to

    def _fio(self, staff):
        """Собирает ФИО из полей surname и name.

        Args:
            staff: Экземпляр Staff.

        Returns:
            Строка вида "Фамилия Имя" или пустая строка.
        """
        parts = []
        if getattr(staff, "surname", None):
            parts.append(staff.surname)
        if getattr(staff, "name", None):
            parts.append(staff.name)
        return " ".join(parts).strip() or ""

    def _build_location_searcher(self):
        """Создаёт поисковик локаций по координатам ClassLocation (Haversine + KDTree).

        Всегда возвращает объект с методом find_nearest(lat, lon, radius): при наличии
        записей ClassLocation с координатами — utils.LocationSearcher (точный поиск
        по Haversine в заданном радиусе), иначе заглушку, возвращающую «Unknown Area».

        Returns:
            Экземпляр utils.LocationSearcher или _StubLocationSearcher.
        """
        locations = list(
            models.ClassLocation.objects.filter(
                latitude__isnull=False,
                longitude__isnull=False,
            ).values("name", "latitude", "longitude")
        )
        if not locations:
            self.stdout.write(
                self.style.WARNING(
                    "No ClassLocation with coordinates; location will be «Unknown Area»."
                )
            )
            return _StubLocationSearcher()
        try:
            return utils.LocationSearcher(locations)
        except Exception as e:
            self.stdout.write(
                self.style.WARNING(
                    f"LocationSearcher init failed: {e}; using stub («Unknown Area»)."
                )
            )
            return _StubLocationSearcher()

    def _default_output_path(self) -> str:
        """Возвращает путь по умолчанию для XLSX (out/suspicious_attendance_<сегодня>.xlsx)."""
        today = datetime.date.today().isoformat()
        return str(Path.cwd() / "out" / f"suspicious_attendance_{today}.xlsx")

    def _resolve_photo_path_raw(self, raw_path: str) -> Optional[Path]:
        """Разрешает строковый путь к файлу фото в абсолютный Path.

        Проверяет: абсолютный путь, затем ATTENDANCE_ROOT и MEDIA_ROOT.
        Статические заглушки (/static/, no-avatar) игнорируются.

        Args:
            raw_path: Значение staff_image_path (относительный или абсолютный).

        Returns:
            Path к существующему файлу или None.
        """
        raw = (raw_path or "").strip()
        if not raw or "/static/" in raw or raw == "/static/media/images/no-avatar.png":
            return None
        p = Path(raw)
        if p.is_absolute() and p.exists():
            return p
        for root in (getattr(settings, "ATTENDANCE_ROOT", None), getattr(settings, "MEDIA_ROOT", None)):
            if not root:
                continue
            candidate = Path(root) / raw.lstrip("/")
            if candidate.exists():
                return candidate
            candidate = Path(root) / p.name
            if candidate.exists():
                return candidate
        if not p.is_absolute():
            candidate = Path.cwd() / raw
            if candidate.exists():
                return candidate
        return None

    def _resolve_photo_paths_batch(self, path_strings: set[str]) -> dict[str, Optional[Path]]:
        """Разрешает множество путей к фото за один проход (без повторных обращений к ФС).

        Для каждого уникального staff_image_path вызывается _resolve_photo_path_raw
        один раз; результат кэшируется в словаре.

        Args:
            path_strings: Множество строк путей (значения staff_image_path).

        Returns:
            Словарь: ключ — исходная строка пути, значение — Path или None.
        """
        return {p: self._resolve_photo_path_raw(p) for p in path_strings}

    def _archive_arcname(self, rec, staff_pin_short: str, used: set) -> str:
        """Формирует имя файла в архиве: <день>/<staff_pin>_<ISO-время>.ext.

        Время в локальном часовом поясе, двоеточия заменены на минус.
        При коллизии добавляется суффикс _1, _2, ...

        Args:
            rec: Запись LessonAttendance.
            staff_pin_short: PIN без обёртки S/T.
            used: Множество уже занятых имён в архиве (модифицируется).

        Returns:
            Строка пути внутри ZIP (например, 2026-03-18/24254_2026-03-18T14-30-00.jpg).
        """
        day_dir = rec.date_at.strftime("%Y-%m-%d") if rec.date_at else "unknown"
        if rec.first_in:
            local_dt = timezone.localtime(rec.first_in)
            time_part = local_dt.strftime("%Y-%m-%dT%H-%M-%S")
        else:
            time_part = "no-time"
        base = f"{staff_pin_short}_{time_part}"
        ext = Path(rec.staff_image_path or "").suffix or ".jpg"
        if not ext.startswith("."):
            ext = "." + ext
        arcname = f"{day_dir}/{base}{ext}"
        suffix = 0
        while arcname in used:
            suffix += 1
            arcname = f"{day_dir}/{base}_{suffix}{ext}"
        used.add(arcname)
        return arcname

    def _write_archive(self, archive_records: list, zip_path: Path) -> int:
        """Создаёт ZIP с фото: папки по дням (YYYY-MM-DD), имена — staff_pin_ISO-время.

        Пути к файлам разрешаются один раз по множеству уникальных staff_image_path,
        затем каждая запись использует готовый результат из словаря.

        Args:
            archive_records: Список кортежей (LessonAttendance, staff_pin_short).
            zip_path: Куда записать ZIP-файл.

        Returns:
            Количество добавленных в архив файлов.
        """
        unique_paths = {rec.staff_image_path for rec, _ in archive_records if rec.staff_image_path}
        resolved = self._resolve_photo_paths_batch(unique_paths)

        added = 0
        used_arcnames = set()
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for rec, staff_pin_short in archive_records:
                src = resolved.get(rec.staff_image_path) if rec.staff_image_path else None
                if not src or not src.is_file():
                    continue
                arcname = self._archive_arcname(rec, staff_pin_short, used_arcnames)
                try:
                    zf.write(os.fspath(src), arcname)
                    added += 1
                except (OSError, zipfile.LargeZipFile) as e:
                    self.stderr.write(self.style.WARNING(f"Skip photo {src}: {e}\n"))
        return added

    def _write_xlsx(self, rows, path: Path):
        """Пишет XLSX-отчёт с жирными заголовками (отдел, ФИО, ID, дата, время, локация)."""
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Suspicious"
        # Порядок: отдел → кто (ФИО, ID) → когда/где (дата, время, локация)
        headers = ["Отдел", "ФИО", "student_id / tutor_id", "Дата", "Время", "Локация"]
        header_font = Font(bold=True)
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
        keys = ["department", "fio", "student_id_tutor_id", "date", "time", "location"]
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, key in enumerate(keys, start=1):
                val = row.get(key, "")
                ws.cell(row=row_idx, column=col_idx, value=val)
        wb.save(path)
